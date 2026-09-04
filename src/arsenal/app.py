#!/usr/bin/env python3
# Copyright (C) 2024-2026 Василий Валерьевич Шадров
# Лицензия GPL-3.0-or-later

import csv
import json
import logging
import os
import platform
import re
import shutil
import subprocess
import sys
import tempfile
import textwrap
import threading
import time
import unicodedata
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import ClassVar

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rich.style import Style
from rich.text import Text
from scipy import stats
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import auc, roc_auc_score, roc_curve
from sklearn.preprocessing import StandardScaler
from textual import events, on
from textual.app import App, ComposeResult, active_app
from textual.binding import Binding
from textual.containers import Container, Horizontal, Vertical, VerticalScroll
from textual.css.query import NoMatches
from textual.message import Message
from textual.reactive import reactive
from textual.screen import ModalScreen, Screen
from textual.scrollbar import ScrollBar, ScrollBarRender
from textual.theme import Theme
from textual.widgets import (
    Button,
    Footer,
    Input,
    Label,
    ListItem,
    ListView,
    RadioButton,
    RadioSet,
    Static,
    TextArea,
)
import io
from contextlib import redirect_stdout
import plotext

# Настройка логгера для отладки
logger = logging.getLogger(__name__)

matplotlib.use('Agg')

# Виджет для добавления маркера в списки
class ListLabel(Static):
    # always_update=True гарантирует вызов watch метода в любой ситуации
    is_highlighted = reactive(False, always_update=True)

    def __init__(self, clean_text: str, **kwargs):
        # ОТКЛЮЧАЕМ РАЗМЕТКУ - это ключевое исправление!
        kwargs.setdefault('markup', False)
        super().__init__(**kwargs)
        self.clean_text = clean_text
        # Устанавливаем базовое состояние текста сразу при создании
        self.renderable = f"  {self.clean_text}"

    def watch_is_highlighted(self, value: bool) -> None:
        if value:
            self.update(f"█ {self.clean_text}")
        else:
            self.update(f"{self.clean_text}")  # Два пробела, чтобы текст не прыгал

# Задаем хранение и обработку данных
class DataManager:
    def __init__(self):
        # Определяем базовую папку (кроссплатформенно)
        self.base_dir = Path.home() / ".arsenal_data"
        self.db_path = self.base_dir / "arsenal_database.json"
        self.reports_dir = self.base_dir / "arsenal_reports"
        self.config_path = self.base_dir / "config.json"

        # Создаем структуру папок сразу при запуске
        self.base_dir.mkdir(exist_ok=True)
        self.reports_dir.mkdir(exist_ok=True)
        self.data = self.load_data()

    def normalize_filename(self, filename: str) -> str:
        """
        Нормализует имя файла для кроссплатформенной совместимости.
        Использует форму NFC (Canonical Composition) как универсальный стандарт.
        """
        # Приводим к NFC форме (более универсальная для обмена)
        normalized = unicodedata.normalize('NFC', filename)
        return normalized

    def normalize_patient_name(self, text: str) -> str:
        """
        Нормализует ФИО пациента: первая буква заглавная, остальные строчные,
        плюс нормализация Unicode.
        """
        if not text:
            return text

        # Нормализуем Unicode
        text = unicodedata.normalize('NFC', text.strip())

        # Разбиваем на слова, каждое слово с заглавной буквы
        words = text.split()
        normalized_words = []
        for word in words:
            if word:
                normalized_words.append(word[0].upper() + word[1:].lower())

        return ' '.join(normalized_words)

    def get_normalized_filename(self, last_name: str, first_name: str, patronymic: str, birth_year: str) -> str:
        """
        Формирует нормализованное имя файла из данных пациента.
        """
        # Собираем части имени
        parts = [last_name, first_name, patronymic, birth_year]
        # Фильтруем пустые
        non_empty = [p for p in parts if p]
        # Объединяем с пробелами
        base = " ".join(non_empty)
        # Заменяем пробелы на подчеркивания
        safe = base.replace(" ", "_")
        # Нормализуем Unicode
        normalized = self.normalize_filename(safe)
        return f"{normalized}.txt"

    def load_data(self):
        if self.db_path.exists():
            try:
                with open(self.db_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except (json.JSONDecodeError, OSError):
                return []
        return []

    def save_assessment(self, patient_info, assessment_details):
        last_name = patient_info.get('last_name', '')
        first_name = patient_info.get('first_name', '')
        patronymic = patient_info.get('patronymic', '')
        birth_year = patient_info.get('birth_year', '')

        # Нормализуем ФИО
        last_name = self.normalize_patient_name(last_name)
        first_name = self.normalize_patient_name(first_name)
        patronymic = self.normalize_patient_name(patronymic)

        fio = f"{last_name} {first_name} {patronymic}".strip()
        fiogr = f"{fio} {birth_year}".strip()

        # Поиск пациента по ВСЕМ полям (включая отчество)
        patient = next((p for p in self.data if
                    p.get('last_name', '') == last_name and
                    p.get('first_name', '') == first_name and
                    p.get('patronymic', '') == patronymic and
                    p.get('birth_year', '') == birth_year), None)

        if not patient:
            patient = {
                "uid": str(uuid.uuid4())[:8],
                "last_name": last_name,
                "first_name": first_name,
                "patronymic": patronymic,
                "birth_year": birth_year,
                "fiogr": fiogr,
                "assessments": []
            }
            self.data.append(patient)

        assessment_details["timestamp"] = uuid.uuid4().hex[:6]
        patient["assessments"].append(assessment_details)

        with open(self.db_path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=4)

        return patient

    # --- Методы для работы с исходами ---

    def get_patient_by_uid(self, uid: str) -> dict:
        """Получает пациента по UID"""
        for patient in self.data:
            if patient.get("uid") == uid:
                return patient
        return None

    def get_patient_by_fiogr(self, last_name: str, first_name: str,
                             patronymic: str = "", birth_year: str = "") -> dict:
        """Находит пациента по ФИО и году рождения (с нормализацией)"""
        last_name_norm = self.normalize_patient_name(last_name)
        first_name_norm = self.normalize_patient_name(first_name)
        patronymic_norm = self.normalize_patient_name(patronymic) if patronymic else ""

        for patient in self.data:
            if (patient.get("last_name", "") == last_name_norm and
                patient.get("first_name", "") == first_name_norm and
                patient.get("patronymic", "") == patronymic_norm and
                str(patient.get("birth_year", "")) == str(birth_year)):
                return patient
        return None

    def get_or_create_patient(self, last_name: str, first_name: str,
                             patronymic: str = "", birth_year: str = "") -> dict:
        """Получает существующего пациента или создает нового"""
        patient = self.get_patient_by_fiogr(last_name, first_name, patronymic, birth_year)

        if patient:
            return patient

        # Создаем нового пациента
        last_name_norm = self.normalize_patient_name(last_name)
        first_name_norm = self.normalize_patient_name(first_name)
        patronymic_norm = self.normalize_patient_name(patronymic) if patronymic else ""

        patient = {
            "uid": str(uuid.uuid4())[:8],
            "last_name": last_name_norm,
            "first_name": first_name_norm,
            "patronymic": patronymic_norm,
            "birth_year": int(birth_year) if birth_year else 0,
            "fiogr": f"{last_name_norm} {first_name_norm} {patronymic_norm} {birth_year}".strip(),
            "assessments": [],
            "outcomes": []
        }
        self.data.append(patient)
        self._save_data()
        return patient

    def add_outcome(self, patient_uid: str, outcome_type: str, date: str,
                   description: str = "", source: str = "ручной ввод") -> dict:
        """Добавляет исход пациенту"""
        patient = self.get_patient_by_uid(patient_uid)
        if not patient:
            raise ValueError(f"Пациент с UID {patient_uid} не найден")

        # Проверяем, нет ли уже такого исхода (по дате и типу)
        outcomes = patient.get("outcomes", [])
        for outcome in outcomes:
            if outcome.get("date") == date and outcome.get("type") == outcome_type:
                raise ValueError(f"Исход {outcome_type} от {date} уже существует")

        outcome = {
            "id": f"out_{uuid.uuid4().hex[:6]}",
            "type": outcome_type,
            "date": date,
            "description": description,
            "source": source,
            "created_at": datetime.now(timezone.utc).isoformat()
        }

        if "outcomes" not in patient:
            patient["outcomes"] = []
        patient["outcomes"].append(outcome)
        self._save_data()
        return outcome

    def get_outcomes(self, patient_uid: str) -> list:
        """Получает все исходы пациента"""
        patient = self.get_patient_by_uid(patient_uid)
        if not patient:
            return []
        return patient.get("outcomes", [])

    def get_all_outcomes(self) -> list:
        """Получает все исходы всех пациентов с информацией о пациенте"""
        all_outcomes = []
        for patient in self.data:
            for outcome in patient.get("outcomes", []):
                all_outcomes.append({
                    "patient": patient,
                    "outcome": outcome
                })
        return all_outcomes

    def delete_outcome(self, patient_uid: str, outcome_id: str) -> bool:
        """Удаляет исход по ID"""
        patient = self.get_patient_by_uid(patient_uid)
        if not patient:
            return False

        outcomes = patient.get("outcomes", [])
        initial_length = len(outcomes)
        patient["outcomes"] = [o for o in outcomes if o.get("id") != outcome_id]

        if len(patient["outcomes"]) < initial_length:
            self._save_data()
            return True
        return False

    def update_outcome(self, patient_uid: str, outcome_id: str,
                      outcome_type: str | None = None, date: str | None = None,
                      description: str | None = None) -> bool:
        """Обновляет исход по ID"""
        patient = self.get_patient_by_uid(patient_uid)
        if not patient:
            return False

        outcomes = patient.get("outcomes") or []
        for outcome in outcomes:
            if outcome.get("id") == outcome_id:
                if outcome_type is not None:
                    outcome["type"] = outcome_type
                if date is not None:
                    outcome["date"] = date
                if description is not None:
                    outcome["description"] = description
                self._save_data()
                return True
        return False

    def get_patients_without_outcomes(self) -> list:
        """Возвращает пациентов, у которых нет исходов"""
        return [p for p in self.data if not p.get("outcomes")]

    def get_patients_with_outcomes(self) -> list:
        """Возвращает пациентов, у которых есть исходы"""
        return [p for p in self.data if p.get("outcomes")]

    def _save_data(self) -> None:
        """Сохраняет данные в файл"""
        with open(self.db_path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=4)

    def load_saved_theme(self, default_theme: str, available_themes: dict) -> str:
        """Безопасно читает сохраненную тему из ~/.arsenal_data/config.json"""
        if self.config_path.exists():
            try:
                with open(self.config_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    saved_theme = data.get("theme", default_theme)
                    # Валидация: проверяем, что тема есть в списке доступных
                    return saved_theme if saved_theme in available_themes else default_theme
            except (OSError, json.JSONDecodeError):
                return default_theme
        return default_theme

    def save_theme(self, theme_name: str) -> None:
        """Записывает выбранную тему в ~/.arsenal_data/config.json"""
        try:
            with open(self.config_path, "w", encoding="utf-8") as f:
                json.dump({"theme": theme_name}, f, indent=4)
        except OSError:
            pass

    def add_opinion(self, patient_uid: str, expert: str, risk_score: int,
                    date: str, comment: str = "") -> dict:
        """Добавляет прогноз специалиста пациенту"""
        patient = self.get_patient_by_uid(patient_uid)
        if not patient:
            raise ValueError(f"Пациент с UID {patient_uid} не найден")

        if "opinions" not in patient:
            patient["opinions"] = []

        # Проверяем, нет ли уже такого прогноза (по дате и специалисту)
        for opinion in patient["opinions"]:
            if opinion.get("date") == date and opinion.get("expert") == expert:
                raise ValueError(f"Прогноз от {expert} от {date} уже существует")

        opinion = {
            "id": f"op_{uuid.uuid4().hex[:6]}",
            "expert": expert,
            "risk_score": risk_score,
            "date": date,
            "comment": comment,
            "created_at": datetime.now(timezone.utc).isoformat()
        }

        patient["opinions"].append(opinion)
        self._save_data()
        return opinion

    def get_opinions(self, patient_uid: str) -> list:
        """Получает все прогнозы пациента"""
        patient = self.get_patient_by_uid(patient_uid)
        if not patient:
            return []
        return patient.get("opinions", [])

    def delete_opinion(self, patient_uid: str, opinion_id: str) -> bool:
        """Удаляет прогноз по ID"""
        patient = self.get_patient_by_uid(patient_uid)
        if not patient:
            return False

        opinions = patient.get("opinions", [])
        initial_length = len(opinions)
        patient["opinions"] = [o for o in opinions if o.get("id") != opinion_id]

        if len(patient["opinions"]) < initial_length:
            self._save_data()
            return True
        return False

    @staticmethod
    def validate_date(date_str: str) -> tuple[bool, str]:
        """
        Проверяет корректность даты.
        Возвращает (True, "") если дата корректна,
        или (False, сообщение_об_ошибке) если нет.
        """
        if not date_str:
            return False, "Дата не указана"

        date_str = date_str.strip()

        # Проверяем формат
        if not re.match(r'^\d{4}\.\d{2}\.\d{2}$', date_str):
            return False, "Дата должна быть в формате ГГГГ.ММ.ДД"

        try:
            year, month, day = map(int, date_str.split('.'))

            # Проверяем диапазон года (от 1900 до текущего)
            current_year = datetime.now().year
            if year < 1900 or year > current_year:
                return False, f"Год должен быть между 1900 и {current_year}"

            # Проверяем месяц
            if month < 1 or month > 12:
                return False, "Месяц должен быть от 1 до 12"

            # Проверяем день с учетом месяца и високосного года
            days_in_month = [31, 29 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 28,
                             31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
            if day < 1 or day > days_in_month[month - 1]:
                return False, f"В {month} месяце не может быть {day} дней"

            # Проверяем, что дата не позже сегодняшней
            today = datetime.now().date()
            date_obj = datetime(year, month, day).date()
            if date_obj > today:
                return False, f"Дата {date_str} позже сегодняшней ({today.strftime('%Y.%m.%d')})"

            return True, ""

        except ValueError:
            return False, "Некорректный формат даты"
        except Exception as e:
            return False, f"Ошибка проверки даты: {e}"


# Темы

DEFAULT_THEME = "themeLT"

THEMES = {
    "themeLT": Theme(
        name="themeLT", # Стяжкин
        dark=False,
        primary="#008080",
        secondary="#666666", # для рамок и кнопок вне фокуса
        accent="#008080", # teal
        background="#FFFFFF",
        surface="#FFFFFF",
        panel="#DCDCDC",
    ),
    "themeLB": Theme(
        name="themeLB", # Голубев
        dark=False,
        primary="#0000FF",
        secondary="#666666",
        accent="#0000FF", # blue
        background="#FFFFFF",
        surface="#FFFFFF",
        panel="#b5b5b5",
    ),
    "themeLS": Theme(
        name="themeLS", # Тарасевич
        dark=False,
        primary="#4682B4",
        secondary="#666666",
        accent="#4682B4", # steelblue
        background="#FFFFFF",
        surface="#FFFFFF",
        panel="#DCDCDC",
    ),
    "themeLM": Theme(
        name="themeLM", # Чекунова
        dark=False,
        primary="#000000",
        secondary="#404040",
        accent="#000000",
        background="#DCDCDC",
        surface="#DCDCDC",
        panel="#b5b5b5",
    ),
    "themeDQ": Theme(
        name="themeDQ", # Макушев
        dark=True,
        primary="#D2B48C",
        secondary="#DCDCDC",
        accent="#D2B48C", # tan
        background="#2F4F4F",
        surface="#2F4F4F",
        panel="#2c2c2c",
    ),
    "themeDM": Theme(
        name="themeDM", # Шалек
        dark=True,
        primary="#DAA520",
        secondary="#DCDCDC",
        accent="#DAA520",
        background="#3b3b3b",
        surface="#3b3b3b",
        panel="#2c2c2c",
    ),
    "themeDC": Theme(
        name="themeDC", # Львов
        dark=True,
        primary="#FFFFFF",
        secondary="#FFFFFF",
        accent="#FFFFFF", # white
        background="#111111",
        surface="#111111",
        panel="#111111",
    ),
    "themeKL": Theme(
        name="themeKL", # Колесник
        dark=False,
        primary="#0363a2",
        secondary="#404040",
        accent="#0363a2",
        background="#DCDCDC",
        surface="#DCDCDC",
        panel="#b5b5b5",
    )
}

class SlimScrollBarRender(ScrollBarRender):
    def render_bar(
        self,
        size: int,
        virtual_size: int,
        window_size: int,
        position: float,
        vertical: bool,
        **kwargs
    ) -> Text:
        # Если контент влезает - рисуем пустоту
        if virtual_size <= window_size:
            return Text(" " * size)

        # 1. ГЕОМЕТРИЯ
        bar_size = max(1, int(size * window_size / virtual_size))
        max_scroll = max(1, virtual_size - window_size)
        max_track = size - bar_size
        start_pos = int(max_track * (position / max_scroll))

        # 2. БЕЗОПАСНОЕ ПОЛУЧЕНИЕ ЦВЕТА ИЗ ТЕМЫ
        theme_accent = "#D3D3D3"  # Дефолтный цвет (фолбек)

        try:
            # Получаем доступ к приложению через контекст потока
            app = active_app.get()
            # Проверяем, что у приложения уже инициализированы переменные темы
            if app and hasattr(app, "theme_variables") and app.theme_variables:
                theme_accent = app.theme_variables.get("accent", "#D3D3D3")
        except LookupError:
            # Сработает, если Textual считает полосу прокрутки до полной инициализации контекста
            pass

        # 3. ОТРИСОВКА
        style_bar = Style(color=theme_accent, bgcolor=None)
        style_bg = Style(bgcolor=None)

        char = "│" if vertical else "─"

        res = Text("", style=style_bg)
        res.append(" " * start_pos)
        res.append(char * bar_size, style=style_bar)
        res.append(" " * (size - start_pos - bar_size))

        return res[:size]

# Применяем КЛАСС
ScrollBar.renderer = SlimScrollBarRender


class CustomNotification(Static):
    """Свое уведомление"""

    class Dismiss(Message):
        pass

    def __init__(self, message: str, severity: str = "info"):
        super().__init__(message)
        self.severity = severity
        self.set_classes(f"notification {severity}")


class NotificationLayer(Vertical):
    """Слой для отображения уведомлений"""
    DEFAULT_LAYER = "notifications"  # Устанавливаем слой по умолчанию

class HoverButton(Button):
    """Кастомная кнопка с обработкой наведения мыши и перемещением фокуса"""

    def on_mouse_move(self, event: events.MouseMove) -> None:
        """При наведении мыши на кнопку"""
        # Находим родительский экран
        screen = self.app.screen
        # Обновляем подсказку, если у экрана есть метод _update_hint
        if hasattr(screen, '_update_hint'):
            screen._update_hint(self.id)
        # Устанавливаем фокус
        if screen.focused != self:
            self.focus()

def open_file_externally(filepath: str | Path) -> bool:
    """Открывает файл в системном приложении.
       Возвращает True в случае успеха, False при ошибке.
    """
    system = platform.system()
    fpath = str(Path(filepath).absolute())

    try:
        if system == "Windows":
            os.startfile(fpath)
        elif system == "Darwin":  # macOS
            subprocess.run(["open", fpath], check=True)
        else:  # Linux
            try:
                # Popen не блокирует поток интерфейса Textual
                subprocess.Popen(["libreoffice", "--writer", fpath])
            except FileNotFoundError:
                subprocess.run(["xdg-open", fpath], check=True)
        return True

    except (OSError, subprocess.SubprocessError) as e:
        logger.error(f"Не удалось открыть файл {fpath}: {e}")
        return False

class ConfirmDeleteDialog(Screen):
    """Модальное окно подтверждения удаления"""
    def __init__(self, filename: str):
        super().__init__()
        self.filename = filename

    def compose(self) -> ComposeResult:
        # Используем Grid для центрирования элементов внутри диалога
        with Vertical(id="dialog_confirm"):
            yield Label(f"Удалить файл пациента?\n\n{self.filename}", id="confirm_msg")
            with Horizontal(id="confirm_buttons"):
                yield Button("Да, удалить", variant="error", id="btn_yes")
                yield Button("Нет", variant="primary", id="btn_no")

    def on_mount(self) -> None:
        self.query_one("#btn_no").focus()

    def on_key(self, event: events.Key) -> None:
        if event.key == "left":
            self.focus_previous()
        elif event.key == "right":
            self.focus_next()

    @on(Button.Pressed, "#btn_yes")
    def confirm(self):
        self.dismiss(True) # Закрываем и возвращаем True

    @on(Button.Pressed, "#btn_no")
    def cancel(self):
        self.dismiss(False) # Закрываем и возвращаем False

# --- Окно выхода ---
class QuitScreen(ModalScreen[bool]):
    def compose(self) -> ComposeResult:
        with Vertical(id="dialog"):
            yield Label("Выйти из программы?", id="question")
            with Horizontal(id="buttons"):
                yield Button("Да", variant="error", id="yes")
                yield Button("Нет", variant="primary", id="no")

    def on_mount(self) -> None:
        self.query_one("#no").focus()

    def on_key(self, event: events.Key) -> None:
        if event.key == "left":
            self.focus_previous()
        elif event.key == "right":
            self.focus_next()

    def on_button_pressed(self, event: Button.Pressed) -> None:
        self.dismiss(event.button.id == "yes")

class ConfirmSaveDialog(ModalScreen[bool]):
    """Диалог подтверждения сохранения оценки"""
    def compose(self) -> ComposeResult:
        with Vertical(id="dialog"):
            yield Label("Завершить оценку и сохранить данные?", id="question")
            with Horizontal(id="buttons"):
                yield Button("Да", variant="primary", id="yes")
                yield Button("Нет", variant="error", id="no")

    def on_mount(self) -> None:
        # Фокус на кнопке "Да" по умолчанию
        self.query_one("#yes").focus()

    def on_key(self, event: events.Key) -> None:
        if event.key == "left":
            self.focus_previous()
        elif event.key == "right":
            self.focus_next()

    def on_button_pressed(self, event: Button.Pressed) -> None:
        self.dismiss(event.button.id == "yes")

class ConfirmTemplateDialog(ModalScreen[bool]):
    """Диалог подтверждения создания нового шаблона"""
    
    CSS = """
    ConfirmTemplateDialog {
        align: center middle;
        background: transparent;
    }

    ConfirmTemplateDialog #dialog {
        width: 70;
        height: auto;
        max-height: 80%;
        background: $background;
        border: round $accent;
        padding: 2 3;
    }

    ConfirmTemplateDialog #question {
        width: 100%;
        height: auto;
        content-align: center middle;
        color: $text;
        text-style: bold;
        margin-bottom: 1;
        text-wrap: wrap;
    }

    ConfirmTemplateDialog #buttons {
        width: 100%;
        height: 3;
        align: center middle;
        margin-top: 1;
    }

    ConfirmTemplateDialog Button {
        width: 18;
        height: 3;
        margin: 0 2;
        background: $background;
        color: $secondary;
        border: none;
    }

    ConfirmTemplateDialog Button:focus {
        border: round $accent;
        color: $accent;
        background: transparent;
        text-style: bold;
    }
    """

    BINDINGS = [
        Binding("escape", "cancel", "Отмена", show=False),
        Binding("left", "focus_previous", show=False),
        Binding("right", "focus_next", show=False),
    ]

    def compose(self) -> ComposeResult:
        with Vertical(id="dialog"):
            yield Label(
                "У вас в папке Документы уже есть шаблон для данных.\n\n"
                "Если вы внесли в него данные, но не импортировали их в программу, "
                "они будут потеряны.\n\n"
                "Создать новый шаблон сейчас?",
                id="question"
            )
            with Horizontal(id="buttons"):
                yield Button("Да", variant="primary", id="yes")
                yield Button("Нет", variant="error", id="no")

    def on_mount(self) -> None:
        self.query_one("#no").focus()

    def on_key(self, event: events.Key) -> None:
        if event.key == "left":
            self.focus_previous()
            event.stop()
        elif event.key == "right":
            self.focus_next()
            event.stop()

    def on_button_pressed(self, event: Button.Pressed) -> None:
        self.dismiss(event.button.id == "yes")

    def action_cancel(self) -> None:
        self.dismiss(False)

class MyRadioButton(RadioButton):
    """Радиокнопки с рамкой при фокусе"""

    def __init__(self, label: str, value: bool = False) -> None:
        super().__init__(label, value)
        self._checked_symbol = "◉"
        self._unchecked_symbol = "○"

    def render(self) -> str:
        """Кастомная отрисовка"""
        if self.value:
            return f"{self._checked_symbol} {self.label}"
        else:
            return f"{self._unchecked_symbol} {self.label}"

# --- Страницы руководства ---

PAGES = [
    {"part": "Общие сведения - назначение",   "text": "Методика Арсенал разработана для оценки динамики риска опасного поведения и совершения повторных общественно опасных деяний (далее - ООД) пациентами с психическими расстройствами, в первую очередь - проходящими принудительное лечение и госпитализированными в недобровольном порядке. Основным фокусом внимания методики является выявление мишеней для терапевтического воздействия и отслеживание динамики тех изменений, которых удалось добиться в ходе этого воздействия."},
    {"part": "Общие сведения - особенности",   "text": "Поскольку особое внимание уделяется именно отслеживанию изменений, в методику не включаются неизменяемые факторы риска, такие как возраст совершения первого ООД, наличие того или иного диагноза или неблагоприятные условия воспитания, которые имеют хорошо подтвержденную связь с высоким риском противоправного поведения и традиционно используются для оценки общего риска повторного совершения ООД. Методику не целесообразно применять для одномоментного выявления среди пациентов лиц с большим или меньшим риском повторения опасного поведения. Для этого следует учитывать наиболее значимые данные из истории пациента такие как тяжесть и повторность совершенных ООД, тяжесть и характер течения заболевания, данные о злоупотреблении ПАВ и другие. На учете и ранжировании таких факторов риска строятся также многие хорошо зарекомендовавшие себя формализованные инструменты оценки риска опасного поведения. При этом, такие инструменты дают мало информации о том, над чем специалисту следует работать для снижения риска. Например, факт совершения пациентом убийства всегда будет являться фактором повышенного риска совершения нового ООД или проявления агрессии при определенных обстоятельствах, однако сам по себе он не может стать мишенью терапевтического воздействия. Специалисту придется работать с симптомами, которые способствовали совершению убийства, работать над выработкой критического отношения пациента к своему заболеванию и совершенному, над пониманием необходимости лечения, но не над самим фактом из прошлого. Поэтому, методика Арсенал нацелена на то, чтобы из всех доступных для определения факторов риска сосредоточиться на тех, которые должны стать предметом работы специалистов психиатрического профиля и по которым у пациента может быть определен некоторый прогресс.\n\nТаким образом, для определения общего риска опасного поведения у пациента следует использовать анамнестический метод либо стандартизованные методики, учитывающие большее количество прогностически надежных факторов риска, в то время как изменения, зафиксированные с помощью методики Арсенал, отражают то, насколько пациент смог продвинуться в преодолении своего проблемного поведения и установок на определенном этапе лечения. Значительное снижение показателей риска при повторной оценке по этой методике говорит о существенной вовлеченности пациента в реабилитационный процесс, в то время как показатели общего риска по другим методикам могут снизиться не столь существенно, так как они могут опираться на неизменяемые факторы риска."},
    {"part": "Общие сведения - ограничения",    "text": "Методика Арсенал не применима в случаях, когда основным фактором опасного поведения является патология влечения, то есть в большинстве случаев, когда пациентом совершено сексуальное насилие. Риск совершения ООД сексуального характера следует определять специально для этого разработанными методиками."},
    {"part": "Общие сведения - структура", "text": "Методика Арсенал включает в себя первичную и повторную оценки по восьми факторам. По каждой из этих оценок составляется заключение.\n\nПервичная оценка предполагает определение значимости (представленности) каждого фактора риска по шкале от 0 до 3 в соответствии с руководством по оценке. Для каждого выявленного фактора (оценка 1 и выше) определяется степень и характер критического восприятия самим пациентом этого фактора, для оценки которого используются так называемые стадии изменения, предложенные Prochaska и DiClemente. Поскольку каждый выявленный фактор риска представляет собой определенную проблемную область в жизни пациента, на планомерные адаптивные изменения в этой сфере могут быть направлены усилия помогающих специалистов - психиатра, психотерапевта и психолога - таким образом выявленный фактор риска и субъективное его восприятие пациентом принимается в качестве терапевтической мишени. Заключение специалиста по результатам первичной оценки строится на описании индивидуальных особенностей, конкретных характеристик выявленных проблемных сфер, их обширности и того, воспринимает ли сам пациент эти сферы как проблемные и стремится ли к преодолению проблем.\n\nПовторная оценка проводится после завершения некоторого этапа терапевтического воздействия на пациента и фиксирует в первую очередь прогресс по выявленным ранее факторам риска в понятиях стадий изменения. Такой подход к оценке прогресса в ходе лечения успешно применен Wong и Gordon, разработавшими не только оригинальную методику оценки риска насилия, но и развитую программу по снижению риска. Заключение по результатам повторной оценки включает в себя с одной стороны анализ тех сфер, в которых достигнуты определенные результаты, а с другой - выявление причин, по которым в других сферах результата достичь не удалось. Более детальные описания методики первичной и повторной оценки, а также составления заключений содержатся в соответствующих разделах данного руководства."},
    {"part": "Общие сведения - цели", "text": "Методика Арсенал предполагает, что после проведения первичной оценки на основе вынесенного заключения пациенту будет проведено лечение, направленное на улучшение в тех проблемных сферах, которые были выявлены. Это лечение может включать как фармакотерапию и психотерапию, так и психокоррекционные и социальные вмешательства - в зависимости от показаний, а кроме того, специальные лечебно-реабилитационные программы, направленные на работу с опасными пациентами.\n\nОчевидно, что усилия специалистов должны быть в первую очередь направлены на те сферы, которые вносят наибольший вклад в поддержание высокого риска, однако менее проблемные области также не стоит оставлять без внимания. При этом, помогающие специалисты должны осторожно подходить к удовлетворению некриминогенных потребностей пациента, то есть направлению терапевтических усилий на дисфункции, не связанные с риском опасного поведения, так как на это может быть потрачено много усилий без значимой динамики снижения риска.\n\nБез реализации конкретных целенаправленных терапевтических мероприятий проведение повторной оценки пациента по методике Арсенал имеет меньше оснований, однако может быть использовано для отслеживания самопроизвольной динамики риска."},
    {"part": "Общие сведения - применение", "text": "Предполагается, что методика Арсенал чувствительна к изменениям повышенного и высокого риска опасного поведения у пациентов, в том числе проходящих принудительное лечение в психиатрических стационарах, то есть в тех ситуациях, когда несмотря на успехи в лечении пациента, другими методами риск по-прежнему определяется как высокий.\n\nЯвляясь инструментом структурированного профессионального суждения методика может использоваться для поддержки решений о продлении или изменении вида принудительного лечения, оценки эффективности тех или иных методов и этапов лечения относительно риска опасного поведения, но безусловно должна применяться в сочетании с другими методами оценки динамики состояния пациента. Кроме того, внедрение методики в практику работы полипрофессиональной бригады специалистов должно способствовать выработке общих терминов, целей и критериев этой работы, облегчая передачу наиболее значимой структурированной информации об опасном пациенте между специалистами."},
    {"part": "Факторы - информация", "text": "Каждый из восьми факторов, включенных в методику Арсенал, представляет собой информацию о пациенте, которая в значительной степени связана с риском опасного или криминального поведения, проявления агрессии, совершения пациентом новых ООД и нарушений лечебного режима. \n\nЭта информация должна собираться из по возможности большего числа источников, среди которых наибольшую ценность представляют документы, содержащие объективные данные о поведении и психическом состоянии пациента, такие как постановления суда, заключения экспертиз, выписки из медицинских карт. По каждому фактору следует провести также расспрос пациента, особое внимание при котором уделяется расхождению субъективной и объективной трактовки обсуждаемых событий из прошлого пациента, его способность принять на себя ответственность за происходящее и стремление преодолеть проблемные формы поведения."},
    {"part": "Факторы - подход", "text": "Из всего объема данных наибольшее значение следует придавать событиям недавнего прошлого, так как они в большей степени отражаются на актуальном состоянии пациента, а также выявленным стойким паттернам поведения, которые прослеживаются на протяжении всей жизни. Из событий далекого прошлого стоит обратить внимание на совершенные ООД, в том числе не получившие правовой оценки, начало и характер течения психического заболевания и события, отражающие грубо дезадаптивное поведение. Стоит также указать на те сферы жизни пациента, в которых ранее определялись существенные трудности, но к настоящему времени они в значительной степени преодолены. Это может дать информацию об используемых пациентом способах адаптации, которые можно будет развить и, возможно, распространить на другие проблемные сферы.\n\nОтдельно следует проанализировать ситуацию, когда у пациента выявляется малое количество, но чрезвычайно надежных факторов риска. К таким ситуациям можно отнести, например, упорные угрозы совершения ООД, стойкие бредовые идеи, направленные на конкретных лиц, сексуальное влечение к детям и другие криминогенные расстройства влечений. При этом, скорее всего, специалисту придется констатировать невозможность значительного снижения риска опасного поведения несмотря на возможный прогресс пациента по другим факторам риска в результате лечения."},
    {"part": "Факторы - в целом", "text": "В целом, специалист должен постараться определить, отмечались ли когда-либо у пациента проблемы по каждому из факторов, насколько они выражены и актуальны сейчас, насколько явно они связаны с криминальным и опасным поведением пациента и какими путями удавалось эти проблемы решить.\n\nПри анализе конкретных событий из прошлого, особенно обстоятельств совершенных ООД, рекомендуется проследить вклад каждого из трех компонентов, на которые указывает Кондратьев Ф. В. - личность, синдром, ситуация. Также, одним из критериев оценки степени представленности факторов риска может служить принцип, в соответствии с которым чем больше усилий потребовалось для того, чтобы прекратить какое-либо поведение, тем больше риск повторения такого поведения.\n\nДля раскрытия сущности опасного поведения может оказаться полезным рассматривать его как дезадаптивный (криминальный или рискованный) способ удовлетворения пациентом своих потребностей. В таком случае можно говорить о криминогенных потребностях, сохранение которых будет означать сохранение риска опасного поведения, а поиск адаптивного пути удовлетворения этих потребностей может составить основу терапевтического вмешательства."},
    {"part": "Факторы - оценка", "text": "При проведении первичной оценки представленность каждого фактора методики на основе собранных объективных данных и целенаправленной беседы с пациентом оценивается специалистом по шкале от 0 до 3. Соответствующий раздел данного руководства содержит более детальные описания принципов оценки каждого фактора.\n\nВ целом же, [bold $accent]оценка 0[/] означает отсутствие проблем и подлежащих исправлению форм поведения в указанной сфере, как и отсутствие видимой связи между опасным поведением и оцениваемым фактором.\n\n[bold $accent]Оценка 1[/] указывает на наличие не грубо выраженных проблемных поведенческих паттернов, недостаточную саморегуляцию в заявленной сфере, наличие данных о существенных проблемах в прошлом, которые, однако, в настоящее время не актуальны. В некоторых случаях оценка 1 может быть выставлена по общему принципу: \"Не так хорошо, как 0\".\n\n[bold $accent]Оценка 2[/] указывает на наличие вполне определенных проблем в заявленной сфере, явную неспособность пациента самостоятельно преодолеть эти проблемы и надолго изменить свое поведение, однако, при некоторых обстоятельствах такое поведение не повторяется, пациенту принципиально доступны иные, более адаптивные модели поведения. В некоторых случаях оценка 2 может быть выставлена по общему принципу: \"Не так плохо, как 3\".\n\n[bold $accent]Оценка 3[/] выставляется в случаях, когда по данному фактору определяется закрепившееся явно дезадаптивное поведение, которое не пересматривается пациентом даже при возникновении отчетливых негативных для него самого последствий, имеются указания на определенную связь проблем в заявленной сфере с опасным, противоправным поведением, проявлением агрессии или аутоагрессии, значительное влияние проблем в заявленной сфере на жизнь пациента в целом.\n\nОценка выраженности (представленности) факторов должна опираться исключительно на сведения о фактических событиях прошлого, а не на то, как сам пациент воспринимает эти события и принимает ли он за них на себя ответственность. Субъективная трактовка пациентом своего поведения, желание или нежелание его изменить рассматривается на следующем этапе первичной и при повторной оценке по методике Арсенал, на котором определяется так называемая стадия изменения по каждому выявленному фактору, то есть получившему оценку 1, 2 или 3."},
    {"part": "Стадии изменения - понятие", "text": "Стадии изменения - термин предложенный Prochaska и DiClemente в рамках разработанной ими Транстеоретической модели изменения, которая изначально предлагалась для оценки и описания динамики аддиктивного и других проблемных видов поведения, таких как курение, злоупотребление алкоголем, переедание, домашнее насилие. Центральной концепцией этой модели является утверждение, что любой человек в процессе изменения своего проблемного поведения проходит одни и те же стадии: предобдумывания, обдумывания, подготовки, действия и удержания.\n\nАвторы назвали эту модель транстеоретической для указания на то, что данные стадии остаются одинаковыми независимо от того, на каких теоретических основах описывается поведение человека и какие методы воздействия применяются. В дальнейшем многими специалистами эта модель была применена к самым разнообразным видам подлежащего изменению поведения, в том числе Wong и Gordon успешно реализовали ее принципы в Шкале риска насилия (VRS) для оценки изменения риска опасного поведения, а также в версии VRS:SO для оценки изменения риска совершения преступлений сексуальной направленности."},
    {"part": "Стадии изменения - в целом", "text": "Каждая стадия изменения представляет собой определенное отношение пациента к своему прошлому или нынешнему поведению, а также представление о том, нужно ли ему что-либо предпринимать в этом отношении в ближайшее время.\n\nСледует иметь в виду, что смена стадий может происходить с одной стороны независимо от того, проводится ли целенаправленная работа со специалистом, а с другой стороны - она может происходить в любом направлении, как на следующую стадию, так и на предыдущую (откат, срыв стадии, рецидив проблемного поведения). В какой-то степени многократные самопроизвольные переходы пациента между несколькими стадиями изменения в обоих направлениях являются обычной ситуацией как до его вовлечения в планомерную работу со специалистом, так и уже в ходе этой работы.\n\nТаким образом, целью терапевтического воздействия становится снижение частоты перехода на предыдущую стадию изменения, консолидация и развитие усилий пациента для установления более планомерного достижения желаемого поведения.\n\nДалее мы рассмотрим стадии изменения в том понимании, в котором они применяются в методике Арсенал."},
    {"part": "Стадия предобдумывания", "text": "Стадия предобдумывания (\"Не готов действовать\") характеризуется тем, что пациент не осознает своих проблем по существу и не проявляет стремления изменить свое поведение в ближайшем будущем. На этой стадии пациент может совершенно отрицать отмечавшееся ранее поведение или не признавать это поведение как неправильное, нежелательное, болезненное или создающее проблемы для него самого или окружающих. Если проблемы признаются, то ответственность за них перекладывается на других лиц или сложившиеся обстоятельства, возможные действия по их решению откладываются на необозримое будущее. Характерна ложная убежденность в том, что если нужно будет избежать проблемного поведения, то это легко можно будет сделать. Пациент на стадии предобдумывания склонен переоценивать аргументы против изменения своего поведения и недооценивать аргументы за, поэтому одним из возможных способов воздействия на пациента с целью вовлечь его в работу является выявление и подчеркивание негативных последствий проблемного поведения и обсуждение возможных позитивных изменений после отказа от такого поведения."},
    {"part": "Стадия обдумывания", "text": "Стадия обдумывания (\"Готовится действовать\") характеризуется главным образом тем, что пациент признает свое поведение как создающее проблемы, в какой-то степени принимает на себя ответственность за него и заявляет о своем желании изменить это поведение в обозримом будущем, хотя реальных действий в этом направлении не предпринимает. Пациент скорее соглашается с тем, что не всегда мог сам контролировать свое проблемное поведение, не мог вовремя остановиться, однако может не воспринимать это как однозначно негативный аспект. В целом характерно приблизительно равнозначное восприятие аргументов за и против активного изменения, стремление больше обсуждать желательное и нежелательное поведение, наблюдение и обсуждение того и другого поведения на примере других людей, может сформироваться некоторое представление о том, кем бы мог стать пациент, если бы смог изменить свое поведение и преодолеть связанные с ним трудности."},
    {"part": "Стадия подготовки", "text": "Стадия подготовки (\"Готов действовать\") отличается от предыдущей в первую очередь появлением объективно наблюдаемых изменений в проблемной сфере. К таким изменениям можно отнести первые практические шаги, которые по мнению пациента помогут привести к изменению проблемного поведения, вовлечение в лечебные или реабилитационные мероприятия и сообщества, нацеленные на преодоление проблемного поведения, обращение за помощью к специалисту или близким, попытки оценить свой первый, пусть и неудачный, опыт изменения. Проблемные сферы вполне ясно определяются пациентом и сформировано намерение действовать с целью преодоления проблем в этих сферах, хотя реальные достижения могут быть очень неустойчивы как по времени, в течение которого они наблюдаются, так и по набору ситуаций, в которых поведение изменилось либо остается проблемным. Пациент осознает, что для достижения желаемого поведения придется приложить значительные усилия и чем лучше он подготовится, тем вероятнее успех. Сомнения в успехе очень характерны. Можно сказать, что на этой стадии аргументы за изменение поведения начинают преобладать в представлении пациента над аргументами против, но исход этой стадии по большей части зависит от того, сможет ли пациент найти понимание окружающих и какую-либо внешнюю поддержку своим намерениям."},
    {"part": "Стадия действия", "text": "Стадия действия (\"Действует\") характеризуется тем, что пациент активно преобразует свое поведение, установки, отношения и окружающие обстоятельства с целью преодоления проблем в заявленной сфере. Поведение пациента изменилось нужным образом, в достаточной степени и в течение значительного времени, однако поддержание этого нового поведения требует постоянных активных усилий со стороны пациента и устойчивость этих изменений не распространяется или не была проверена в соответствующих ситуациях высокого риска (избегание этих ситуаций характерно и вызвано страхом рецидива). Пациент настроен обсуждать полученный им опыт со специалистом, близкими людьми или другими пациентами, накапливает собственный набор примеров ситуаций, в которых проявилось или не проявилось изменение его поведения или реакций. Срыву стадии может способствовать представление пациента о том, что он взял на себя больше обязательств, чем может выполнить. Для многих пациентов оказывается неожиданным то, насколько трудно удерживать желаемое поведение и насколько легко может произойти рецидив, таким образом основным вызовом этой стадии становится именно борьба за стабильность достигнутых результатов."},
    {"part": "Стадия удержания", "text": "Стадия удержания (\"Отслеживает\") - стадия на которой пациент длительное время удерживает желаемое поведение в заявленной сфере (относительно того времени, в течение которого ранее проявлялось проблемное поведение) в том числе в ситуациях высокого риска, при этом удержание этого поведения уже не требует постоянных усилий со стороны пациента, он не сомневается в том, что успешное и стабильное преодоление проблем достижимо и уже пользуется определенными положительными результатами своих достижений. Усилия пациента на этой стадии направлены на недопущение рецидива, для чего используются индивидуальные техники, выработанные на стадии действия, отслеживание обстоятельств, которые могут к нему привести. Характерно стабильное преобладание в представлении пациента аргументов за желаемое поведение с учетом полученного опыта - того, насколько трудно было преодолеть проблемное поведение и насколько важны результаты его преодоления."},
    {"part": "Стадии изменения - перспектива", "text": "Кроме этого, авторами Транстеоретической модели изменения выделена еще одна стадия - прекращения, на которой от пациента не требуется никаких усилий для сохранения желаемого поведения и проблемы в заявленной сфере можно считать полностью разрешенными. В методике Арсенал эта стадия не учитывается, так как лежит за границами терапевтического вмешательства. По этой же причине не устанавливается стадия изменения по тем факторам, по которым не выявлено проблем, то есть фактор риска оценивается как отсутствующий (оценка 0).\n\nВ целом, планомерное продвижение пациента от предыдущей стадии изменения к последующей принимается как индикатор улучшения в заявленной проблемной сфере, а соответственно и уменьшение участия этой сферы в формировании риска опасного поведения. Поэтому, эффективность проводимых с пациентом вмешательств, направленных на снижение риска опасного поведения, измеряется не простой оценкой по каждому фактору до и после вмешательства, а именно сопоставлением стадий изменения.\n\nОтдельные особенности и рекомендации по оценке стадий изменения применительно к конкретным факторам методики приводятся в следующем разделе данного руководства."},
    {"part": "Агрессия - определение", "text": "Данный фактор риска характеризует то, насколько часто пациент проявляет агрессию, в том числе вербальную, совершает насильственные, опасные действия и то, в каких обстоятельствах это происходит. Оценивать это следует в первую очередь за время, проведенное пациентом в обществе, когда он предоставлен сам себе. Совершение опасных действий в ограничивающих свободу условиях, например во время лечения в психиатрическом стационаре или отбывания наказания, должно рассматриваться более детально. Так, например, если начавшийся конфликт был быстро пресечен персоналом, его следует рассматривать как более серьезное проявление агрессии, чем если бы сам пациент смог прервать развитие конфликта без внешнего вмешательства. К агрессии может быть отнесено добровольное оставление пациентом своих детей. Аутоагрессивные действия и тенденции не учитываются, если только аутоагрессия не используется как инструмент грубого манипулятивного воздействия на окружающих."},
    {"part": "Агрессия - выраженность", "text": "[bold $accent]Оценка 0[/] ставится в том случае, когда пациент не прибегает к насилию в конфликтах, не проявляет агрессии к другим людям.\n\n[bold $accent]Оценка 1[/] может быть выставлена в том случае, если пациент легко вовлекается в конфликты, склонен браниться, угрожать; с недоверием и опаской относится к другим людям, ожидает обмана или агрессии от окружающих, однако при этом не проявляет тяжелых форм насилия с причинением физического вреда.\n\n[bold $accent]Оценка 2[/] должна быть установлена при наличии данных о совершении пациентом опасных действий, связанных с насилием, причинением физического вреда, частом не вынужденном вовлечении в конфликты, неспособности пациента выйти из уже начавшегося конфликта. В то же время, агрессивное поведение не носит систематического постоянного характера, провоцируется субъективно значимыми, \"психологически понятными\" событиями.\n\n[bold $accent]Оценка 3[/] ставится в тех случаях, когда пациент проявляет агрессию в тяжелых формах, без видимых причин, либо когда агрессивное поведение преобладает и становится своеобразным образом жизни. Может быть учтено систематическое проявление агрессии без причинения физического вреда. Сюда же следует относить совершение насильственных действий ради наживы, удовлетворения сексуальных потребностей, участие в организованных преступных сообществах."},
    {"part": "Агрессия - стадии", "text": "Стадия [bold $accent]предобдумывания[/] в отношении агрессии характеризуется отрицанием пациентом каких-либо проблем, связанных с его агрессивным, опасным поведением, отказ от принятия на себя ответственности за это поведение, убежденностью в его правомерности, адекватности или вынужденности, а также представлением о том, что он легко сможет избежать проявления агрессии в будущем, если это будет нужно. Пациент не видит необходимости что-то менять в этом отношении.\n\nСтадия [bold $accent]обдумывания[/] отличается тем, что пациент признает, что бывает агрессивным, ведет себя опасно или враждебно и это создает для него определенные проблемы. Пациент признает, что изменить это поведение было бы желательно, но не предпринимает и в обозримом будущем не планирует предпринимать каких-либо действий в этом отношении.\n\nНа стадии [bold $accent]подготовки[/] пациент предпринимает целенаправленные усилия для того, чтобы попытаться изменить свое агрессивное поведение, делится этим намерением с близкими или обращается за помощью к специалисту, обсуждает и обдумывает альтернативные способы разрешения конфликтных ситуаций, а также те выгоды, которые он мог бы извлечь, если бы изменил свое поведение. Появляется понимание того, что изменить свое поведение не так уж просто и для этого придется приложить определенные усилия. Могут отмечаться рецидивы агрессивного поведения.\n\nНа стадии [bold $accent]действия[/] пациенту уже удается значительное время избегать проявления агрессии, однако это требует от него значительных усилий, он сам понимает, что может столкнуться с такими ситуациями, в которых удержаться от вступления в конфликт будет чрезвычайно трудно. Способность пациента удержаться от опасного поведения не распространяется или еще не была проверена в ситуациях высокого риска, то есть в первую очередь в тех, в которых ранее он проявлял агрессию.\n\nСтадия [bold $accent]удержания[/] характеризуется устойчивым изменением поведения в плане сознательного отхода от проявлений агрессии, конфликтности. Способность пациента удержаться от опасного поведения была проверена в ситуациях высокого риска и не требует значительных целенаправленных усилий. При этом пациент осознает необходимость отслеживать определенные индивидуальные условия, которые позволяют ему сохранять желаемое поведение, и необходимость избегать обстоятельств, которые могут привести к рецидиву."},
    {"part": "Когнитивные и другие симптомы - определение", "text": "Данный фактор среди всех симптомов психического заболевания определяет те, которые либо напрямую обусловливают опасное поведение, либо препятствуют установлению продуктивного контакта с пациентом, что в свою очередь существенно осложняет работу, направленную на сознательную трансформацию поведения, ценностных установок и жизненных обстоятельств пациента. К таким симптомам относятся дисфории, нарушения мышления - непоследовательность, непродуктивность, разноплановость, аутичность, актуальные бредовые идеи, параноидная настроенность, резидуальные бредовые установки, врожденное и приобретенное слабоумие, значительные нарушения памяти."},
    {"part": "Когнитивные и другие симптомы - выраженность", "text": "[bold $accent]Оценка 0[/] ставится при отсутствии таких симптомов.\n\n[bold $accent]Оценка 1[/] может быть установлена в случае, когда указанные симптомы присутствуют, но не выражены значительно, либо были выражены, но в настоящее время скомпенсированы. Влияние наблюдаемых симптомов на опасное поведение пациента косвенное, не постоянное, эти симптомы не значительно затрудняют продуктивный контакт специалиста с пациентом.\n\n[bold $accent]Оценка 2[/] устанавливается при отчетливом влиянии указанных симптомов на опасное либо враждебное поведение пациента и/или на возможность установления продуктивного терапевтического контакта с ним. Симптомы определяются в настоящее время и значительно затрудняют вовлечение пациента в работу над его поведением, ухудшают его когнитивные и критические способности.\n\n[bold $accent]Оценка 3[/] должна выставляться в случае, когда указанные симптомы определяются в настоящее время и выражены настолько, что делают практически невозможным сознательный анализ пациентом своего поведения, установление контакта со специалистом, либо лишают его способности усваивать инструкции и соблюдать основные общественные нормы."},
    {"part": "Когнитивные и другие симптомы - стадии", "text": "Стадия [bold $accent]предобдумывания[/] применительно к симптомам психического заболевания характеризуется отсутствием понимания сущности имеющегося расстройства, его основных, самых тяжелых проявлений, способности оценить наличие либо отсутствие симптомов у себя в настоящее время. Если наличие симптомов признается пациентом, то значительно недооцениваются их тяжесть и влияние на поведение. Отсутствует стремление предпринимать какие-либо действия для того, чтобы снизить интенсивность имеющихся симптомов, узнать что-то о методах лечения. Преобладает представление, что болезненные явления - это случайность, они прекратятся сами собой и не будут повторяться, не способны повлиять на поведение и самообладание пациента.\n\nСтадия [bold $accent]обдумывания[/] характеризуется признанием наличия определенных проблем с психическим здоровьем, мешающих жизни симптомов. Проявлениями психического расстройства может объясняться или даже оправдываться необычное или агрессивное поведение в прошлом. Пациент также может признавать то, что наличие у него психического заболевания приносит в его жизнь некоторые проблемы, по крайней мере попадание в психиатрическую больницу. При этом, несмотря на заявляемое стремление избавиться от имеющихся симптомов, каких либо конкретных действий для этого не предпринимается. Характерна недооценка риска возможного ухудшения психического состояния при определенных обстоятельствах, стремление искать и обсуждать скорее психологические, нежели биологические причины заболевания, объяснять его внешними обстоятельствами.\n\nНа стадии [bold $accent]подготовки[/] у пациента формируется представление о болезненной природе наблюдаемых симптомов и стремление от них избавиться. Он готов обсуждать со специалистом возможные методы лечения, согласен лечиться, хотя очень вероятны несогласие со специалистом и сомнения относительно того, какие именно методы считать наиболее подходящими и применять в первую очередь. Представление о сущности и прогнозе заболевания могут быть весьма искаженными, но формируется убежденность в том, что симптомы заболевания мешают жить и их нужно лечить.\n\nСтадия [bold $accent]действия[/] характеризуется тем, что пациент фактически принимает назначенные ему лечебные мероприятия - лекарства, психотерапию, тренинги и прочее - с целью преодоления симптомов заболевания, стремится отслеживать и обсуждать со специалистом эффект этих мероприятий, пытается выработать свои собственные способы отслеживания наличия симптомов. Выраженность болезненных проявлений существенно снижена, однако говорить об устойчивости этих изменений преждевременно. Характерно избегание ситуаций, которые ранее провоцировали ухудшение состояния или тех, в которых предъявляются требования к \"слабым сторонам\" психики и способности пациента снижены. Возможен страх рецидива и вызванное им стремление скрыть ранние признаки ухудшения состояния.\n\nНа стадии [bold $accent]удержания[/] достигнут устойчивый и длительный контроль над указанными симптомами, пациент располагает достаточными когнитивными, прогностическими и критическими способностями для контроля своего поведения, преодоления трудностей, разрешения конфликтов."},
    {"part": "Контроль над эмоциями - определение", "text": "Данный фактор определяет способность пациента проявлять адекватные эмоциональные реакции, а также отслеживать, рефлексировать и совладать со своими эмоциями. В первую очередь эта способность определяется как длительно наблюдаемое свойство личности, а не как текущее проявление психического расстройства. Особое внимание следует уделить ситуациям, в которых неадекватность реакций или неспособность продуктивно и сознательно отреагировать свои эмоции приводила пациента к опасным действиям, необдуманным поступкам и другим негативным последствиям."},
    {"part": "Контроль над эмоциями - выраженность", "text": "[bold $accent]Оценка 0[/] ставится если пациент проявляет преимущественно адекватные по степени эмоциональные реакции, безопасными способами утилизирует негативные эмоции и не создает для себя и окружающих проблем, вызванных неприемлемым проявлением своих эмоций.\n\n[bold $accent]Оценка 1[/] характеризует пациента, который имеет определенные проблемы, связанные с управлением эмоциями, по крайней мере некоторые обстоятельства вызывают у него чрезмерную реакцию или он оказывается неспособен справиться с эмоциями при определенных условиях, например в состоянии опьянения. Однако, пациенту принципиально доступны приемлемые способы отреагирования и в некоторых случаях эмоциональные реакции можно считать адекватными.\n\n[bold $accent]Оценка 2[/] ставится в случаях, когда для пациента можно считать характерной неспособность контролировать и отслеживать свои эмоциональные реакции. Они часто чрезмерны, утрированы, не позволяют в должной мере управлять происходящим, затрудняют контакты с окружающими, мешают достигать поставленных целей.\n\n[bold $accent]Оценка 3[/] должна выставляться в случаях, когда пациент длительное время испытывает отчетливые трудности контроля проявления эмоций, его реакции часто чрезмерны, непродуктивны, мешают ему трезво оценивать ситуацию, приводят к необдуманным поступкам, опасным, агрессивным, шантажным действиям. Эмоции с трудом распознаются, не доступны приемлемые способы отреагирования негативных переживаний."},
    {"part": "Контроль над эмоциями - стадии", "text": "Стадия [bold $accent]предобдумывания[/] применительно к контролю над эмоциями характеризуется тем, что пациент не готов признать неадекватность своих реакций, наличие каких-либо проблем с управлением эмоциями. Пациент считает, что в имеющихся обстоятельствах реагировать иначе невозможно или было бы неправильно, и если бы он считал нужным не проявлять своих эмоций, то легко смог бы это сделать. Пациент не планирует предпринимать каких-либо действий для изменения своего стереотипа реагирования.\n\nСтадия [bold $accent]обдумывания[/] характеризуется тем, что несмотря на признание проблем, связанных с контролем над эмоциями, пациент не предпринимает каких-либо действий для изменения ситуации. Степень неадекватности эмоциональных реакций значительно преуменьшается или воспринимается как данность: \"Я такой, не надо меня злить\". Пациент может признавать, что не всегда контролирует свои эмоции и, пожалуй, было бы лучше иметь возможность их контролировать, однако как этого добиться неясно и можно это отложить.\n\nНа стадии [bold $accent]подготовки[/] у пациента формируется стремление преодолеть проблемы, связанные с недостаточным контролем над эмоциями, он оценивает по крайней мере некоторые свои реакции как неадекватные и ищет способы сглаживания этих реакций. Чаще всего определить эту стадию можно по тому, что пациент обращается за помощью к специалисту с прямым или косвенным запросом на преодоление этих проблем, однако стремление к изменениям могут проявляться и в других формах, например, - попытках найти соответствующие лекарства или альтернативные методы \"успокоения\", которые пациент определяет как возможно эффективные.\n\nСтадия [bold $accent]действия[/] определяется тем, что пациент активно применяет те способы развития контроля над эмоциями, которые ему удалось найти и освоить. Пациент настроен обсуждать ситуации, в которых ему удалось отследить свои эмоции и отсечь непродуктивные или неадекватные способы реагирования. В то же время, характерно осознание того, насколько трудно оказывается достижение желаемого результата и пациент может чувствовать себя скованным, ограниченным в свободе проявлений, что может выражаться в боязни проявлять какие-либо эмоции вообще. Характерен также постепенный, довольно длительный переход от того, что пациент в большей степени скрывает свои эмоции, к тому, что он действительно управляет своим реагированием и связанным с этим поведением. В любом случае, на стадии действия определяется значительное улучшение контроля над эмоциями, что, однако, наблюдается не столь длительное время или не было проверено в ситуациях высокого риска.\n\nСтадия [bold $accent]удержания[/] характеризуется длительным сохранением достаточного контроля над эмоциями, в том числе в ситуациях высокого риска. Это не требует от пациента постоянных усилий и он не чувствует себя ограниченным в проявлениях. При этом, пациент способен предугадать, какие обстоятельства несут в себе наибольший риск в плане эмоционального напряжения и старается их избегать."},
    {"part": "Контроль над поведением - определение", "text": "Этот фактор по принципам его оценки во многом схож с предыдущим и определяет способность пациента контролировать и оценивать свое поведение. Эта способность определяется как длительно наблюдаемое свойство личности и затрагивает эпизоды поведения, приводящего к проблемам и трудностям, при том, что это поведение не было напрямую вызвано симптомами заболевания, например бредом (что отражается фактором 2) или неадекватной эмоциональной реакцией (фактор 3). Рассматриваемое здесь проблемное поведение преимущественно связано с низкой способностью переживать фрустрацию, прогнозировать развитие ситуации, неразвитыми социальными навыками, низким сдерживающим влиянием высших эмоций и ценностей, своеобразием влечений, повышенной внушаемостью."},
    {"part": "Контроль над поведением - выраженность", "text": "[bold $accent]Оценка 0[/] ставится при отсутствии проблем, вызванных указанными нарушениями поведения и контроля над ним.\n\n[bold $accent]Оценка 1[/] выставляется, когда пациент испытывает значительные затруднения в том, чтобы проследить истинные мотивы своих действий, не всегда понимает, почему повел себя именно таким образом в конкретной ситуации. Эти особенности по крайней мере в некоторых случаях имели для пациента неблагоприятные последствия. Ему часто приходится извиняться за свое поведение; через какое-то время, особенно получив нежелательный результат, он понимает, что повел себя странно или неадекватно и что лучше было бы поступить иначе.\n\n[bold $accent]Оценка 2[/] выставляется в том случае, если у пациента длительно прослеживаются отчетливые нарушения поведения, он явно не способен планомерно придерживаться установленных правил, самостоятельно извлекать уроки из негативного опыта и корректировать свое поведение, вызывающее проблемы. Пациент не склонен задумываться, почему ведет себя именно так, хотя и не достигает желаемого результата.\n\n[bold $accent]Оценка 3[/] характеризует пациента практически не способного контролировать свое поведение, корректировать его в зависимости от обстоятельств, отслеживать и оценивать его. Поведение пациента явно приносит ему и окружающим значительные трудности, бывает агрессивным, конфликтным, нарушающим законы, правила и договоренности."},
    {"part": "Контроль над поведением - стадии", "text": "На стадии [bold $accent]предобдумывания[/] пациент не замечает каких-либо трудностей, вызванных его поведением, как и своей неспособности это поведение полноценно контролировать. Доводы окружающих о необходимости изменить свое поведение пациентом не рассматриваются или не воспринимаются всерьез, негативные последствия объясняются какими-то посторонними факторами, либо такое поведение считается вынужденным.\n\nСтадия [bold $accent]обдумывания[/] характеризуется признанием наличия определенных трудностей в том, чтобы контролировать свое поведение, как и того, что поведение пациента создает некоторые проблемы для него самого или окружающих. Пациент соглашается, что в некоторых ситуациях действует по наитию, не задумываясь о последствиях, иногда сам не может объяснить своих действий. При этом, каких планов по развитию способности контролировать свое поведение пациент не строит, считает, что это ненужно или невозможно.\n\nНа стадии [bold $accent]подготовки[/] пациент вполне четко осознает проблемы, вызванные его неспособностью контролировать и управлять своим поведением, может привести примеры такого рода ситуаций из своей жизни. У пациента сформировано намерение преодолеть эти проблемы и ему ясны по крайней мере первые шаги, которые можно сделать в этом направлении. Можно отметить реальные попытки пациента разобрать проблемные ситуации, отследить свои ошибки, подумать прежде, чем делать.\n\nСтадия [bold $accent]действия[/] характеризуется значительным вовлечением пациента в работу по развитию способности контролировать свое поведение и в этой работе можно зафиксировать определенные успехи. Пациент прилагает усилия для того, чтобы его поведение не создавало конфликтов или других нежелательных ситуаций, готов обсуждать и усваивать практические навыки самообладания. Если значительную роль в воздействии на поведение пациента играет фармакотерапия, важно определить, что пациент знает лекарства, которые принимает, и отслеживает их влияние на поведение.\n\nСтадия [bold $accent]удержания[/] определяется достижением достаточного и длительного контроля над поведением. Пациент не испытывает трудностей в осознании мотивов своих действий и их последствий, ему удается избегать проблем, связанных с неадекватным поведением даже в трудных ситуациях высокого риска."},
    {"part": "Злоупотребление веществами - определение", "text": "Данный фактор риска считается присутствующим, когда употребление алкоголя или других психоактивных веществ выходит из-под безусловного контроля со стороны пациента, приводит к трудностям и нежелательным последствиям или приобретает другие черты зависимости. Оценка данного фактора в то время, когда пациент находится в условиях, исключающих употребление, проводится на основании данных о прошлых эпизодах употребления и их последствий, отношения пациента к употреблению в прошлом и в будущем, а также на основе анализа замещающего поведения, например попыток получить эйфоризирующий или другой особый эффект от употребления доступных веществ или методов (крепкого чая или кофе, избранных лекарств, управляемой гипоксии мозга и т. п.) и другого поведения, имитирующего употребление."},
    {"part": "Злоупотребление веществами - выраженность", "text": "[bold $accent]Оценка 0[/] ставится в случае, когда пациент либо совсем не употребляет психоактивные вещества, либо это употребление касается только разрешенных веществ, объективно легко и полноценно им контролируется, не приводит к каким-либо трудностям и проблемам со здоровьем. Пациент без колебаний отказывается от употребления, если видит, что обстоятельства для этого не вполне подходящие.\n\n[bold $accent]Оценка 1[/] обозначает наличие признаков злоупотребления психоактивными веществами, снижение количественного и ситуационного контроля, однако употребление не приводило к существенным трудностям, таким как нарушение трудовой дисциплины, семейные проблемы, проблемы со здоровьем, совершение опасных действий.\n\n[bold $accent]Оценка 2[/] выставляется при выявлении существенных проблем, связанных с употреблением ПАВ, таких как нарушение дисциплины, потеря работы, семейные проблемы, проблемы со здоровьем, привлечение к административной ответственности, совершение не тяжких, не направленных против личности общественно опасных действий в состоянии опьянения.\n\n[bold $accent]Оценка 3[/] ставится при наличии подтвержденного диагноза синдрома зависимости или других психических и поведенческих расстройств вследствие употребления ПАВ, а также в случае выявления личностных изменений по зависимому типу, когда употребление во многом определяет образ жизни пациента и является частым источником его проблем в различных сферах жизни. Кроме того, оценка 3 выставляется в случаях, когда употребление ПАВ способствовало совершению пациентом общественно опасных деяний, направленных против личности (жизни, здоровья, половой неприкосновенности)."},
    {"part": "Злоупотребление веществами - стадии", "text": "Стадия [bold $accent]предобдумывания[/] определяется отрицанием злоупотребления и/или связанных с ним проблем, отсутствием стремления что-либо предпринимать в этом отношении. Характерна убежденность, что от употребления можно легко оказаться, если это будет нужно. Зависимость отрицается или значительно преуменьшается, употребление связывается с внешними обстоятельствами.\n\nНа стадии [bold $accent]обдумывания[/] употребление ПАВ определяется пациентом как создающее проблемы, возможно, им уже предпринимались попытки отказаться от употребления, но это вызвало значительные трудности. Наличие зависимости либо отрицается, либо не воспринимается как нечто опасное. Пациент скорее стремится избавиться от проблем, вызванных употреблением, чем от самого употребления. Необходимость отказаться от употребления признается, но постоянно откладывается, возможен расчет на внешние обстоятельства, которые заставят его это сделать.\n\nСтадия [bold $accent]подготовки[/] характеризуется убежденностью в необходимости отказаться от употребления ПАВ и наличием объективно наблюдаемых попыток это сделать. Зависимость признается и определяется как самостоятельная проблема. Обычно у пациента в какой-то степени сформирован образ себя свободного от употребления. Пациент охотно обсуждает возможные способы избавления от зависимости, готов принимать советы и помощь от других людей, однако употребляющее окружение легко может убедить его в отсутствии необходимости отказываться от употребления. Пациент обычно сам может спрогнозировать ситуацию, в которой ему будет сложно не сорваться, сомнения в возможности преодолеть зависимость очень характерны.\n\nСтадия [bold $accent]действия[/] определяется значительным успехом в отказе от употребления, который наблюдается значительное время, хотя не был проверен в ситуациях высокого риска. Пациент сам определяет это как существенное достижение, может отмечать снижение тяги к употреблению и другие положительные изменения, однако понимает, что рецидив возможен и для его предотвращения все еще требуются значительные усилия и, возможно, посторонняя помощь.\n\nСтадия [bold $accent]удержания[/] характеризуется способностью пациента не возвращаться к употреблению ПАВ даже в ситуациях высокого риска, значительным снижением тяги к употреблению. Воздержание не требует от пациента значительных постоянных усилий."},
    {"part": "Приверженность режиму и лечению - определение", "text": "Этот фактор касается готовности и способности пациента следовать рекомендациям специалиста, к которому он обратился за помощью, и приверженности своим собственным положительным установкам относительно желаемого образа жизни и поведения. В какой-то степени оценка данного фактора может строиться на общем анализе способности пациента находить способы решения своих проблем и следовать им, в том числе способности принимать помощь окружающих."},
    {"part": "Приверженность режиму и лечению - выраженность", "text": "[bold $accent]Оценка 0[/] обозначает устойчивую приверженность пациента назначениям и рекомендациям специалистов, реалистичную субъективную картину болезни, последовательное представление о том, какой стратегии следует придерживаться для преодоления своих проблем.\n\n[bold $accent]Оценка 1[/] выставляется в случае выявления у пациента значимых сомнений в необходимости следования назначениям и рекомендациям, деградации приверженности лечению после выпадения из-под контроля специалиста, стремления как можно скорее отказаться от приема лекарств после улучшения состояния. Характерна недооценка вероятности ухудшения состояния после самостоятельного видоизменения рекомендаций или отказа от их исполнения. В представлении о болезни и причинах проблемного поведения могут выявляться значительные противоречия.\n\n[bold $accent]Оценка 2[/] выставляется в случае, когда у пациента определяется явное недовольство необходимостью следовать рекомендациям специалиста, сомнение в их эффективности, преувеличение негативных последствий выполнения рекомендаций. Это может касаться как лекарственных назначений, так и рекомендаций по посещению специалиста, профилактических и диагностических мероприятий, участия в занятиях, тренингах, изменению образа жизни, отказу от употребления ПАВ и прочих. Пациент руководствуется собственными представлениями о том, что ему нужно делать, и стремится видоизменить назначения в соответствии с ними даже находясь под наблюдением специалистов.\n\n[bold $accent]Оценка 3[/] должна выставляться в случаях, когда пациент оказывает активное сопротивление выполнению назначений, открыто противодействует диагностическим и лечебным мероприятиям."},
    {"part": "Приверженность режиму и лечению - стадии", "text": "Стадия [bold $accent]предобдумывания[/] характеризуется значительной недооценкой необходимости какой-либо целенаправленной деятельности направленной на изменение своего поведения и решение проблем. Пациент отказывается признавать, что ухудшение его состояния или поведения в прошлом могли быть вызваны отказом от лечения, недобросовестным выполнением рекомендаций специалистов. Пациент в целом не настроен принимать помощь со стороны и придерживаться режима, способствующего стабилизации и улучшению его состояния.\n\nНа стадии [bold $accent]обдумывания[/] пациент осознает, что несоблюдение предписаний и советов специалистов, нарушение режима терапии его заболевания может привести или приводило ранее к ухудшению его состояния и вызывало определенные проблемы; понимает, что в значительной степени нуждается в помощи и должен быть способен ее принять. Такое понимание, однако, не отражается явным образом на поведении пациента и он не предпринимает конкретных действий по развитию своей приверженности рекомендациям.\n\nНа стадии [bold $accent]подготовки[/] отмечаются изменения в поведении пациента, которые можно посчитать попытками осознанно придерживаться рекомендаций специалистов в сфере медикаментозного лечения или режима. Пациент стремится разобраться в способах реализации этих рекомендаций, например, запоминает названия назначенных лекарств, их дозы, ожидаемый эффект, возможные нежелательные явления, способы получения, возможные альтернативы. В беседах со специалистами пытается обсудить, каким образом ему лучше всего встроить выполнение рекомендаций в текущий или предполагаемый в будущем образ жизни. Пациент стремится сохранить контакт со специалистом, настроен на длительное взаимодействие с ним. При этом, время, в течение которого пациент придерживается такой установки, не достаточно продолжительно, чтобы специалист и сам пациент могли убедиться в ее устойчивости и отследить положительные результаты.\n\nСтадия [bold $accent]действия[/] характеризуется установлением относительно длительного сознательного соблюдения пациентом данных ему рекомендаций по лечению и режиму. Пациент охотно контактирует со специалистами, понимает необходимость принимать лекарства и выполнять другие предписания, осознает возможные последствия отказа от лечения и нарушения режима. От общих представлений о полезном и вредном пациент переходит к анализу конкретных ситуаций в своей жизни, в которых соблюдение или нарушение режима играет ключевую роль. При этом, удержание такого поведения требует от пациента определенных усилий и еще не было проверено в ситуациях высокого риска.\n\nСтадия [bold $accent]удержания[/] определяется длительным сознательным соблюдением режима и выполнением рекомендаций специалистов, в том числе в ситуациях высокого риска, которое приобретает черты устоявшегося образа жизни и не требует постоянных усилий со стороны пациента."},
    {"part": "Личностные установки - определение", "text": "Этот фактор отражает устоявшиеся ценностные ориентиры пациента и особенности его личности. Среди множества возможных характеристик особое внимание следует уделить общей незрелости, дезадаптирующим чертам личности, криминальным установкам, отношению к ценности своих и чужих жизни, здоровья и благополучия, отношению к личным границам, принятым в обществе формальным и неформальным правилам поведения, представлению о собственной роли и роли окружающих в отношениях, семье, коллективе, обществе, высокому уровню недоверия, либо напротив повышенной зависимости и ведомости. Подлежать оценке также может повторение одних и тех же неблагоприятных жизненных циклов или стереотипов, как, например, повторение общественно опасных действий по одному и тому же сценарию, неспособность справиться с одними и теми же жизненными ситуациями с неблагоприятными последствиями."},
    {"part": "Личностные установки - выраженность", "text": "[bold $accent]Оценка 0[/] ставится при отсутствии отчетливых дезадаптирующих личностных установок, зрелом представлении о своем и общественном благополучии. В жизни пациента можно выделить длительный период, в течение которого он был самостоятельно адаптирован и не проявлял опасного или противоправного поведения. Прослеживается способность пациента извлекать уроки из собственного опыта.\n\n[bold $accent]Оценка 1[/] отражает наличие умеренно выраженного своеобразия ценностных ориентиров, которое может проявляться в стремлении снять с себя любую ответственность, неспособности поддерживать длительные доверительные отношения, неадекватной самооценке, недоверии окружающим, восприятии мира как враждебного либо ощущении исключенности себя из мира, лживости, стремлении к манипулированию.\n\n[bold $accent]Оценка 2[/] характеризует пациента с отчетливым дезадаптирующим личностным радикалом. Его ценности и представление о самом себе значительно затрудняют взаимодействие с окружающими, ухудшают социальное функционирование, например, приводят к частым конфликтам, неспособности удержаться в трудовом коллективе, создать семью. Пациент склонен нарушать договоренности, пренебрежительно относиться к установленным правилам и благополучию окружающих.\n\n[bold $accent]Оценка 3[/] должна указывать на грубо дезадаптирующие, прежде всего - антисоциальные установки, которые проявляются практически при любых обстоятельствах и являются постоянным источником проблем. Такие пациенты эгоцентричны, лживы, циничны, склонны оставаться невосприимчивыми к чувствам и переживаниям окружающих, хотя могут демонстрировать эмоциональную вовлеченность, если преследуют какие-либо цели при взаимодействии с конкретными людьми или группами. Эта же оценка должна выставляться при выявлении грубо нетерпимого отношения к каким-либо социальным, национальным и т. п. группам или отдельным лицам, особенно если это уже приводило к проявлению агрессии."},
    {"part": "Личностные установки - стадии", "text": "Стадия [bold $accent]предобдумывания[/] характеризуется тем, что пациент воспринимает свои дезадаптирующие установки как единственно верные или вынужденные, считает, что связанные с этим проблемы исходят не от него, а от других людей, которые не способны его понять и оценить. Отсутствует какое-либо представление о необходимости пересмотреть свои ценностные ориентиры и отношения, их возможное изменение воспринимается как потеря себя и своей уникальности. Пациент, как правило, не обращается за помощью и не понимает, зачем она может быть ему нужна.\n\nНа стадии [bold $accent]обдумывания[/] пациент признает, что его \"характер\" и ценностные ориентиры отличаются от \"обычных\" людей и что это доставляет ему определенные жизненные трудности. Это, однако, не приводит его к убеждению, что он должен как-то измениться, что ему необходима помощь специалистов, имеющиеся проблемы воспринимаются как неизбежные и часто недооцениваются.\n\nСтадия [bold $accent]подготовки[/] представляет значительные трудности при ее определении. В первую очередь это связано с тем, что пациенты с такого рода личностными установками, особенно при вынужденном взаимодействии со специалистом, как, например, при недобровольной госпитализации или принудительном лечении, склонны очень легко соглашаться на предложенные им психотерапевтические мероприятия, работу с психологом, а часто - и на лекарственную терапию, однако определить, насколько это согласие свидетельствует о действительном намерении пересмотреть свои установки, крайне сложно и обычно требует длительного наблюдения. Иными словами, оказавшись в вынужденной ситуации, когда его дальнейшая судьба зависит от того, как его состояние оценивается специалистами, пациент может демонстрировать согласие с любыми требованиями, которые к нему предъявляются, что может быть ошибочно принято за формирование критического отношения к дезадаптирующим личностным чертам. В действительности может оказаться, что пациент убедился не в необходимости изменить свои установки, а в необходимости их скрывать. Имея в виду эти обстоятельства, стадия подготовки должна устанавливаться в случаях, когда пациент заявляет о своей готовности уже сейчас работать со специалистом с целью преодоления проблем, вызванных своеобразием его ценностных ориентиров и черт личности, и при этом нет явных признаков того, что эта готовность исключительно показная.\n\nНа стадии [bold $accent]действия[/] пациент фактически вовлечен в назначенные ему коррекционные и терапевтические мероприятия, направленные на преодоление проблем, вызванных своеобразием его ценностных ориентиров и дезадаптирующими чертами личности, и уже можно определить некоторые положительные результаты этих мероприятий. При этом, как и на стадии подготовки, существует высокий риск того, что внутренние мотивы пациента остаются неизменными, сам материал психотерапевтической работы используется им для того, чтобы еще убедительнее демонстрировать переосмысление своих установок; пациент присваивает язык специалистов, точнее определяет, по каким критериям его оценивают, и тем самым лишь совершенствует свою способность вводить их в заблуждение, тогда как сами психопатические черты лишь усугубляются. В связи с этим, специалисту следует учитывать, что личность является устойчивой структурой и меняется крайне медленно, поэтому достигнутые на этой стадии изменения должны оцениваться крайне осторожно, желательно - с наблюдением за пациентом в различных условиях. Кроме того, очень высок риск срыва стадии, особенно при изменении условий, в которых находится пациент, он может полностью выпасть из-под влияния специалиста, нарушить договоренности и вернуться к прежнему образу жизни так же легко, как до этого вовлекся в терапию.\n\nСтадия [bold $accent]удержания[/] по данному фактору вряд ли может быть установлена во время нахождения пациента в стационаре, так как в таких условиях не выполняется критерий проверки достигнутых изменений в ситуациях высокого риска. Такой ситуацией должно считаться самостоятельное проживание. Если этот критерий удается выполнить, то стадия удержания устанавливается при стабильном изменении ценностных ориентиров пациента, существенном сглаживании дезадаптирующих черт личности, что уже не требует постоянных усилий со стороны пациента и специалиста. Сам пациент осознает произошедшие с ним изменения по сравнению с проблемным периодом, понимает, за счет чего они произошли и какие это имеет последствия, признает сложность и ценность этих изменений."},
    {"part": "Окружение, быт и планы - определение", "text": "Данный фактор в основном определяет способность пациента создать для себя приемлемые бытовые условия как в материальном плане, так и в плане построения отношений с теми людьми, которые его окружают, а также осложняющие адаптацию внешние по отношению к пациенту условия и объективные обстоятельства, в которых он находится или будет находиться после выписки из стационара. Если пациент извлечен из своего обычного окружения, например, находится в стационаре, большее внимание следует уделить оценке реалистичности его планов на будущее и оценки собственных бытовых возможностей и способностей. Важными аспектами для оценки являются наличие места жительства, источника средств к существованию, трудоспособности, а также особенности взаимоотношений со значимыми близкими и поддержки с их стороны."},
    {"part": "Окружение, быт и планы - выраженность", "text": "[bold $accent]Оценка 0[/] устанавливается при отсутствии значимых проблем, связанных с обеспечением собственного бытового благополучия, при нормальных поддерживающих отношениях с проживающими совместно членами семьи, реалистичных планах на будущее в отношении места проживания, доходов, продолжения лечения, понимании пациентом собственных ограничений.\n\n[bold $accent]Оценка 1[/] должна устанавливаться при наличии не грубо выраженных проблем в отношениях со значимыми близкими, чрезмерном полагании на помощь окружающих, трудностей в обеспечении базовых жизненных потребностей. Особое внимание следует уделить возможной связи данных аспектов с совершенными ранее опасными действиями, например, высокий риск вовлечения в криминальную компанию, хищения как способ пропитания и т. п. Сюда же могут быть отнесены созависимые и другие дезадаптивные отношения или, например, непонимание со стороны кого-то из близких пациента необходимости его лечения.\n\n[bold $accent]Оценка 2[/] обозначает явное неблагополучие в повседневной жизни, которое может выражаться, например, в необходимости проживания с конфликтными, криминальными или злоупотребляющими родственниками, неспособности пациента разорвать близкие грубо дисгармоничные или дезадаптирующие отношения, неспособности обеспечить себе приемлемые бытовые условия, рационально распоряжаться доходами. Эта же оценка выставляется при отсутствии у пациента опыта самостоятельной жизни, явной некомпетентности в бытовых вопросах, отсутствии социальных связей.\n\n[bold $accent]Оценка 3[/] должна выставляться в тех случаях, когда неблагополучие в повседневной жизни имеет явную связь с совершением пациентом общественно опасных деяний. К таким случаям можно отнести необходимость проживания с родственниками, в отношении которых пациент ранее проявлял агрессию, при сохранении конфликтных отношений с ними; бездомность, полное отсутствие средств к существованию либо другие тяжелые обстоятельства, в сравнении с которыми нахождение в больнице или местах лишения свободы является для пациента более благоприятным вариантом. Сюда же могут быть отнесены случаи, когда планы пациента грубо не соответствуют его реальным возможностям и способностям, предполагают отказ от необходимого лечения и наблюдения либо прямо указывают на возможность совершения новых общественно опасных деяний."},
    {"part": "Окружение, быт и планы - стадии", "text": "На стадии [bold $accent]предобдумывания[/] пациент игнорирует имеющиеся бытовые или материальные проблемы, не видит необходимости разрешать длительно существующие конфликты. Планы на дальнейшую жизнь расплывчаты, легковесны, лишены конкретики, не учитывают имеющиеся у пациента ограничения, либо пациент вовсе не строит планов на будущее или не готов делиться ими со специалистом.\n\nСтадия [bold $accent]обдумывания[/] характеризуется признанием пациентом имеющихся у него проблем в повседневной жизни, бытовых условиях или отношениях с совместно проживающими родственниками, однако пациент считает эти проблемы несущественными, либо уже разрешенными, либо напротив принципиально неразрешимыми. Он соглашается с тем, что его жизнь могла бы быть более благополучной, но не имеет представления о том, как этого добиться, не понимает, чем ему может помочь специалист.\n\nНа стадии [bold $accent]подготовки[/] пациент понимает необходимость и имеет сформированное намерение разрешить имеющиеся у него проблемы в повседневной жизни, охотно обсуждает их со специалистом, предпринимает первые шаги в выбранном направлении. Планы на будущее учитывают имеющиеся у пациента ограничения, он стремится заранее обдумать свои действия в случае неблагоприятного развития событий.\n\nСтадия [bold $accent]действия[/] обозначает планомерную и длительную вовлеченность пациента в работу по разрешению имеющихся у него проблем в повседневной жизни и отношениях с близкими, которая уже приносит положительные результаты. К таким результатам можно отнести восстановление контактов и доверительных отношений с родственниками, с которыми пациент проживает или планирует проживать; формирование навыков бытового самообслуживания и взаимодействия, использования денег и т. п. Планы на будущее реалистичны, детализированы, учитывают действительные способности и возможности пациента, необходимость продолжения лечения и наблюдения у специалиста, предполагают в том числе негативные сценарии развития событий.\n\nСтадия [bold $accent]удержания[/] не может быть установлена во время нахождения пациента в стационаре, т. к. при этом невозможно оценить устойчивость достигнутых изменений в ситуациях высокого риска, но может быть определена во время амбулаторного лечения, если пациент длительное время проживает самостоятельно или при поддержке родственников, справляется с удовлетворением своих базовых бытовых потребностей, готов к разрешению предполагаемых трудностей."},
    {"part": "Составление заключений - содержание", "text": "По результатам проведения первичной оценки по методике Арсенал составляется заключение, которое включает в себя данные о пациенте и специалисте, дату проведения, числовые оценки выраженности каждого из восьми факторов риска (от 0 до 3), общую сумму этих оценок (от 0 до 24), установленные стадии изменения по присутствующим факторам, графическое представление этих оценок (формируется программой), а также - комментарии к оценке по каждому фактору и заключение по оценке в целом.\n\nКак уже было указано, оценка выраженности фактора риска производится в первую очередь по объективным данным о прошлом пациента, его поведении и рассмотрению конкретных событий, а оценка стадий изменения основывается на нынешнем отношении пациента к этим событиям, его намерениях и наблюдении за тем, как поведение пациента на момент оценки отличается от наблюдавшегося ранее и приводящего к проблемам. При выставлении оценки 0 по какому либо фактору в комментарии приводится краткое обоснование такой оценки и стадия изменения не определяется."},
    {"part": "Составление заключений - комментарии", "text": "В комментариях по оценке каждого фактора в первую очередь указываются конкретные данные из истории пациента, данные, полученные при его непосредственном обследовании, и обстоятельства, которые послужили основанием для выставления той или иной оценки. Кроме того, комментарий должен содержать краткий обобщенный анализ затрагиваемой данным фактором проблемной сферы в жизни пациента, оценку потенциала изменений в этой сфере, возможные сценарии дальнейшего развития событий и, по возможности, предлагаемые специалистом методы терапевтического воздействия. Также в комментарий могут быть включены указания на невозможность убедительной оценки фактора, необходимость сбора дополнительной информации, сомнения специалиста в достоверности и полноте имеющейся информации или в возможности дать точную оценку в настоящий момент. Стоит уделить внимание тому, в течение какого времени наблюдалось проблемное поведение и какие обстоятельства влияли на его изменение."},
    {"part": "Составление заключений - сумма баллов", "text": "Общая сумма оценок по всем факторам, которая может принимать значение от 0 до 24, на текущем этапе разработки методики не имеет конкретного способа ранжирования, то есть еще не определено какое количество набранных баллов можно отнести к категориям низкого или высокого риска и какие именно выводы должны соответствовать полученному таким образом конкретному числу. Однако, есть основания считать, что высокий суммарный балл по первичной оценке предполагает наличие у пациента сразу нескольких проблемных жизненных сфер, неблагополучие в которых с одной стороны ведет к повышению риска совершения им общественно опасных деяний, а с другой - может быть преодолено или значительно скомпенсировано при успешных терапевтических вмешательствах.\n\nУпрощенно можно представить подход к оценке этого показателя так, что чем выше суммарный балл, тем выше риск совершения новых опасных деяний и тем больший объем лечебных и коррекционных мероприятий должен быть успешно реализован для снижения этого риска. При этом следует учитывать, что методика эксплуатирует небольшое количество отобранных для своих целей факторов риска и поэтому, при получении низкого суммарного балла, например 1 или 2, чтобы сделать вывод о действительно низком риске совершения пациентом опасных действий нужно убедиться также в отсутствии других значимых факторов риска, которые методика не затрагивает.\n\nДля определения пороговых значений и категорий риска, которые можно применить к вашей выборке пациентов, необходимо проведение ROC-анализа. Подробная информация об этом содержится в разделе руководства ROC-анализ."},
    {"part": "Составление заключений - заключение", "text": "В комментарии на последнем этапе первичной оценки, то есть в заключении по всей оценке в целом, специалисту следует обобщить обработанную в ходе оценки информацию о пациенте и указать на наиболее важные аспекты, наблюдения и \"находки\", касающиеся риска опасного поведения.\n\nВ некоторых случаях можно выделить один или несколько наиболее значимых факторов риска или проблемных особенностей пациента, которые играют главную роль в формировании опасного поведения или определяют неэффективность проводимых ранее мероприятий по снижению риска. В других случаях почти все факторы приходится считать присутствующими, но среди них можно выделить те, над которыми пациент уже способен работать, и те, которые им радикально отрицаются. Эта информация позволит определить первые этапы терапевтического вмешательства.\n\nВ заключении специалист может предложить возможные и подходящие в данном случае методы воздействия на пациента, попытаться определить перспективу этих воздействий, вероятные трудности при их реализации, общий настрой пациента на работу со специалистом и его понимание и признание необходимости каких-либо изменений.\n\nТакже может быть указано на недостаток какой-либо информации, необходимость прояснения отдельных аспектов или дальнейшего наблюдения для уточнения оценки."},
    {"part": "Составление заключений - выводы", "text": "Таким образом, первичная оценка по методике Арсенал в основном отвечает на вопрос, какую роль играют основные динамические факторы в общей структуре риска опасного поведения у данного пациента и на что в первую очередь стоит направить усилия специалистов для снижения этого риска. Отчасти, как и любая методика структурированного профессионального суждения методика Арсенал срабатывает на универсальном принципе - какое впечатление формируется у специалиста после рассмотрения конкретного случая по заданной методике суждения. Немалую ценность также представляет возможность обмена собранной по общим принципам информацией о пациенте в виде заключения между специалистами."},
    {"part": "Мишени", "text": "Поскольку методика использует только динамические факторы риска, каждый фактор, определенный как присутствующий, то есть тот, по которому выставлена оценка больше чем 0, может быть принят специалистом как возможная мишень для терапевтической проработки. Выбор методов вмешательства зависит как от самого содержания фактора риска, например, возможность воздействия медикаментозными или немедикаментозными методами, так и от возможностей специалиста - имеется ли установленный контакт с пациентом и запрос с его стороны на вмешательство, есть ли возможность собрать несколько пациентов для групповой работы, есть ли возможность привлечь к работе разных специалистов и так далее. В любом случае логика методики предполагает, что над выявленными проблемными сферами жизни пациента может быть проведена работа, эффективность которой может быть оценена при повторных оценках."},
    {"part": "Методы вмешательства", "text": "Сложно судить о том, какие методы терапевтического вмешательства будут наиболее эффективны в данном направлении работы и возможно ли каким-то одним методом добиться улучшения по всем выявленным мишеням. Скорее всего, лучших результатов удастся добиться сочетая разнородные методы - психофармакотерапию, психотерапию, образование, меры социальной поддержки.\n\nПредставляется возможным включение в психотерапевтическую или психокоррекционную программу обсуждения с самим пациентом представления о сущности выявленных факторов риска, возможных путей преодоления связанных с ними проблем, стадий изменения.\n\nСтоит отметить, что в рамках Транстеоретической модели изменения, описывающей указанные стадии, ее авторы сформулировали также так называемые \"10 процессов изменения\", представляющие собой действия, которые предлагаются ими для продвижения по стадиям. Среди них - получение необходимой информации, переоценка себя, переоценка обстоятельств, эмоциональное облегчение, помогающие отношения и другие. Целый ряд авторов присоединились к развитию этого направления и существенно расширили представление о возможных \"процессах изменения\"."},
    {"part": "Повторные оценки - подход", "text": "Независимо от выбранных методов работы с пациентом, если эта работа была целенаправленной и существенно затрагивала выявленные у пациента факторы риска, достигнутые изменения должны отразиться на представлениях пациента о его проблемах и путях их преодоления и, таким образом, могут быть представлены в виде продвижения по стадиям изменения. На этом принципе основываются повторные оценки по методике Арсенал. Поскольку оценка выраженности фактора риска строится на анализе событий прошлого пациента, нет содержательного смысла повторно оценивать эту выраженность через небольшой промежуток времени по тем же критериям, которые были применены при первичной оценке. Вместо этого динамику риска предлагается оценивать определив повторно только текущую (достигнутую) стадию изменения по тем факторам, которые при первичной оценке определены как присутствующие.\n\nКак уже было сказано, стадии изменения не обязательно последовательно сменяют друг друга в одном направлении. Вполне возможно зафиксировать длительную задержку на той же стадии по одним факторам при положительной динамике по другим, так же как и регресс (срыв, откат) стадии, особенно при изменении условий, в которых находится пациент."},
    {"part": "Повторные оценки - расчет", "text": "Выражение динамики риска в точных цифрах представляется весьма условным, однако на текущем этапе разработки методики предлагается следующий алгоритм. Достижение пациентом стадии удержания по какому-либо фактору уменьшает первичную оценку выраженности этого фактора на 2. Достижение стадии действия уменьшает первичную оценку на 1. Достижение стадии подготовки не позволяет провести переоценку выраженности фактора, но свидетельствует об успешном преодолении первичного сопротивления со стороны пациента и указывает на факторы, которые должны стать первоочередными мишенями для вмешательства, т. к. пациент к такому вмешательству уже готов. Переход от стадии предобдумывания к стадии обдумывания не предполагает каких-либо реальных действий со стороны пациента, поэтому также не приводит к переоценке выраженности фактора, и, вероятно, указывает на неэффективность принятых мер.\n\nСтоит заметить, что если первично по фактору установлена оценка 3, то снижение ее возможно только до 1, в остальных случаях возможно снижение до 0. Полученные скорректированные баллы по всем факторам суммируются и эта сумма условно сравнивается с суммой баллов первичной оценки. Если проводится несколько повторных оценок, то независимо от того, произошел ли срыв стадии или отмечено улучшение, первичная оценка корректируется в зависимости от вновь установленной стадии."},
    {"part": "Условия для оценок", "text": "Как правило, первичная оценка проводится однократно. Исключением могут стать случаи, когда за время взаимодействия с пациентом была получена дополнительная информация, которая заставляет специалиста пересмотреть оценки, установленные изначально. Повторных же оценок можно провести несколько, чтобы отследить изменения после отдельных этапов лечения. Может оказаться полезным проведение оценок одного пациента несколькими специалистами. Сам характер изменений, которые должна отслеживать методика, предполагает, что даже при самых интенсивных вмешательствах повторная оценка должна проводиться по крайней мере через несколько месяцев после первичной.\n\nПри каждой повторной оценке специалисту необходимо сначала ознакомиться с заключением по первичной оценке.\n\nВ том случае, если у пациента произошло значимое ухудшение по какому-либо из факторов, например совершено новое ООД или обнаружились новые проблемы по факторам, которые первично оценены как отсутствующие, то стоит еще раз провести первичную оценку и заново определить представленность всех факторов и стадий изменения по ним (для упрощения переноса данных найдите уже проведенную оценку в разделе \"[bold $accent]5. Все оценки[/]\", откройте ее, сохраните как черновик, а затем откройте черновик через меню \"[bold $accent]7. Работа с данными[/]\" или \"[bold $accent]2. Первичная оценка[/]\" и внесите изменившиеся данные).\n\nВ противном случае, когда такого значимого ухудшения нет, достаточно провести повторную оценку только стадий изменения по факторам, первично определенным как присутствующие, на основании данных о текущем состоянии и поведении пациента и его прицельного расспроса. Каждая такая оценка сопровождается комментарием, в котором следует указать, на основании каких наблюдений она установлена."},
    {"part": "Программа - структура заключения", "text": "В файле пациента автоматически формируется следующее представление оценок, данных по каждому фактору.\n\n(Если схема отображается некорректно, сделайте окно программы шире или уменьшите размер шрифта с помощью сочетания клавиш [bold $accent]Ctrl -[/] (на MacOS [bold $accent]Cmd -[/]) так, чтобы линия ниже умещалась на одной строке.)\n├──────────────────────────────────────────────────────────────────────────────┤\n\nПри первичной оценке:\n                                              [$accent]╭дата первичной оценки[/]\n                                        [$accent]╭─────┴────╮[/]\n     М О Д Р О П   А р с е н а л    ╭────2026.02.04───╮\nФакторы                             │ оцен  стад      │ [$accent]← заголовки[/]\n\\[1] Агрессия                        │ 2 ▓▓  ▁▂▄   под │\\[1] [$accent]← оценки по фактору[/]\n[$accent]╰┬╯╰───────────────┬───────────────╯ ╰──┬──╯╰───┬────╯ ╰┬╯[/]\n[$accent] ╰номер фактора    ╰название фактора    │       │       ╰номер фактора[/]\n[$accent]                                        │       ╰стадия изменения**[/]\n[$accent]                                        ╰оценка выраженности*[/]\n[$accent]                            ╭дата первичной оценки[/]\n\\[9] Всего из 24       [$accent]╭─────┴────╮[/]  ╭────┬────────────────────────╮\n                       2026.02.04   │ 14 │██████████████          │\n                                    ╰────┴────────────────────────╯\n[$accent]                                    ╰─┬─╯╰┬───────────────────────╯\n                                      │   ╰шкала представления суммы оценок\n                                      ╰сумма оценок по всем факторам[/]\n\nПри повторной оценке:\n[$accent]                    дата первичной оценки╮       дата повторной оценки╮\n                                        ╭┴─────────╮        ╭─────────┴╮[/]\n     М О Д Р О П   А р с е н а л    ╭────2025.01.28───╮╭─────2026.03.19─────╮\nФакторы                             │ оцен  стад      ││ стад      изм оцен │\n\\[1] Агрессия                        │ 3 ███ ▁     пре ││ ▁▂▄▆  дей +++ 2 ▓▓ │\\[1]\n[$accent]╰┬╯╰───────────────┬───────────────╯ ╰──┬──╯╰───┬────╯  ╰───┬────╯╰─┬─╯╰─┬─╯ ╰┬╯\n ╰номер фактора    ╰название фактора    │       │           │       │    │    │\n  первичная оценка выраженности фактора*╯       │           │       │    │    │\n         первичная стадия изменения по фактору**╯           │       │    │    │\n                         новая стадия изменения по фактору**╯       │    │    │\n   динамика стадии изменения между первичной и повторной оценками***╯    │    │\n                       новая (пересчитанная) оценка выраженности фактора*╯    │\n                                                                 номер фактора╯\n     сумма первичных оценок выраженности и ее шкала╮\n       дата первичной оценки╮       ╭──────────────┴──────────────╮[/]\n\\[9] Всего из 24       [$accent]╭─────┴────╮[/]  ╭────┬────────────────────────╮\n                       2025.01.28   │ 22 │██████████████████████  │\n                                    ├────┼────────────────────────┤\n                       2026.03.19   │ 16 │████████████████        │\n                      [$accent]╰─────┬────╯[/]  ╰────┴────────────────────────╯\n[$accent]       дата повторной оценки╯       ╰─────────────────────┬───────╯\nсумма новых (пересчитанных) оценок выраженности и ее шкала╯[/]\n\n[$accent]*[/] Оценка выраженности фактора может иметь следующие обозначения:\n    0 ▏   - оценка 0,\n    1 ▒   - оценка 1,\n    2 ▓▓  - оценка 2,\n    3 ███ - оценка 3.\n\n[$accent]**[/] Стадия изменения может иметь следующие обозначения:\n    ▁     пре - предобдумывание,\n    ▁▂    обд - обдумывание,\n    ▁▂▄   под - подготовка,\n    ▁▂▄▆  дей - действие,\n    ▁▂▄▆█ уде - удержание.\n\n[$accent]***[/] Представление динамики стадии изменения между первичной и повторной оценками может иметь следующие обозначения:\n    +++ - продвижение на три или четыре стадии,\n    ++  - продвижение на две стадии,\n    +   - продвижение на одну стадию,\n        - отсутствие динамики,\n    -   - откат на одну стадию,\n    --  - откат на две стадии,\n    --- - откат на три или четыре стадии."},
    {"part": "Программа - хранение данных", "text": "Программа для работы с методикой Арсенал сохраняет данные по завершению каждой проведенной оценки. Одни и те же данные - сведения о пациенте, специалисте, вид оценки и дата ее проведения, оценки выраженности факторов, стадии изменения, комментарии и заключение - сохраняются в двух видах: в файл пациента и в базу данных.\n\nФайлы пациентов доступны для просмотра через функцию \"[bold $accent]4. Файлы пациентов[/]\", так же доступно открытие этих файлов через внешний редактор для вывода на печать и их удаление. Файл пациента именуется по схеме Фамилия_Имя_Отчество_год рождения пациента.txt, все оценки, проведенные одному пациенту, сохраняются последовательно в один файл.\n\nЗаписи об оценках, сохраненные в базу данных, доступны для просмотра через функцию \"[bold $accent]5. Все оценки[/]\", так же доступно открытие этих записей через внешний редактор для вывода на печать и восстановление файла пациента по этим данным. В случае, если выбрана повторная оценка, то заключение по ней формируется с использованием данных о последней первичной оценке.\n\nПри импорте данных с компьютеров на других операционных системах могут возникать ошибки кодировки, при которых внешне одинаковые имена файлов пациентов в действительности различаются, так как используют разные символы. В таком случае, если вы, например, импортировали данные о первичной оценке, а затем провели повторную, данные об этой повторной оценке могут записаться в новый файл пациента, а не в тот, который был импортирован, то есть появятся два файла одного пациента с внешне одинаковыми названиями. Если вы столкнулись с такой проблемой, можете удалить оба эти файла, затем в меню \"[bold $accent]5. Все оценки[/]\" последовательно открыть оценки этого пациента и воспользоваться действием \"[bold $accent]2. Записать в файл пациента[/]\" для внесения данных в новый файл, который будет создан в вашей операционной системе и объединит все данные. То же самое можно сделать с данными об исходах и прогнозах специалистов через меню \"[bold $accent]1. Просмотр исходов/прогнозов[/]\" действие \"[bold $accent]3. Записать в файл пациента[/]\".\n\nЕсли вы импортировали данные об оценках, проведенных в ранних исполнениях первой версии программы, то полный текст заключения по таким оценкам с комментариями по каждому фактору может быть доступен только в файле пациента, который именуется по схеме ФамилияИО1990.txt. При этом в базу данных из всех комментариев сохраняется только заключение по оценке в целом. На основе таких первичных оценок возможно проведение повторной оценки, но комментарии по факторам из первичной оценки будут недоступны. Полная функциональная интеграция данных из ранних версий программы в текущую возможна путем правки файла Журнал.txt старой версии. Если необходимо это сделать, обратитесь к разработчику."},
    {"part": "Программа - работа с данными", "text": "Пользователю предоставляется возможность управления данными об оценках, внесенными в файлы пациентов, в то время как изменение данных, сохраненных в базу данных, не рекомендуется и напрямую из программы не доступно. База данных, таким образом, представляет собой неизменяемый журнал всех оценок, проведенных на данном компьютере и импортированных с других компьютеров.\n\n[bold $accent]ВНИМАНИЕ! Не открывайте одновременно несколько окон программы на одном компьютере. Это может привести к некорректной записи и потере данных.[/]\n\nВсе проведенные оценки автоматически последовательно записываются в файл пациента. При необходимости в этот файл можно внести правки открыв его во внешнем редакторе (нажмите \"[bold $accent]е[/]\") - удалить продублированные при импорте данные, внести дополнительные заметки и т. д. Эти правки не затронут сведения об оценке, внесенные в базу данных.\n\nВ настройках вашей операционной системы вы можете установить, какой именно редактор будет использоваться, определив программу по умолчанию для открытия файлов с расширением .txt. Учитывайте, что для нормального отображения файлов с оценками необходимо применять моноширинный шрифт.\n\nЕсли необходимо внести правки не только в файл пациента, но и в саму запись об оценке, сохраненную в базе данных, например, если пересмотрены оценки факторов или стадии изменения, то для этого в меню \"[bold $accent]5. Все оценки[/]\" нужно открыть оценку, подлежащую правке, и воспользоваться действием \"[bold $accent]3. Сохранить как черновик[/]\". Затем - открыть этот черновик через меню \"[bold $accent]7. Работа с данными[/]\" и, перемещаясь по шагам оценки кнопками \"назад\" и \"вперед\", внести правки, после чего перейти к последнему шагу и завершить оценку. Запись об этой оценке будет с новой датой сохранена в базу данных вместе с предыдущей и автоматически внесена в файл пациента.\n\nВ случае, если в файл пациента внесены ошибочные или избыточные сведения, одна и та же оценка продублирована несколько раз, то этот файл можно удалить, а затем в меню \"[bold $accent]5. Все оценки[/]\" последовательно открыть необходимые оценки и воспользоваться действием \"[bold $accent]2. Записать в файл пациента[/]\" для внесения верных данных в новый файл, который будет создан автоматически.\n\nВо время проведения оценки уже внесенные данные можно сохранить в виде черновика по кнопке \"[bold $accent]F12 Сохранить черновик[/]\", после чего продолжить или прервать проведение оценки. Сохраненные черновики доступны для продолжения работы в меню \"[bold $accent]7. Работа с данными[/]\" либо на первом шаге первичной или повторной оценки по кнопке \"[bold $accent]F10 Открыть черновик[/]\".\n\nЧерновики хранятся локально на том компьютере, на котором были созданы, и не передаются при экспорте данных."},
    {"part": "ROC-анализ", "text": "ROC-анализ (Receiver Operating Characteristic) — это статистический метод, позволяющий оценить прогностическую способность диагностического или прогностического инструмента. Применительно к методике Арсенал ROC-анализ сопоставляет результаты проведенных оценок с тем, было ли отмечено у данного пациента опасное поведение в последующие 6 месяцев.\n\nГрафически ROC-анализ представляет собой кривую в координатах, где по горизонтальной оси откладывается доля ложноположительных результатов (1 — специфичность), а по вертикальной — доля истинноположительных результатов (чувствительность). Чем ближе кривая к левому верхнему углу графика, тем выше прогностическая способность инструмента.\n\nТаким образом, ROC-анализ позволяет ответить на вопрос, насколько хорошо оценки по методике Арсенал, предсказывают реальное опасное поведение пациентов в будущем?\n\nВ программе для этого сопоставляются по отдельности три показателя:\n\n  [bold $accent]1[/]. Сумма баллов по каждой оценке (Оценки)\n\n  [bold $accent]2[/]. Условно исчисленная сумма стадий (Стадии) — для этих целей определенная стадия изменения по каждому фактору пересчитывается в баллы: стадия предобдумывания - 4, обдумывания - 3, подготовки - 2, действия - 1, удержания - 0.\n\n  [bold $accent]3[/]. Сумма первых двух показателей (Сумма) - пытаемся одновременно учесть оценки выраженности факторов и стадии изменения.\n\nДополнительно проводится логистическая регрессия для построения комбинированного предиктора.\n\nДля проведения ROC-анализа вам необходимо:\n\n  [bold $accent]1[/]. Иметь достаточное количество оценок. Минимальный объем выборки для надежного анализа зависит от многих показателей. Для простоты можно считать \"Чем больше, тем надежнее\", минимум 30 оценок для получения ориентировочных результатов, а лучше - больше 100.\n\n  [bold $accent]2[/]. Внести данные об опасных проявлениях пациентов (исходах) через соответствующий раздел. Исход считается положительным, если он наступил в течение 6 месяцев после проведения оценки. Для ROC-анализа необходимо достаточное количество как положительных, так и отрицательных исходов. Имейте в виду, что программа в данной версии не отслеживает, прошло ли 6 месяцев с даты оценки к моменту проведения ROC-анализа, таким образом, если оценка проведена месяц назад и за это время не наступило исхода, он будет считаться отрицательным, хотя время отслеживания составило всего один месяц вместо необходимых шести. Качество анализа напрямую зависит от полноты и достоверности данных об исходах.\n\nВ данной версии программы за опасные проявления пациентов предлагается считать совершенные ими общественно опасные деяния и необходимость применения к ним мер физического ограничения, что более применимо к пациентам, находящимся в психиатрическом стационаре. Если к вашим пациентам применимы другие критерии опасного поведения, вносите данные об исходах используя эти критерии.\n\nПри проведении ROC-анализа программа строит ROC-кривые для каждого из трех показателей, рассчитывает AUC (Area Under the Curve - площадь под кривой) с 95% доверительными интервалами, определяет оптимальные пороговые значения по критерию Юдена, строит модель логистической регрессии и выводит формулу расчета риска.\n\n[bold $accent]Интерпретация результатов\n\nAUC[/] (площадь под ROC-кривой) - это основной показатель прогностической способности:\n\n  AUC = 0.5 - прогноз не лучше случайного угадывания\n\n  AUC = 0.7–0.8 - удовлетворительная прогностическая способность\n\n  AUC = 0.8–0.9 - хорошая прогностическая способность\n\n  AUC > 0.9 - отличная прогностическая способность\n\nДоверительные интервалы показывают диапазон, в котором с 95% вероятностью находится истинное значение AUC.\n\n[bold $accent]Оптимальный порог[/] - это значение показателя, при котором достигается наилучшее соотношение чувствительности и специфичности (максимальный критерий Юдена). Этот порог можно использовать для разделения пациентов в вашей выборке на две группы - высокого и низкого риска.\n\nВ отчете по ROC-анализу найдите [bold $accent]значение оптимального порога по оценкам╮[/]\n-----------------------------------------------------------------------[bold $accent]│[/]--------\nОптимальные пороги (критерий Юдена) [bold $accent]╭──────────────────────────────────╯[/]                             \n------------------------------[bold $accent]╭──╮[/]--[bold $accent]│[/]--------------------------------[bold $accent]╭────────╮[/]-\nОценки                │ Порог:[bold $accent]│[/]17[bold $accent]├──╯[/]увств.: 0.667 │ Специф.: 0.909 │[bold $accent]│[/]J: 0.576[bold $accent]├╮[/]\nСтадии                │ Порог:[bold $accent]╰──╯[/]│ Чувств.: 0.833 │ Специф.: 0.515 │[bold $accent]╰────────╯│[/]\nСумма                 │ Порог: 40 │ Чувств.: 0.833 │ Специф.: 0.848 │ J: 0.682 [bold $accent]│\n              ╭────────────────────────────────────────────────────────────────╯\nКритерий Юдена╯[/] показывает, насколько эффективно методика разделяет пациентов на группы высокого и низкого риска с учетом ложных срабатываний:\n\n  J выше 0.7: Отличный тест. Методика очень четко разделяет пациентов на группы\n\n  J от 0.5 до 0.7: Хороший тест. Ошибки есть, но методика имеет высокую разделительную точность\n\n  J от 0.3 до 0.5: Удовлетворительный тест. Разделительная способность средняя\n\n  J ниже 0.3: Слабый тест. Ошибок слишком много, методика не позволяет достоверно говорить о высоком или низком риске\n\nПолученные пороговые значения являются специфичными для вашей выборки пациентов и могут отличаться от таковых в других учреждениях. Данные об оптимальных порогах на больших разнородных выборках еще не получены.\n\nТехнически, эти показатели позволяют вам сказать, что на основе данных о ваших пациентах при оценке по методике Арсенал набравшие баллов больше указанного порога с высокой вероятностью проявят опасное поведение в ближайшие полгода после оценки, и наоборот, набравшие меньше указанного порога с высокой вероятностью не проявят опасного поведения. Точность этого предсказания тем выше, чем ближе критерий Юдена J к единице.\n\nПрограмма позволяет вам сравнить полученные оценки риска по методике Арсенал с прогнозами специалистов, участвующих в оказании помощи этим пациентам. Для этого специалисту предлагается по своему усмотрению оценить риск проявления пациентом опасного поведения в ближайшие полгода по шкале от 1 до 10, где 1 - минимальный риск, а 10 - максимальный. Эти данные вносятся в программу в соответствующем разделе с указанием даты и в том случае, если таких прогнозов достаточное количество, их прогностическая точность (ROC-кривая, AUC, оптимальный порог, критерий Юдена и др.) будет рассчитана программой наряду с оценками по методике Арсенал, чтобы их можно было сопоставить.\n\nФормула расчета риска\n\nНа основе логистической регрессии программа выводит формулу для расчета индивидуального риска:\n\nZ = β₁ × Оценки + β₂ × Стадии + β₃ × Сумма + β₀\n\nгде β₁, β₂, β₃ - коэффициенты, β₀ - константа (intercept)\n\nРиск (P) в процентах рассчитывается по формуле:\n\nP = 1 / (1 + e^(-Z)) × 100%\n\nКатегории риска установлены простым делением на пять секторов:\n\n< 20% - Низкий риск\n\n20–40% - Средний риск\n\n40–60% - Повышенный риск\n\n60–80% - Высокий риск\n\n> 80% - Критический риск\n\nПрограммой формируется таблица, в которой для каждой оценки указываются ее результаты (оценки, стадии и их сумма), рассчитанный риск, категория риска и знаком ! отмечается, был ли в течение следующих шести месяцев зафиксирован исход (опасные проявления).\n\nТакая же таблица формируется по данным о прогнозах специалистов, если они есть."},
    {"part": "Об авторе", "text": "Методика Арсенал разработана в 2024 году в ФКУ \"Санкт-Петербургская ПБСТИН\" Минздрава России Шадровым В. В.\n\nПо всем вопросам, касающимся проведения оценок по методике Арсенал, настройки и работы с программой методики, приветствуется обращение к автору по адресам электронной почты shadrov@pbstin.ru, shadrovv@gmail.com.\n\n[bold $accent]Автор выражает глубочайшую благодарность всем специалистам, оказавшим содействие в создании и развитии методики Арсенал и программы для нее и не может не перечислить следующих людей с особой признательностью.[/]\n\n[bold $accent]Александр Николаевич Колесник [/](Санкт-Петербургская ПБСТИН)\n[bold $accent]Анастасия Александровна Ульянич [/](Санкт-Петербургский государственный университет)\n[bold $accent]Анна Сергеевна Шадрова  [/](Санкт-Петербургская ПБСТИН)\n[bold $accent]Гаянэ Аршалуисовна Вартанян [/](Санкт-Петербургский государственный университет)\n[bold $accent]Иван Станиславович Григорьев [/](ПБ Святого Николая Чудотворца)\n[bold $accent]Игорь Иванович Чижиков [/](Санкт-Петербургская ПБСТИН)\n[bold $accent]Лидия Николаевна Казакова [/](Санкт-Петербургский государственный университет)\n[bold $accent]Талия Станиславовна Богомолова [/](Санкт-Петербургская ПБСТИН)\n[bold $accent]Audrey Gordon [/](University of Saskatchewan)\n[bold $accent]Carlo C. DiClemente [/](University of Rhode Island, UMBC)\n[bold $accent]James O. Prochaska [/](University of Rhode Island)\n[bold $accent]Liang Wenfeng [/](DeepSeek AI)\n[bold $accent]Stephen Wong [/](University of Nottingham, University of Saskatchewan)\n[bold $accent]Will McGugan [/](Textualize)"},
]

# Словарь для маппинга частей руководства к шагам оценки
GUIDE_MAPPING = {
    # Личные данные
    "personal": "[bold $accent]Начало первичной оценки.[/]\n\n"
                "Введите данные о пациенте и специалисте, проводящем оценку, в соответствующие поля.\n\n"
                "Если пациенту ранее уже проводилась первичная оценка, убедитесь, что вносите данные о нем точно так же, как в прошлый раз, особенно если какая-то часть ФИО состоит из нескольких слов, имеет дефисы и т. п. Иностранные имена рекомендуется вносить русскими буквами. Если у пациента нет отчества, оставьте поле пустым.\n\nВ поле \"Год рождения\" введите четыре цифры. Методика не применима к пациентам младше 14 лет, поэтому введенный год рождения должен попадать в допустимый диапазон.\n\nПереход к следующему полю происходит по нажатию [bold $accent]Enter[/] или [bold $accent]Tab[/]. Перемещаясь между шагами с помощью клавиш [bold $accent]F2[/] (назад) и [bold $accent]F3[/] (вперед) вы можете вносить исправления, которые сохраняются автоматически.\n\n[bold $accent]ВНИМАНИЕ![/] В текущей версии программы при очень частом нажатии мышью на кнопки перехода между шагами программа может аварийно завершиться и введенные данные будут потеряны. Если вам нужно пролистать несколько шагов подряд, перед каждым нажатием кнопки дождитесь загрузки экрана.",

    # Описания факторов (выраженность)
    "factor_1_score": "Агрессия - выраженность",
    "factor_2_score": "Когнитивные и другие симптомы - выраженность",
    "factor_3_score": "Контроль над эмоциями - выраженность",
    "factor_4_score": "Контроль над поведением - выраженность",
    "factor_5_score": "Злоупотребление веществами - выраженность",
    "factor_6_score": "Приверженность режиму и лечению - выраженность",
    "factor_7_score": "Личностные установки - выраженность",
    "factor_8_score": "Окружение, быт и планы - выраженность",

    # Описания стадий
    "factor_1_stage": "Агрессия - стадия",
    "factor_2_stage": "Когнитивные и другие симптомы - стадия",
    "factor_3_stage": "Контроль над эмоциями - стадия",
    "factor_4_stage": "Контроль над поведением - стадия",
    "factor_5_stage": "Злоупотребление веществами - стадия",
    "factor_6_stage": "Приверженность режиму и лечению - стадия",
    "factor_7_stage": "Личностные установки - стадия",
    "factor_8_stage": "Окружение, быт и планы - стадия",

    # Заключение
    "conclusion": "Составление заключений",
}

# Словарь с фактическими индексами в PAGES
GUIDE_INDICES = {
    "personal": None,  # Текст задан явно выше
    "factor_1_defin": 18,
    "factor_2_defin": 21,
    "factor_3_defin": 24,
    "factor_4_defin": 27,
    "factor_5_defin": 30,
    "factor_6_defin": 33,
    "factor_7_defin": 36,
    "factor_8_defin": 39,
    "factor_1_score": 19,
    "factor_2_score": 22,
    "factor_3_score": 25,
    "factor_4_score": 28,
    "factor_5_score": 31,
    "factor_6_score": 34,
    "factor_7_score": 37,
    "factor_8_score": 40,
    "factor_1_stage": 20,
    "factor_2_stage": 23,
    "factor_3_stage": 26,
    "factor_4_stage": 29,
    "factor_5_stage": 32,
    "factor_6_stage": 35,
    "factor_7_stage": 38,
    "factor_8_stage": 41,
    "factor_comment": 43,
    "conclusion_part_1": 42,
    "conclusion_part_2": 45,
    "conclusion_part_3": 44,
}

# Названия факторов для отображения
FACTOR_NAMES = {
    1: "Агрессия",
    2: "Когнитивные и другие симптомы",
    3: "Контроль над эмоциями",
    4: "Контроль над поведением",
    5: "Злоупотребление веществами",
    6: "Приверженность режиму и лечению",
    7: "Личностные установки",
    8: "Окружение, быт и планы",
}

# --- Экран руководства ---
class ManualScreen(Screen):
    BINDINGS: ClassVar[list[Binding]] = [
        Binding("escape", "back", "Назад в меню"),
        Binding("left", "focus_list", show=False),
        Binding("right", "focus_content", show=False),
    ]

    def action_focus_list(self) -> None:
        self.query_one("#page_list").focus()

    def action_focus_content(self) -> None:
        # Фокусируем саму панель или контент внутри
        self.query_one("#detail_panel").focus()

    def action_back(self) -> None:
        self.app.pop_screen()

    def compose(self) -> ComposeResult:
        with Horizontal():
            with Vertical(id="list_panel") as list_col:
                list_col.border_title = "Руководство по МОДРОП Арсенал"
                yield ListView(
                    ListItem(Static(""), classes="spacer", disabled=True),
                    # ВМЕСТО Static используем наш ListLabel
                    *[ListItem(ListLabel(f"{c['part']}"), id=f"c_{i}") for i, c in enumerate(PAGES)],
                    ListItem(Static(""), classes="spacer", disabled=True),
                    id="page_list"
                )
            with VerticalScroll(id="detail_panel") as detail_col:
                detail_col.border_title = "Руководство по МОДРОП Арсенал"
                yield Static("Выберите раздел", id="details")
        yield Footer(show_command_palette=False)

    @on(ListView.Highlighted)
    def on_list_view_highlighted(self, event: ListView.Highlighted) -> None:
        if event.item and event.item.id:
            try:
                # 1. Обновляем текст в правой панели
                idx = int(event.item.id.split("_")[1])
                self.query_one("#details", Static).update(PAGES[idx]['text'])

                detail_panel = self.query_one("#detail_panel")
                detail_panel.border_title = PAGES[idx]['part']

                for label in self.query("ListLabel"):
                    label.is_highlighted = False

                # Включаем маркер только у текущего выделенного пункта
                if hasattr(event.item, "children") and event.item.children:
                    current_label = event.item.children[0]
                    if isinstance(current_label, ListLabel):
                        current_label.is_highlighted = True

            except (ValueError, IndexError):
                pass


    def on_mount(self) -> None:
        self.query_one("#page_list").focus()

class Rate1Screen(Screen):
    """Экран первичной оценки"""

    BINDINGS = [
        Binding("escape", "go_back", "Выход", show=True, priority=True),
        Binding("pagedown", "scroll_guide_down", "Листать вниз", show=True, priority=True),
        Binding("pageup", "scroll_guide_up", "Листать вверх", show=True, priority=True),
        Binding("f2", "prev_step", "Назад", show=True, priority=True),
        Binding("f3", "next_step", "Вперед", show=True),
        Binding("f10", "open_draft", "Открыть черновик", show=True, priority=True),
        Binding("f12", "save_draft", "Сохранить черновик", show=True, priority=True),
    ]

    def action_scroll_guide_up(self):
        """Прокрутка руководства вверх"""
        guide_scroll = self.query_one("#guide_scroll")
        guide_scroll.scroll_up()

    def action_scroll_guide_down(self):
        """Прокрутка руководства вниз"""
        guide_scroll = self.query_one("#guide_scroll")
        guide_scroll.scroll_down()

    def on_key(self, event: events.Key) -> None:
        """Обработчик нажатия клавиш"""

        # Обработка F10 на первом шаге
        if event.key == "f10" and self.step_index == 0:
            self.action_open_draft()
            event.prevent_default()
            event.stop()
            return

        # Для F2 и F3 добавляем защиту от множественных срабатываний
        if event.key in ["f2", "f3"]:
            import time
            current_time = time.time()

            # Проверяем, не слишком ли часто нажимают
            if hasattr(self, '_last_f_key') and self._last_f_key == event.key:
                if current_time - getattr(self, '_last_f_time', 0) < 0.3:
                    event.prevent_default()
                    event.stop()
                    return

            # Запоминаем время и клавишу
            self._last_f_key = event.key
            self._last_f_time = current_time

            # Вызываем соответствующий action с задержкой
            if event.key == "f2":
                self.call_after_refresh(self.action_prev_step)
            else:  # f3
                self.call_after_refresh(self.action_next_step)

            event.prevent_default()
            event.stop()
            return

        if event.key == "enter":
            # Определяем текущий шаг и тип виджета в фокусе
            focused = self.focused

            # Шаг 0: Личные данные
            if self.step_index == 0:
                if isinstance(focused, Input):
                    # Переходим к следующему полю ввода
                    self.focus_next()
                elif isinstance(focused, Button):
                    # Нажатие на кнопку "Далее"
                    self.call_after_refresh(self.action_next_step)

            # Шаги с факторами (1-24)
            elif 1 <= self.step_index <= 24:
                if self.current_substep in [0, 1]:  # Шаги с радиокнопками

                    if isinstance(focused, RadioSet):
                        # Получаем выбранную радиокнопку
                        if hasattr(focused, 'pressed_index'):
                            current_index = focused.pressed_index

                            # Если индекс изменился, обновляем и НЕ переходим дальше
                            if current_index != self._last_radio_index:
                                self._last_radio_index = current_index
                                # Предотвращаем стандартное поведение Enter
                                event.prevent_default()
                            else:
                                # Тот же индекс - переходим дальше
                                if self._last_radio_index is not None:
                                    event.prevent_default()
                                    self.call_after_refresh(self.action_next_step)

                    elif isinstance(focused, (RadioButton, MyRadioButton)):
                        # Если фокус на кнопке внутри сета, переносим фокус на сам RadioSet
                        if focused.parent and isinstance(focused.parent, RadioSet):
                            self.set_focus(focused.parent)
                            event.prevent_default()
                            return

                    elif isinstance(focused, Button):
                        event.prevent_default()
                        self.call_after_refresh(self.action_next_step)

                elif self.current_substep == 2:  # Шаг с комментарием
                    if isinstance(focused, TextArea):
                        # В TextArea Enter должен переходить дальше
                        # Shift+Enter для новой строки
                        if event.is_shift_down:
                            # Разрешаем стандартное поведение (новая строка)
                            pass
                        else:
                            # Обычный Enter - переходим дальше
                            event.prevent_default()
                            event.stop()
                            self.call_after_refresh(self.action_next_step)
                    elif isinstance(focused, Button):
                        event.prevent_default()
                        self.call_after_refresh(self.action_next_step)

            # Шаг 25: Заключение
            elif self.step_index == 25:
                if isinstance(focused, TextArea):
                    if event.is_shift_down:  # Shift+Enter для новой строки
                        pass
                    else:  # Обычный Enter для сохранения
                        event.prevent_default()
                        self.call_after_refresh(self.action_save_final)
                elif isinstance(focused, Button):
                    event.prevent_default()
                    self.call_after_refresh(self.action_save_final)

        # Обработка стрелок для радиокнопок
        elif event.key in ["up", "down"]:
            if 1 <= self.step_index <= 24 and self.current_substep in [0, 1]:
                if isinstance(self.focused, RadioSet):
                    # При перемещении сбрасываем запомненный индекс
                    self._last_radio_index = None

    def action_open_draft(self) -> None:
        """Открывает диалог выбора черновика первичной оценки"""
        # Проверяем, что мы на первом шаге
        if self.step_index != 0:
            self.app.custom_notify("Открыть черновик можно только в начале оценки", severity="info")
            return

        drafts = self.app.drafts.get_drafts_by_type("первичная")

        if not drafts:
            self.app.custom_notify("Нет сохраненных черновиков для первичной оценки", severity="info")
            return

        def handle_draft(draft_data):
            if draft_data:
                # Закрываем текущий экран и открываем новый с данными черновика
                self.app.pop_screen()
                self.app.push_screen(Rate1Screen(draft_data=draft_data))

        self.app.push_screen(OpenDraftDialog(drafts, "первичной"), handle_draft)

    def action_save_draft(self) -> None:
        """Действие по F12 - сохранение черновика"""
        # Проверяем, что мы не на первом шаге
        if self.step_index == 0:
            self.app.custom_notify("Сохранить черновик можно после ввода данных пациента", severity="info")
            return
        self.save_draft()

    def save_draft(self) -> None:
        """Сохраняет текущее состояние как черновик"""
        try:
            # Сохраняем текущее состояние из виджетов
            self.save_current_state()

            # Формируем имя пациента для отображения
            patient_name = f"{self.form_data.get('last_name', '')} {self.form_data.get('first_name', '')}"
            if not patient_name.strip():
                patient_name = "Без имени"

            # Создаем данные черновика
            draft_data = {
                "assessment_type": "первичная",
                "step_index": self.step_index,
                "current_factor": self.current_factor,
                "current_substep": self.current_substep,
                "form_data": self.form_data,
                "patient_info": {
                    "last_name": self.form_data.get("last_name", ""),
                    "first_name": self.form_data.get("first_name", ""),
                    "patronymic": self.form_data.get("patronymic", ""),
                    "birth_year": self.form_data.get("birth_year", "")
                }
            }

            # Если есть ID черновика, сохраняем его
            if self.draft_data and self.draft_data.get("draft_id"):
                draft_data["draft_id"] = self.draft_data["draft_id"]

            # Сохраняем черновик
            draft_id = self.app.drafts.save_draft(draft_data)

            # Обновляем draft_id в текущем объекте
            if not self.draft_data:
                self.draft_data = {}
            self.draft_data["draft_id"] = draft_id

            # Показываем диалог
            self.app.push_screen(SaveDraftDialog(patient_name), self._after_save_draft)

        except Exception as e:
            self.app.custom_notify(f"Ошибка сохранения черновика: {e}", severity="error")
            import traceback
            traceback.print_exc()

    def _after_save_draft(self, result: str) -> None:
        """Обработка результата диалога сохранения черновика"""
        if result == "exit":
            # Выходим в главное меню
            self.app.pop_screen()

    def __init__(self, draft_data: dict = None):
        super().__init__()
        self.draft_data = draft_data
        self.step_index = 0  # 0-25: 0 - ФИО, 1-24 - факторы, 25 - заключение
        self.current_factor = 1
        self.current_substep = 0  # 0-выраженность, 1-стадия, 2-комментарий
        self._last_radio_index = None  # Для отслеживания повторного нажатия на радиокнопку
        self._last_f_key = None  # Последняя нажатая клавиша F2/F3
        self._last_f_time = 0  # Время последнего нажатия F2/F3
        self._last_button_time = 0  # Время последнего нажатия кнопки
        self._last_action_time = 0  # Время последнего действия
        # Данные формы
        self.form_data = {
            "last_name": "", "first_name": "", "patronymic": "", "birth_year": "",
            "rater": "",
            "factors": {f"f{i}": {"score": None, "stage": None, "comment": ""} for i in range(1, 9)},
            "conclusion": ""
        }
        self.load_rater_data() # Попытка загрузить предыдущего специалиста
        self.last_saved_file = None # Сохраняем имя файла, в который сохранили заключение
        if draft_data:
            self.load_draft_data()

    def load_rater_data(self):
        """Загружает данные последнего специалиста"""
        rater_file = Path.home() / ".arsenal_data" / "last_rater.txt"
        if rater_file.exists():
            try:
                self.form_data["rater"] = rater_file.read_text(encoding="utf-8").strip()
            except OSError:
                pass

    def save_rater_data(self):
        """Сохраняет данные специалиста"""
        if self.form_data["rater"]:
            rater_file = Path.home() / ".arsenal_data" / "last_rater.txt"
            try:
                rater_file.parent.mkdir(exist_ok=True)
                rater_file.write_text(self.form_data["rater"], encoding="utf-8")
            except OSError:
                pass

    def load_draft_data(self):
        """Загружает данные из черновика"""
        if not self.draft_data:
            return

        # Загружаем данные формы
        self.form_data = self.draft_data.get("form_data", self.form_data)

        # Устанавливаем шаг, на котором остановились
        self.step_index = self.draft_data.get("step_index", 0)
        self.current_factor = self.draft_data.get("current_factor", 1)
        self.current_substep = self.draft_data.get("current_substep", 0)

    def compose(self) -> ComposeResult:
        with Horizontal():
            # Левая панель - руководство
            with Vertical(id="guide_panel", classes="panel-left") as guide_panel:
                guide_panel.border_title = "Данные и руководство"
                with VerticalScroll(id="guide_scroll"):
                    yield Static("", id="guide_text", markup=True)

            # Правая панель - ввод данных
            with Vertical(id="input_panel", classes="panel-right") as input_panel:
                input_panel.border_title = "Оценка"
                with VerticalScroll(id="input_scroll"):
                    yield Vertical(id="input_container")
        yield Footer(show_command_palette=False)

    def on_mount(self) -> None:
        """При монтировании показываем первый шаг"""
        self.update_step()

    def update_step(self) -> None:
        """Обновление содержимого в зависимости от шага"""
        try:
            container = self.query_one("#input_container")
        except NoMatches:
            # Если контейнер не найден, выходим
            return

        # Полностью очищаем контейнер
        try:
            container.remove_children()
        except NoMatches:
            return

        # Принудительно обновляем значения для текущего шага из form_data
        if 1 <= self.step_index <= 24:
            f_key = f"f{self.current_factor}"
            if self.current_substep == 0:  # Шаг с оценкой
                # Убеждаемся, что в form_data есть значение
                if self.form_data["factors"][f_key]["score"] is None:
                    self.form_data["factors"][f_key]["score"] = 0
            elif self.current_substep == 1:  # Шаг со стадией
                # Убеждаемся, что в form_data есть значение
                if self.form_data["factors"][f_key]["stage"] is None:
                    self.form_data["factors"][f_key]["stage"] = 0

        try:
            guide_panel = self.query_one("#guide_panel")
            guide_text = self.query_one("#guide_text")
            input_panel = self.query_one("#input_panel")
        except NoMatches:
            return

        # Шаг 0: Личные данные
        if self.step_index == 0:
            guide_panel.border_title = "Первичная оценка по МОДРОП Арсенал"
            guide_text.update(GUIDE_MAPPING["personal"])
            self.show_personal_data()

        # Шаги 1-24: Факторы (8 факторов * 3 подшага)
        elif 1 <= self.step_index <= 24:
            self.current_factor = (self.step_index - 1) // 3 + 1
            self.current_substep = (self.step_index - 1) % 3
            f_key = f"f{self.current_factor}"

            # Название фактора
            factor_name = FACTOR_NAMES.get(self.current_factor, f"Фактор {self.current_factor}")

            if self.current_substep == 0:  # Оценка выраженности
                guide_panel.border_title = f"\\[{self.current_factor}] {factor_name}"
                input_panel.border_title = "Оценка выраженности"

                # Получаем индекс для определения фактора
                defin_idx = GUIDE_INDICES.get(f"factor_{self.current_factor}_defin")
                defin_text = PAGES[defin_idx]["text"] if defin_idx is not None and defin_idx < len(PAGES) else ""

                # Получаем индекс для описания выраженности
                score_idx = GUIDE_INDICES.get(f"factor_{self.current_factor}_score")
                score_text = PAGES[score_idx]["text"] if score_idx is not None and score_idx < len(PAGES) else ""

                # Комбинируем тексты: определение, пустая строка, выраженность
                if defin_text and score_text:
                    combined_text = f"{defin_text}\n\n{score_text}"
                elif defin_text:
                    combined_text = defin_text
                elif score_text:
                    combined_text = score_text
                else:
                    combined_text = "Описание недоступно"

                guide_text.update(combined_text)

                self.show_factor_score(f_key, factor_name)

            elif self.current_substep == 1:  # Определение стадии
                # Проверяем, нужно ли показывать стадию (если оценка > 0)
                score = self.form_data["factors"][f_key]["score"]

                if score is not None and score > 0:
                    guide_panel.border_title = f"\\[{self.current_factor}] {factor_name}"
                    input_panel.border_title = "Стадия изменения"

                    # Показываем описание стадий для этого фактора
                    stage_idx = GUIDE_INDICES.get(f"factor_{self.current_factor}_stage")

                    if stage_idx is not None and stage_idx < len(PAGES):
                        specific_stage_text = PAGES[stage_idx]["text"]
                    else:
                        specific_stage_text = "Описание стадий изменения недоступно."

                    guide_text.update(specific_stage_text)

                    self.show_factor_stage(f_key, factor_name)
                else:
                    # Если оценка 0, показываем комментарий
                    self.show_factor_comment(f_key)

            elif self.current_substep == 2:  # Комментарий
                score = self.form_data["factors"][f_key]["score"]
                stage = self.form_data["factors"][f_key]["stage"]

                # Формируем превью текущих оценок
                stage_names = ["предобдумывание", "обдумывание", "подготовка", "действие", "удержание"]

                if score is not None and score > 0:
                    stage_name = stage_names[stage] if stage is not None else "не выбрана"
                else:
                    stage_name = "нет"

                comment_idx = GUIDE_INDICES.get("factor_comment")
                comment_text = PAGES[comment_idx]["text"] if comment_idx is not None and comment_idx < len(PAGES) else ""

                preview = (f"Вы дали такие оценки по фактору:\n"
                        f"  Оценка выраженности: [bold $accent]{score if score is not None else 'не определена'}[/]\n"
                        f"  Стадия изменения: [bold $accent]{stage_name}[/]\n\n"
                        f"Введите комментарий в поле справа.\n\n"
                        f"{comment_text}")

                guide_panel.border_title = f"\\[{self.current_factor}] {factor_name}"
                guide_text.update(preview)
                input_panel.border_title = "Комментарий"

                self.show_factor_comment(f_key)

        # Шаг 25: Заключение
        elif self.step_index == 25:
            guide_panel.border_title = "Первичная оценка по МОДРОП Арсенал"
            input_panel.border_title = "\\[9] Заключение"

            # Получаем данные пациента из form_data (УЖЕ ОБРАБОТАННЫЕ)
            last_name = self.form_data["last_name"]
            first_name = self.form_data["first_name"]
            patronymic = self.form_data["patronymic"]
            birth_year = self.form_data["birth_year"]
            rater = self.form_data["rater"]
            current_date = datetime.now().strftime('%Y.%m.%d')

            # Формируем сводку по оценкам
            summary = []

            summary.append("     [bold $accent]↓ Пролистайте вниз для просмотра внесенных данных ↓[/]")
            summary.append("")

            # Часть руководства - заключение
            part2_idx = GUIDE_INDICES.get("conclusion_part_2")
            if part2_idx is not None and part2_idx < len(PAGES):
                summary.append(PAGES[part2_idx]["text"])
                summary.append("")

            # Часть руководства - сумма баллов
            part3_idx = GUIDE_INDICES.get("conclusion_part_3")
            if part3_idx is not None and part3_idx < len(PAGES):
                summary.append(PAGES[part3_idx]["text"])
                summary.append("")

            # Информация о пациенте и специалисте
            summary.append("                                ***")
            summary.append("")
            summary.append("Вы дали следующие оценки и комментарии. Если нужно внести изменения, вернитесь на соответствующий шаг с помощью [bold $accent]F2[/], затем вернитесь к заключению с помощью [bold $accent]F3[/]. После сохранения заключения первичная оценка будет завершена.")
            summary.append("")

            # Используем данные из form_data без дополнительной обработки
            summary.append(f"[bold $accent]Пациент {last_name} {first_name} {patronymic} {birth_year} г.р.[/]")
            summary.append("╭──────────────────────────────────────────────────────────────────────────────╮")
            summary.append(f"Первичную оценку провел(а) {rater} {current_date}.")
            summary.append("")
            summary.append("     М О Д Р О П   А р с е н а л    ╭────" + current_date + "───╮")
            summary.append("Факторы                             │ оцен  стад      │")

            # Формируем строки для каждого фактора
            factor_names = [
                "\\[1] Агрессия                        ",
                "\\[2] Когнитивные и другие симптомы   ",
                "\\[3] Контроль над эмоциями           ",
                "\\[4] Контроль над поведением         ",
                "\\[5] Злоупотребление веществами      ",
                "\\[6] Приверженность режиму и лечению ",
                "\\[7] Личностные установки            ",
                "\\[8] Окружение, быт и планы          "
            ]

            # Факторы
            for i in range(8):
                f_key = f"f{i+1}"
                score = self.form_data["factors"][f_key]["score"]
                stage = self.form_data["factors"][f_key]["stage"]

                score_bar = self.format_score_bar(score) if score is not None else "?    "
                stage_bar = self.format_stage_bar(stage) if stage is not None else "         "

                summary.append(f"{factor_names[i]}│ {score_bar} {stage_bar} │\\[{i+1}]")
            summary.append("                                    ╰─────────────────╯")

            # Общий балл
            total_score = self.calculate_total_score()
            total_bar = "█" * total_score + " " * (24 - total_score)

            summary.append("\\[9] Всего из 24                     ╭────┬────────────────────────╮")
            summary.append(f"                       {current_date}   │[bold $accent]{total_score:3}[/] │{total_bar}│")
            summary.append("                                    ╰────┴────────────────────────╯")

            # Комментарии к факторам
            summary.append("Комментарии к факторам:")
            for i in range(1, 9):
                comment = self.form_data["factors"][f"f{i}"]["comment"]
                if comment:  # Добавляем комментарий только если он не пустой
                    # Разбиваем длинные строки
                    wrapped = textwrap.fill(comment, width=76)
                    summary.append(f"\\[[bold $accent]{i}[/]] {wrapped}")
                else:
                    summary.append(f"\\[[bold $accent]{i}[/]] Нет комментария")

            # Добавляем информацию о поле для ввода заключения
            summary.append("")
            summary.append("Введите ваше заключение в поле справа.")

            # Объединяем все в одну строку
            guide_text.update("\n".join(summary))

            # Показываем поле для ввода заключения
            self.show_conclusion()

    def show_personal_data(self):
        """Показывает форму для ввода личных данных"""
        container = self.query_one("#input_container")

        # Фамилия
        container.mount(Label("Фамилия:"))
        last_name_input = Input(
            value=self.form_data["last_name"],
            placeholder="Фамилия",
            id=f"in_last_name_{self.step_index}"
        )
        container.mount(last_name_input)

        # Имя
        container.mount(Label("Имя:"))
        first_name_input = Input(
            value=self.form_data["first_name"],
            placeholder="Имя",
            id=f"in_first_name_{self.step_index}"
        )
        container.mount(first_name_input)

        # Отчество
        container.mount(Label("Отчество:"))
        patronymic_input = Input(
            value=self.form_data["patronymic"],
            placeholder="Отчество",
            id=f"in_patronymic_{self.step_index}"
        )
        container.mount(patronymic_input)

        # Год рождения
        container.mount(Label("Год рождения:"))
        year_input = Input(
            value=self.form_data["birth_year"],
            placeholder="1990",
            max_length=4,
            id=f"in_year_{self.step_index}"
        )
        container.mount(year_input)

        # Специалист
        container.mount(Label("Специалист:"))
        rater_input = Input(
            value=self.form_data["rater"],
            placeholder="Фамилия, инициалы",
            id=f"in_rater_{self.step_index}"
        )
        container.mount(rater_input)

        # Кнопка далее
        next_btn = Button("Далее (Enter или F3)", variant="primary", id=f"btn_next_{self.step_index}")
        container.mount(next_btn)

        # Устанавливаем фокус на поле Фамилия
        self.call_after_refresh(lambda: last_name_input.focus())

    def validate_year(self, year_str: str) -> bool:
        """Проверяет корректность года рождения"""
        try:
            if not year_str or not year_str.isdigit() or len(year_str) != 4:
                return False

            year = int(year_str)
            current_year = datetime.now().year
            if year < current_year - 113 or year > current_year - 14:
                return False

            return True
        except (ValueError, TypeError):
            return False

    def show_factor_score(self, f_key: str, factor_name: str):
        """Показывает выбор оценки выраженности фактора"""
        container = self.query_one("#input_container")

        current_score = self.form_data["factors"][f_key]["score"]
        if current_score is None:
            current_score = 0

        # Создаем RadioSet с кастомной отрисовкой
        radio_buttons = [
            MyRadioButton("Все хорошо.    Оценка 0", value=(current_score == 0)),
            MyRadioButton("Есть проблемы. Оценка 1", value=(current_score == 1)),
            MyRadioButton("Плохо.         Оценка 2", value=(current_score == 2)),
            MyRadioButton("Совсем плохо.  Оценка 3", value=(current_score == 3)),
        ]

        rs = RadioSet(*radio_buttons, id=f"rs_score_{self.current_factor}_{self.step_index}")
        container.mount(rs)

        container.mount(Button("Далее (Enter или F3)", variant="primary",
                            id=f"btn_next_{self.step_index}"))

        self.reset_radio_index()
        self.call_after_refresh(lambda: rs.focus())

    def show_factor_stage(self, f_key: str, factor_name: str):
        """Показывает выбор стадии изменения"""
        container = self.query_one("#input_container")

        current_stage = self.form_data["factors"][f_key]["stage"]
        if current_stage is None:
            current_stage = 0

        # container.mount(Label(f"{factor_name}"))
        # container.mount(Label("Стадия изменения:"))

        # Создаем RadioSet
        radio_buttons = [
            MyRadioButton("Предобдумывание", value=(current_stage == 0)),
            MyRadioButton("Обдумывание", value=(current_stage == 1)),
            MyRadioButton("Подготовка", value=(current_stage == 2)),
            MyRadioButton("Действие", value=(current_stage == 3)),
            MyRadioButton("Удержание", value=(current_stage == 4)),
        ]

        rs = RadioSet(*radio_buttons, id=f"rs_stage_{self.current_factor}_{self.step_index}")
        container.mount(rs)

        # Кнопка далее
        container.mount(Button("Далее (Enter или F3)", variant="primary",
                            id=f"btn_next_{self.step_index}"))

        # Сбрасываем запомненный индекс
        self.reset_radio_index()

        # Устанавливаем фокус на RadioSet
        self.call_after_refresh(lambda: rs.focus())


    @on(RadioSet.Changed)
    def handle_radio_changed(self, event: RadioSet.Changed) -> None:
        """Сбрасываем запомненный индекс при изменении выбора мышью"""
        self._last_radio_index = event.radio_set.pressed_index
        # Не сбрасываем полностью, а обновляем, чтобы знать текущий выбор

    def reset_radio_index(self) -> None:
        """Сбрасывает запомненный индекс радиокнопки"""
        self._last_radio_index = None

    def show_factor_comment(self, f_key: str):
        """Показывает ввод комментария к фактору"""
        container = self.query_one("#input_container")

        # container.mount(Label(f"Комментарий:"))

        current_comment = self.form_data["factors"][f_key]["comment"] or ""

        # Создаем TextArea
        ta = TextArea(
            text=current_comment,
            id=f"ta_comment_{self.current_factor}_{self.step_index}",
            classes="comment-area"
        )
        container.mount(ta)

        # Кнопка далее
        container.mount(Button("Далее (F3)", variant="primary", id=f"btn_next_{self.step_index}"))

        # Устанавливаем фокус на TextArea
        self.call_after_refresh(lambda: ta.focus())

    def get_radio_value(self, radio_id: str) -> int:
        """Получает значение из RadioSet по ID"""
        try:
            rs = self.query_one(f"#{radio_id}", RadioSet)
            # В Textual, pressed_index возвращает индекс выбранной кнопки
            if hasattr(rs, 'pressed_index'):
                return rs.pressed_index
            return 0
        except Exception:
            return 0

    def show_conclusion(self):
        """Показывает ввод итогового заключения"""
        container = self.query_one("#input_container")

        # Создаем TextArea
        ta = TextArea(
            text=self.form_data["conclusion"],
            id=f"ta_conclusion_{self.step_index}",
            classes="conclusion-area"
        )
        container.mount(ta)

        # Кнопка сохранения
        save_btn = Button("Сохранить (F3)", variant="success",
                        id=f"btn_save_{self.step_index}")
        container.mount(save_btn)

        # Устанавливаем фокус на TextArea
        self.call_after_refresh(lambda: ta.focus())

    def save_current_state(self) -> None:
        """Сохраняет данные из текущих виджетов"""
        try:
            # Шаг 0: Личные данные
            if self.step_index == 0:
                try:
                    # Проверяем существование виджетов перед использованием
                    try:
                        last_name_input = self.query_one(f"#in_last_name_{self.step_index}", Input)
                        first_name_input = self.query_one(f"#in_first_name_{self.step_index}", Input)
                        year_input = self.query_one(f"#in_year_{self.step_index}", Input)
                        rater_input = self.query_one(f"#in_rater_{self.step_index}", Input)
                    except NoMatches:
                        # Если виджеты не найдены, значит мы уже перешли
                        return

                    # Получаем значения и удаляем пробелы по краям
                    last_name_raw = last_name_input.value.strip()
                    first_name_raw = first_name_input.value.strip()

                    # ОБРАБОТКА ФАМИЛИИ - каждое слово должно начинаться с заглавной буквы,
                    # но остальные буквы оставляем как ввел пользователь
                    last_name_parts = last_name_raw.split()
                    last_name_processed_parts = []
                    for part in last_name_parts:
                        if part:  # если часть не пустая
                            # Делаем первую букву заглавной, остальные оставляем как есть
                            last_name_processed_parts.append(part[0].upper() + part[1:])
                    self.form_data["last_name"] = ' '.join(last_name_processed_parts)

                    # ОБРАБОТКА ИМЕНИ - каждое слово с заглавной буквы, остальное как есть
                    first_name_parts = first_name_raw.split()
                    first_name_processed_parts = []
                    for part in first_name_parts:
                        if part:
                            first_name_processed_parts.append(part[0].upper() + part[1:])
                    self.form_data["first_name"] = ' '.join(first_name_processed_parts)

                    # ОТЧЕСТВО - первое слово с заглавной, остальные как есть
                    try:
                        patronymic_input = self.query_one(f"#in_patronymic_{self.step_index}", Input)
                        patronymic_raw = patronymic_input.value.strip()

                        if patronymic_raw:
                            patronymic_parts = patronymic_raw.split()
                            if patronymic_parts:
                                # Первое слово делаем с заглавной буквы, остальное как есть
                                first_part = patronymic_parts[0]
                                patronymic_parts[0] = first_part[0].upper() + first_part[1:]
                                # Остальные оставляем как есть
                                self.form_data["patronymic"] = ' '.join(patronymic_parts)
                            else:
                                self.form_data["patronymic"] = ""
                        else:
                            self.form_data["patronymic"] = ""
                    except NoMatches:
                        self.form_data["patronymic"] = ""

                    # ГОД РОЖДЕНИЯ - просто сохраняем
                    self.form_data["birth_year"] = year_input.value.strip()

                    # СПЕЦИАЛИСТ - только убираем пробелы по краям, регистр не меняем
                    self.form_data["rater"] = rater_input.value.strip()

                    # Валидация
                    if not self.form_data["last_name"]:
                        self.app.custom_notify("Введите фамилию!", severity="error")
                        raise ValueError("Фамилия обязательна")
                    if not self.form_data["first_name"]:
                        self.app.custom_notify("Введите имя!", severity="error")
                        raise ValueError("Имя обязательно")
                    if not self.form_data["birth_year"]:
                        self.app.custom_notify("Введите год рождения!", severity="error")
                        raise ValueError("Год рождения обязателен")
                    if not self.form_data["rater"]:
                        self.app.custom_notify("Укажите специалиста!", severity="error")
                        raise ValueError("Специалист обязателен")

                    # Проверка года рождения
                    try:
                        year = int(self.form_data["birth_year"])
                        current_year = datetime.now().year
                        if year < current_year - 113 or year > current_year - 14:
                            self.app.custom_notify(f"Год рождения должен быть между {current_year-113} и {current_year-14}",
                                    severity="error")
                            raise ValueError("Некорректный год рождения")
                    except ValueError:
                        self.app.custom_notify("Введите корректный год рождения (4 цифры)", severity="error")
                        raise
                except Exception as e:
                    if isinstance(e, ValueError):
                        raise
                    print(f"Ошибка в личных данных: {e}")
                    pass

            # Шаги с факторами
            elif 1 <= self.step_index <= 24:
                f_key = f"f{self.current_factor}"

                if self.current_substep == 0:  # Сохраняем балл
                    try:
                        # Проверяем существование виджета
                        try:
                            rs = self.query_one(f"#rs_score_{self.current_factor}_{self.step_index}", RadioSet)
                        except NoMatches:
                            return

                        if hasattr(rs, 'pressed_index') and rs.pressed_index is not None:
                            old_score = self.form_data["factors"][f_key]["score"]
                            new_score = rs.pressed_index
                            self.form_data["factors"][f_key]["score"] = new_score

                            # Если оценка изменилась с >0 на 0, очищаем стадию
                            if old_score is not None and old_score > 0 and new_score == 0:
                                self.form_data["factors"][f_key]["stage"] = None
                                # Комментарий сохраняем
                        else:
                            # Если pressed_index не определен, пробуем найти выбранную кнопку другим способом
                            for i, child in enumerate(rs.children):
                                if hasattr(child, 'value') and child.value:
                                    self.form_data["factors"][f_key]["score"] = i
                                    break
                    except Exception as e:
                        print(f"Ошибка сохранения балла: {e}")
                        pass

                elif self.current_substep == 1:  # Сохраняем стадию
                    try:
                        # Проверяем существование виджета
                        try:
                            rs = self.query_one(f"#rs_stage_{self.current_factor}_{self.step_index}", RadioSet)
                        except NoMatches:
                            return

                        if hasattr(rs, 'pressed_index') and rs.pressed_index is not None:
                            self.form_data["factors"][f_key]["stage"] = rs.pressed_index
                        else:
                            # Альтернативный поиск
                            for i, child in enumerate(rs.children):
                                if hasattr(child, 'value') and child.value:
                                    self.form_data["factors"][f_key]["stage"] = i
                                    break
                    except Exception as e:
                        print(f"Ошибка сохранения стадии: {e}")
                        pass

                elif self.current_substep == 2:  # Сохраняем комментарий
                    try:
                        # Проверяем существование виджета
                        try:
                            ta = self.query_one(f"#ta_comment_{self.current_factor}_{self.step_index}", TextArea)
                        except NoMatches:
                            return

                        self.form_data["factors"][f_key]["comment"] = ta.text
                    except Exception as e:
                        print(f"Ошибка сохранения комментария: {e}")
                        pass

            # Шаг 25: Заключение
            elif self.step_index == 25:
                try:
                    # Проверяем существование виджета
                    try:
                        ta = self.query_one(f"#ta_conclusion_{self.step_index}", TextArea)
                    except NoMatches:
                        return

                    self.form_data["conclusion"] = ta.text
                except Exception as e:
                    print(f"Ошибка сохранения заключения: {e}")
                    pass

        except ValueError:
            raise
        except Exception as e:
            self.app.custom_notify(f"Ошибка сохранения: {e}", severity="error")
            raise

    def calculate_total_score(self) -> int:
        """Вычисляет сумму баллов по всем факторам"""
        total = 0
        for i in range(1, 9):
            score = self.form_data["factors"][f"f{i}"]["score"]
            if score is not None:
                total += score
        return total

    def format_stage_bar(self, stage: int) -> str:
        """Форматирует стадию в виде графической полоски"""
        bars = ["▁     пре", "▁▂    обд", "▁▂▄   под", "▁▂▄▆  дей", "▁▂▄▆█ уде"]
        if stage is not None and 0 <= stage <= 4:
            return bars[stage]
        return "         "

    def format_score_bar(self, score: int) -> str:
        """Форматирует оценку в виде графической полоски"""
        if score is None:
            return "?    "
        bars = ["0 ▏  ", "1 ▒  ", "2 ▓▓ ", "3 ███"]
        if 0 <= score <= 3:
            return bars[score]
        return "?    "

    def save_report(self) -> None:
        try:
            last_name = self.form_data["last_name"]
            first_name = self.form_data["first_name"]
            patronymic = self.form_data["patronymic"]
            birth_year = self.form_data["birth_year"]
            rater = self.form_data["rater"]

            # Получаем нормализованное имя файла через DataManager
            filename = self.app.results.get_normalized_filename(
                last_name, first_name, patronymic, birth_year
            )
            filepath = self.app.results.reports_dir / filename

            # Собираем оценки
            scores = []
            stages = []
            for i in range(1, 9):
                f_key = f"f{i}"
                scores.append(self.form_data["factors"][f_key]["score"] or 0)
                stages.append(self.form_data["factors"][f_key]["stage"])

            total_score = sum(scores)

            # Генерируем содержимое отчета
            content = self.generate_report_content(last_name, first_name, patronymic,
                                                birth_year, rater, scores, stages, total_score)

            # Запись в файл (в конец, если файл существует)
            with open(filepath, "a", encoding="utf-8") as f:
                f.write(content)

            # Сохраняем в JSON
            patient_info = {
                "last_name": last_name,
                "first_name": first_name,
                "patronymic": patronymic,
                "birth_year": int(birth_year)  # Преобразуем в число
            }

            # Формируем данные для оценки
            assessment = {
                "assessment_id": datetime.now().strftime("%Y%m%d_%H%M"),
                "type": "первичная",
                "date": datetime.now().strftime("%Y.%m.%d"),
                "rater": rater,
                "total_score": total_score,
                "factors": {},
                "conclusion": self.form_data["conclusion"]
            }

            # Заполняем факторы
            for i in range(1, 9):
                f_key = f"f{i}"
                assessment["factors"][f_key] = {
                    "score": scores[i-1],
                    "stage": stages[i-1] if stages[i-1] is not None else 5,
                    "comment": self.form_data["factors"][f_key]["comment"]
                }

            # Сохраняем через DataManager
            self.app.results.save_assessment(patient_info, assessment)

            # Сохраняем данные специалиста
            self.save_rater_data()
            # Сохраняем имя файла для последующего выделения в списке
            self.last_saved_file = filename

        except Exception as e:
            self.app.custom_notify(f"Ошибка сохранения: {e}", severity="error")
            import traceback
            traceback.print_exc()


    def generate_report_content(self, last_name, first_name, patronymic, birth_year,
                                rater, scores, stages, total_score) -> str:
        """Генерирует содержимое отчета"""

        content = []

        # Проверяем, существует ли файл
        filename_parts = [last_name, first_name, patronymic, birth_year]
        filename_base = " ".join(filter(None, filename_parts))
        filename_safe = filename_base.replace(" ", "_")
        filepath = self.app.results.reports_dir / f"{filename_safe}.txt"

        # Если файл НЕ существует, добавляем информацию о пациенте в начале
        if not filepath.exists():
            # Используем данные как есть, без дополнительной обработки
            content.append(f"Пациент {last_name} {first_name} {patronymic} {birth_year} г.р.")
        content.append("╭──────────────────────────────────────────────────────────────────────────────╮")
        content.append(f"Первичную оценку провел(а) {rater} {datetime.now().strftime('%Y.%m.%d')}.")
        content.append("")
        content.append("     М О Д Р О П   А р с е н а л    ╭────" + datetime.now().strftime('%Y.%m.%d') + "───╮")
        content.append("Факторы                             │ оцен  стад      │")

        # Факторы 1-7
        factor_names = [
            "[1] Агрессия                        ",
            "[2] Когнитивные и другие симптомы   ",
            "[3] Контроль над эмоциями           ",
            "[4] Контроль над поведением         ",
            "[5] Злоупотребление веществами      ",
            "[6] Приверженность режиму и лечению ",
            "[7] Личностные установки            ",
            "[8] Окружение, быт и планы          "
        ]

        for i in range(8):
            score_bar = self.format_score_bar(scores[i])
            stage_bar = self.format_stage_bar(stages[i])
            content.append(f"{factor_names[i]}│ {score_bar} {stage_bar} │[{i+1}]")
        content.append("                                    ╰─────────────────╯")

        # Графическая шкала общего балла
        content.append("[9] Всего из 24                     ╭────┬────────────────────────╮")
        total_bar = "█" * total_score + " " * (24 - total_score)
        content.append(f"                       {datetime.now().strftime('%Y.%m.%d')}   │{total_score:3} │{total_bar}│")
        content.append("                                    ╰────┴────────────────────────╯")

        # Комментарии к факторам
        content.append("Заключение по первичной оценке:")
        for i in range(1, 9):
            comment = self.form_data["factors"][f"f{i}"]["comment"]
            # Разбиваем длинные строки
            wrapped = textwrap.fill(comment, width=76)
            content.append(f"[{i}] {wrapped}")

        # Общее заключение
        if self.form_data["conclusion"]:
            wrapped = textwrap.fill(self.form_data["conclusion"], width=76)
            content.append(f"[9] {wrapped}")

        content.append("╰──────────────────────────────────────────────────────────────────────────────╯")

        # Если файл существует, добавляем пустую строку для разделения оценок
        if filepath.exists():
            content.insert(0, "")  # Пустая строка сверху для разделения

        return "\n".join(content)

    def action_next_step(self) -> None:
        """Переход к следующему шагу"""
        # Добавляем защиту от множественных вызовов
        import time
        current_time = time.time()

        if hasattr(self, '_last_action_time'):
            if current_time - self._last_action_time < 0.3:
                return

        self._last_action_time = current_time

        try:
            self.save_current_state()

            # Проверка обязательных полей на определенных шагах
            if self.step_index == 0:
                # Проверки уже выполнены в save_current_state
                pass

            # Для шагов с оценкой фактора
            elif 1 <= self.step_index <= 24:
                f_key = f"f{self.current_factor}"

                if self.current_substep == 0:  # Проверка выбора балла
                    score = self.form_data["factors"][f_key]["score"]
                    if score is None:
                        self.app.custom_notify("Выберите оценку!", severity="warning")
                        return

                    # Если оценка 0, пропускаем стадию
                    if score == 0:
                        # Стадию оставляем как None
                        self.form_data["factors"][f_key]["stage"] = None
                        # Переходим сразу к комментарию (+2 шага)
                        if self.step_index + 2 <= 24:
                            self.step_index += 2
                            self.update_step()
                            return

                elif self.current_substep == 1:  # Проверка выбора стадии
                    score = self.form_data["factors"][f_key]["score"]
                    # Проверяем, что стадия выбрана только если оценка > 0
                    if score is not None and score > 0:
                        if self.form_data["factors"][f_key]["stage"] is None:
                            self.app.custom_notify("Выберите стадию!", severity="warning")
                            return

                elif self.current_substep == 2:  # Проверка комментария
                    if not self.form_data["factors"][f_key]["comment"]:
                        self.app.custom_notify("Введите комментарий!", severity="warning")
                        return

            # Шаг 25: проверка заключения
            elif self.step_index == 25:
                if not self.form_data["conclusion"]:
                    self.app.custom_notify("Введите заключение!", severity="warning")
                    return
                self.action_save_final()
                return

            # Переход к следующему шагу (только здесь!)
            if self.step_index < 25:
                self.step_index += 1
                self.update_step()

        except ValueError:
            pass  # Ошибка уже показана в save_current_state
        except NoMatches:
            # Если какой-то виджет не найден, просто выходим
            pass
        except Exception as e:
            self.app.custom_notify(f"Ошибка: {e}", severity="error")

    def action_prev_step(self) -> None:
        """Переход к предыдущему шагу"""
        # Добавляем защиту от множественных вызовов
        import time
        current_time = time.time()

        if hasattr(self, '_last_action_time'):
            if current_time - self._last_action_time < 0.3:
                return

        self._last_action_time = current_time

        # Сохраняем текущее состояние
        try:
            self.save_current_state()
        except NoMatches:
            pass
        except Exception as e:
            print(f"Error saving state: {e}")

        if self.step_index > 0:
            self.step_index -= 1

            # Обновляем current_factor и current_substep для нового шага
            if 1 <= self.step_index <= 24:
                self.current_factor = (self.step_index - 1) // 3 + 1
                self.current_substep = (self.step_index - 1) % 3
            elif self.step_index == 0:
                self.current_factor = 1
                self.current_substep = 0

            # Очищаем и обновляем
            try:
                container = self.query_one("#input_container")
                container.remove_children()
                self.update_step()
            except NoMatches:
                pass

            # Сбрасываем индекс радиокнопки
            self._last_radio_index = None

    def action_save_final(self) -> None:
        """Диалог подтверждения перед сохранением"""
        try:
            # Сначала проверяем, что заключение введено
            if not self.form_data["conclusion"]:
                self.app.custom_notify("Введите заключение!", severity="warning")
                return

            # Показываем диалог подтверждения
            self.app.push_screen(ConfirmSaveDialog(), self._after_save_confirmation)
        except Exception as e:
            self.app.custom_notify(f"Ошибка: {e}", severity="error")

    def _after_save_confirmation(self, confirmed: bool) -> None:
        """Обработка результата диалога подтверждения"""
        if confirmed:
            try:
                self.save_current_state()
                self.save_report()

                # Удаляем черновик, если он был
                if self.draft_data and self.draft_data.get("draft_id"):
                    self.app.drafts.delete_draft(self.draft_data["draft_id"])

                self.dismiss()
                self.app.push_screen(ListScreen(highlight_file=self.last_saved_file))
            except Exception as e:
                self.app.custom_notify(f"Ошибка сохранения: {e}", severity="error")
        else:
            pass

    def action_go_back(self) -> None:
        """Возврат в меню с подтверждением"""
        def check_back(do_back: bool) -> None:
            if do_back:
                self.app.pop_screen()
        self.app.push_screen(ConfirmDialog("Прервать оценку? Данные будут потеряны"), check_back)

    def safe_button_handler(self, button_id: str, handler_func):
        """Безопасный вызов обработчика кнопки"""
        try:
            # Проверяем, существует ли еще кнопка
            button = self.query_one(f"#{button_id}")
            if button:
                # Проверяем, что кнопка все еще имеет screen
                if hasattr(button, 'screen') and button.screen is not None:
                    handler_func()
        except NoMatches:
            # Кнопка уже удалена - игнорируем
            pass
        except Exception as e:
            print(f"Error in safe button handler: {e}")

    @on(Button.Pressed)
    def handle_button(self, event: Button.Pressed) -> None:
        """Обработчик нажатия на любую кнопку"""
        button_id = event.button.id

        # Добавляем защиту от множественных срабатываний
        import time
        current_time = time.time()

        # Проверяем, не слишком ли часто нажимают
        if hasattr(self, '_last_button_time'):
            if current_time - self._last_button_time < 0.5:  # 0.5 секунды задержки
                event.stop()
                return

        self._last_button_time = current_time

        # Проверяем, содержит ли ID кнопки информацию о шаге
        if button_id and button_id.startswith("btn_next_"):
            # Извлекаем номер шага из ID кнопки
            try:
                btn_step = int(button_id.split("_")[-1])
                # Если кнопка не соответствует текущему шагу, игнорируем
                if btn_step != self.step_index:
                    event.stop()
                    return
            except (ValueError, IndexError):
                pass

            # Вызываем действие с задержкой
            self.call_after_refresh(self.action_next_step)

        elif button_id and button_id.startswith("btn_save_"):
            # Извлекаем номер шага из ID кнопки
            try:
                btn_step = int(button_id.split("_")[-1])
                # Если кнопка не соответствует текущему шагу, игнорируем
                if btn_step != self.step_index:
                    event.stop()
                    return
            except (ValueError, IndexError):
                pass

            # ДЛЯ КНОПКИ СОХРАНЕНИЯ: сначала сохраняем текущее состояние
            try:
                # Сохраняем текст из TextArea в form_data
                if self.step_index == 25:
                    try:
                        ta = self.query_one(f"#ta_conclusion_{self.step_index}", TextArea)
                        self.form_data["conclusion"] = ta.text
                    except NoMatches:
                        pass
            except Exception as e:
                print(f"Ошибка при сохранении заключения: {e}")

            # Затем вызываем сохранение
            self.call_after_refresh(self.action_save_final)

class Rate2Screen(Screen):
    """Экран повторной оценки"""

    BINDINGS = [
        Binding("escape", "go_back", "Выход", show=True, priority=True),
        Binding("pagedown", "scroll_guide_down", "Листать вниз", show=True, priority=True),
        Binding("pageup", "scroll_guide_up", "Листать вверх", show=True, priority=True),
        Binding("f2", "prev_step", "Назад", show=True, priority=True),
        Binding("f3", "next_step", "Вперед", show=True),
        Binding("f10", "open_draft", "Открыть черновик", show=True, priority=True),
        Binding("f12", "save_draft", "Сохранить черновик", show=True, priority=True),
    ]

    def action_scroll_guide_up(self):
        """Прокрутка руководства вверх"""
        try:
            guide_scroll = self.query_one("#guide_scroll")
            guide_scroll.scroll_up()
        except NoMatches:
            pass

    def action_scroll_guide_down(self):
        """Прокрутка руководства вниз"""
        try:
            guide_scroll = self.query_one("#guide_scroll")
            guide_scroll.scroll_down()
        except NoMatches:
            pass

    def on_key(self, event: events.Key) -> None:
        """Обработчик нажатия клавиш"""

        if event.key == "f10" and self.step_index == 0:
            self.action_open_draft()
            event.prevent_default()
            event.stop()
            return

        # Для F2 и F3 добавляем защиту от множественных срабатываний
        if event.key in ["f2", "f3"]:
            import time
            current_time = time.time()

            # Проверяем, не слишком ли часто нажимают
            if hasattr(self, '_last_f_key') and self._last_f_key == event.key:
                if current_time - getattr(self, '_last_f_time', 0) < 0.3:
                    event.prevent_default()
                    event.stop()
                    return

            # Запоминаем время и клавишу
            self._last_f_key = event.key
            self._last_f_time = current_time

            # Вызываем соответствующий action с задержкой
            if event.key == "f2":
                self.call_after_refresh(self.action_prev_step)
            else:  # f3
                self.call_after_refresh(self.action_next_step)

            event.prevent_default()
            event.stop()
            return

        if event.key == "enter":
            # Шаг 0: Выбор пациента
            if self.step_index == 0:
                if isinstance(self.focused, ListView):
                    # Нажатие на ListView - переходим к следующему шагу
                    self.call_after_refresh(self.action_next_step)
                elif isinstance(self.focused, Button):
                    self.call_after_refresh(self.action_next_step)

            # Шаг 1: Ввод специалиста
            elif self.step_index == 1:
                if isinstance(self.focused, Input):
                    # Переходим к следующему полю ввода
                    self.focus_next()
                elif isinstance(self.focused, Button):
                    # Нажатие на кнопку "Далее"
                    self.call_after_refresh(self.action_next_step)

            # Шаги с факторами
            elif 2 <= self.step_index <= self.max_steps + 1:
                if self.current_substep == 0:  # Шаг с радиокнопками для стадии
                    if isinstance(self.focused, RadioSet):
                        if hasattr(self.focused, 'pressed_index'):
                            current_index = self.focused.pressed_index

                            if current_index != self._last_radio_index:
                                self._last_radio_index = current_index
                                event.prevent_default()
                            else:
                                if self._last_radio_index is not None:
                                    event.prevent_default()
                                    self.call_after_refresh(self.action_next_step)

                    elif isinstance(self.focused, (RadioButton, MyRadioButton)):
                        if self.focused.parent and isinstance(self.focused.parent, RadioSet):
                            self.set_focus(self.focused.parent)
                            event.prevent_default()
                            return

                    elif isinstance(self.focused, Button):
                        event.prevent_default()
                        self.call_after_refresh(self.action_next_step)

                elif self.current_substep == 1:  # Шаг с комментарием
                    if isinstance(self.focused, TextArea):
                        if event.is_shift_down:
                            pass
                        else:
                            event.prevent_default()
                            event.stop()
                            self.call_after_refresh(self.action_next_step)
                    elif isinstance(self.focused, Button):
                        event.prevent_default()
                        self.call_after_refresh(self.action_next_step)

            # Последний шаг: Заключение
            elif self.step_index == self.max_steps + 2:
                if isinstance(self.focused, TextArea):
                    if event.is_shift_down:
                        pass
                    else:
                        event.prevent_default()
                        self.call_after_refresh(self.action_save_final)
                elif isinstance(self.focused, Button):
                    event.prevent_default()
                    self.call_after_refresh(self.action_save_final)

        # Обработка стрелок для радиокнопок
        elif event.key in ["up", "down"]:
            if 2 <= self.step_index <= self.max_steps + 1 and self.current_substep == 0:
                if isinstance(self.focused, RadioSet):
                    self._last_radio_index = None

    def _after_save_confirmation(self, confirmed: bool) -> None:
        """Обработка результата диалога подтверждения"""
        if confirmed:
            try:
                self.save_current_state()
                self.save_report()

                # Удаляем черновик, если он был
                if self.draft_data and self.draft_data.get("draft_id"):
                    draft_id = self.draft_data["draft_id"]
                    print(f"DEBUG: Удаляем черновик с ID: {draft_id}")  # Отладка
                    self.app.drafts.delete_draft(draft_id)
                    self.draft_data = None
                else:
                    print(f"DEBUG: Нет draft_data или draft_id. draft_data={self.draft_data}")  # Отладка

                self.dismiss()
                self.app.push_screen(ListScreen(highlight_file=self.last_saved_file))
            except Exception as e:
                self.app.custom_notify(f"Ошибка сохранения: {e}", severity="error")
                import traceback
                traceback.print_exc()
        else:
            pass

    def __init__(self, draft_data: dict = None):
        super().__init__()

        # Отладка
        if draft_data:
            print(f"DEBUG: Rate2Screen получил draft_data с ID: {draft_data.get('draft_id')}")
            print(f"DEBUG: draft_data keys: {draft_data.keys()}")

        self.draft_data = draft_data
        self.step_index = 0
        self.current_factor_index = 0
        self.current_substep = 0
        self._last_radio_index = None
        self._last_f_key = None
        self._last_f_time = 0
        self._last_button_time = 0
        self._last_action_time = 0
        self._is_mounted = False
        self._is_navigating = False
        self._last_next_time = 0
        self._last_prev_time = 0

        # Данные
        self.primary_assessments = []
        self.selected_assessment = None
        self.active_factors = []
        self.form_data = {
            "patient": {},
            "primary": {},
            "rater": "",
            "factors": {},
            "new_scores": {},
            "conclusion": ""
        }
        self.last_saved_file = None
        self.load_rater_data()

        # Загружаем данные черновика ТОЛЬКО если он есть
        if self.draft_data:
            self.load_draft_data()

    def load_draft_data(self):
        """Загружает данные из черновика"""
        if not self.draft_data:
            print("DEBUG: load_draft_data: нет draft_data")
            return

        # СОХРАНЯЕМ draft_id - это критично!
        draft_id = self.draft_data.get("draft_id")
        print(f"DEBUG: load_draft_data: получен draft_id = {draft_id}")

        # Загружаем основные данные
        self.form_data = self.draft_data.get("form_data", self.form_data)
        self.active_factors = self.draft_data.get("active_factors", [])
        self.step_index = self.draft_data.get("step_index", 0)
        self.current_factor_index = self.draft_data.get("current_factor_index", 0)
        self.current_substep = self.draft_data.get("current_substep", 0)

        # Восстанавливаем draft_id
        if draft_id:
            if not self.draft_data:
                self.draft_data = {}
            self.draft_data["draft_id"] = draft_id
            print(f"DEBUG: draft_id восстановлен: {self.draft_data['draft_id']}")
        else:
            print("DEBUG: ВНИМАНИЕ! draft_id не найден в загружаемых данных!")

        # Восстанавливаем выбранную первичную оценку по ключу
        primary_key = self.draft_data.get("primary_assessment_key")
        if primary_key:
            self.load_primary_assessments()

            patient_uid = primary_key.get("patient_uid")
            assessment_id = primary_key.get("assessment_id")

            for ass in self.primary_assessments:
                if (ass.get("patient", {}).get("uid") == patient_uid and
                    ass.get("assessment", {}).get("assessment_id") == assessment_id):
                    self.selected_assessment = ass
                    self._select_assessment(ass)
                    break

    def action_open_draft(self) -> None:
        """Открывает диалог выбора черновика повторной оценки"""
        # Проверяем, что мы на первом шаге
        if self.step_index != 0:
            self.app.custom_notify("Открыть черновик можно только в начале оценки", severity="info")
            return

        drafts = self.app.drafts.get_drafts_by_type("повторная")

        if not drafts:
            self.app.custom_notify("Нет сохраненных черновиков для повторной оценки", severity="info")
            return

        def handle_draft(draft_data):
            if draft_data:
                self.app.pop_screen()
                self.app.push_screen(Rate2Screen(draft_data=draft_data))

        self.app.push_screen(OpenDraftDialog(drafts, "повторной"), handle_draft)

    def action_save_draft(self) -> None:
        """Действие по F12 - сохранение черновика"""
        # Проверяем, что мы не на первом шаге
        if self.step_index == 0:
            self.app.custom_notify("Сохранить черновик можно после ввода данных", severity="info")
            return
        self.save_draft()

    def save_draft(self) -> None:
        """Сохраняет текущее состояние как черновик"""
        try:
            self.save_current_state()

            print(f"DEBUG: save_draft: self.draft_data = {self.draft_data}")

            patient_name = "Без имени"
            if self.selected_assessment:
                patient = self.selected_assessment.get("patient", {})
                patient_name = f"{patient.get('last_name', '')} {patient.get('first_name', '')}"

            draft_data = {
                "assessment_type": "повторная",
                "step_index": self.step_index,
                "current_factor_index": self.current_factor_index,
                "current_substep": self.current_substep,
                "form_data": self.form_data,
                "primary_assessment_key": {
                    "patient_uid": self.selected_assessment.get("patient", {}).get("uid", ""),
                    "assessment_id": self.selected_assessment.get("assessment", {}).get("assessment_id", ""),
                    "primary_date": self.selected_assessment.get("assessment", {}).get("date", "")
                } if self.selected_assessment else None,
                "active_factors": self.active_factors,
                "patient_info": self.selected_assessment.get("patient", {}) if self.selected_assessment else {}
            }

            # Сохраняем существующий ID, если есть
            if self.draft_data and self.draft_data.get("draft_id"):
                draft_data["draft_id"] = self.draft_data["draft_id"]
                print(f"DEBUG: Обновляем черновик с ID: {self.draft_data['draft_id']}")
            else:
                print("DEBUG: Создаем новый черновик (нет draft_id)")

            draft_id = self.app.drafts.save_draft(draft_data)
            print(f"DEBUG: save_draft вернул ID: {draft_id}")

            # Обновляем draft_id в текущем объекте
            if not self.draft_data:
                self.draft_data = {}
            self.draft_data["draft_id"] = draft_id
            print(f"DEBUG: self.draft_data['draft_id'] установлен в: {self.draft_data['draft_id']}")

            self.app.push_screen(SaveDraftDialog(patient_name), self._after_save_draft)

        except Exception as e:
            self.app.custom_notify(f"Ошибка сохранения черновика: {e}", severity="error")
            import traceback
            traceback.print_exc()

    def _after_save_draft(self, result: str) -> None:
        """Обработка результата диалога сохранения черновика"""
        if result == "exit":
            self.app.pop_screen()

    def _get_selected_assessment(self) -> dict:
        """Получает выбранную оценку из списка"""
        try:
            list_view = self.query_one("#assessment_list", ListView)
            if list_view.index is not None and list_view.index < len(list_view.children):
                item = list_view.children[list_view.index]
                # Получаем данные из item
                if hasattr(item, 'assessment_data'):
                    return item.assessment_data
        except NoMatches:
            pass
        return None

    def _select_assessment(self, assessment: dict) -> None:
        """Выбирает оценку и обновляет данные"""
        if not assessment:
            return

        self.selected_assessment = assessment
        self.form_data["patient"] = assessment["patient"]
        self.form_data["primary"] = assessment["assessment"]

        # Заполняем активные факторы (с оценкой > 0)
        self.active_factors = []
        primary_factors = assessment["assessment"].get("factors", {})
        for i in range(1, 9):
            f_key = f"f{i}"
            score = primary_factors.get(f_key, {}).get("score", 0)
            if score > 0:
                stage = primary_factors.get(f_key, {}).get("stage")
                if stage == 5:
                    stage = None
                comment = primary_factors.get(f_key, {}).get("comment", "")
                self.active_factors.append({
                    "factor": i,
                    "primary_score": score,
                    "primary_stage": stage,
                    "primary_comment": comment
                })
                if f_key not in self.form_data["factors"]:
                    self.form_data["factors"][f_key] = {"stage": stage, "comment": ""}

    def load_rater_data(self):
        """Загружает данные последнего специалиста"""
        rater_file = Path.home() / ".arsenal_data" / "last_rater.txt"
        if rater_file.exists():
            try:
                self.form_data["rater"] = rater_file.read_text(encoding="utf-8").strip()
            except OSError:
                pass

    def save_rater_data(self):
        """Сохраняет данные специалиста"""
        if self.form_data["rater"]:
            rater_file = Path.home() / ".arsenal_data" / "last_rater.txt"
            try:
                rater_file.parent.mkdir(exist_ok=True)
                rater_file.write_text(self.form_data["rater"], encoding="utf-8")
            except OSError:
                pass

    def compose(self) -> ComposeResult:
        with Horizontal():
            # Левая панель - руководство
            with Vertical(id="guide_panel", classes="panel-left") as guide_panel:
                guide_panel.border_title = "Данные и руководство"
                with VerticalScroll(id="guide_scroll"):
                    yield Static("", id="guide_text", markup=True)

            # Правая панель - ввод данных
            with Vertical(id="input_panel", classes="panel-right") as input_panel:
                input_panel.border_title = "Оценка"
                with VerticalScroll(id="input_scroll"):
                    yield Vertical(id="input_container")
        yield Footer(show_command_palette=False)

    def on_mount(self) -> None:
        """При монтировании загружаем список первичных оценок"""
        self._is_mounted = True
        self.load_primary_assessments()
        self.update_step()

    def load_primary_assessments(self):
        """Загружает все первичные оценки из базы данных, у которых есть хотя бы один фактор с оценкой > 0"""
        data = self.app.results.data

        for patient in data:
            for assessment in patient.get("assessments", []):
                if assessment.get("type") == "первичная":
                    # Проверяем, есть ли хоть один фактор с оценкой > 0
                    factors = assessment.get("factors", {})
                    has_positive_score = False
                    total_score = 0

                    for i in range(1, 9):
                        score = factors.get(f"f{i}", {}).get("score", 0)
                        if score > 0:
                            has_positive_score = True
                            total_score += score

                    # Добавляем только если есть хотя бы один фактор с оценкой > 0
                    if has_positive_score:
                        # Формируем отображаемое имя
                        last_name = patient.get("last_name", "")
                        first_name = patient.get("first_name", "")
                        patronymic = patient.get("patronymic", "")
                        birth_year = patient.get("birth_year", "")
                        date = assessment.get("date", "")
                        rater = assessment.get("rater", "")

                        # Формируем строку с указанием специалиста
                        if rater:
                            display = f"{last_name} {first_name} {patronymic} {birth_year} г.р. - оценка от {date} ({rater})"
                        else:
                            display = f"{last_name} {first_name} {patronymic} {birth_year} г.р. - оценка от {date}"

                        # Сохраняем информацию для отображения
                        self.primary_assessments.append({
                            "patient": {
                                "last_name": last_name,
                                "first_name": first_name,
                                "patronymic": patronymic,
                                "birth_year": birth_year,
                                "uid": patient.get("uid", "")
                            },
                            "assessment": assessment,
                            "display": display,
                            "total_score": total_score,
                            # Добавляем сортировочные ключи
                            "_sort_key": (
                                last_name.lower(),
                                first_name.lower(),
                                patronymic.lower(),
                                date  # новые даты будут сверху при reverse=True
                            )
                        })

        # Сортируем: сначала по ФИО (алфавит), затем по дате (новые сверху)
        self.primary_assessments.sort(
            key=lambda x: (
                x["patient"]["last_name"].lower(),
                x["patient"]["first_name"].lower(),
                x["patient"]["patronymic"].lower(),
                x["assessment"].get("date", "")
            ),
            reverse=False  # Сначала по ФИО в алфавитном порядке
        )

    @property
    def max_steps(self) -> int:
        """Количество шагов с факторами"""
        return len(self.active_factors) * 2  # Каждый фактор: стадия + комментарий

    def update_step(self) -> None:
        """Обновление содержимого в зависимости от шага"""
        try:
            container = self.query_one("#input_container")
        except NoMatches:
            return

        # Полностью очищаем контейнер
        try:
            container.remove_children()
        except NoMatches:
            return

        try:
            guide_panel = self.query_one("#guide_panel")
            guide_text = self.query_one("#guide_text")
            input_panel = self.query_one("#input_panel")
        except NoMatches:
            return

        # Шаг 0: Выбор пациента
        if self.step_index == 0:
            guide_panel.border_title = "Повторная оценка"

            if not self.primary_assessments:
                guide_text.update("В базе данных нет первичных оценок или при первичных оценках ни один фактор не был отмечен как присутствующий.\n\nСначала проведите первичную оценку.")
            else:
                guide_text.update("В списке справа выберите первичную оценку, на основе которой вы хотите провести повторную оценку.\n\nДля продолжения нажмите Enter на выбранном элементе или кнопку Далее.")

            self.show_patient_selection()

        # Шаг 1: Ввод специалиста
        elif self.step_index == 1:
            if self.selected_assessment:
                patient = self.selected_assessment["patient"]
                primary = self.selected_assessment["assessment"]

                last_name = patient.get("last_name", "")
                first_name = patient.get("first_name", "")
                patronymic = patient.get("patronymic", "")
                birth_year = patient.get("birth_year", "")
                primary_date = primary.get("date", "")

                guide_panel.border_title = "Повторная оценка по МОДРОП Арсенал"
                guide_text.update(
                    "Начало повторной оценки.\n\n"
                    f"Выбран пациент: [bold $accent]{last_name} {first_name} {patronymic} {birth_year} г.р.[/]\n"
                    f"Основываемся на первичной оценке от: {primary_date}\n\n"
                    "Введите данные специалиста, проводящего повторную оценку, в поле справа.\n\n"
                    "Перемещаясь между шагами с помощью клавиш [bold $accent]F2[/] (назад) и [bold $accent]F3[/] (вперед) вы можете вносить исправления, которые сохраняются автоматически.\n\n[bold $accent]ВНИМАНИЕ![/] В текущей версии программы при очень частом нажатии мышью на кнопки перехода между шагами программа может аварийно завершиться и введенные данные будут потеряны. Если вам нужно пролистать несколько шагов подряд, перед каждым нажатием кнопки дождитесь загрузки экрана.\n"
                    "Для продолжения нажмите Enter, F3 или кнопку Далее."
                )
                self.show_rater_input_step()
            else:
                # Если почему-то нет выбранной оценки, возвращаемся на шаг 0
                self.step_index = 0
                self.update_step()

        # Шаги с факторами
        elif 2 <= self.step_index <= self.max_steps + 1 and self.active_factors:
            self.current_factor_index = (self.step_index - 2) // 2
            self.current_substep = (self.step_index - 2) % 2

            if self.current_factor_index < len(self.active_factors):
                factor_info = self.active_factors[self.current_factor_index]
                factor_num = factor_info["factor"]
                factor_name = FACTOR_NAMES.get(factor_num, f"Фактор {factor_num}")
                f_key = f"f{factor_num}"

                # Получаем данные из первичной оценки
                primary_score = factor_info["primary_score"]
                primary_stage = factor_info["primary_stage"]
                primary_comment = factor_info["primary_comment"]

                stage_names = ["предобдумывание", "обдумывание", "подготовка", "действие", "удержание"]
                primary_stage_name = stage_names[primary_stage] if primary_stage is not None and primary_stage < 5 else "не определена"

                if self.current_substep == 0:  # Выбор новой стадии
                    guide_panel.border_title = f"\\[{factor_num}] {factor_name}"
                    input_panel.border_title = "Стадия изменения"

                    # Показываем описание стадий для этого фактора
                    stage_idx = GUIDE_INDICES.get(f"factor_{factor_num}_stage")
                    if stage_idx is not None and stage_idx < len(PAGES):
                        specific_stage_text = PAGES[stage_idx]["text"]
                    else:
                        specific_stage_text = "Описание стадий изменения недоступно."

                    # Добавляем информацию из первичной оценки
                    preview = (f"Данные о первичной оценке:\n"
                              f"  Оценка выраженности: [bold $accent]{primary_score}[/]\n"
                              f"  Стадия изменения: [bold $accent]{primary_stage_name}[/]\n"
                              f"  Комментарий, данный специалистом по данному фактору при первичной оценке:\n{primary_comment}\n\n"
                              f"Руководство по установлению стадии изменения:\n\n"
                              f"{specific_stage_text}")

                    guide_text.update(preview)
                    self.show_factor_stage(factor_num, f_key, factor_name)

                elif self.current_substep == 1:  # Ввод нового комментария
                    guide_panel.border_title = f"\\[{factor_num}] {factor_name}"
                    input_panel.border_title = "Комментарий"

                    # Получаем новую стадию
                    new_stage = self.form_data["factors"].get(f_key, {}).get("stage", primary_stage)

                    # Пересчитываем оценку
                    new_score = self.calculate_new_score(factor_num, new_stage)
                    new_stage_name = stage_names[new_stage] if new_stage is not None and new_stage < 5 else "не определена"

                    stage_change = 0
                    if new_stage is not None and primary_stage is not None:
                        stage_change = new_stage - primary_stage

                    change_symbol = "+" if stage_change > 0 else "" if stage_change == 0 else "-"

                    comment_idx = GUIDE_INDICES.get("factor_comment")
                    comment_text = PAGES[comment_idx]["text"] if comment_idx is not None and comment_idx < len(PAGES) else ""

                    preview = (f"Введите новый комментарий по фактору в поле справа.\n\n"
                              f"Текущая повторная оценка:\n"
                              f"  Стадия изменения: [bold $accent]{new_stage_name} ({change_symbol}{abs(stage_change)} стад.)[/]\n"
                              f"  Пересчитанная оценка выраженности: [bold $accent]{new_score}[/]\n\n"
                              f"Данные о первичной оценке:\n"
                              f"  Оценка выраженности: [bold $accent]{primary_score}[/]\n"
                              f"  Стадия изменения: [bold $accent]{primary_stage_name}[/]\n"
                              f"  Комментарий, данный специалистом по данному фактору при первичной оценке:\n{primary_comment}\n\n"
                              f"{comment_text}")

                    guide_text.update(preview)
                    self.show_factor_comment(factor_num, f_key)

        # Последний шаг: Заключение
        elif self.step_index == self.max_steps + 2 and self.active_factors:
            guide_panel.border_title = "Повторная оценка по МОДРОП Арсенал"
            input_panel.border_title = "\\[9] Заключение"

            self.show_conclusion_preview(guide_text)

    def show_patient_selection(self):
        """Показывает список доступных первичных оценок"""
        try:
            container = self.query_one("#input_container")
        except NoMatches:
            return

        if not self.primary_assessments:
            container.mount(Label("Нет доступных первичных оценок"))
            container.mount(Button("Назад", variant="primary", id="btn_back"))
            return

        list_view = ListView(id="assessment_list")
        container.mount(list_view)

        # Сохраняем индекс выделенного элемента
        selected_idx = self.draft_data.get("selected_idx", 0) if self.draft_data else 0

        for i, ass in enumerate(self.primary_assessments):
            # ВМЕСТО Static используем ваш новый класс ListLabel
            label = ListLabel(ass['display'])

            # Если этот элемент должен быть выбран изначально — сразу активируем маркер
            if i == selected_idx:
                label.is_highlighted = True

            item = ListItem(label)
            item.assessment_data = ass
            list_view.append(item)

        if selected_idx < len(self.primary_assessments):
            list_view.index = selected_idx

        container.mount(Button("Далее (Enter или F3)", variant="primary", id="btn_next"))
        self.call_after_refresh(lambda: list_view.focus())

    @on(ListView.Highlighted, "#assessment_list")
    def handle_assessment_list_highlighted(self, event: ListView.Highlighted) -> None:
        """Обновляет отображение элементов списка при навигации"""
        if self.step_index != 0:
            return

        if event.item:
            try:
                # Находим текущий список
                list_view = self.query_one("#assessment_list", ListView)

                # Сбрасываем выделение ТОЛЬКО у элементов этого списка
                for child in list_view.children:
                    if hasattr(child, "children") and child.children:
                        label = child.children[0]
                        if isinstance(label, ListLabel):
                            # Если это текущий выделенный пункт, ставим True, иначе False
                            label.is_highlighted = (child == event.item)

            except Exception:
                pass

    def show_rater_input_step(self):
        """Показывает поле для ввода специалиста"""
        try:
            container = self.query_one("#input_container")
        except NoMatches:
            return

        # Специалист
        container.mount(Label("Специалист:"))
        rater_input = Input(
            value=self.form_data["rater"],
            placeholder="Фамилия, инициалы",
            id="in_rater_1"
        )
        container.mount(rater_input)

        # Кнопка далее
        next_btn = Button("Далее (Enter или F3)", variant="primary", id="btn_next_1")
        container.mount(next_btn)

        # Устанавливаем фокус на поле ввода
        self.call_after_refresh(lambda: rater_input.focus())

    def show_factor_stage(self, factor_num: int, f_key: str, factor_name: str):
        """Показывает выбор новой стадии изменения"""
        try:
            container = self.query_one("#input_container")
        except NoMatches:
            return

        # Получаем текущее значение
        current_stage = self.form_data["factors"].get(f_key, {}).get("stage")
        if current_stage is None:
            # Берем первичную стадию как значение по умолчанию
            for factor in self.active_factors:
                if factor["factor"] == factor_num:
                    current_stage = factor["primary_stage"]
                    break

        # Создаем RadioSet
        radio_buttons = [
            MyRadioButton("Предобдумывание", value=(current_stage == 0)),
            MyRadioButton("Обдумывание", value=(current_stage == 1)),
            MyRadioButton("Подготовка", value=(current_stage == 2)),
            MyRadioButton("Действие", value=(current_stage == 3)),
            MyRadioButton("Удержание", value=(current_stage == 4)),
        ]

        rs = RadioSet(*radio_buttons, id=f"rs_stage_{factor_num}_{self.step_index}")
        container.mount(rs)

        # Кнопка далее
        container.mount(Button("Далее (Enter или F3)", variant="primary",
                            id=f"btn_next_{self.step_index}"))

        self._last_radio_index = None
        self.call_after_refresh(lambda: rs.focus())

    def show_factor_comment(self, factor_num: int, f_key: str):
        """Показывает ввод нового комментария"""
        try:
            container = self.query_one("#input_container")
        except NoMatches:
            return

        current_comment = self.form_data["factors"].get(f_key, {}).get("comment", "")

        ta = TextArea(
            text=current_comment,
            id=f"ta_comment_{factor_num}_{self.step_index}",
            classes="comment-area"
        )
        container.mount(ta)

        container.mount(Button("Далее (F3)", variant="primary", id=f"btn_next_{self.step_index}"))

        self.call_after_refresh(lambda: ta.focus())

    def show_conclusion_preview(self, guide_text):
        """Показывает превью заключения в левой панели"""
        patient = self.form_data["patient"]
        primary = self.form_data["primary"]
        current_date = datetime.now().strftime('%Y.%m.%d')
        primary_date = primary.get("date", "")

        last_name = patient.get("last_name", "")
        first_name = patient.get("first_name", "")
        patronymic = patient.get("patronymic", "")
        birth_year = patient.get("birth_year", "")
        rater = self.form_data["rater"]

        summary = []
        summary.append("Введите ваше заключение по повторной оценке в поле справа.")
        summary.append("")
        summary.append("Вы дали следующие оценки и комментарии. Если нужно внести изменения, вернитесь на соответствующий шаг с помощью [bold $accent]F2[/], затем вернитесь к заключению с помощью [bold $accent]F3[/]. После сохранения заключения повторная оценка будет завершена.")
        summary.append("")
        summary.append("(Если схема отображается некорректно, сделайте окно программы шире или уменьшите размер шрифта с помощью сочетания клавиш [bold $accent]Ctrl -[/] (на MacOS [bold $accent]Cmd -[/]) так, чтобы линия ниже умещалась на одной строке.)")
        summary.append("")
        summary.append(f"[bold $accent]Пациент {last_name} {first_name} {patronymic} {birth_year} г.р.[/]")
        summary.append("╭──────────────────────────────────────────────────────────────────────────────╮")
        summary.append(f"Повторную оценку провел(а) {rater} {current_date}.")
        summary.append("")
        summary.append("     М О Д Р О П   А р с е н а л    ╭────" + primary_date + "───╮╭─────" + current_date + "─────╮")
        summary.append("Факторы                             │ оцен  стад      ││ стад      изм оцен │")

        # Формируем строки для каждого фактора
        factor_names = [
            "\\[1] Агрессия                        ",
            "\\[2] Когнитивные и другие симптомы   ",
            "\\[3] Контроль над эмоциями           ",
            "\\[4] Контроль над поведением         ",
            "\\[5] Злоупотребление веществами      ",
            "\\[6] Приверженность режиму и лечению ",
            "\\[7] Личностные установки            ",
            "\\[8] Окружение, быт и планы          "
        ]

        # Собираем все данные для отображения
        all_factors_data = {}
        for factor_info in self.active_factors:
            f_num = factor_info["factor"]
            f_key = f"f{f_num}"

            primary_score = factor_info["primary_score"]
            primary_stage = factor_info["primary_stage"]
            new_stage = self.form_data["factors"].get(f_key, {}).get("stage", primary_stage)
            new_score = self.calculate_new_score(f_num, new_stage)

            stage_change = 0
            if new_stage is not None and primary_stage is not None:
                stage_change = new_stage - primary_stage

            if stage_change <= -3:
                change_display = "---"
            elif stage_change == -2:
                change_display = "-- "
            elif stage_change == -1:
                change_display = "-  "
            elif stage_change == 0:
                change_display = "   "
            elif stage_change == 1:
                change_display = "+  "
            elif stage_change == 2:
                change_display = "++ "
            elif stage_change >= 3:
                change_display = "+++"
            else:
                change_display = "   "

            all_factors_data[f_num] = {
                "primary_score": primary_score,
                "primary_stage": primary_stage,
                "new_stage": new_stage,
                "new_score": new_score,
                "change": change_display
            }

        # Выводим все факторы
        for i in range(1, 9):
            if i in all_factors_data:
                data = all_factors_data[i]
                primary_score_bar = self.format_score_bar(data["primary_score"])
                primary_stage_bar = self.format_stage_bar(data["primary_stage"])
                new_stage_bar = self.format_stage_bar(data["new_stage"])
                new_score_bar = self.format_score_bar(data["new_score"])

                summary.append(f"{factor_names[i-1]}│ {primary_score_bar} {primary_stage_bar} ││ {new_stage_bar} {data['change']} {new_score_bar}│\\[{i}]")
            else:
                summary.append(f"{factor_names[i-1]}│ 0 ▏             ││                    │\\[{i}]")

        summary.append("                                    ╰─────────────────╯╰────────────────────╯")

        # Общий балл первичной оценки
        primary_total = 0
        for i in range(1, 9):
            primary_total += primary.get("factors", {}).get(f"f{i}", {}).get("score", 0)

        summary.append("\\[9] Всего из 24                     ╭────┬────────────────────────╮")
        summary.append(f"                       {primary_date}   │[bold $accent]{primary_total:3}[/] │{self.format_total_bar(primary_total)}│")
        summary.append("                                    ├────┼────────────────────────┤")
        # Общий балл повторной оценки
        new_total = sum(data["new_score"] for data in all_factors_data.values())
        summary.append(f"                       {current_date}   │[bold $accent]{new_total:3}[/] │{self.format_total_bar(new_total)}│")
        summary.append("                                    ╰────┴────────────────────────╯")

        # Комментарии
        summary.append("Новые комментарии к факторам:")
        for factor_info in self.active_factors:
            f_num = factor_info["factor"]
            f_key = f"f{f_num}"
            comment = self.form_data["factors"].get(f_key, {}).get("comment", "")
            if comment:
                wrapped = textwrap.fill(comment, width=76)
                summary.append(f"\\[[bold $accent]{f_num}[/]] {wrapped}")
            else:
                summary.append(f"\\[[bold $accent]{f_num}[/]] Нет комментария")

        guide_text.update("\n".join(summary))

        # Показываем поле для ввода заключения
        self.show_conclusion_input()

    def show_conclusion_input(self):
        """Показывает поле для ввода заключения"""
        try:
            container = self.query_one("#input_container")
        except NoMatches:
            return

        ta = TextArea(
            text=self.form_data["conclusion"],
            id=f"ta_conclusion_{self.step_index}",
            classes="conclusion-area"
        )
        container.mount(ta)

        save_btn = Button("Сохранить (F3)", variant="success",
                        id=f"btn_save_{self.step_index}")
        container.mount(save_btn)

        self.call_after_refresh(lambda: ta.focus())

    @on(RadioSet.Changed)
    def handle_radio_changed(self, event: RadioSet.Changed) -> None:
        """Сбрасываем запомненный индекс при изменении выбора мышью"""
        self._last_radio_index = event.radio_set.pressed_index

    @on(ListView.Selected)
    def handle_list_selected(self, event: ListView.Selected) -> None:
        """Обработка выбора пациента из списка (Enter или клик)"""
        if self.step_index == 0 and event.item:
            # Получаем данные прямо из item
            if hasattr(event.item, 'assessment_data'):
                selected = event.item.assessment_data
                self._select_assessment(selected)
                # Автоматически переходим к следующему шагу
                self.call_after_refresh(self.action_next_step)
            else:
                event.stop()

    def save_current_state(self) -> None:
        """Сохраняет данные из текущих виджетов"""
        try:
            # Шаг 0: Выбор пациента
            if self.step_index == 0:
                try:
                    list_view = self.query_one("#assessment_list", ListView)
                except NoMatches:
                    return

                if list_view.index is not None and list_view.index < len(list_view.children):
                    item = list_view.children[list_view.index]
                    if hasattr(item, 'assessment_data'):
                        selected = item.assessment_data
                        self.selected_assessment = selected
                        self.form_data["patient"] = selected["patient"]
                        self.form_data["primary"] = selected["assessment"]

                        # Заполняем активные факторы (с оценкой > 0)
                        self.active_factors = []
                        primary_factors = selected["assessment"].get("factors", {})
                        for i in range(1, 9):
                            f_key = f"f{i}"
                            score = primary_factors.get(f_key, {}).get("score", 0)
                            if score > 0:
                                stage = primary_factors.get(f_key, {}).get("stage")
                                # В JSON stage=5 означает отсутствие стадии (оценка 0)
                                if stage == 5:
                                    stage = None
                                comment = primary_factors.get(f_key, {}).get("comment", "")
                                self.active_factors.append({
                                    "factor": i,
                                    "primary_score": score,
                                    "primary_stage": stage,
                                    "primary_comment": comment
                                })
                                # Инициализируем данные для нового комментария
                                if f_key not in self.form_data["factors"]:
                                    self.form_data["factors"][f_key] = {"stage": stage, "comment": ""}

            # Шаг 1: Ввод специалиста
            elif self.step_index == 1:
                try:
                    rater_input = self.query_one("#in_rater_1", Input)
                    self.form_data["rater"] = rater_input.value.strip()
                except NoMatches:
                    pass

            # Шаги с факторами
            elif 2 <= self.step_index <= self.max_steps + 1 and self.active_factors:
                if self.current_factor_index < len(self.active_factors):
                    factor_info = self.active_factors[self.current_factor_index]
                    factor_num = factor_info["factor"]
                    f_key = f"f{factor_num}"

                    if self.current_substep == 0:  # Сохраняем стадию
                        try:
                            rs = self.query_one(f"#rs_stage_{factor_num}_{self.step_index}", RadioSet)
                            if hasattr(rs, 'pressed_index') and rs.pressed_index is not None:
                                if f_key not in self.form_data["factors"]:
                                    self.form_data["factors"][f_key] = {"stage": None, "comment": ""}
                                self.form_data["factors"][f_key]["stage"] = rs.pressed_index
                        except NoMatches:
                            pass

                    elif self.current_substep == 1:  # Сохраняем комментарий
                        try:
                            ta = self.query_one(f"#ta_comment_{factor_num}_{self.step_index}", TextArea)
                            if f_key not in self.form_data["factors"]:
                                self.form_data["factors"][f_key] = {"stage": None, "comment": ""}
                            self.form_data["factors"][f_key]["comment"] = ta.text
                        except NoMatches:
                            pass

            # Последний шаг: Заключение
            elif self.step_index == self.max_steps + 2 and self.active_factors:
                try:
                    ta = self.query_one(f"#ta_conclusion_{self.step_index}", TextArea)
                    self.form_data["conclusion"] = ta.text
                except NoMatches:
                    pass

        except Exception as e:
            print(f"Ошибка сохранения: {e}")
            pass

    def calculate_new_score(self, factor_num: int, new_stage: int) -> int:
        """Пересчитывает оценку выраженности на основе новой стадии"""
        # Находим первичную оценку для этого фактора
        primary_score = None
        for factor in self.active_factors:
            if factor["factor"] == factor_num:
                primary_score = factor["primary_score"]
                break

        if primary_score is None or primary_score == 0:
            return 0

        if new_stage is None:
            return primary_score

        # Правила расчета:
        # - Стадия 3 (Действие): уменьшаем на 1
        # - Стадия 4 (Удержание): уменьшаем на 2
        # - Иначе: оставляем как есть
        if new_stage == 3:  # Действие
            new_score = max(0, primary_score - 1)
        elif new_stage == 4:  # Удержание
            new_score = max(0, primary_score - 2)
        else:
            new_score = primary_score

        return new_score

    def format_score_bar(self, score: int) -> str:
        """Форматирует оценку в виде графической полоски"""
        if score is None:
            return "?    "
        bars = ["0 ▏  ", "1 ▒  ", "2 ▓▓ ", "3 ███"]
        if 0 <= score <= 3:
            return bars[score]
        return "?    "

    def format_stage_bar(self, stage: int) -> str:
        """Форматирует стадию в виде графической полоски"""
        if stage is None:
            return "         "
        bars = ["▁     пре", "▁▂    обд", "▁▂▄   под", "▁▂▄▆  дей", "▁▂▄▆█ уде"]
        if 0 <= stage <= 4:
            return bars[stage]
        return "         "

    def format_total_bar(self, total: int) -> str:
        """Форматирует общую сумму в виде полоски"""
        if total < 0 or total > 24:
            return "│" + " " * 24 + "│"
        return "█" * total + " " * (24 - total)

    def action_next_step(self) -> None:
        """Переход к следующему шагу - с защитой от падений"""
        import time
        current_time = time.time()

        if hasattr(self, '_last_next_time'):
            if current_time - self._last_next_time < 0.3:
                self._release_navigation_lock()
                return
        self._last_next_time = current_time

        try:
            if not hasattr(self, '_is_mounted') or not self._is_mounted:
                self._release_navigation_lock()
                return

            self.save_current_state()

            # Проверка обязательных полей
            if self.step_index == 0:
                # Проверяем, что оценка выбрана
                if not self.selected_assessment:
                    self.app.custom_notify("Выберите первичную оценку!", severity="warning")
                    self._release_navigation_lock()
                    return

            elif self.step_index == 1:
                if not self.form_data["rater"]:
                    self.app.custom_notify("Введите данные специалиста!", severity="warning")
                    self._release_navigation_lock()
                    return

            elif 2 <= self.step_index <= self.max_steps + 1 and self.active_factors:
                if self.current_factor_index < len(self.active_factors):
                    factor_info = self.active_factors[self.current_factor_index]
                    factor_num = factor_info["factor"]
                    f_key = f"f{factor_num}"

                    if self.current_substep == 0:  # Проверка выбора стадии
                        stage = self.form_data["factors"].get(f_key, {}).get("stage")
                        if stage is None:
                            self.app.custom_notify("Выберите стадию!", severity="warning")
                            return

                    elif self.current_substep == 1:  # Проверка комментария
                        comment = self.form_data["factors"].get(f_key, {}).get("comment", "")
                        if not comment:
                            self.app.custom_notify("Введите комментарий!", severity="warning")
                            return

            elif self.step_index == self.max_steps + 2 and self.active_factors:
                if not self.form_data["conclusion"]:
                    self.app.custom_notify("Введите заключение!", severity="warning")
                    return
                self.action_save_final()
                return

            max_total_steps = self.max_steps + 2
            if self.step_index < max_total_steps:
                self.step_index += 1
                self.call_after_refresh(self._safe_update_step)

        except Exception as e:
            self.app.custom_notify(f"Ошибка: {e}", severity="error")
            self._release_navigation_lock()

    def _safe_update_step(self):
        """Безопасное обновление шага"""
        try:
            container = self.query_one("#input_container")
            if container:
                for child in list(container.children):
                    try:
                        child.remove()
                    except Exception:
                        pass
                self.update_step()
        except NoMatches:
            pass
        except Exception as e:
            print(f"Error updating step: {e}")
        finally:
            self._release_navigation_lock()

    def action_prev_step(self) -> None:
        """Переход к предыдущему шагу - НЕ ВЫХОДИМ НА ШАГ ВЫБОРА ПАЦИЕНТА"""
        import time
        current_time = time.time()

        if hasattr(self, '_last_action_time'):
            if current_time - self._last_action_time < 0.3:
                return

        self._last_action_time = current_time

        try:
            self.save_current_state()
        except Exception:
            pass

        # НЕ ПОЗВОЛЯЕМ ВОЗВРАЩАТЬСЯ НА ШАГ 0
        # Минимальный шаг - 1 (ввод специалиста)
        if self.step_index > 1:  # Меняем условие с >0 на >1
            self.step_index -= 1

            try:
                container = self.query_one("#input_container")
                container.remove_children()
                self.update_step()
            except NoMatches:
                pass

            self._last_radio_index = None

    def action_save_final(self) -> None:
        """Диалог подтверждения перед сохранением"""
        try:
            if not self.form_data["conclusion"]:
                self.app.custom_notify("Введите заключение!", severity="warning")
                return

            self.app.push_screen(ConfirmSaveDialog(), self._after_save_confirmation)
        except Exception as e:
            self.app.custom_notify(f"Ошибка: {e}", severity="error")

    def _after_save_confirmation(self, confirmed: bool) -> None:
        """Обработка результата диалога подтверждения"""
        if confirmed:
            try:
                self.save_current_state()
                self.save_report()

                # Удаляем черновик, если он был
                print(f"DEBUG: _after_save_confirmation: self.draft_data = {self.draft_data}")

                if self.draft_data:
                    draft_id = self.draft_data.get("draft_id")
                    print(f"DEBUG: Получен draft_id = {draft_id}")
                    if draft_id:
                        print(f"DEBUG: Вызываем delete_draft с ID: {draft_id}")
                        self.app.drafts.delete_draft(draft_id)
                        print("DEBUG: delete_draft выполнен")
                    else:
                        print("DEBUG: draft_id не найден в self.draft_data")
                else:
                    print("DEBUG: self.draft_data is None")

                self.dismiss()
                self.app.push_screen(ListScreen(highlight_file=self.last_saved_file))
            except Exception as e:
                self.app.custom_notify(f"Ошибка сохранения: {e}", severity="error")
                import traceback
                traceback.print_exc()
        else:
            pass

    def save_report(self) -> None:
        """Сохраняет заключение в файл и в базу данных"""
        try:
            patient = self.form_data["patient"]
            primary = self.form_data["primary"]

            last_name = patient.get("last_name", "")
            first_name = patient.get("first_name", "")
            patronymic = patient.get("patronymic", "")
            birth_year = str(patient.get("birth_year", ""))
            rater = self.form_data["rater"]

            # Формируем имя файла
            filename_parts = [last_name, first_name, patronymic, birth_year]
            filename_base = " ".join(filter(None, filename_parts))
            filename_safe = filename_base.replace(" ", "_")
            filename = f"{filename_safe}.txt"
            filepath = self.app.results.reports_dir / filename

            # Собираем данные для отчета
            all_factors_data = {}
            for factor_info in self.active_factors:
                f_num = factor_info["factor"]
                f_key = f"f{f_num}"

                primary_score = factor_info["primary_score"]
                primary_stage = factor_info["primary_stage"]
                primary_comment = factor_info["primary_comment"]
                new_stage = self.form_data["factors"].get(f_key, {}).get("stage", primary_stage)
                new_comment = self.form_data["factors"].get(f_key, {}).get("comment", "")
                new_score = self.calculate_new_score(f_num, new_stage)

                stage_change = 0
                if new_stage is not None and primary_stage is not None:
                    stage_change = new_stage - primary_stage

                all_factors_data[f_num] = {
                    "primary_score": primary_score,
                    "primary_stage": primary_stage,
                    "primary_comment": primary_comment,
                    "new_stage": new_stage,
                    "new_comment": new_comment,
                    "new_score": new_score,
                    "change": stage_change
                }

            # Генерируем содержимое отчета
            content = self.generate_report_content(
                last_name, first_name, patronymic, birth_year,
                rater, primary, all_factors_data
            )

            # Запись в файл - теперь всегда добавляем пустую строку перед новым содержанием,
            # если файл существует
            if filepath.exists():
                # Читаем текущее содержимое
                current_content = filepath.read_text(encoding="utf-8")
                # Если файл не заканчивается пустой строкой, добавляем её
                if current_content and not current_content.endswith('\n'):
                    with open(filepath, "a", encoding="utf-8") as f:
                        f.write('\n')

                # Записываем новое содержание с пустой строкой перед ним
                with open(filepath, "a", encoding="utf-8") as f:
                    f.write('\n' + content)
            else:
                # Если файла нет, просто записываем
                with open(filepath, "w", encoding="utf-8") as f:
                    f.write(content)

            # Сохраняем в JSON
            patient_info = {
                "last_name": last_name,
                "first_name": first_name,
                "patronymic": patronymic,
                "birth_year": int(birth_year) if birth_year else 0
            }

            # Формируем данные для оценки
            assessment = {
                "assessment_id": datetime.now().strftime("%Y%m%d_%H%M"),
                "type": "повторная",
                "date": datetime.now().strftime("%Y.%m.%d"),
                "primary_date": primary.get("date", ""),
                "rater": rater,
                "total_score": sum(data["new_score"] for data in all_factors_data.values()),
                "factors": {},
                "conclusion": self.form_data["conclusion"]
            }

            # Заполняем факторы
            for f_num, data in all_factors_data.items():
                f_key = f"f{f_num}"
                assessment["factors"][f_key] = {
                    "score": data["new_score"],
                    "stage": data["new_stage"] if data["new_stage"] is not None else 5,
                    "comment": data["new_comment"],
                    "change": data["change"]
                }

            # Сохраняем через DataManager
            self.app.results.save_assessment(patient_info, assessment)

            # Сохраняем данные специалиста
            self.save_rater_data()
            self.last_saved_file = filename

        except Exception as e:
            self.app.custom_notify(f"Ошибка сохранения: {e}", severity="error")
            import traceback
            traceback.print_exc()

    def generate_report_content(self, last_name, first_name, patronymic, birth_year,
                                rater, primary, factors_data) -> str:
        """Генерирует содержимое отчета"""

        content = []
        current_date = datetime.now().strftime('%Y.%m.%d')
        primary_date = primary.get("date", "")

        # Формируем имя файла для проверки существования
        filename_parts = [last_name, first_name, patronymic, birth_year]
        filename_base = " ".join(filter(None, filename_parts))
        filename_safe = filename_base.replace(" ", "_")
        filepath = self.app.results.reports_dir / f"{filename_safe}.txt"

        # Если файл НЕ существует, добавляем информацию о первичной оценке в начале
        if not filepath.exists():
            content.append(f"Пациент {last_name} {first_name} {patronymic} {birth_year} г.р.")
            content.append("╭──────────────────────────────────────────────────────────────────────────────╮")
            content.append(f"Первичную оценку провел(а) {primary.get('rater', '')} {primary_date}.")
            content.append("")
            content.append("     М О Д Р О П   А р с е н а л    ╭────" + primary_date + "───╮")
            content.append("Факторы                             │ оцен  стад      │")

            # Выводим первичные факторы
            factor_names = [
                "[1] Агрессия                        ",
                "[2] Когнитивные и другие симптомы   ",
                "[3] Контроль над эмоциями           ",
                "[4] Контроль над поведением         ",
                "[5] Злоупотребление веществами      ",
                "[6] Приверженность режиму и лечению ",
                "[7] Личностные установки            ",
                "[8] Окружение, быт и планы          "
            ]

            for i in range(1, 9):
                f_key = f"f{i}"
                score = primary.get("factors", {}).get(f_key, {}).get("score", 0)
                stage = primary.get("factors", {}).get(f_key, {}).get("stage")
                if stage == 5:
                    stage = None

                score_bar = self.format_score_bar(score)
                stage_bar = self.format_stage_bar(stage)
                content.append(f"{factor_names[i-1]}│ {score_bar} {stage_bar} │[{i}]")

            content.append("                                    ╰─────────────────╯")

            # Общий балл первичной оценки
            primary_total = 0
            for i in range(1, 9):
                primary_total += primary.get("factors", {}).get(f"f{i}", {}).get("score", 0)

            content.append("[9] Всего из 24                     ╭────┬────────────────────────╮")
            content.append(f"                       {primary_date}   │{primary_total:3} │{self.format_total_bar(primary_total)}│")
            content.append("                                    ╰────┴────────────────────────╯")

            # Комментарии к первичной оценке
            content.append("Заключение по первичной оценке:")
            for i in range(1, 9):
                comment = primary.get("factors", {}).get(f"f{i}", {}).get("comment", "")
                if comment:
                    wrapped = textwrap.fill(comment, width=76)
                    content.append(f"[{i}] {wrapped}")

            if primary.get("conclusion"):
                wrapped = textwrap.fill(primary.get("conclusion", ""), width=76)
                content.append(f"[9] {wrapped}")

            content.append("╰──────────────────────────────────────────────────────────────────────────────╯")

        else:
            # Если файл существует, добавляем пустую строку для разделения
            content.append("")  # Добавляем пустую строку в начало новой оценки

        # Добавляем повторную оценку
        content.append("╭──────────────────────────────────────────────────────────────────────────────╮")
        content.append(f"Повторную оценку провел(а) {rater} {current_date}.")
        content.append("")
        content.append("     М О Д Р О П   А р с е н а л    ╭────" + primary_date + "───╮╭─────" + current_date + "─────╮")
        content.append("Факторы                             │ оцен  стад      ││ стад      изм оцен │")

        # Выводим все факторы
        factor_names = [
            "[1] Агрессия                        ",
            "[2] Когнитивные и другие симптомы   ",
            "[3] Контроль над эмоциями           ",
            "[4] Контроль над поведением         ",
            "[5] Злоупотребление веществами      ",
            "[6] Приверженность режиму и лечению ",
            "[7] Личностные установки            ",
            "[8] Окружение, быт и планы          "
        ]

        for i in range(1, 9):
            if i in factors_data:
                data = factors_data[i]
                primary_score_bar = self.format_score_bar(data["primary_score"])
                primary_stage_bar = self.format_stage_bar(data["primary_stage"])
                new_stage_bar = self.format_stage_bar(data["new_stage"])
                new_score_bar = self.format_score_bar(data["new_score"])

                change = data["change"]
                if change <= -3:
                    change_display = "---"
                elif change == -2:
                    change_display = "-- "
                elif change == -1:
                    change_display = "-  "
                elif change == 0:
                    change_display = "   "
                elif change == 1:
                    change_display = "+  "
                elif change == 2:
                    change_display = "++ "
                elif change >= 3:
                    change_display = "+++"
                else:
                    change_display = "   "

                content.append(f"{factor_names[i-1]}│ {primary_score_bar} {primary_stage_bar} ││ {new_stage_bar} {change_display} {new_score_bar}│[{i}]")
            else:
                content.append(f"{factor_names[i-1]}│ 0 ▏             ││                    │[{i}]")

        content.append("                                    ╰─────────────────╯╰────────────────────╯")

        # Общий балл первичной оценки
        primary_total = 0
        for i in range(1, 9):
            primary_total += primary.get("factors", {}).get(f"f{i}", {}).get("score", 0)

        content.append("[9] Всего из 24                     ╭────┬────────────────────────╮")
        content.append(f"                       {primary_date}   │{primary_total:3} │{self.format_total_bar(primary_total)}│")
        content.append("                                    ├────┼────────────────────────┤")
        # Общий балл повторной оценки
        new_total = sum(data["new_score"] for data in factors_data.values())
        content.append(f"                       {current_date}   │{new_total:3} │{self.format_total_bar(new_total)}│")
        content.append("                                    ╰────┴────────────────────────╯")

        # Комментарии к повторной оценке
        content.append("Заключение по повторной оценке:")
        for i in range(1, 9):
            if i in factors_data:
                comment = factors_data[i]["new_comment"]
                if comment:
                    wrapped = textwrap.fill(comment, width=76)
                    content.append(f"[{i}] {wrapped}")

        if self.form_data["conclusion"]:
            wrapped = textwrap.fill(self.form_data["conclusion"], width=76)
            content.append(f"[9] {wrapped}")

        content.append("╰──────────────────────────────────────────────────────────────────────────────╯")

        return "\n".join(content)

    def action_go_back(self) -> None:
        """Возврат в меню с подтверждением"""
        def check_back(do_back: bool) -> None:
            if do_back:
                self.app.pop_screen()

        # Если мы уже на шаге выбора пациента, просто выходим без подтверждения
        if self.step_index == 0 and not self.selected_assessment:
            self.app.pop_screen()
        else:
            self.app.push_screen(ConfirmDialog("Прервать оценку? Данные будут потеряны"), check_back)

    @on(Button.Pressed)
    def handle_button(self, event: Button.Pressed) -> None:
        """Обработчик нажатия на любую кнопку - с полной блокировкой"""
        button_id = event.button.id

        import time
        current_time = time.time()

        # Блокируем все нажатия, если идет навигация
        if self._is_navigating:
            event.stop()
            return

        if hasattr(self, '_last_button_time'):
            if current_time - self._last_button_time < 0.5:
                event.stop()
                return

        self._last_button_time = current_time

        # Обработка кнопки "Назад" на шаге 0
        if button_id == "btn_back" and self.step_index == 0:
            self.action_go_back()
            return

        # Обработка кнопки "Далее" на шаге 0
        if button_id == "btn_next" and self.step_index == 0:
            # Получаем выбранную оценку
            try:
                list_view = self.query_one("#assessment_list", ListView)
                if list_view.index is not None and list_view.index < len(list_view.children):
                    item = list_view.children[list_view.index]
                    if hasattr(item, 'assessment_data'):
                        selected = item.assessment_data
                        self._select_assessment(selected)
                        self._is_navigating = True
                        self.call_after_refresh(self._safe_action_next_step)
                    else:
                        self.app.custom_notify("Выберите первичную оценку из списка", severity="warning")
                else:
                    self.app.custom_notify("Выберите первичную оценку из списка", severity="warning")
            except NoMatches:
                self.app.custom_notify("Выберите первичную оценку из списка", severity="warning")
            return

        # Обработка других кнопок с номером шага
        if button_id and button_id.startswith("btn_next_"):
            try:
                btn_step = int(button_id.split("_")[-1])
                if btn_step != self.step_index:
                    event.stop()
                    return
            except (ValueError, IndexError):
                pass

            self._is_navigating = True
            self.call_after_refresh(self._safe_action_next_step)

        elif button_id and button_id.startswith("btn_save_"):
            try:
                btn_step = int(button_id.split("_")[-1])
                if btn_step != self.step_index:
                    event.stop()
                    return
            except (ValueError, IndexError):
                pass

            if self.step_index == self.max_steps + 2:
                try:
                    ta = self.query_one(f"#ta_conclusion_{self.step_index}", TextArea)
                    self.form_data["conclusion"] = ta.text
                except NoMatches:
                    pass

            self._is_navigating = True
            self.call_after_refresh(self._safe_action_save_final)

        elif button_id == "btn_back":
            self.action_go_back()

    def _safe_action_next_step(self):
        """Безопасное выполнение перехода вперед"""
        try:
            self.action_next_step()
        finally:
            self.call_after_refresh(self._release_navigation_lock)

    def _safe_action_save_final(self):
        """Безопасное выполнение сохранения"""
        try:
            self.action_save_final()
        finally:
            self.call_after_refresh(self._release_navigation_lock)

    def _release_navigation_lock(self):
        """Снимаем блокировку навигации"""
        self._is_navigating = False

class DraftManager:
    """Управление черновиками оценок"""

    def __init__(self, data_manager: DataManager):
        self.base_dir = data_manager.base_dir
        self.drafts_path = self.base_dir / "drafts.json"
        self.drafts = self.load_drafts()

    def load_drafts(self) -> list:
        """Загружает список черновиков"""
        if self.drafts_path.exists():
            try:
                with open(self.drafts_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except (json.JSONDecodeError, Exception):
                return []
        return []

    def save_drafts(self) -> None:
        """Сохраняет черновики в файл"""
        with open(self.drafts_path, "w", encoding="utf-8") as f:
            json.dump(self.drafts, f, ensure_ascii=False, indent=4)

    def save_draft(self, draft_data: dict) -> str:
        """Сохраняет новый черновик или обновляет существующий"""
        draft_id = draft_data.get("draft_id")

        if draft_id:
            # Обновляем существующий черновик
            for i, draft in enumerate(self.drafts):
                if draft.get("draft_id") == draft_id:
                    self.drafts[i] = draft_data
                    self.save_drafts()
                    return draft_id

        # Создаем новый черновик
        draft_id = str(uuid.uuid4())[:8]
        draft_data["draft_id"] = draft_id
        draft_data["created_at"] = datetime.now().isoformat()
        self.drafts.append(draft_data)
        self.save_drafts()
        return draft_id

    def delete_draft(self, draft_id: str) -> None:
        """Удаляет черновик по ID"""
        print(f"DEBUG: DraftManager.delete_draft вызван с ID: '{draft_id}'")

        if not draft_id:
            print("DEBUG: delete_draft вызван с пустым ID")
            return

        original_count = len(self.drafts)
        print(f"DEBUG: Было черновиков: {original_count}")

        # Выводим все ID для отладки
        for i, d in enumerate(self.drafts):
            print(f"DEBUG: Черновик {i}: ID = {d.get('draft_id')}")

        self.drafts = [d for d in self.drafts if d.get("draft_id") != draft_id]

        if len(self.drafts) < original_count:
            self.save_drafts()
            print(f"DEBUG: Черновик удален, осталось: {len(self.drafts)}")
        else:
            print(f"DEBUG: Черновик с ID '{draft_id}' не найден")

    def get_drafts_by_type(self, assessment_type: str) -> list:
        """Возвращает черновики определенного типа (первичная/повторная)"""
        return [d for d in self.drafts if d.get("assessment_type") == assessment_type]

    def get_all_drafts(self) -> list:
        """Возвращает все черновики"""
        # Возвращаем копию, чтобы избежать проблем с мутацией
        return [d.copy() for d in self.drafts if d.get("assessment_type") in ["первичная", "повторная"]]

class SaveDraftDialog(ModalScreen[str]):
    """Диалог подтверждения сохранения черновика"""

    CSS = """
    SaveDraftDialog {
        align: center middle;
        background: transparent;
    }

    #draft_dialog {
        width: 60;
        height: auto;
        background: transparent;
        border: round $accent;
        padding: 1 2;
    }

    #draft_message {
        width: 100%;
        height: auto;
        content-align: center middle;
        color: $text;
        background: transparent;
        text-style: bold;
        margin-bottom: 1;
    }

    #draft_buttons {
        width: 100%;
        height: 3;
        background: transparent;
        align: center middle;
    }

    #draft_buttons Button {
        width: 18;
        height: 3;
        margin: 0 2;
        background: transparent;
        color: $secondary;
        border: none;
    }

    #draft_buttons Button:focus {
        border: round $accent;
        background: transparent;
        color: $accent;
        text-style: bold;
    }
    """

    BINDINGS = [
        Binding("escape", "cancel", "Отмена", show=False),
    ]

    def __init__(self, patient_name: str):
        super().__init__()
        self.patient_name = patient_name

    def compose(self) -> ComposeResult:
        with Vertical(id="draft_dialog"):
            yield Label(f"Черновик {self.patient_name} сохранен.\nВыйти из текущей оценки?",
                       id="draft_message")
            with Horizontal(id="draft_buttons"):
                yield Button("Выйти", variant="error", id="btn_exit")
                yield Button("Продолжить", variant="primary", id="btn_continue")

    def on_mount(self) -> None:
        self.query_one("#btn_continue").focus()

    def on_key(self, event: events.Key) -> None:
        if event.key == "left":
            self.focus_previous()
            event.stop()
        elif event.key == "right":
            self.focus_next()
            event.stop()

    @on(Button.Pressed, "#btn_exit")
    def handle_exit(self, event: Button.Pressed) -> None:
        self.dismiss("exit")

    @on(Button.Pressed, "#btn_continue")
    def handle_continue(self, event: Button.Pressed) -> None:
        self.dismiss("continue")

    def action_cancel(self) -> None:
        self.dismiss("continue")

class OpenDraftDialog(ModalScreen[dict]):
    """Диалог выбора черновика"""

    CSS = """
    OpenDraftDialog {
        align: center middle;
        background: transparent;
    }

    #draft_list_dialog {
        width: 70;
        height: 20;
        background: transparent;
        border: round $accent;
        padding: 1 2;
    }

    #draft_list_title {
        width: 100%;
        height: 1;
        content-align: center middle;
        color: $text;
        text-style: bold;
        margin-bottom: 1;
    }

    #draft_list_view {
        width: 100%;
        height: 13;
        border: none;
        background: transparent;
    }

    #draft_actions {
        width: 100%;
        height: 3;
        align: center middle;
        margin-top: 1;
    }

    #draft_actions Button {
        width: 20;
        height: 3;
        margin: 0 2;
        background: transparent;
        color: $secondary;
        border: none;
    }

    #draft_actions Button:focus {
        border: round $accent;
        color: $accent;
        background: transparent;
        text-style: bold;
    }

    #draft_list_view ListItem {
        padding-left: 1;
        background: transparent;
        color: $text;
    }

    #draft_list_view > ListItem.--highlight {
        background: transparent;
    }

    #draft_list_view > ListItem.--highlight > Static {
        color: $accent;
        text-style: bold;
    }

    #no_drafts_message {
        width: 100%;
        height: 13;
        content-align: center middle;
        color: $text;
        border: round $secondary;
        background: transparent;
    }
    """

    BINDINGS = [
        Binding("escape", "cancel", "Отмена", show=False),
        Binding("enter", "select_draft", "Открыть", show=False),
    ]

    def __init__(self, drafts: list, assessment_type: str = None):
        super().__init__()
        self.drafts = drafts
        self.assessment_type = assessment_type

    def compose(self) -> ComposeResult:
        with Vertical(id="draft_list_dialog"):
            title = "Выберите черновик" if not self.assessment_type else f"Черновики для {self.assessment_type} оценки"
            yield Label(title, id="draft_list_title")

            if self.drafts:
                yield ListView(id="draft_list_view")
            else:
                # Используем простой текст без разметки
                yield Label("Нет сохраненных черновиков", id="no_drafts_message")

            with Horizontal(id="draft_actions"):
                yield Button("Открыть", variant="primary", id="btn_open")
                yield Button("Отмена", variant="default", id="btn_cancel")

    def on_mount(self) -> None:
        """Заполняем список после монтирования"""
        if not self.drafts:
            try:
                self.query_one("#btn_open").disabled = True
            except NoMatches:
                pass
            try:
                self.query_one("#btn_cancel").focus()
            except NoMatches:
                pass
            return

        list_view = self.query_one("#draft_list_view", ListView)

        for i, draft in enumerate(self.drafts):
            patient_info = draft.get("patient_info", {})
            display = (f"{patient_info.get('last_name', '')} "
                    f"{patient_info.get('first_name', '')} "
                    f"{patient_info.get('patronymic', '')} "
                    f"{patient_info.get('birth_year', '')} г.р. - "
                    f"{draft.get('created_at', '')[:10]}")

            label = ListLabel(display)

            # Если это самый первый элемент — сразу активируем маркер █
            if i == 0:
                label.is_highlighted = True

            item = ListItem(label)
            item.draft_data = draft
            list_view.append(item)

        # Фокус первого элемента
        if list_view.children:
            list_view.focus()
            list_view.index = 0

    @on(ListView.Highlighted, "#draft_list_view")
    def handle_draft_list_highlighted(self, event: ListView.Highlighted) -> None:
        """Обновляет отображение элементов списка черновиков при навигации"""
        if event.item:
            try:
                list_view = self.query_one("#draft_list_view", ListView)

                for child in list_view.children:
                    if hasattr(child, "children") and child.children:
                        label = child.children[0]
                        if isinstance(label, ListLabel):
                            # Если элемент является текущим выделенным — True, иначе — False
                            label.is_highlighted = (child == event.item)

            except NoMatches:
                pass

    def on_key(self, event: events.Key) -> None:
        """Обработка клавиш для навигации"""
        if event.key == "left":
            focused = self.focused
            if hasattr(focused, 'id') and focused.id == "draft_list_view":
                self.query_one("#btn_open").focus()
                event.stop()
                event.prevent_default()
            else:
                self.focus_previous()
                event.stop()
                event.prevent_default()
        elif event.key == "right":
            focused = self.focused
            if isinstance(focused, Button):
                self.query_one("#draft_list_view").focus()
                event.stop()
                event.prevent_default()
            else:
                self.focus_next()
                event.stop()
                event.prevent_default()
        elif event.key == "enter":
            focused = self.focused
            # Если фокус на списке, выбираем элемент
            if hasattr(focused, 'id') and focused.id == "draft_list_view":
                self.action_select_draft()
                event.stop()
                event.prevent_default()

    @on(ListView.Selected)
    def handle_list_selected(self, event: ListView.Selected) -> None:
        """Обработка выбора элемента списка (двойной клик или Enter)"""
        if event.item and hasattr(event.item, 'draft_data'):
            print(f"DEBUG: OpenDraftDialog выбран черновик с ID: {event.item.draft_data.get('draft_id')}")
            self.dismiss(event.item.draft_data)

    def action_select_draft(self) -> None:
        """Выбор черновика по Enter"""
        if not self.drafts:
            return

        try:
            list_view = self.query_one("#draft_list_view", ListView)
            if list_view.index is not None and list_view.index < len(list_view.children):
                item = list_view.children[list_view.index]
                if hasattr(item, 'draft_data'):
                    print(f"DEBUG: OpenDraftDialog выбран черновик с ID: {item.draft_data.get('draft_id')}")
                    self.dismiss(item.draft_data)
        except (NoMatches, IndexError):
            pass

    @on(Button.Pressed, "#btn_open")
    def handle_open(self, event: Button.Pressed) -> None:
        """Обработка нажатия кнопки Открыть"""
        self.action_select_draft()

    @on(Button.Pressed, "#btn_cancel")
    def handle_cancel(self, event: Button.Pressed) -> None:
        """Обработка нажатия кнопки Отмена"""
        self.dismiss(None)

    def action_cancel(self) -> None:
        self.dismiss(None)

class ConfirmDialog(ModalScreen[bool]):
    """Простой диалог подтверждения"""

    CSS = """
    ConfirmDialog {
        align: center middle;
        background: transparent;
    }

    ConfirmDialog #dialog {
        width: 60;
        height: auto;           /* Автоматическая высота */
        max-height: 80%;        /* Не больше 80% терминала */
        background: $background;
        border: round $accent;
        padding: 1 2;
    }

    ConfirmDialog #question {
        width: 100%;
        height: auto;           /* Автоматическая высота */
        content-align: center middle;
        color: $text;
        text-style: bold;
        margin-bottom: 1;
        text-wrap: wrap;        /* Перенос текста */
    }

    ConfirmDialog #buttons {
        width: 100%;
        height: 3;
        align: center middle;
        margin-top: 1;
    }

    ConfirmDialog Button {
        width: 14;
        height: 3;
        margin: 0 2;
        background: $background;
        color: $secondary;
        border: none;
    }

    ConfirmDialog Button:focus {
        border: round $accent;
        color: $accent;
        background: transparent;
        text-style: bold;
    }
    """

    def __init__(self, message: str):
        super().__init__()
        self.message = message

    def compose(self) -> ComposeResult:
        with Vertical(id="dialog"):
            yield Label(self.message, id="question")
            with Horizontal(id="buttons"):
                yield Button("Да", variant="error", id="yes")
                yield Button("Нет", variant="primary", id="no")

    def on_mount(self) -> None:
        self.query_one("#no").focus()

    def on_key(self, event: events.Key) -> None:
        """Обработка нажатий клавиш для навигации между кнопками"""
        if event.key == "left":
            self.focus_previous()
            event.stop()
        elif event.key == "right":
            self.focus_next()
            event.stop()

    def on_button_pressed(self, event: Button.Pressed) -> None:
        self.dismiss(event.button.id == "yes")

class ListScreen(Screen):
    BINDINGS = [
        Binding("escape", "app.pop_screen", "Назад"),
        Binding("e", "open_office", "Открыть в редакторе"),
        Binding("t", "open_office", "Открыть в Офисе", show=False),
        Binding("у", "open_office", "Открыть в Офисе", show=False),
        Binding("е", "open_office", "Открыть в Офисе", show=False),
        Binding("left", "focus_list", show=False),
        Binding("right", "focus_preview", show=False),
        Binding("delete", "delete_report", "Удалить"),
    ]

    def __init__(self, highlight_file: str = None):
        super().__init__()
        self.highlight_file = highlight_file  # Имя файла для выделения

    def compose(self) -> ComposeResult:
        with Horizontal():
            with Vertical(id="list_panel") as list_col:
                list_col.border_title = "Пациенты"
                yield ListView(id="reports_list")

            with VerticalScroll(id="detail_panel") as detail_col:
                detail_col.border_title = "Просмотр"
                yield Static("", id="details", markup=False)
        yield Footer(show_command_palette=False)

    # --- Навигация стрелками ---
    def action_focus_list(self) -> None:
        self.query_one("#reports_list").focus()

    def action_focus_preview(self) -> None:
        self.query_one("#detail_panel").focus()

    # --- Логика работы с файлами ---
    def on_mount(self) -> None:
        self.refresh_list()

        # Если нужно выделить конкретный файл
        if self.highlight_file:
            list_view = self.query_one("#reports_list", ListView)
            for i, item in enumerate(list_view.children):
                if hasattr(item, "name") and item.name == self.highlight_file:
                    # Устанавливаем выделение на этот элемент
                    list_view.index = i
                    break

        self.query_one("#reports_list").focus()

    def refresh_list(self) -> None:
        path = self.app.results.reports_dir
        files = sorted([f.name for f in path.glob("*.txt")])
        list_view = self.query_one("#reports_list", ListView)
        list_view.clear()

        # Определяем, какой индекс файла должен быть выделен
        target_index = 0
        if self.highlight_file and self.highlight_file in files:
            target_index = files.index(self.highlight_file)

        for i, f_name in enumerate(files):
            label = ListLabel(f_name)

            # Если этот файл должен быть выделен — сразу активируем маркер
            if i == target_index:
                label.is_highlighted = True

            list_view.append(ListItem(label, name=f_name))

        # Устанавливаем индекс в самом ListView (если файлы вообще есть)
        if files:
            list_view.index = target_index

    @on(ListView.Highlighted, "#reports_list")
    def handle_list_highlighted(self, event: ListView.Highlighted) -> None:
        """Обновляет отображение элементов списка при навигации"""
        try:
            list_view = self.query_one("#reports_list", ListView)

            # Переключаем маркеры, используя индекс элемента в цикле
            for i, child in enumerate(list_view.children):
                if hasattr(child, "children") and child.children:
                    label = child.children[0]
                    if isinstance(label, ListLabel):
                        # Включаем маркер, если индекс элемента совпадает с текущим выбранным индексом списка
                        label.is_highlighted = (i == list_view.index)

        except NoMatches:
            pass

    def on_list_view_highlighted(self, event: ListView.Highlighted) -> None:
        """Обновление превью и заголовка (первая строка файла) при навигации"""
        detail_panel = self.query_one("#detail_panel")
        static_details = self.query_one("#details", Static)

        # Проверяем выбор
        if event.item and hasattr(event.item, "name") and event.item.name:
            file_name = event.item.name
            file_path = self.app.results.reports_dir / file_name

            try:
                # Читаем содержимое файла
                content = file_path.read_text(encoding="utf-8")

                # Извлекаем первую строку для заголовка
                lines = content.splitlines()
                # Если файл не пуст, берем 1-ю строку, иначе имя файла
                first_line = lines[0].strip() if lines else file_name

                # Обновляем заголовок панели
                detail_panel.border_title = f"{first_line}"

                # Обновляем текст превью
                static_details.update(content)

            except Exception as e:
                detail_panel.border_title = "Ошибка"
                static_details.update(f"Не удалось прочитать файл: {e}")
        else:
            # Если список пуст или ничего не выбрано
            detail_panel.border_title = "Просмотр"
            static_details.update("Выберите файл из списка слева")

    def action_open_office(self) -> None:
        """Открывает выделенный файл во внешнем редакторе"""
        selected = self.query_one("#reports_list").highlighted_child
        if selected and hasattr(selected, "name"):
            full_path = self.app.results.reports_dir / selected.name
            open_file_externally(full_path)

    def action_delete_report(self) -> None:
        """Срабатывает по клавише Delete"""
        selected = self.query_one("#reports_list").highlighted_child
        if selected and hasattr(selected, "name"):
            file_name = selected.name

            # Вызываем диалог и передаем функцию, которая сработает после закрытия
            def check_answer(should_delete: bool) -> None:
                if should_delete:
                    file_path = self.app.results.reports_dir / file_name
                    try:
                        file_path.unlink()
                        self.refresh_list()
                        self.query_one("#details", Static).update("Файл удален")
                        self.query_one("#detail_panel").border_title = "Просмотр"
                    except Exception as e:
                        self.app.custom_notify(f"Ошибка: {e}", severity="error")

            self.app.push_screen(ConfirmDeleteDialog(file_name), check_answer)


class LookScreen(Screen):
    """Экран поиска и просмотра оценок из базы данных"""

    BINDINGS = [
        Binding("escape", "go_back", "Назад", show=True, priority=True),
        Binding("pageup", "scroll_view_up", "Вверх", show=True),
        Binding("pagedown", "scroll_view_down", "Вниз", show=True),
        Binding("enter", "select_item", "Выбрать", show=True),
        Binding("1", "press_open", "Открыть", show=False),
        Binding("2", "press_restore", "Записать", show=False),
        Binding("3", "press_draft", "Черновик", show=False),
    ]

    def action_press_open(self):
        """Нажатие 1 - Открыть в редакторе"""
        if self.mode == "view":
            self.handle_open()
        else:
            self.app.custom_notify("Сначала выберите оценку", severity="warning")

    def action_press_restore(self):
        """Нажатие 2 - Записать в файл пациента"""
        if self.mode == "view":
            self.handle_restore()
        else:
            self.app.custom_notify("Сначала выберите оценку", severity="warning")

    def action_press_draft(self):
        """Нажатие 3 - Сохранить как черновик"""
        if self.mode == "view":
            self.handle_save_as_draft()
        else:
            self.app.custom_notify("Сначала выберите оценку", severity="warning")

    def __init__(self):
        super().__init__()
        # Все оценки из базы данных
        self.all_assessments = []
        # Отфильтрованные оценки для отображения
        self.filtered_assessments = []
        # Индекс выбранной оценки
        self.selected_index = 0
        # Текущий режим: "search" (поиск) или "view" (просмотр)
        self.mode = "search"
        # Данные для текущего просмотра
        self.current_view_data = None
        self.current_view_type = None
        self.current_view_assessment = None
        self.current_view_primary = None
        self.current_temp_file = None
        # Поисковые поля
        self._search_last_name = ''
        self._search_first_name = ''
        self._search_patronymic = ''
        self._search_year = ''
        self._search_rater = ''
        # Список ID полей ввода для навигации
        self.input_fields = []
        self.current_field_index = 0
        # Список кнопок в режиме просмотра
        self.view_buttons = []
        self.current_button_index = 0
        # Счетчик для уникальных ID
        self._form_counter = 0
        # Флаг: фокус на левой панели
        self._is_on_list = True

    def on_mouse_move(self, event: events.MouseMove) -> None:
        """При движении мыши подсвечиваем кнопку под курсором в режиме просмотра"""
        if self.mode != "view":
            return

        # Получаем виджет под курсором
        result = self.get_widget_at(event.screen_x, event.screen_y)

        if result is None:
            return

        if isinstance(result, tuple):
            widget = result[0]
        else:
            widget = result

        # Ищем родителя-кнопку
        current = widget
        button = None
        while current:
            if isinstance(current, Button) and current.id and current.id in self.view_buttons:
                button = current
                break
            current = current.parent

        if button and self.focused != button:
            button.focus()

    def compose(self) -> ComposeResult:
        with Horizontal():
            # Левая панель - список оценок
            with Vertical(id="look_list_panel", classes="panel-left") as list_panel:
                list_panel.border_title = "Проведенные оценки"
                with VerticalScroll(id="list_scroll"):
                    yield ListView(id="assessments_list")

            # Правая панель - поиск или просмотр
            with Vertical(id="look_content_panel", classes="panel-right") as content_panel:
                content_panel.border_title = "Поиск оценок"
                with VerticalScroll(id="content_scroll") as self.content_scroll:
                    yield Vertical(id="content_container")
        yield Footer(show_command_palette=False)

    def on_mount(self) -> None:
        """Загружаем все оценки при монтировании"""
        self.load_all_assessments()
        self.filtered_assessments = self.all_assessments.copy()
        self.show_search_form()
        self.update_list()  # Обновляем список сразу после создания формы

        # Устанавливаем фокус на список (левое окно)
        try:
            list_view = self.query_one("#assessments_list", ListView)
            if list_view.children:
                list_view.focus()
                list_view.index = 0
                self._is_on_list = True
        except NoMatches:
            pass

    def _get_unique_id(self, base_id: str) -> str:
        """Генерирует уникальный ID для формы"""
        self._form_counter += 1
        return f"{base_id}_{self._form_counter}"

    def load_all_assessments(self):
        """Загружает все оценки из базы данных"""
        data = self.app.results.data

        for patient in data:
            patient_data = {
                "last_name": patient.get("last_name", ""),
                "first_name": patient.get("first_name", ""),
                "patronymic": patient.get("patronymic", ""),
                "birth_year": str(patient.get("birth_year", "")),
                "uid": patient.get("uid", "")
            }

            # Сортируем оценки по дате (старые сверху, новые снизу)
            assessments = sorted(
                patient.get("assessments", []),
                key=lambda a: a.get("date", ""),
                reverse=False
            )

            for assessment in assessments:
                assessment_type = assessment.get("type", "первичная")
                rater = assessment.get("rater", "")

                # Формируем отображаемую строку с указанием специалиста
                if rater:
                    display = (f"{patient_data['last_name']} {patient_data['first_name']} "
                               f"{patient_data['patronymic']} {patient_data['birth_year']} г.р. - "
                               f"{assessment_type} от {assessment.get('date', '')} ({rater})")
                else:
                    display = (f"{patient_data['last_name']} {patient_data['first_name']} "
                               f"{patient_data['patronymic']} {patient_data['birth_year']} г.р. - "
                               f"{assessment_type} от {assessment.get('date', '')}")

                # Для повторной оценки нужно найти первичную
                primary_assessment = None
                if assessment_type == "повторная":
                    # Ищем последнюю первичную оценку этого пациента
                    for a in assessments:
                        if a.get("type") == "первичная":
                            primary_assessment = a
                            break

                self.all_assessments.append({
                    "patient": patient_data.copy(),
                    "assessment": assessment,
                    "primary": primary_assessment,
                    "display": display,
                    "search_text": f"{display} {assessment.get('rater', '')}".lower()
                })

        # Сортируем по алфавиту (по фамилии, имени, отчеству)
        self.all_assessments.sort(key=lambda x: (
            x["patient"]["last_name"].lower(),
            x["patient"]["first_name"].lower(),
            x["patient"]["patronymic"].lower(),
            x["assessment"].get("date", "")
        ))

    def update_list(self):
        """Обновляет список в левой панели"""
        try:
            list_view = self.query_one("#assessments_list", ListView)
            list_view.clear()

            for i, ass in enumerate(self.filtered_assessments):
                label = ListLabel(ass['display'])

                # Устанавливаем маркер только для изначально выбранного элемента
                # (маркеры при навигации будут обновляться через обработчик Highlighted)
                if i == self.selected_index:
                    label.is_highlighted = True

                item = ListItem(label)
                item.db_index = i
                item.assessment_data = ass
                list_view.append(item)

            if self.filtered_assessments:
                list_view.index = self.selected_index

        except NoMatches:
            pass

    def show_search_form(self):
        """Показывает форму поиска в правой панели"""
        self.mode = "search"
        self.current_view_data = None
        self.input_fields = []
        self.view_buttons = []
        self._form_counter = 0  # Сбрасываем счетчик для новой формы

        try:
            container = self.query_one("#content_container")

            # Полностью очищаем контейнер
            for child in list(container.children):
                child.remove()

            content_panel = self.query_one("#look_content_panel")
            content_panel.border_title = "Поиск оценок"

            # Поля ввода с уникальными ID
            container.mount(Label("Фамилия:"))
            last_name_id = self._get_unique_id("search_last_name")
            last_name_input = Input(
                value=self._search_last_name,
                placeholder="Фамилия",
                id=last_name_id
            )
            container.mount(last_name_input)
            self.input_fields.append(last_name_id)

            container.mount(Label("Имя:"))
            first_name_id = self._get_unique_id("search_first_name")
            first_name_input = Input(
                value=self._search_first_name,
                placeholder="Имя",
                id=first_name_id
            )
            container.mount(first_name_input)
            self.input_fields.append(first_name_id)

            container.mount(Label("Отчество:"))
            patronymic_id = self._get_unique_id("search_patronymic")
            patronymic_input = Input(
                value=self._search_patronymic,
                placeholder="Отчество",
                id=patronymic_id
            )
            container.mount(patronymic_input)
            self.input_fields.append(patronymic_id)

            container.mount(Label("Год рождения:"))
            year_id = self._get_unique_id("search_year")
            year_input = Input(
                value=self._search_year,
                placeholder="1990",
                max_length=4,
                id=year_id
            )
            container.mount(year_input)
            self.input_fields.append(year_id)

            container.mount(Label("Специалист:"))
            rater_id = self._get_unique_id("search_rater")
            rater_input = Input(
                value=self._search_rater,
                placeholder="Фамилия, инициалы",
                id=rater_id
            )
            container.mount(rater_input)
            self.input_fields.append(rater_id)

        except NoMatches:
            pass

    def on_key(self, event: events.Key) -> None:
        """Обработчик нажатий клавиш"""

        # В режиме поиска
        if self.mode == "search":
            focused = self.focused

            # --- Навигация между панелями ---
            if event.key == "left":
                # Стрелка влево - переключаемся на список
                if not self._is_on_list:
                    self.action_focus_list()
                    event.stop()
                    event.prevent_default()
                    return

            elif event.key == "right":
                # Стрелка вправо - переключаемся на первое поле ввода
                if self._is_on_list:
                    self.action_focus_first_input()
                    event.stop()
                    event.prevent_default()
                    return

            # --- Навигация внутри панелей ---
            if self._is_on_list:
                # Фокус на списке
                if event.key == "up":
                    # Перемещение вверх по списку
                    self.action_move_up()
                    event.stop()
                    event.prevent_default()
                    return
                elif event.key == "down":
                    # Перемещение вниз по списку
                    self.action_move_down()
                    event.stop()
                    event.prevent_default()
                    return
                elif event.key == "enter":
                    # Выбор элемента в списке
                    self.action_select_item()
                    event.stop()
                    event.prevent_default()
                    return
            else:
                # Фокус на полях ввода
                if isinstance(focused, Input):
                    if event.key == "up":
                        # Переход к предыдущему полю
                        self.action_prev_field()
                        event.stop()
                        event.prevent_default()
                        return
                    elif event.key == "down":
                        # Переход к следующему полю
                        self.action_next_field()
                        event.stop()
                        event.prevent_default()
                        return
                    elif event.key == "enter":
                        # Enter на поле ввода - переходим к следующему полю
                        self.action_next_field()
                        event.stop()
                        event.prevent_default()
                        return

        # В режиме просмотра
        elif self.mode == "view":
            focused = self.focused

            # --- Навигация между панелями ---
            if event.key == "left":
                # Стрелка влево - переключаемся на текст
                if not self._is_on_list:
                    self.action_focus_text()
                    event.stop()
                    event.prevent_default()
                    return

            elif event.key == "right":
                # Стрелка вправо - переключаемся на кнопки
                if self._is_on_list:
                    self.action_focus_buttons()
                    event.stop()
                    event.prevent_default()
                    return

            # --- Навигация внутри панелей ---
            if self._is_on_list:
                # Фокус на тексте просмотра - стрелки вверх/вниз для прокрутки
                if event.key == "up":
                    self.action_scroll_view_up()
                    event.stop()
                    event.prevent_default()
                    return
                elif event.key == "down":
                    self.action_scroll_view_down()
                    event.stop()
                    event.prevent_default()
                    return
            else:
                # Фокус на кнопках
                if isinstance(focused, Button):
                    if event.key == "up":
                        # Переход к предыдущей кнопке
                        self.action_prev_button()
                        event.stop()
                        event.prevent_default()
                        return
                    elif event.key == "down":
                        # Переход к следующей кнопке
                        self.action_next_button()
                        event.stop()
                        event.prevent_default()
                        return

    def action_next_field(self):
        """Переход к следующему полю ввода (для режима поиска)"""
        if self.mode != "search" or not self.input_fields:
            return

        # Если фокус на списке, переходим к первому полю ввода
        focused = self.focused
        if focused and hasattr(focused, 'id') and focused.id == "assessments_list":
            self.current_field_index = 0
            field_id = self.input_fields[self.current_field_index]
            try:
                self.query_one(f"#{field_id}").focus()
                self._is_on_list = False
            except NoMatches:
                pass
            return

        self.current_field_index = (self.current_field_index + 1) % len(self.input_fields)
        field_id = self.input_fields[self.current_field_index]
        try:
            self.query_one(f"#{field_id}").focus()
            self._is_on_list = False
        except NoMatches:
            pass

    def action_prev_field(self):
        """Переход к предыдущему полю ввода (для режима поиска)"""
        if self.mode != "search" or not self.input_fields:
            return

        # Если фокус на первом поле, переходим к списку
        focused = self.focused
        if focused and hasattr(focused, 'id'):
            if focused.id == self.input_fields[0]:
                self.action_focus_list()
                return

        self.current_field_index = (self.current_field_index - 1) % len(self.input_fields)
        field_id = self.input_fields[self.current_field_index]
        try:
            self.query_one(f"#{field_id}").focus()
            self._is_on_list = False
        except NoMatches:
            pass

    def action_next_button(self):
        """Переход к следующей кнопке (для режима просмотра)"""
        if self.mode != "view" or not self.view_buttons:
            return

        self.current_button_index = (self.current_button_index + 1) % len(self.view_buttons)
        button_id = self.view_buttons[self.current_button_index]
        try:
            self.query_one(f"#{button_id}").focus()
            self._is_on_list = False
        except NoMatches:
            pass

    def action_prev_button(self):
        """Переход к предыдущей кнопке (для режима просмотра)"""
        if self.mode != "view" or not self.view_buttons:
            return

        self.current_button_index = (self.current_button_index - 1) % len(self.view_buttons)
        button_id = self.view_buttons[self.current_button_index]
        try:
            self.query_one(f"#{button_id}").focus()
            self._is_on_list = False
        except NoMatches:
            pass

    @on(Input.Changed)
    def handle_input_changed(self, event: Input.Changed) -> None:
        """Обработка изменения текста в полях ввода"""
        if self.mode != "search":
            return

        field_id = event.input.id

        # Определяем базовое имя поля (без суффикса)
        if '_' in field_id:
            base_id = field_id.rsplit('_', 1)[0]
        else:
            base_id = field_id

        if base_id == "search_last_name":
            self._search_last_name = event.value
        elif base_id == "search_first_name":
            self._search_first_name = event.value
        elif base_id == "search_patronymic":
            self._search_patronymic = event.value
        elif base_id == "search_year":
            self._search_year = event.value
        elif base_id == "search_rater":
            self._search_rater = event.value

        # Фильтруем список
        self.filter_assessments()

    def filter_assessments(self):
        """Фильтрует оценки по введенным критериям"""
        last_name = self._search_last_name.lower().strip()
        first_name = self._search_first_name.lower().strip()
        patronymic = self._search_patronymic.lower().strip()
        year = self._search_year.strip()
        rater = self._search_rater.lower().strip()

        filtered = []
        for ass in self.all_assessments:
            patient = ass["patient"]
            assessment = ass["assessment"]

            # Проверяем каждое поле отдельно
            if last_name and last_name not in patient["last_name"].lower():
                continue

            if first_name and first_name not in patient["first_name"].lower():
                continue

            if patronymic and patronymic not in patient["patronymic"].lower():
                continue

            if year and year != patient["birth_year"]:
                continue

            if rater and rater not in assessment.get("rater", "").lower():
                continue

            # Если все проверки пройдены, добавляем в результат
            filtered.append(ass)

        self.filtered_assessments = filtered
        self.selected_index = 0
        self.update_list()

    def action_focus_list(self):
        """Перемещает фокус на список"""
        try:
            list_view = self.query_one("#assessments_list", ListView)
            list_view.focus()
            self._is_on_list = True
        except NoMatches:
            pass

    def action_focus_first_input(self):
        """Перемещает фокус на первое поле ввода"""
        if self.mode == "search" and self.input_fields:
            field_id = self.input_fields[0]
            try:
                self.query_one(f"#{field_id}").focus()
                self._is_on_list = False
            except NoMatches:
                pass

    def action_focus_text(self):
        """Перемещает фокус на текст в режиме просмотра"""
        try:
            view_text = self.query_one("#view_text", Static)
            view_text.focus()
            self._is_on_list = True
        except NoMatches:
            pass

    def action_focus_buttons(self):
        """Перемещает фокус на кнопки в режиме просмотра"""
        if self.mode == "view" and self.view_buttons:
            # Находим индекс кнопки "Назад к поиску" (она последняя)
            # Или ищем по ID, содержащему "btn_back"
            back_button_index = -1
            for i, btn_id in enumerate(self.view_buttons):
                if "btn_back" in btn_id:
                    back_button_index = i
                    break

            # Если кнопка "Назад" найдена - фокусируемся на ней
            # Иначе на первой кнопке
            if back_button_index != -1:
                self.current_button_index = back_button_index
                try:
                    self.query_one(f"#{self.view_buttons[back_button_index]}").focus()
                except NoMatches:
                    pass
            else:
                self.current_button_index = 0
                try:
                    self.query_one(f"#{self.view_buttons[0]}").focus()
                except NoMatches:
                    pass
            self._is_on_list = False

    def action_move_up(self):
        """Перемещение вверх по списку"""
        if self.mode == "search":
            focused = self.focused
            if focused and hasattr(focused, 'id') and focused.id == "assessments_list":
                list_view = focused
                if list_view.index > 0:
                    list_view.index -= 1
                    self.selected_index = list_view.index
                    self._update_list_display()

    def action_move_down(self):
        """Перемещение вниз по списку"""
        if self.mode == "search":
            focused = self.focused
            if focused and hasattr(focused, 'id') and focused.id == "assessments_list":
                list_view = focused
                if list_view.index < len(self.filtered_assessments) - 1:
                    list_view.index += 1
                    self.selected_index = list_view.index
                    self._update_list_display()

    def _update_list_display(self):
        """Обновляет отображение элементов списка"""
        try:
            list_view = self.query_one("#assessments_list", ListView)
            for i, child in enumerate(list_view.children):
                if hasattr(child, 'children') and child.children:
                    label = child.children
                    if isinstance(label, ListLabel):
                        label.is_highlighted = (i == list_view.index)
        except NoMatches:
            pass

    def action_select_item(self):
        """Выбор текущего элемента в списке"""
        if self.mode == "search":
            try:
                list_view = self.query_one("#assessments_list", ListView)
                if list_view.index is not None and list_view.index < len(self.filtered_assessments):
                    self.selected_index = list_view.index
                    self.show_selected_assessment()
            except NoMatches:
                pass

    def action_scroll_view_up(self):
        """Прокрутка панели просмотра вверх"""
        if self.mode == "view":
            try:
                list_scroll = self.query_one("#list_scroll")
                list_scroll.scroll_up()
            except NoMatches:
                pass

    def action_scroll_view_down(self):
        """Прокрутка панели просмотра вниз"""
        if self.mode == "view":
            try:
                list_scroll = self.query_one("#list_scroll")
                list_scroll.scroll_down()
            except NoMatches:
                pass

    @on(ListView.Selected)
    def handle_list_selected(self, event: ListView.Selected):
        self.action_select_item()

    @on(ListView.Highlighted, "#assessments_list")
    def handle_single_assessments_list_highlighted(self, event: ListView.Highlighted) -> None:
        """Обработчик перемещения фокуса для списка оценок"""

        if event.item:
            try:
                list_view = self.query_one("#assessments_list", ListView)

                # ВСЕГДА перемещаем маркер за курсором, независимо от режима (search или обычный)
                for i, child in enumerate(list_view.children):
                    if hasattr(child, 'children') and child.children:
                        label = child.children
                        if isinstance(label, ListLabel):
                            label.is_highlighted = (i == list_view.index)

                # Синхронизируем внутренний индекс класса с реальным выбором на экране
                if list_view.index is not None:
                    self.selected_index = list_view.index

            except NoMatches:
                pass

    @on(ListView.Highlighted, "#assessments_list")
    def handle_assessments_list_highlighted(self, event: ListView.Highlighted) -> None:
        """Обновляет отображение элементов списка при навигации"""
        if self.mode != "search":
            return

        if event.item:
            try:
                list_view = self.query_one("#assessments_list", ListView)

                # Обновляем маркеры у всех элементов
                for i, child in enumerate(list_view.children):
                    if hasattr(child, 'children') and child.children:
                        label = child.children[0]
                        if isinstance(label, ListLabel):
                            # Включаем маркер только для текущего выделенного элемента
                            label.is_highlighted = (i == list_view.index)

                # Синхронизируем внутренний индекс
                if list_view.index is not None:
                    self.selected_index = list_view.index

            except NoMatches:
                pass

    def show_selected_assessment(self):
        """Отображает выбранную оценку - файл слева, кнопки справа"""
        if self.selected_index >= len(self.filtered_assessments):
            return

        selected = self.filtered_assessments[self.selected_index]
        self.current_view_data = selected
        self.current_view_assessment = selected["assessment"]
        self.current_view_primary = selected["primary"]
        self.current_view_type = selected["assessment"]["type"]

        # Генерируем содержимое для просмотра
        content = self.generate_view_content(selected)

        # Создаем временный файл
        temp_dir = Path(tempfile.gettempdir())
        temp_file = temp_dir / f"arsenal_view_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        temp_file.write_text(content, encoding="utf-8")
        self.current_temp_file = temp_file

        # Обновляем левую панель - показываем содержимое файла
        try:
            # Левая панель теперь показывает содержимое файла
            list_panel = self.query_one("#look_list_panel")
            list_panel.border_title = f"{selected['display'][:80]}"

            # Убираем ListView и показываем текст
            list_scroll = self.query_one("#list_scroll")
            for child in list(list_scroll.children):
                child.remove()

            # Добавляем Static с содержимым файла
            view_text = Static(content, markup=False, id="view_text")
            list_scroll.mount(view_text)

            # Правая панель - кнопки действий
            container = self.query_one("#content_container")
            for child in list(container.children):
                child.remove()

            content_panel = self.query_one("#look_content_panel")
            content_panel.border_title = "Действия"

            # Создаем контейнер для кнопок
            buttons_container = Vertical(id="buttons_container")
            container.mount(buttons_container)

            # Добавляем кнопки с уникальными ID
            self.view_buttons = []

            btn_open_id = self._get_unique_id("btn_open")
            btn_open = Button("1. Открыть в редакторе     ", variant="primary", id=btn_open_id)
            buttons_container.mount(btn_open)
            self.view_buttons.append(btn_open_id)

            btn_restore_id = self._get_unique_id("btn_restore")
            btn_restore = Button("2. Записать в файл пациента", variant="success", id=btn_restore_id)
            buttons_container.mount(btn_restore)
            self.view_buttons.append(btn_restore_id)

            btn_draft_id = self._get_unique_id("btn_save_as_draft")
            btn_draft = Button("3. Сохранить как черновик  ", variant="primary", id=btn_draft_id)
            buttons_container.mount(btn_draft)
            self.view_buttons.append(btn_draft_id)

            btn_back_id = self._get_unique_id("btn_back")
            btn_back = Button("Esc. Назад к поиску        ", variant="default", id=btn_back_id)
            buttons_container.mount(btn_back)
            self.view_buttons.append(btn_back_id)

            self.mode = "view"
            self.current_button_index = 0
            self._is_on_list = True

            # Устанавливаем фокус на левую панель с текстом
            view_text.focus()

        except NoMatches:
            pass

    def generate_view_content(self, selected: dict) -> str:
        """Генерирует содержимое для просмотра оценки"""
        patient = selected["patient"]
        assessment = selected["assessment"]
        primary = selected["primary"]
        assessment_type = assessment["type"]

        last_name = patient["last_name"]
        first_name = patient["first_name"]
        patronymic = patient["patronymic"]
        birth_year = patient["birth_year"]

        date = assessment.get("date", "")
        rater = assessment.get("rater", "")

        content = []

        # Информация о пациенте
        content.append(f"Пациент {last_name} {first_name} {patronymic} {birth_year} г.р.")
        if assessment_type == "первичная":
            # Первичная оценка
            content.append("╭──────────────────────────────────────────────────────────────────────────────╮")
            content.append(f"Первичную оценку провел(а) {rater} {date}.")
            content.append("")
            content.append("     М О Д Р О П   А р с е н а л    ╭────" + date + "───╮")
            content.append("Факторы                             │ оцен  стад      │")

            # Факторы 1-8
            factor_names = [
                "[1] Агрессия                        ",
                "[2] Когнитивные и другие симптомы   ",
                "[3] Контроль над эмоциями           ",
                "[4] Контроль над поведением         ",
                "[5] Злоупотребление веществами      ",
                "[6] Приверженность режиму и лечению ",
                "[7] Личностные установки            ",
                "[8] Окружение, быт и планы          "
            ]

            total_score = 0
            for i in range(1, 9):
                f_key = f"f{i}"
                factor = assessment.get("factors", {}).get(f_key, {})
                score = factor.get("score", 0)
                stage = factor.get("stage")
                if stage == 5:
                    stage = None

                total_score += score

                score_bar = self.format_score_bar(score)
                stage_bar = self.format_stage_bar(stage)

                content.append(f"{factor_names[i-1]}│ {score_bar} {stage_bar} │[{i}]")

            content.append("                                    ╰─────────────────╯")

            # Общий балл
            content.append("[9] Всего из 24                     ╭────┬────────────────────────╮")
            total_bar = "█" * total_score + " " * (24 - total_score)
            content.append(f"                       {date}   │{total_score:3} │{total_bar}│")
            content.append("                                    ╰────┴────────────────────────╯")

            # Комментарии
            content.append("Заключение по первичной оценке:")
            for i in range(1, 9):
                f_key = f"f{i}"
                comment = assessment.get("factors", {}).get(f_key, {}).get("comment", "")
                if comment:
                    wrapped = textwrap.fill(comment, width=76)
                    content.append(f"[{i}] {wrapped}")

            if assessment.get("conclusion"):
                wrapped = textwrap.fill(assessment.get("conclusion", ""), width=76)
                content.append(f"[9] {wrapped}")

        else:  # повторная оценка
            if primary:
                primary_date = primary.get("date", "")
                content.append("╭──────────────────────────────────────────────────────────────────────────────╮")
                content.append(f"Повторную оценку провел(а) {rater} {date}.")
                content.append("")
                content.append("     М О Д Р О П   А р с е н а л    ╭────" + primary_date + "───╮╭─────" + date + "─────╮")
                content.append("Факторы                             │ оцен  стад      ││ стад      изм оцен │")

                # Собираем данные для отображения
                factor_names = [
                    "[1] Агрессия                        ",
                    "[2] Когнитивные и другие симптомы   ",
                    "[3] Контроль над эмоциями           ",
                    "[4] Контроль над поведением         ",
                    "[5] Злоупотребление веществами      ",
                    "[6] Приверженность режиму и лечению ",
                    "[7] Личностные установки            ",
                    "[8] Окружение, быт и планы          "
                ]

                primary_total = 0
                new_total = 0

                for i in range(1, 9):
                    f_key = f"f{i}"

                    # Первичные данные
                    primary_factor = primary.get("factors", {}).get(f_key, {})
                    primary_score = primary_factor.get("score", 0)
                    primary_stage = primary_factor.get("stage")
                    if primary_stage == 5:
                        primary_stage = None

                    # Повторные данные
                    new_factor = assessment.get("factors", {}).get(f_key, {})
                    new_score = new_factor.get("score", 0)
                    new_stage = new_factor.get("stage")
                    if new_stage == 5:
                        new_stage = None

                    # Вычисляем изменение
                    change = 0
                    if new_stage is not None and primary_stage is not None:
                        change = new_stage - primary_stage

                    change_display = "   "
                    if change <= -3:
                        change_display = "---"
                    elif change == -2:
                        change_display = "-- "
                    elif change == -1:
                        change_display = "-  "
                    elif change == 1:
                        change_display = "+  "
                    elif change == 2:
                        change_display = "++ "
                    elif change >= 3:
                        change_display = "+++"

                    primary_total += primary_score
                    new_total += new_score

                    primary_score_bar = self.format_score_bar(primary_score)
                    primary_stage_bar = self.format_stage_bar(primary_stage)
                    new_stage_bar = self.format_stage_bar(new_stage)
                    new_score_bar = self.format_score_bar(new_score)

                    content.append(f"{factor_names[i-1]}│ {primary_score_bar} {primary_stage_bar} ││ {new_stage_bar} {change_display} {new_score_bar}│[{i}]")

                content.append("                                    ╰─────────────────╯╰────────────────────╯")

                # Общие баллы
                content.append("[9] Всего из 24                     ╭────┬────────────────────────╮")
                content.append(f"                       {primary_date}   │{primary_total:3} │{self.format_total_bar(primary_total)}│")
                content.append("                                    ├────┼────────────────────────┤")
                content.append(f"                       {date}   │{new_total:3} │{self.format_total_bar(new_total)}│")
                content.append("                                    ╰────┴────────────────────────╯")

                # Комментарии
                content.append("Заключение по повторной оценке:")
                for i in range(1, 9):
                    f_key = f"f{i}"
                    comment = assessment.get("factors", {}).get(f_key, {}).get("comment", "")
                    if comment:
                        wrapped = textwrap.fill(comment, width=76)
                        content.append(f"[{i}] {wrapped}")

                if assessment.get("conclusion"):
                    wrapped = textwrap.fill(assessment.get("conclusion", ""), width=76)
                    content.append(f"[9] {wrapped}")
            else:
                content.append("ОШИБКА: Для повторной оценки не найдена первичная оценка")

        content.append("╰──────────────────────────────────────────────────────────────────────────────╯")

        return "\n".join(content)

    def format_score_bar(self, score: int) -> str:
        """Форматирует оценку в виде графической полоски"""
        if score is None:
            return "?    "
        bars = ["0 ▏  ", "1 ▒  ", "2 ▓▓ ", "3 ███"]
        if 0 <= score <= 3:
            return bars[score]
        return "?    "

    def format_stage_bar(self, stage: int) -> str:
        """Форматирует стадию в виде графической полоски"""
        if stage is None:
            return "         "
        bars = ["▁     пре", "▁▂    обд", "▁▂▄   под", "▁▂▄▆  дей", "▁▂▄▆█ уде"]
        if 0 <= stage <= 4:
            return bars[stage]
        return "         "

    def format_total_bar(self, total: int) -> str:
        """Форматирует общую сумму в виде полоски"""
        if total < 0 or total > 24:
            return "│" + " " * 24 + "│"
        return "█" * total + " " * (24 - total)

    @on(Button.Pressed)
    def handle_button(self, event: Button.Pressed) -> None:
        """Общий обработчик всех кнопок"""
        button_id = event.button.id

        # Определяем базовое имя кнопки (без суффикса)
        if '_' in button_id:
            base_id = button_id.rsplit('_', 1)[0]
        else:
            base_id = button_id

        if base_id == "btn_open":
            self.handle_open()
        elif base_id == "btn_restore":
            self.handle_restore()
        elif base_id == "btn_save_as_draft":
            self.handle_save_as_draft()
        elif base_id == "btn_back":
            self.handle_back_to_search()

    def handle_save_as_draft(self):
        """Сохраняет текущую оценку как черновик"""
        if not self.current_view_data:
            self.app.custom_notify("Нет данных для сохранения", severity="error")
            return

        selected = self.current_view_data
        assessment = selected.get("assessment", {})
        assessment_type = assessment.get("type", "")
        patient = selected.get("patient", {})

        # Проверяем, что это первичная или повторная оценка
        if assessment_type not in ["первичная", "повторная"]:
            self.app.custom_notify(f"Можно сохранить только первичную или повторную оценку (получен тип: {assessment_type})", severity="warning")
            return

        try:
            # Формируем данные черновика
            draft_data = {
                "assessment_type": assessment_type,
                "patient_info": {
                    "last_name": patient.get("last_name", ""),
                    "first_name": patient.get("first_name", ""),
                    "patronymic": patient.get("patronymic", ""),
                    "birth_year": str(patient.get("birth_year", ""))
                }
            }

            if assessment_type == "первичная":
                # Для первичной оценки
                draft_data.update({
                    "step_index": 25,  # Последний шаг - заключение
                    "current_factor": 8,
                    "current_substep": 2,
                    "form_data": {
                        "last_name": patient.get("last_name", ""),
                        "first_name": patient.get("first_name", ""),
                        "patronymic": patient.get("patronymic", ""),
                        "birth_year": str(patient.get("birth_year", "")),
                        "rater": assessment.get("rater", ""),
                        "factors": {},
                        "conclusion": assessment.get("conclusion", "")
                    }
                })

                # Заполняем факторы
                factors = assessment.get("factors", {})
                for i in range(1, 9):
                    f_key = f"f{i}"
                    factor = factors.get(f_key, {})
                    score = factor.get("score", 0)
                    stage = factor.get("stage")

                    # Преобразуем stage=5 (нет стадии) в None
                    if stage == 5:
                        stage = None

                    draft_data["form_data"]["factors"][f_key] = {
                        "score": score,
                        "stage": stage,
                        "comment": factor.get("comment", "")
                    }

            else:  # повторная оценка
                primary = selected.get("primary")
                if not primary:
                    self.app.custom_notify("Для повторной оценки не найдена первичная оценка", severity="error")
                    return

                # Формируем active_factors
                active_factors = []
                primary_factors = primary.get("factors", {})

                for i in range(1, 9):
                    f_key = f"f{i}"
                    primary_factor = primary_factors.get(f_key, {})
                    primary_score = primary_factor.get("score", 0)

                    if primary_score > 0:
                        primary_stage = primary_factor.get("stage")
                        if primary_stage == 5:
                            primary_stage = None

                        factor_info = {
                            "factor": i,
                            "primary_score": primary_score,
                            "primary_stage": primary_stage,
                            "primary_comment": primary_factor.get("comment", "")
                        }
                        active_factors.append(factor_info)

                # Формируем primary_assessments
                primary_assessment_info = {
                    "patient": patient,
                    "assessment": primary,
                    "display": f"{patient.get('last_name', '')} {patient.get('first_name', '')} "
                              f"{patient.get('patronymic', '')} {patient.get('birth_year', '')} г.р. - "
                              f"первичная - {primary.get('date', '')}",
                    "total_score": primary.get("total_score", 0)
                }

                draft_data.update({
                    "step_index": 2 + len(active_factors) * 2,  # Последний шаг
                    "current_factor_index": len(active_factors) - 1 if active_factors else 0,
                    "current_substep": 1,  # Комментарий
                    "form_data": {
                        "patient": patient,
                        "primary": primary,
                        "rater": assessment.get("rater", ""),
                        "factors": {},
                        "new_scores": {},
                        "conclusion": assessment.get("conclusion", "")
                    },
                    "selected_assessment": primary_assessment_info,
                    "primary_assessments": [primary_assessment_info],
                    "active_factors": active_factors
                })

                # Заполняем новые стадии и комментарии
                new_factors = assessment.get("factors", {})
                for factor_info in active_factors:
                    f_num = factor_info["factor"]
                    f_key = f"f{f_num}"
                    new_factor = new_factors.get(f_key, {})

                    new_stage = new_factor.get("stage")
                    if new_stage == 5:
                        new_stage = None

                    draft_data["form_data"]["factors"][f_key] = {
                        "stage": new_stage if new_stage is not None else factor_info["primary_stage"],
                        "comment": new_factor.get("comment", "")
                    }

            # Сохраняем черновик
            self.app.drafts.save_draft(draft_data)

            # Показываем уведомление об успехе
            patient_name = f"{patient.get('last_name', '')} {patient.get('first_name', '')}"
            if not patient_name.strip():
                patient_name = "Без имени"
            self.app.custom_notify(f"Черновик для {patient_name} сохранен", severity="success")

        except Exception as e:
            self.app.custom_notify(f"Ошибка сохранения черновика: {e}", severity="error")
            import traceback
            traceback.print_exc()

    def handle_open(self):
        """Открывает текущий просматриваемый файл во внешнем редакторе"""
        if hasattr(self, 'current_temp_file') and self.current_temp_file and self.current_temp_file.exists():
            open_file_externally(self.current_temp_file)
            self.app.custom_notify("Файл открыт во внешнем редакторе", severity="info")

    def handle_restore(self):
        """Восстанавливает запись об оценке в файл пациента"""
        if not self.current_view_data:
            return

        patient = self.current_view_data["patient"]
        assessment = self.current_view_data["assessment"]

        last_name = patient["last_name"]
        first_name = patient["first_name"]
        patronymic = patient["patronymic"]
        birth_year = patient["birth_year"]

        # Формируем имя файла
        filename_parts = [last_name, first_name, patronymic, birth_year]
        filename_base = " ".join(filter(None, filename_parts))
        filename_safe = filename_base.replace(" ", "_")
        filename = f"{filename_safe}.txt"
        filepath = self.app.results.reports_dir / filename

        # Генерируем содержимое
        content = self.generate_view_content(self.current_view_data)

        # Проверяем, существует ли файл
        if filepath.exists():
            # Читаем текущее содержимое
            current_content = filepath.read_text(encoding="utf-8")

            # Проверяем, нет ли уже такой оценки
            assessment_id = assessment.get("assessment_id", "")
            if assessment_id and assessment_id in current_content:
                self.app.custom_notify("Эта оценка уже есть в файле", severity="warning")
                return

            # Добавляем разделитель и новое содержание
            with open(filepath, "a", encoding="utf-8") as f:
                f.write("\n\n" + content)
        else:
            # Новый файл
            filepath.write_text(content, encoding="utf-8")

        self.app.custom_notify(f"Оценка восстановлена в файл:\n{filename}", severity="success")

    def handle_back_to_search(self):
        """Возврат к режиму поиска"""
        # Восстанавливаем левую панель - список оценок
        try:
            list_panel = self.query_one("#look_list_panel")
            list_panel.border_title = "Проведенные оценки"

            list_scroll = self.query_one("#list_scroll")
            for child in list(list_scroll.children):
                child.remove()

            # Восстанавливаем ListView
            list_view = ListView(id="assessments_list")
            list_scroll.mount(list_view)

            # Обновляем список
            self.update_list()

            # Показываем форму поиска
            self.show_search_form()

            # Устанавливаем фокус на список
            if list_view.children:
                list_view.focus()
                list_view.index = self.selected_index if self.selected_index < len(list_view.children) else 0

        except NoMatches:
            pass

    def action_go_back(self) -> None:
        """Возврат в главное меню"""
        self.app.pop_screen()



class StatsScreen(Screen):
    """Экран статистики по оценкам"""

    CSS = """
        #stats_list_panel {
            width: 70%;
            border: round $secondary;
            background: transparent;
            padding: 1 0 1 0;
            margin: 0;
            border-title-align: left;
            border-title-color: $secondary;
        }

        #stats_content_panel {
            width: 30%;
            border: round $secondary;
            background: transparent;
            padding: 0;
            margin: 0;
            border-title-align: left;
            border-title-color: $secondary;
        }

        #stats_list_panel:focus-within,
        #stats_content_panel:focus-within {
            border: round $accent;
            border-title-color: $accent;
            border-title-style: bold;
            background: transparent;
        }

        #stats_scroll {
            scrollbar-gutter: stable;
            scrollbar-size: 1 1;
            scrollbar-color: $accent 30%;
            scrollbar-color-hover: $accent 30%;
            scrollbar-background: transparent;
            background: transparent;
            padding: 0;
            margin: 0;
        }

        #stats_text {
            padding: 1 2 1 2;
            color: $text;
            text-wrap: wrap;
            width: 100%;
            background: transparent;
        }

        #stats_buttons_container {
            width: 100%;
            height: auto;
            align: left top;
            padding: 1 0 0 1;
        }

        #stats_buttons_container Button {
            width: 90%;
            height: 3;
            margin: 1 1;
            background: $background;
            color: $secondary;
            border: transparent;
            content-align: left middle;
            padding: 0 2;
        }

        #stats_buttons_container Button:focus {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
            content-align: left middle;
        }

        #stats_buttons_container Button:hover {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
        }

        #stats_buttons_container Button.variant-primary {
            border: round $accent;
            color: $accent;
        }

        #stats_buttons_container Button.variant-success {
            border: round green;
            color: green;
        }

        #stats_buttons_container Button.variant-default {
            border: round $secondary;
            color: $secondary;
        }

        #stats_buttons_container Button.spacer {
            height: 1;
            background: transparent;
            border: none;
            margin: 0;
            padding: 0;
            width: 100%;
        }

        #stats_buttons_container Button.spacer:focus {
            border: none;
            background: transparent;
        }

        #stats_buttons_container Button.spacer:hover {
            border: none;
            background: transparent;
        }

        #stats_buttons_container Button > Static {
            width: 100%;
            content-align: left middle;
        }
    """

    BINDINGS = [
        Binding("escape", "go_back", "Назад", show=True, priority=True),
        Binding("right", "focus_buttons", "К кнопкам", show=False),   # вправо - кнопки
        Binding("left", "focus_text", "К тексту", show=False),        # влево - текст
        Binding("up", "move_up", "Вверх", show=False),
        Binding("down", "move_down", "Вниз", show=False),
        Binding("pageup", "scroll_view_up", "Вверх", show=True),
        Binding("pagedown", "scroll_view_down", "Вниз", show=True),
        Binding("enter", "activate_button", "Выбрать", show=False),
        Binding("1", "open_editor", "Открыть", show=False),
        Binding("2", "save_history", "В историю", show=False),
        Binding("3", "open_history", "История", show=False),
    ]

    def __init__(self, show_history: bool = False):
        super().__init__()
        self.show_history = show_history
        self.temp_file = None
        self.stats_content = ""
        self.view_buttons = []
        self.current_button_index = 0
        self._form_counter = 0
        self.history_path = Path.home() / ".arsenal_data" / "arsenal_stat.txt"

    def _get_unique_id(self, base_id: str) -> str:
        """Генерирует уникальный ID для элементов"""
        self._form_counter += 1
        return f"{base_id}_{self._form_counter}"

    def compose(self) -> ComposeResult:
        with Horizontal():
            # Левая панель - содержимое статистики
            with Vertical(id="stats_list_panel", classes="panel-left") as list_panel:
                list_panel.border_title = "Статистика"
                with VerticalScroll(id="stats_scroll"):
                    yield Static("", id="stats_text", markup=False)

            # Правая панель - кнопки действий
            with Vertical(id="stats_content_panel", classes="panel-right") as content_panel:
                content_panel.border_title = "Действия"
                with VerticalScroll(id="stats_content_scroll"):
                    yield Vertical(id="stats_buttons_container")
        yield Footer(show_command_palette=False)

    def on_mount(self) -> None:
        """При монтировании загружаем статистику"""
        if self.show_history:
            self.load_history()
        else:
            self.generate_stats()

        self.show_buttons()

        # Устанавливаем фокус на текст
        try:
            stats_text = self.query_one("#stats_text", Static)
            stats_text.focus()
        except NoMatches:
            pass

    def _format_score_bar(self, score: int) -> str:
        """Форматирует оценку в виде графической полоски"""
        if score is None:
            return "?    "
        bars = ["0 ▏  ", "1 ▒  ", "2 ▓▓ ", "3 ███"]
        if 0 <= score <= 3:
            return bars[score]
        return "?    "

    def _format_stage_bar(self, stage: int) -> str:
        """Форматирует стадию в виде графической полоски"""
        if stage is None:
            return "         "
        bars = ["▁     пре", "▁▂    обд", "▁▂▄   под", "▁▂▄▆  дей", "▁▂▄▆█ уде"]
        if 0 <= stage <= 4:
            return bars[stage]
        return "         "

    def _format_total_bar(self, total: int) -> str:
        """Форматирует общую сумму в виде полоски"""
        if total < 0 or total > 24:
            return "│" + " " * 24 + "│"
        return "█" * total + " " * (24 - total)

    def _generate_comparison_block(self, assessments: list, assessment_type: str) -> list:
        """Генерирует блок сравнения оценок одного типа для одного пациента"""
        lines = []

        if not assessments:
            return lines

        # Сортируем по дате
        sorted_assessments = sorted(assessments, key=lambda a: a.get("date", ""))

        if assessment_type == "первичная":
            lines.append("        Первичные оценки            ╭─────────────────╮")
            lines.append("Факторы                             │ оцен  стад      │")

            # Собираем данные по факторам для сравнения
            factor_data = {}
            for assessment in sorted_assessments:
                rater = assessment.get("rater", "не указан")
                date = assessment.get("date", "")
                factors = assessment.get("factors", {})

                for i in range(1, 9):
                    f_key = f"f{i}"
                    factor = factors.get(f_key, {})
                    score = factor.get("score", 0)
                    stage = factor.get("stage")
                    if stage == 5:
                        stage = None

                    key = f"factor_{i}"
                    if key not in factor_data:
                        factor_data[key] = {
                            "factor_num": i,
                            "assessments": []
                        }

                    factor_data[key]["assessments"].append({
                        "score": score,
                        "stage": stage,
                        "rater": rater,
                        "date": date
                    })

            # Выводим только факторы, по которым есть различия
            factor_names = [
                "[1] Агрессия                        ",
                "[2] Когнитивные и другие симптомы   ",
                "[3] Контроль над эмоциями           ",
                "[4] Контроль над поведением         ",
                "[5] Злоупотребление веществами      ",
                "[6] Приверженность режиму и лечению ",
                "[7] Личностные установки            ",
                "[8] Окружение, быт и планы          "
            ]

            first_factor = True
            for key, data in factor_data.items():
                factor_num = data["factor_num"]
                # Проверяем, есть ли различия в оценках или стадиях
                scores = [a["score"] for a in data["assessments"]]
                stages = [a["stage"] for a in data["assessments"]]

                if len(set(scores)) > 1 or len(set(stages)) > 1:
                    # Добавляем разделитель перед каждым фактором, кроме первого
                    if not first_factor:
                        lines.append("                                    ├─────────────────┤")
                    else:
                        first_factor = False

                    # Выводим все оценки по этому фактору
                    for assessment_data in data["assessments"]:
                        score_bar = self._format_score_bar(assessment_data["score"])
                        stage_bar = self._format_stage_bar(assessment_data["stage"])
                        lines.append(f"{factor_names[factor_num - 1]}│ {score_bar} {stage_bar} │ "
                                     f"{assessment_data['rater']} {assessment_data['date']}")

            lines.append("                                    ╰─────────────────╯")

            # Сравнение общего балла
            lines.append("[9] Всего из 24         ╭────┬────────────────────────╮")
            for assessment in sorted_assessments:
                total_score = assessment.get("total_score", 0)
                rater = assessment.get("rater", "не указан")
                date = assessment.get("date", "")
                total_bar = self._format_total_bar(total_score)
                lines.append(f"                        │{total_score:3} │{total_bar}│ {rater} {date}")
            lines.append("                        ╰────┴────────────────────────╯")

        else:  # повторная оценка
            lines.append("        Повторные оценки            ╭────────────────────╮")
            lines.append("Факторы                             │ стад      изм оцен │")

            # Собираем данные по факторам для сравнения
            factor_data = {}
            for assessment in sorted_assessments:
                rater = assessment.get("rater", "не указан")
                date = assessment.get("date", "")
                factors = assessment.get("factors", {})

                for i in range(1, 9):
                    f_key = f"f{i}"
                    factor = factors.get(f_key, {})
                    stage = factor.get("stage")
                    if stage == 5:
                        stage = None
                    score = factor.get("score", 0)
                    change = factor.get("change", 0)

                    key = f"factor_{i}"
                    if key not in factor_data:
                        factor_data[key] = {
                            "factor_num": i,
                            "assessments": []
                        }

                    factor_data[key]["assessments"].append({
                        "stage": stage,
                        "score": score,
                        "change": change,
                        "rater": rater,
                        "date": date
                    })

            # Выводим только факторы, по которым есть различия
            factor_names = [
                "[1] Агрессия                        ",
                "[2] Когнитивные и другие симптомы   ",
                "[3] Контроль над эмоциями           ",
                "[4] Контроль над поведением         ",
                "[5] Злоупотребление веществами      ",
                "[6] Приверженность режиму и лечению ",
                "[7] Личностные установки            ",
                "[8] Окружение, быт и планы          "
            ]

            first_factor = True
            for key, data in factor_data.items():
                factor_num = data["factor_num"]
                # Проверяем, есть ли различия
                stages = [a["stage"] for a in data["assessments"]]
                scores = [a["score"] for a in data["assessments"]]
                changes = [a["change"] for a in data["assessments"]]

                if len(set(stages)) > 1 or len(set(scores)) > 1 or len(set(changes)) > 1:
                    # Добавляем разделитель перед каждым фактором, кроме первого
                    if not first_factor:
                        lines.append("                                    ├────────────────────┤")
                    else:
                        first_factor = False

                    for assessment_data in data["assessments"]:
                        stage_bar = self._format_stage_bar(assessment_data["stage"])
                        score_bar = self._format_score_bar(assessment_data["score"])

                        change = assessment_data["change"]
                        if change <= -3:
                            change_display = "---"
                        elif change == -2:
                            change_display = "-- "
                        elif change == -1:
                            change_display = "-  "
                        elif change == 0:
                            change_display = "   "
                        elif change == 1:
                            change_display = "+  "
                        elif change == 2:
                            change_display = "++ "
                        elif change >= 3:
                            change_display = "+++"
                        else:
                            change_display = "   "

                        lines.append(f"{factor_names[factor_num - 1]}│ {stage_bar} {change_display} {score_bar}│ "
                                     f"{assessment_data['rater']} {assessment_data['date']}")

            lines.append("                                    ╰────────────────────╯")

            # Сравнение общего балла
            lines.append("[9] Всего из 24            ╭────┬────────────────────────╮")
            for assessment in sorted_assessments:
                total_score = assessment.get("total_score", 0)
                rater = assessment.get("rater", "не указан")
                date = assessment.get("date", "")
                total_bar = self._format_total_bar(total_score)
                lines.append(f"                           │{total_score:3} │{total_bar}│ {rater} {date}")
            lines.append("                           ╰────┴────────────────────────╯")

        return lines

    def generate_stats(self) -> None:
        """Генерирует статистику из базы данных"""
        try:
            db_path = Path.home() / ".arsenal_data" / "arsenal_database.json"

            if not db_path.exists():
                self.stats_content = "База данных не найдена.\n\nПроведите хотя бы одну оценку."
                self.update_stats_display()
                return

            with open(db_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            if not data:
                self.stats_content = "База данных пуста.\n\nПроведите хотя бы одну оценку."
                self.update_stats_display()
                return

            # Собираем все оценки
            all_assessments = []
            patients_count = len(data)

            for patient in data:
                for assessment in patient.get("assessments", []):
                    all_assessments.append({
                        "patient": patient,
                        "assessment": assessment
                    })

            if not all_assessments:
                self.stats_content = "Нет проведенных оценок."
                self.update_stats_display()
                return

            # Сортируем оценки по дате
            all_assessments.sort(key=lambda x: x["assessment"].get("date", ""))

            # --- Формируем статистику ---
            lines = []
            lines.append("╭───────────────────────╮")
            lines.append("│ СТАТИСТИКА ПО ОЦЕНКАМ │")
            lines.append("╰───────────────────────╯")
            lines.append("")
            lines.append(f"Дата формирования: {datetime.now().strftime('%Y.%m.%d %H:%M:%S')}")
            lines.append("")

            # 1. Количество пациентов
            lines.append(f"  • Количество пациентов всего: {patients_count}")

            # 2. Возрастная статистика
            ages = []
            for item in all_assessments:
                patient = item["patient"]
                birth_year = patient.get("birth_year", "")
                assessment_date = item["assessment"].get("date", "")

                if birth_year and assessment_date:
                    try:
                        birth_year_int = int(birth_year) if isinstance(birth_year, (int, str)) and str(birth_year).isdigit() else 0
                        if birth_year_int > 1900:
                            date_parts = assessment_date.split(".")
                            if len(date_parts) == 3:
                                year = int(date_parts[0])
                                age = year - birth_year_int
                                if 14 <= age <= 113:
                                    ages.append({
                                        "age": age,
                                        "patient": patient,
                                        "date": assessment_date
                                    })
                    except (ValueError, TypeError):
                        pass

            if ages:
                avg_age = sum(a["age"] for a in ages) / len(ages)
                lines.append(f"  • Средний возраст пациента на дату оценки: {avg_age:.1f} лет")

                youngest = min(ages, key=lambda x: x["age"])
                lines.append(f"  • Самый молодой пациент: {youngest['patient'].get('last_name', '')} "
                            f"{youngest['patient'].get('first_name', '')} "
                            f"{youngest['patient'].get('patronymic', '')} "
                            f"({youngest['patient'].get('birth_year', '')} г.р.) - {youngest['age']} лет на {youngest['date']}")

                oldest = max(ages, key=lambda x: x["age"])
                lines.append(f"  • Самый возрастной пациент: {oldest['patient'].get('last_name', '')} "
                            f"{oldest['patient'].get('first_name', '')} "
                            f"{oldest['patient'].get('patronymic', '')} "
                            f"({oldest['patient'].get('birth_year', '')} г.р.) - {oldest['age']} лет на {oldest['date']}")
            else:
                lines.append("  • Нет данных для расчета возраста")

            lines.append("")

            # 3. Количество оценок
            total_assessments = len(all_assessments)
            primary_count = sum(1 for a in all_assessments if a["assessment"].get("type") == "первичная")
            repeat_count = sum(1 for a in all_assessments if a["assessment"].get("type") == "повторная")

            lines.append(f"  • Всего оценок: {total_assessments}")
            lines.append(f"    - Первичных: {primary_count}")
            lines.append(f"    - Повторных: {repeat_count}")
            lines.append("")

            # 4. Первая и последняя оценка
            first = all_assessments[0]
            last = all_assessments[-1]

            lines.append("  • Первая оценка:")
            lines.append(f"    - Дата: {first['assessment'].get('date', '')}")
            lines.append(f"    - Пациент: {first['patient'].get('last_name', '')} "
                        f"{first['patient'].get('first_name', '')} "
                        f"{first['patient'].get('patronymic', '')}")
            lines.append(f"    - Специалист: {first['assessment'].get('rater', '')}")
            lines.append("")
            lines.append("  • Последняя оценка:")
            lines.append(f"    - Дата: {last['assessment'].get('date', '')}")
            lines.append(f"    - Пациент: {last['patient'].get('last_name', '')} "
                        f"{last['patient'].get('first_name', '')} "
                        f"{last['patient'].get('patronymic', '')}")
            lines.append(f"    - Специалист: {last['assessment'].get('rater', '')}")
            lines.append("")

            # 5. Топ пациентов по максимальному баллу оценки
            lines.append("╭───────────────────────────────────╮")
            lines.append("│ ПАЦИЕНТЫ С МАКСИМАЛЬНЫМИ ОЦЕНКАМИ │")
            lines.append("╰───────────────────────────────────╯")
            lines.append("")

            patient_scores = {}
            for item in all_assessments:
                patient = item["patient"]
                uid = patient.get("uid", "")
                score = item["assessment"].get("total_score", 0)

                if uid not in patient_scores:
                    patient_scores[uid] = {
                        "patient": patient,
                        "max_score": score,
                        "assessment_date": item["assessment"].get("date", "")
                    }
                else:
                    # Если текущая оценка больше сохраненной, обновляем
                    if score > patient_scores[uid]["max_score"]:
                        patient_scores[uid]["max_score"] = score
                        patient_scores[uid]["assessment_date"] = item["assessment"].get("date", "")

            sorted_by_score = sorted(
                patient_scores.values(),
                key=lambda x: x["max_score"],
                reverse=True
            )

            for i, p in enumerate(sorted_by_score[:10], 1):
                patient = p["patient"]
                lines.append(f"  {i:2}. {patient.get('last_name', '')} {patient.get('first_name', '')} "
                            f"{patient.get('patronymic', '')} ({patient.get('birth_year', '')}) - "
                            f"сумма баллов {p['max_score']} (оценка от {p['assessment_date']})")

            lines.append("")

            # 6. Пациенты с минимальным суммарным баллом
            lines.append("╭──────────────────────────────────╮")
            lines.append("│ ПАЦИЕНТЫ С МИНИМАЛЬНЫМИ ОЦЕНКАМИ │")
            lines.append("╰──────────────────────────────────╯")
            lines.append("")

            # Используем те же данные patient_scores, сортируем по возрастанию
            sorted_by_score_asc = sorted(
                patient_scores.values(),
                key=lambda x: x["max_score"]
            )

            for i, p in enumerate(sorted_by_score_asc[:10], 1):
                patient = p["patient"]
                lines.append(f"  {i:2}. {patient.get('last_name', '')} {patient.get('first_name', '')} "
                            f"{patient.get('patronymic', '')} ({patient.get('birth_year', '')}) - "
                            f"сумма баллов {p['max_score']} (оценка от {p['assessment_date']})")

            lines.append("")

            # 7. Пациенты с лучшими стадиями изменения
            lines.append("╭───────────────────────────────────────╮")
            lines.append("│ ПАЦИЕНТЫ С ЛУЧШИМИ СТАДИЯМИ ИЗМЕНЕНИЯ │")
            lines.append("╰───────────────────────────────────────╯")
            lines.append("")

            patient_stages = {}
            stage_names = ["предобдумывание", "обдумывание", "подготовка", "действие", "удержание"]

            for item in all_assessments:
                patient = item["patient"]
                uid = patient.get("uid", "")

                if uid not in patient_stages:
                    patient_stages[uid] = {
                        "patient": patient,
                        "total_stage": 0,
                        "stages_count": 0
                    }

                factors = item["assessment"].get("factors", {})
                for i in range(1, 9):
                    f_key = f"f{i}"
                    factor = factors.get(f_key, {})
                    stage = factor.get("stage")
                    if stage is not None and stage != 5:
                        patient_stages[uid]["total_stage"] += stage
                        patient_stages[uid]["stages_count"] += 1

            patient_avg_stages = []
            for uid, data in patient_stages.items():
                if data["stages_count"] > 0:
                    avg = data["total_stage"] / data["stages_count"]
                    patient_avg_stages.append({
                        "patient": data["patient"],
                        "avg_stage": avg,
                        "stages_count": data["stages_count"]
                    })

            sorted_by_stage = sorted(
                patient_avg_stages,
                key=lambda x: x["avg_stage"],
                reverse=True
            )

            lines.append("  (усредненные стадии по всем факторам всех оценок)")
            lines.append("  (0 - предобдумывание, 4 - удержание)")
            lines.append("")
            for i, p in enumerate(sorted_by_stage[:10], 1):
                patient = p["patient"]
                stage_name = stage_names[int(round(p["avg_stage"]))] if 0 <= round(p["avg_stage"]) <= 4 else "-"
                lines.append(f"  {i:2}. {patient.get('last_name', '')} {patient.get('first_name', '')} "
                            f"{patient.get('patronymic', '')} ({patient.get('birth_year', '')}) - "
                            f"{p['avg_stage']:.2f} ({stage_name})")

            lines.append("")

            # 8. Статистика по специалистам
            lines.append("╭────────────────────────────╮")
            lines.append("│ СТАТИСТИКА ПО СПЕЦИАЛИСТАМ │")
            lines.append("╰────────────────────────────╯")
            lines.append("")

            rater_stats = {}
            for item in all_assessments:
                rater = item["assessment"].get("rater", "не указан")
                if rater not in rater_stats:
                    rater_stats[rater] = {
                        "assessments": [],
                        "primary_count": 0,
                        "repeat_count": 0
                    }
                rater_stats[rater]["assessments"].append(item)
                if item["assessment"].get("type") == "первичная":
                    rater_stats[rater]["primary_count"] += 1
                else:
                    rater_stats[rater]["repeat_count"] += 1

            for rater, stats in sorted(rater_stats.items()):
                assessments = stats["assessments"]
                assessments.sort(key=lambda x: x["assessment"].get("date", ""))

                total = len(assessments)
                primary = stats["primary_count"]
                repeat = stats["repeat_count"]

                lines.append(f"  • {rater}:")
                lines.append(f"    - Всего оценок: {total}")
                lines.append(f"    - Первичных: {primary}")
                lines.append(f"    - Повторных: {repeat}")

                if assessments:
                    first = assessments[0]
                    last = assessments[-1]
                    lines.append(f"    - Первая оценка: {first['assessment'].get('date', '')} "
                                f"({first['patient'].get('last_name', '')} {first['patient'].get('first_name', '')} {first['patient'].get('patronymic', '')})")
                    lines.append(f"    - Последняя оценка: {last['assessment'].get('date', '')} "
                                f"({last['patient'].get('last_name', '')} {last['patient'].get('first_name', '')} {last['patient'].get('patronymic', '')})")

                    if len(assessments) >= 2:
                        try:
                            first_date = first['assessment'].get('date', '')
                            last_date = last['assessment'].get('date', '')

                            if first_date and last_date:
                                first_parts = first_date.split('.')
                                last_parts = last_date.split('.')
                                if len(first_parts) == 3 and len(last_parts) == 3:
                                    first_dt = datetime(int(first_parts[0]), int(first_parts[1]), int(first_parts[2]))
                                    last_dt = datetime(int(last_parts[0]), int(last_parts[1]), int(last_parts[2]))
                                    days_diff = (last_dt - first_dt).days
                                    # if days_diff > 0:
                                    #     weeks = days_diff / 7
                                    #     avg_per_week = total / weeks
                                        # lines.append(f"    - Среднее количество оценок в неделю: {avg_per_week:.2f}")
                                    # else:
                                        # lines.append(f"    - Среднее количество оценок в неделю: {total:.2f} (все в один день)")
                        except (ValueError, TypeError):
                            pass
                lines.append("")

            # 9. Пациенты с несколькими специалистами
            lines.append("╭───────────────────────────────╮")
            lines.append("│ СРАВНЕНИЕ ОЦЕНОК СПЕЦИАЛИСТОВ │")
            lines.append("╰───────────────────────────────╯")
            lines.append("")
            patient_rater_stats = {}
            for item in all_assessments:
                patient = item["patient"]
                uid = patient.get("uid", "")
                rater = item["assessment"].get("rater", "не указан")

                if uid not in patient_rater_stats:
                    patient_rater_stats[uid] = {
                        "patient": patient,
                        "raters": set(),
                        "assessments_count": 0,
                        "primary_assessments": [],
                        "repeat_assessments": []
                    }
                patient_rater_stats[uid]["raters"].add(rater)
                patient_rater_stats[uid]["assessments_count"] += 1

                # Сортируем оценки по типу
                if item["assessment"].get("type") == "первичная":
                    patient_rater_stats[uid]["primary_assessments"].append(item["assessment"])
                else:
                    patient_rater_stats[uid]["repeat_assessments"].append(item["assessment"])

            multi_rater_patients = [
                p for p in patient_rater_stats.values()
                if len(p["raters"]) > 1
            ]

            if multi_rater_patients:
                lines.append(
                    f"  Пациентов, оценки которым проводили несколько специалистов: {len(multi_rater_patients)}")
                lines.append("")
                for p in multi_rater_patients:
                    patient = p["patient"]
                    raters_list = sorted(list(p["raters"]))
                    lines.append(f"  • {patient.get('last_name', '')} {patient.get('first_name', '')} "
                                 f"{patient.get('patronymic', '')} ({patient.get('birth_year', '')}) - "
                                 f"количество оценок: {p['assessments_count']}")
                    lines.append(f"     Специалисты: {', '.join(raters_list)}")
                    lines.append(" ")

                    # Детальное сравнение первичных оценок
                    if len(p["primary_assessments"]) > 1:
                        primary_block = self._generate_comparison_block(p["primary_assessments"], "первичная")
                        for line in primary_block:
                            lines.append(f"     {line}")
                        lines.append("")

                    # Детальное сравнение повторных оценок
                    if len(p["repeat_assessments"]) > 1:
                        repeat_block = self._generate_comparison_block(p["repeat_assessments"], "повторная")
                        for line in repeat_block:
                            lines.append(f"     {line}")
                        lines.append("")
            else:
                lines.append("  Нет пациентов, которым оценки проводили несколько специалистов")

            self.stats_content = "\n".join(lines)

            # Создаем временный файл
            temp_dir = Path(tempfile.gettempdir())
            self.temp_file = temp_dir / f"arsenal_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            self.temp_file.write_text(self.stats_content, encoding="utf-8")

        except Exception as e:
            self.stats_content = f"Ошибка при формировании статистики:\n\n{str(e)}"
            import traceback
            traceback.print_exc()

        # Отображаем в левой панели
        self.update_stats_display()

    def load_history(self) -> None:
        """Загружает историю статистики из файла"""
        try:
            if self.history_path.exists():
                self.stats_content = self.history_path.read_text(encoding="utf-8")
            else:
                self.stats_content = "Файл истории статистики не найден.\n\n"
                self.stats_content += "Сначала создайте статистику и запишите её в историю."
        except Exception as e:
            self.stats_content = f"Ошибка при чтении истории:\n\n{str(e)}"

        self.update_stats_display()

    def update_stats_display(self) -> None:
        """Обновляет отображение статистики в левой панели"""
        try:
            stats_text = self.query_one("#stats_text", Static)
            stats_text.update(self.stats_content)
        except NoMatches:
            pass

    def show_buttons(self) -> None:
        """Показывает кнопки действий в правой панели"""
        try:
            container = self.query_one("#stats_buttons_container")

            # Очищаем контейнер
            for child in list(container.children):
                child.remove()

            self.view_buttons = []

            if self.show_history:
                # В режиме просмотра истории - кнопка открыть файл и назад
                btn_open_id = self._get_unique_id("btn_open")
                btn_open = Button("1. Открыть в редакторе     ", variant="primary", id=btn_open_id)
                container.mount(btn_open)
                self.view_buttons.append(btn_open_id)

                btn_back_id = self._get_unique_id("btn_back")
                btn_back = Button("Esc. Назад                 ", variant="default", id=btn_back_id)
                container.mount(btn_back)
                self.view_buttons.append(btn_back_id)
            else:
                # Кнопки для режима статистики
                btn_open_id = self._get_unique_id("btn_open")
                btn_open = Button("1. Открыть в редакторе     ", variant="primary", id=btn_open_id)
                container.mount(btn_open)
                self.view_buttons.append(btn_open_id)

                btn_save_id = self._get_unique_id("btn_save_history")
                btn_save = Button("2. Записать в историю      ", variant="success", id=btn_save_id)
                container.mount(btn_save)
                self.view_buttons.append(btn_save_id)

                btn_history_id = self._get_unique_id("btn_open_history")
                btn_history = Button("3. Открыть историю         ", variant="primary", id=btn_history_id)
                container.mount(btn_history)
                self.view_buttons.append(btn_history_id)

                # Отступ
                spacer_id = self._get_unique_id("spacer")
                container.mount(Button("", disabled=True, classes="spacer", id=spacer_id))

                btn_back_id = self._get_unique_id("btn_back")
                btn_back = Button("Esc. Назад                 ", variant="default", id=btn_back_id)
                container.mount(btn_back)
                self.view_buttons.append(btn_back_id)

            self.current_button_index = 0

        except NoMatches:
            pass

    # --- Навигация ---

    def action_focus_buttons(self) -> None:
        """Переключение фокуса на кнопки"""
        if self.view_buttons:
            self.current_button_index = 0
            try:
                self.query_one(f"#{self.view_buttons[0]}").focus()
            except NoMatches:
                pass

    def action_focus_text(self) -> None:
        """Переключение фокуса на текст"""
        try:
            stats_text = self.query_one("#stats_text", Static)
            stats_text.focus()
        except NoMatches:
            pass

    def action_move_up(self) -> None:
        """Перемещение вверх по кнопкам"""
        focused = self.focused
        if isinstance(focused, Button) and self.view_buttons:
            current_index = -1
            for i, btn_id in enumerate(self.view_buttons):
                try:
                    if self.query_one(f"#{btn_id}").has_focus:
                        current_index = i
                        break
                except NoMatches:
                    pass

            if current_index >= 0:
                self.current_button_index = (current_index - 1) % len(self.view_buttons)
                try:
                    self.query_one(f"#{self.view_buttons[self.current_button_index]}").focus()
                except NoMatches:
                    pass

    def action_move_down(self) -> None:
        """Перемещение вниз по кнопкам"""
        focused = self.focused
        if isinstance(focused, Button) and self.view_buttons:
            current_index = -1
            for i, btn_id in enumerate(self.view_buttons):
                try:
                    if self.query_one(f"#{btn_id}").has_focus:
                        current_index = i
                        break
                except NoMatches:
                    pass

            if current_index >= 0:
                self.current_button_index = (current_index + 1) % len(self.view_buttons)
                try:
                    self.query_one(f"#{self.view_buttons[self.current_button_index]}").focus()
                except NoMatches:
                    pass

    def action_scroll_view_up(self) -> None:
        """Прокрутка текста вверх"""
        try:
            scroll = self.query_one("#stats_scroll")
            scroll.scroll_up()
        except NoMatches:
            pass

    def action_scroll_view_down(self) -> None:
        """Прокрутка текста вниз"""
        try:
            scroll = self.query_one("#stats_scroll")
            scroll.scroll_down()
        except NoMatches:
            pass

    def action_activate_button(self) -> None:
        """Активация текущей кнопки по Enter"""
        focused = self.focused
        if isinstance(focused, Button):
            event = Button.Pressed(focused)
            self.handle_button(event)

    def action_open_editor(self) -> None:
        """Открывает соответствующий файл во внешнем редакторе"""
        if self.show_history:
            # В режиме истории открываем файл arsenal_stat.txt
            if self.history_path.exists():
                open_file_externally(self.history_path)
                self.app.custom_notify("Файл истории статистики открыт во внешнем редакторе", severity="info")
            else:
                self.app.custom_notify("Файл истории статистики не найден", severity="warning")
        else:
            # В обычном режиме открываем временный файл с текущей статистикой
            if self.temp_file and self.temp_file.exists():
                open_file_externally(self.temp_file)
                self.app.custom_notify("Файл статистики открыт во внешнем редакторе", severity="info")
            else:
                self.app.custom_notify("Сначала сформируйте статистику", severity="warning")

    def action_save_history(self) -> None:
        """Сохраняет статистику в историю"""
        try:
            self.history_path.parent.mkdir(exist_ok=True)

            timestamp = datetime.now().strftime('%Y.%m.%d %H:%M:%S')
            separator = "─" * 80

            entry = f"\n\n{separator}\n"
            entry += f"ЗАПИСЬ СТАТИСТИКИ ОТ {timestamp}\n"
            entry += f"{separator}\n\n"
            entry += self.stats_content

            if self.history_path.exists():
                with open(self.history_path, "a", encoding="utf-8") as f:
                    f.write(entry)
            else:
                with open(self.history_path, "w", encoding="utf-8") as f:
                    f.write(entry)

            self.app.custom_notify("Статистика записана в историю", severity="success")

        except Exception as e:
            self.app.custom_notify(f"Ошибка сохранения истории: {e}", severity="error")
            import traceback
            traceback.print_exc()

    def action_open_history(self) -> None:
        """Открывает экран просмотра истории"""
        self.app.push_screen(StatsScreen(show_history=True))

    def action_go_back(self) -> None:
        """Возврат в ResearchScreen"""
        self.app.pop_screen()

    @on(Button.Pressed)
    def handle_button(self, event: Button.Pressed) -> None:
        """Обработчик нажатия кнопок"""
        button_id = event.button.id

        # Определяем базовое имя кнопки
        if '_' in button_id:
            base_id = button_id.rsplit('_', 1)[0]
        else:
            base_id = button_id

        if base_id == "btn_open":
            self.action_open_editor()
        elif base_id == "btn_save_history":
            self.action_save_history()
        elif base_id == "btn_open_history":
            self.action_open_history()
        elif base_id == "btn_back":
            self.action_go_back()

    def on_key(self, event: events.Key) -> None:
        """Обработчик нажатий клавиш"""
        if event.key == "enter":
            focused = self.focused
            if isinstance(focused, Button):
                # Кнопка сама обработает Enter
                pass
            elif self.view_buttons:
                # Если фокус на тексте, переключаемся на первую кнопку
                self.action_focus_buttons()
                event.stop()
                event.prevent_default()

        elif event.key == "up" or event.key == "down":
            focused = self.focused
            if isinstance(focused, Button) and self.view_buttons:
                if event.key == "up":
                    self.action_move_up()
                else:
                    self.action_move_down()
                event.stop()
                event.prevent_default()

        elif event.key == "right":  # Стрелка вправо - на кнопки
            self.action_focus_buttons()
            event.stop()
            event.prevent_default()

        elif event.key == "left":   # Стрелка влево - на текст
            self.action_focus_text()
            event.stop()
            event.prevent_default()


class RocAnalysis:
    """Класс для проведения ROC-анализа на основе данных из arsenal_database.json"""

    def __init__(self, data_manager):
        self.data_manager = data_manager
        self.results = None
        self.coefs = None
        self.intercept = None
        self.roc_data = None
        self.roc_data_with_risk = None

    def prepare_data(self, months: int = 6) -> pd.DataFrame:
        """Готовит данные для ROC-анализа из базы данных.
        Исход считается положительным, если он наступил в течение `months` месяцев ПОСЛЕ оценки.
        """
        data = self.data_manager.data
        rows = []

        for patient in data:
            uid = patient.get('uid', '')
            last_name = patient.get('last_name', '')
            first_name = patient.get('first_name', '')
            patronymic = patient.get('patronymic', '')
            birth_year = patient.get('birth_year', '')
            fiogr = f"{last_name} {first_name} {patronymic} {birth_year}".strip()

            outcomes = patient.get('outcomes', [])

            for assessment in patient.get('assessments', []):
                assessment_dt = self._parse_date(assessment.get('date', ''))
                if assessment_dt is None:
                    continue

                has_outcome = False
                for outcome in outcomes:
                    outcome_dt = self._parse_date(outcome.get('date', ''))
                    if outcome_dt is None:
                        continue

                    # Исход в течение 6 месяцев ПОСЛЕ оценки
                    if assessment_dt < outcome_dt:
                        delta_months = (
                            (outcome_dt.year - assessment_dt.year) * 12
                            + (outcome_dt.month - assessment_dt.month)
                        )
                        if delta_months <= months:
                            has_outcome = True
                            break

                factors = assessment.get('factors', {})
                total_score = assessment.get('total_score', 0)

                # Расчет Total Stage - ПРАВИЛЬНЫЙ МАППИНГ
                stages = []
                for i in range(1, 9):
                    factor_key = f'f{i}'
                    if factor_key in factors:
                        stage = factors[factor_key].get('stage', 0)
                    else:
                        stage = 0
                    stages.append(stage)

                # stage: 0=предобдумывание(4), 1=обдумывание(3), 2=подготовка(2),
                #        3=действие(1), 4=удержание(0), 5=не определена(0)
                stage_to_risk = {
                    0: 4,  # предобдумывание ➜ высокий риск
                    1: 3,  # обдумывание
                    2: 2,  # подготовка
                    3: 1,  # действие
                    4: 0,  # удержание ➜ низкий риск
                    5: 0   # не определена
                }
                converted_values = [stage_to_risk.get(stage, 0) for stage in stages]
                total_stage = sum(converted_values)
                arsenal_index = total_score + total_stage

                rows.append({
                    'uid': uid,
                    'fiogr': fiogr,
                    'type': assessment.get('type', ''),
                    'date': assessment.get('date', ''),
                    'total_score': total_score,
                    'total_stage': total_stage,
                    'arsenal_index': arsenal_index,
                    'outcome': 1 if has_outcome else 0,
                })

        self.roc_data = pd.DataFrame(rows)
        return self.roc_data

    @staticmethod
    def _parse_date(date_str):
        """Разбирает дату из YYYY.MM.DD или других форматов."""
        if not date_str:
            return None
        date_str = str(date_str).strip()
        for fmt in ("%Y.%m.%d", "%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        return None

    def run_analysis(self, months: int = 6) -> dict:
        """Запускает полный ROC-анализ, включая прогнозы специалистов."""
        df = self.prepare_data(months=months)
        opinion_df = self.prepare_opinion_data(months=months)

        self.opinion_data = opinion_df

        if len(df) == 0:
            return {'error': 'Нет данных для анализа'}

        positive = int(df['outcome'].sum())
        negative = len(df) - positive

        if positive == 0 or negative == 0:
            return {
                'error': (
                    f"Для ROC-анализа нужны оба класса исходов. "
                    f"Сейчас положительных исходов: {positive}, отрицательных: {negative}. "
                    "Проверьте данные об исходах (должны быть даты в пределах 6 месяцев после оценки)."
                )
            }

        results = {}

        for display_name, column in (
            ('Total Score', 'total_score'),
            ('Total Stage', 'total_stage'),
            ('Arsenal Index', 'arsenal_index'),
        ):
            result = self.calculate_roc_with_bootstrap(df, column)
            if result[0] is not None:
                results[display_name] = {
                    'auc': result[3],
                    'ci': result[4],
                    'threshold': result[5],
                    'sensitivity': result[6],
                    'specificity': result[7],
                    'fpr': result[0],
                    'tpr': result[1],
                }

        # Анализ прогнозов специалистов
        if len(opinion_df) > 0:
            opinion_positive = int(opinion_df['outcome'].sum())
            opinion_negative = len(opinion_df) - opinion_positive

            if opinion_positive > 0 and opinion_negative > 0:
                result = self.calculate_roc_with_bootstrap(opinion_df, 'risk_score')
                if result[0] is not None:
                    results['Прогнозы специалистов'] = {
                        'auc': result[3],
                        'ci': result[4],
                        'threshold': result[5],
                        'sensitivity': result[6],
                        'specificity': result[7],
                        'fpr': result[0],
                        'tpr': result[1],
                        'n_opinions': len(opinion_df),
                        'n_positive': opinion_positive,
                    }

        # Логистическая регрессия
        try:
            combined_pred, coefs_raw, intercept_raw, scaler, lr_model = self.create_combined_predictor(df)
            self.coefs = coefs_raw
            self.intercept = intercept_raw

            fprs_lr, tprs_lr, thresholds_lr = roc_curve(df['outcome'], combined_pred)
            auc_lr = auc(fprs_lr, tprs_lr)

            boot_aucs_lr = []
            n = len(df)
            for _ in range(500):
                idx = np.random.choice(n, n, replace=True)
                boot_data = df.iloc[idx]
                try:
                    pred_boot, _, _, _, _ = self.create_combined_predictor(boot_data)
                    boot_auc = roc_auc_score(boot_data['outcome'], pred_boot)
                    boot_aucs_lr.append(boot_auc)
                except Exception:
                    continue

            if boot_aucs_lr:
                ci_lr = (np.percentile(boot_aucs_lr, 2.5), np.percentile(boot_aucs_lr, 97.5))
            else:
                ci_lr = (auc_lr, auc_lr)

            results['Logistic Regression'] = {
                'auc': auc_lr,
                'ci': ci_lr,
                'coefs': coefs_raw,
                'intercept': intercept_raw,
                'combined_pred': combined_pred,
                'fpr': fprs_lr,
                'tpr': tprs_lr,
            }

            self.roc_data_with_risk = self.calculate_individual_risks(df, coefs_raw, intercept_raw)
        except Exception as e:
            print(f"Ошибка при построении логистической регрессии: {e}")

        self.results = results
        return results

    def calculate_roc_with_bootstrap(self, df, score_column, outcome_column='outcome', n_bootstrap=1000):
        """Расчет ROC-кривой с бутстреп-интервалами для AUC."""
        valid_data = df[[score_column, outcome_column]].dropna()

        if len(valid_data) == 0:
            return None, None, None, 0, (0, 0), 0, 0, 0

        fprs, tprs, thresholds = roc_curve(valid_data[outcome_column], valid_data[score_column])
        roc_auc = auc(fprs, tprs)

        boot_aucs = []
        n = len(valid_data)
        for _ in range(n_bootstrap):
            idx = np.random.choice(n, n, replace=True)
            boot_data = valid_data.iloc[idx]
            try:
                boot_auc = roc_auc_score(boot_data[outcome_column], boot_data[score_column])
                boot_aucs.append(boot_auc)
            except Exception:
                continue

        if boot_aucs:
            ci_lower = np.percentile(boot_aucs, 2.5)
            ci_upper = np.percentile(boot_aucs, 97.5)
        else:
            ci_lower, ci_upper = roc_auc, roc_auc

        youden_j = tprs - fprs
        optimal_idx = np.argmax(youden_j)
        optimal_threshold = thresholds[optimal_idx] if len(thresholds) > optimal_idx else 0
        optimal_sensitivity = tprs[optimal_idx] if len(tprs) > optimal_idx else 0
        optimal_specificity = 1 - fprs[optimal_idx] if len(fprs) > optimal_idx else 0

        return fprs, tprs, thresholds, roc_auc, (ci_lower, ci_upper), optimal_threshold, optimal_sensitivity, optimal_specificity

    def create_combined_predictor(self, df):
        """Создает комбинированный предиктор с помощью логистической регрессии."""
        X = df[['total_score', 'total_stage', 'arsenal_index']].values
        y = df['outcome'].values

        if len(set(y)) < 2:
            raise ValueError("Для логистической регрессии нужны оба класса исходов")

        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        lr = LogisticRegression(penalty='l2', C=1.0, max_iter=1000, random_state=42)
        lr.fit(X_scaled, y)
        y_pred_prob = lr.predict_proba(X_scaled)[:, 1]

        lr_raw = LogisticRegression(penalty='l2', C=1.0, max_iter=1000, random_state=42)
        lr_raw.fit(X, y)
        coefs_raw = dict(zip(['total_score', 'total_stage', 'arsenal_index'], lr_raw.coef_[0]))
        intercept_raw = lr_raw.intercept_[0]

        return y_pred_prob, coefs_raw, intercept_raw, scaler, lr

    def calculate_individual_risks(self, df, coefs, intercept):
        """Рассчитывает индивидуальную вероятность риска для каждого пациента."""
        df_copy = df.copy()

        probabilities = []
        for _, row in df_copy.iterrows():
            z = (coefs['total_score'] * row['total_score'] +
                 coefs['total_stage'] * row['total_stage'] +
                 coefs['arsenal_index'] * row['arsenal_index'] +
                 intercept)
            prob = 1 / (1 + np.exp(-z)) * 100
            probabilities.append(prob)

        df_copy['risk_probability'] = probabilities

        def get_risk_category(prob):
            if prob < 20:
                return "Низкий"
            elif prob < 40:
                return "Средний"
            elif prob < 60:
                return "Повышенный"
            elif prob < 80:
                return "Высокий"
            else:
                return "Критический"

        df_copy['risk_category'] = df_copy['risk_probability'].apply(get_risk_category)
        return df_copy

    def generate_report(self, months: int = 6) -> str:
        """Генерирует текстовый отчет с результатами анализа и ASCII-графиком."""
        if self.roc_data is None:
            return "Сначала проведите ROC-анализ."

        lines = []
        lines.append("─" * 80)
        lines.append("Отчет по ROC-анализу")
        lines.append(f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("─" * 80)
        lines.append("Период учета исходов: 6 месяцев после оценки")
        lines.append("")
        lines.append("ROC-анализ сопоставляет результаты проведенных оценок с тем, было ли отмечено")
        lines.append("у данного пациента опасное поведение в последующие 6 месяцев.")
        lines.append("Для этого сопоставления принимаются по отдельности:")
        lines.append("   1. Сумма баллов по каждой оценке (Оценки)")
        lines.append("   2. Условно исчисленная сумма стадий (Стадии)")
        lines.append("        Для этих целей определенная стадия изменения по каждому фактору")
        lines.append("        пересчитывается в баллы: стадия предобдумывания - 4, обдумывания - 3,")
        lines.append("        подготовки - 2, действия - 1, удержания - 0.")
        lines.append("   3. Сумма первых двух показателей Оценки + Стадии (Сумма)")
        lines.append("Дополнительно проводится логистическая регрессия.")
        lines.append("")
        lines.append(f"Объем выборки: {len(self.roc_data)}")
        lines.append(f"Положительные исходы: {int(self.roc_data['outcome'].sum())}")
        lines.append(f"Отрицательные исходы: {len(self.roc_data) - int(self.roc_data['outcome'].sum())}")
        # Информация о прогнозах специалистов
        if hasattr(self, 'opinion_data') and self.opinion_data is not None:
            if not self.opinion_data.empty and 'outcome' in self.opinion_data.columns:
                lines.append("")
                lines.append("Прогнозы специалистов")
                lines.append(f"Количество прогнозов: {len(self.opinion_data)}")
                lines.append(f"Положительные исходы: {int(self.opinion_data['outcome'].sum())}")
                lines.append(
                    f"Отрицательные исходы: {len(self.opinion_data) - int(self.opinion_data['outcome'].sum())}")
            else:
                lines.append("")
                lines.append("Прогнозы специалистов: нет данных")
        lines.append("")

        if self.results:
            lines.append("-" * 80)
            lines.append("AUC метрики с 95% доверительными интервалами")
            lines.append("-" * 80)

            # Словарь для перевода названий на русский
            name_mapping = {
                'Total Score': 'Оценки',
                'Total Stage': 'Стадии',
                'Arsenal Index': 'Сумма',
                'Logistic Regression': 'Логистическая регрессия',
            }

            sorted_results = sorted(
                [(name, data) for name, data in self.results.items()
                 if isinstance(data, dict) and 'auc' in data],
                key=lambda x: x[1]['auc'],
                reverse=True,
            )

            for name, data in sorted_results:
                display_name = name_mapping.get(name, name)
                ci_low, ci_high = data.get('ci', (data['auc'], data['auc']))
                lines.append(f"{display_name:30} │ AUC: {data['auc']:.4f} │ 95% CI: [{ci_low:.4f}, {ci_high:.4f}]")

            if len(sorted_results) > 1:
                lines.append("")
                lines.append("-" * 80)
                lines.append("Сравнение показателей (разница AUC)")
                lines.append("-" * 80)

                best_name, best_data = sorted_results[0]
                best_display = name_mapping.get(best_name, best_name)
                for name, data in sorted_results[1:]:
                    display_name = name_mapping.get(name, name)
                    diff = best_data['auc'] - data['auc']
                    lines.append(f"{best_display} vs {display_name}: {diff:+.4f}")

            # Оптимальные пороги
            threshold_rows = [(name, data) for name, data in self.results.items()
                              if isinstance(data, dict) and 'threshold' in data]
            if threshold_rows:
                lines.append("")
                lines.append("-" * 80)
                lines.append("Оптимальные пороги (критерий Юдена)")
                lines.append("-" * 80)

                # Используем русские названия в таблице порогов
                for name, data in threshold_rows:
                    display_name = name_mapping.get(name, name)
                    youden = data.get('sensitivity', 0) + data.get('specificity', 0) - 1
                    lines.append(
                        f"{display_name:21} │ Порог: {int(data['threshold']):>2} │ "
                        f"Чувств.: {data['sensitivity']:.3f} │ "
                        f"Специф.: {data['specificity']:.3f} │ J: {youden:.3f}"
                    )

            # Формула риска
            if self.coefs is not None and self.intercept is not None:
                lines.append("")
                lines.append("-" * 80)
                lines.append("Формула расчета риска")
                lines.append("-" * 80)

                formula_z = f"Z = {self.coefs['total_score']:.4f} × Оценки"
                if self.coefs['total_stage'] >= 0:
                    formula_z += f" + {self.coefs['total_stage']:.4f} × Стадии"
                else:
                    formula_z += f" - {abs(self.coefs['total_stage']):.4f} × Стадии"
                if self.coefs['arsenal_index'] >= 0:
                    formula_z += f" + {self.coefs['arsenal_index']:.4f} × Сумма"
                else:
                    formula_z += f" - {abs(self.coefs['arsenal_index']):.4f} × Сумма"
                formula_z += f" + ({self.intercept:.4f})"

                lines.append("")
                lines.append("1. Промежуточный индекс Z:")
                lines.append(f"   {formula_z}")
                lines.append("")
                lines.append("2. Риск (P) в процентах:")
                lines.append("   P = 1 / (1 + e^(-Z)) × 100%")

                lines.append("")
                lines.append("-" * 80)
                lines.append("Интерпретация коэффициентов:")
                lines.append("-" * 80)

                # Русские названия для коэффициентов
                coef_mapping = {
                    'total_score': 'Оценки',
                    'total_stage': 'Стадии',
                    'arsenal_index': 'Сумма',
                }
                for name, coef in self.coefs.items():
                    display_name = coef_mapping.get(name, name.replace('_', ' ').title())
                    or_value = float(np.exp(coef))
                    lines.append(f"  {display_name}:")
                    lines.append(f"    Коэффициент: {coef:.4f}")
                    lines.append(f"    Отношение шансов (OR): {or_value:.4f}")

                lines.append("")
                lines.append("-" * 80)
                lines.append("Категории риска:")
                lines.append("-" * 80)
                lines.append("  < 20%  - Низкий")
                lines.append("  20-40% - Средний")
                lines.append("  40-60% - Повышенный")
                lines.append("  60-80% - Высокий")
                lines.append("  > 80%  - Критический")

            if self.roc_data_with_risk is not None:
                lines.append("")
                lines.append("-" * 80)
                lines.append("Показатели риска по пациентам")
                lines.append("-" * 80)

                sorted_data = self.roc_data_with_risk.sort_values('risk_probability', ascending=False)

                # Таблица
                lines.append(f"{'№':<3} {'ФИО':<35} {'Оцен':<5} {'Стад':<5} {'Сумм':<6} {'Риск':<6} {'Кат':<10} {'Исх'}")
                lines.append("-" * 80)

                for idx, (_, row) in enumerate(sorted_data.iterrows(), 1):
                    name = row['fiogr']
                    if len(name) > 35:
                        name = name[:32] + "..."
                    outcome_marker = "!" if row['outcome'] == 1 else ""
                    lines.append(
                        f"{idx:<3} {name:<35} {row['total_score']:<5} {row['total_stage']:<5} "
                        f"{row['arsenal_index']:<5} {row['risk_probability']:>5.1f}%  "
                        f"{row['risk_category']:<11} {outcome_marker}"
                    )

                lines.append("")
                lines.append("-" * 80)
                lines.append("Статистика по группам риска:")
                lines.append("-" * 80)

                risk_stats = self.roc_data_with_risk.groupby('risk_category').agg({
                    'fiogr': 'count',
                    'outcome': 'sum',
                }).rename(columns={'fiogr': 'count', 'outcome': 'positive'})
                risk_stats['positive_pct'] = (risk_stats['positive'] / risk_stats['count'] * 100).round(1)

                category_order = ['Низкий', 'Средний', 'Повышенный', 'Высокий', 'Критический']
                for category in category_order:
                    if category in risk_stats.index:
                        row = risk_stats.loc[category]
                        lines.append(
                            f"  {category:12} : {int(row['count']):3} оценок, "
                            f"из них с исходом: {int(row['positive']):3} ({row['positive_pct']:.1f}%)"
                        )

            # Получаем данные прогнозов
            if hasattr(self, 'opinion_data') and self.opinion_data is not None:
                if not self.opinion_data.empty:
                    opinion_lines = self.generate_opinions_table(self.opinion_data)
                    lines.extend(opinion_lines)

            # --- ASCII-график ROC-кривых ---
            lines.append("")
            lines.append("-" * 80)
            lines.append("ROC-кривые")
            lines.append("-" * 80)

            # Формируем данные для графика с русскими названиями
            curves = {}
            markers = {
                'Оценки': '▒',
                'Стадии': '▓',
                'Сумма': '█',
                'Логистическая регрессия': '#',
            }

            for name, data in self.results.items():
                if isinstance(data, dict) and 'fpr' in data and 'tpr' in data:
                    display_name = name_mapping.get(name, name)
                    curves[display_name] = {
                        'fpr': data['fpr'],
                        'tpr': data['tpr'],
                        'auc': data.get('auc', 0),
                        'color': 'white',
                        'marker': markers.get(display_name, '•'),
                    }

            graph = build_roc_ascii_graph(curves)
            if graph:
                lines.append(graph)

        lines.append("")
        lines.append("-" * 80)

        return "\n".join(lines)

    def prepare_opinion_data(self, months: int = 6) -> pd.DataFrame:
        """Готовит данные для ROC-анализа прогнозов специалистов."""
        data = self.data_manager.data
        rows = []

        for patient in data:
            uid = patient.get('uid', '')
            last_name = patient.get('last_name', '')
            first_name = patient.get('first_name', '')
            patronymic = patient.get('patronymic', '')
            birth_year = patient.get('birth_year', '')
            fiogr = f"{last_name} {first_name} {patronymic} {birth_year}".strip()

            outcomes = patient.get('outcomes', [])
            opinions = patient.get('opinions', [])

            for opinion in opinions:
                opinion_dt = self._parse_date(opinion.get('date', ''))
                if opinion_dt is None:
                    continue

                risk_score = opinion.get('risk_score', 0)

                has_outcome = False
                for outcome in outcomes:
                    outcome_dt = self._parse_date(outcome.get('date', ''))
                    if outcome_dt is None:
                        continue

                    if opinion_dt < outcome_dt:
                        delta_months = (
                            (outcome_dt.year - opinion_dt.year) * 12
                            + (outcome_dt.month - opinion_dt.month)
                        )
                        if delta_months <= months:
                            has_outcome = True
                            break

                rows.append({
                    'uid': uid,
                    'fiogr': fiogr,
                    'expert': opinion.get('expert', ''),
                    'date': opinion.get('date', ''),
                    'risk_score': risk_score,
                    'outcome': 1 if has_outcome else 0,
                })

        self.opinion_data = pd.DataFrame(rows)
        return self.opinion_data

    def prepare_opinion_data_for_table(self, months: int = 6) -> pd.DataFrame:
        """
        Подготавливает данные прогнозов специалистов для таблицы.
        """
        data = self.data_manager.data
        rows = []

        for patient in data:
            uid = patient.get('uid', '')
            last_name = patient.get('last_name', '')
            first_name = patient.get('first_name', '')
            patronymic = patient.get('patronymic', '')
            birth_year = patient.get('birth_year', '')
            fiogr = f"{last_name} {first_name} {patronymic} {birth_year}".strip()

            outcomes = patient.get('outcomes', [])
            opinions = patient.get('opinions', [])

            for opinion in opinions:
                opinion_dt = self._parse_date(opinion.get('date', ''))
                if opinion_dt is None:
                    continue

                risk_score = opinion.get('risk_score', 0)
                expert = opinion.get('expert', '')

                # Проверяем, был ли исход в течение months месяцев ПОСЛЕ прогноза
                has_outcome = False
                for outcome in outcomes:
                    outcome_dt = self._parse_date(outcome.get('date', ''))
                    if outcome_dt is None:
                        continue

                    if opinion_dt < outcome_dt:
                        delta_months = (
                                (outcome_dt.year - opinion_dt.year) * 12
                                + (outcome_dt.month - opinion_dt.month)
                        )
                        if delta_months <= months:
                            has_outcome = True
                            break

                rows.append({
                    'uid': uid,
                    'fiogr': fiogr,
                    'expert': expert,
                    'risk_score': risk_score,
                    'outcome': 1 if has_outcome else 0,
                })

        return pd.DataFrame(rows)

    def calculate_opinion_risks(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Рассчитывает вероятности риска на основе прогнозов специалистов.
        Использует логистическую регрессию.
        """
        if len(df) == 0:
            return pd.DataFrame()

        df_copy = df.copy()

        try:
            # Используем логистическую регрессию для прогнозов специалистов
            X = df_copy[['risk_score']].values
            y = df_copy['outcome'].values

            # Проверяем, что есть оба класса
            if len(set(y)) < 2:
                # Если только один класс, просто возвращаем данные без рисков
                df_copy['risk_probability'] = df_copy['risk_score'] * 10  # Примерная шкала
                df_copy['risk_category'] = df_copy['risk_score'].apply(self._get_risk_category_from_score)
                return df_copy

            # Стандартизируем данные
            scaler = StandardScaler()
            X_scaled = scaler.fit_transform(X)

            # Обучаем модель
            lr = LogisticRegression(penalty='l2', C=1.0, max_iter=1000, random_state=42)
            lr.fit(X_scaled, y)

            # Получаем вероятности
            probabilities = lr.predict_proba(X_scaled)[:, 1] * 100

            df_copy['risk_probability'] = probabilities
            df_copy['risk_category'] = df_copy['risk_probability'].apply(self._get_risk_category)

            # Сохраняем коэффициенты для отчета
            self.opinion_coefs = {
                'coef': lr.coef_[0][0],
                'intercept': lr.intercept_[0]
            }

        except Exception as e:
            print(f"Ошибка при расчете рисков прогнозов: {e}")
            # Если модель не обучилась, используем простую шкалу
            df_copy['risk_probability'] = df_copy['risk_score'] * 10
            df_copy['risk_category'] = df_copy['risk_score'].apply(self._get_risk_category_from_score)

        return df_copy

    def _get_risk_category_from_score(self, score: int) -> str:
        """Определяет категорию риска на основе оценки специалиста (1-10)"""
        if score <= 2:
            return "Низкий"
        elif score <= 4:
            return "Средний"
        elif score <= 6:
            return "Повышенный"
        elif score <= 8:
            return "Высокий"
        else:
            return "Критический"

    def _get_risk_category(self, prob: float) -> str:
        if prob < 20:
            return "Низкий"
        elif prob < 40:
            return "Средний"
        elif prob < 60:
            return "Повышенный"
        elif prob < 80:
            return "Высокий"
        else:
            return "Критический"

    def generate_opinions_table(self, df: pd.DataFrame) -> list:
        """
        Генерирует таблицу прогнозов специалистов для отчета.
        """
        lines = []

        if len(df) == 0:
            lines.append("Нет данных о прогнозах специалистов.")
            return lines

        # Рассчитываем риски
        df_with_risks = self.calculate_opinion_risks(df)

        if len(df_with_risks) == 0:
            lines.append("Нет данных для расчета рисков прогнозов.")
            return lines

        # Сортируем по вероятности риска (по убыванию)
        sorted_df = df_with_risks.sort_values('risk_probability', ascending=False)

        lines.append("")
        lines.append("-" * 80)
        lines.append("Показатели риска по прогнозам специалистов")
        lines.append("-" * 80)

        # Заголовки
        lines.append(f"{'№':<3} {'ФИО':<35} {'Прогноз':<8} {'Риск':<6} {'Категория':<12} {'Исх'}")
        lines.append("-" * 80)

        # Данные
        for idx, (_, row) in enumerate(sorted_df.iterrows(), 1):
            # ФИО
            name = row['fiogr']
            if len(name) > 35:
                name = name[:32] + "..."

            # Прогноз специалиста
            risk_score = row['risk_score']
            if pd.isna(risk_score):
                risk_score_display = "нет"
            else:
                risk_score_display = f"{int(risk_score)}/10"

            # Риск в процентах
            risk_pct = row['risk_probability']
            if pd.isna(risk_pct):
                risk_pct_display = "  -   "
            else:
                risk_pct_display = f"{risk_pct:>5.1f}%"

            # Категория риска
            category = row['risk_category']
            if pd.isna(category):
                category = "Не определено"

            # Исход
            outcome_marker = "!" if row.get('outcome', 0) == 1 else ""

            lines.append(
                f"{idx:<3} {name:<35} {risk_score_display:<8} {risk_pct_display:<6} {category:<12} {outcome_marker}"
            )

        # Статистика по группам риска
        lines.append("")
        lines.append("-" * 80)
        lines.append("Статистика по группам риска (прогнозы специалистов):")
        lines.append("-" * 80)

        risk_stats = sorted_df.groupby('risk_category').agg({
            'fiogr': 'count',
            'outcome': 'sum',
        }).rename(columns={'fiogr': 'count', 'outcome': 'positive'})

        if len(risk_stats) > 0:
            risk_stats['positive_pct'] = (risk_stats['positive'] / risk_stats['count'] * 100).round(1)

            category_order = ['Низкий', 'Средний', 'Повышенный', 'Высокий', 'Критический']
            for category in category_order:
                if category in risk_stats.index:
                    row = risk_stats.loc[category]
                    lines.append(
                        f"  {category:12} : {int(row['count']):3} прогнозов, "
                        f"из них с исходом: {int(row['positive']):3} ({row['positive_pct']:.1f}%)"
                    )
        else:
            lines.append("  Нет данных для группировки по категориям риска")

        return lines


def clean_ansi_codes(text: str) -> str:
    """Убирает ANSI-последовательности из текста."""
    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
    return ansi_escape.sub('', text)


def build_roc_ascii_graph(curves: dict) -> str:
    """Строит ROC-кривые как ASCII-текст (plotext)."""
    if not curves:
        return ""

    try:
        import plotext
        import io
        from contextlib import redirect_stdout

        # Универсальная очистка для разных версий plotext
        if hasattr(plotext, 'clear'):
            plotext.clear()
        elif hasattr(plotext, 'clf'):
            plotext.clf()
        elif hasattr(plotext, 'reset'):
            plotext.reset()

        # Случайный выбор - пунктирная линия из точек.
        plotext.plot([0, 1], [0, 1], marker=".", color="white")

        # Маркеры для разных типов кривых
        markers = {
            'Оценки': '▒',
            'Стадии': '▓',
            'Сумма': '█',
            'Логистическая регрессия': '▢',
            'Прогнозы специалистов': '●',
        }

        for name, data in curves.items():
            fpr = np.asarray(data.get("fpr", []), dtype=float)
            tpr = np.asarray(data.get("tpr", []), dtype=float)
            if len(fpr) == 0:
                continue

            marker = markers.get(name, '•')
            plotext.plot(
                fpr,
                tpr,
                marker=marker,
                color="white"
            )

        plotext.title("ROC-кривые с 95% доверительными интервалами")
        plotext.xlabel("1 - специфичность")
        plotext.ylabel("Чувствительность")
        plotext.plotsize(80, 32)
        plotext.theme("clear")

        buf = io.StringIO()
        with redirect_stdout(buf):
            plotext.show()

        graph = clean_ansi_codes(buf.getvalue())

        # Собираем подписи для легенды
        legend_items = [". Случайный выбор"]
        for name, data in curves.items():
            marker = markers.get(name, '•')
            # Добавляем информацию о количестве прогнозов для специалистов
            if name == 'Прогнозы специалистов' and 'n_opinions' in data:
                legend_items.append(f"{marker} {name} (n={data['n_opinions']})")
            else:
                legend_items.append(f"{marker} {name}")

        graph = _place_legend_bottom_right(graph, legend_items)

        # Добавляем информацию об AUC под графиком
        auc_lines = ["Значения AUC:"]
        for name, data in curves.items():
            if name == 'Прогнозы специалистов' and 'n_opinions' in data:
                auc_lines.append(f"  {name:24}│ {data.get('auc', 0):.3f} (n={data['n_opinions']})")
            else:
                auc_lines.append(f"  {name:24}│ {data.get('auc', 0):.3f}")

        return graph + "\n" + "\n".join(auc_lines)

    except Exception as e:
        return f"(Не удалось построить график: {e})"


def _place_legend_bottom_right(graph: str, legend_items: list) -> str:
    """Размещает легенду в правом нижнем углу ASCII-графика."""
    lines = graph.split("\n")
    if not lines or not legend_items:
        return graph

    width = max(len(line) for line in lines)
    legend_width = max(len(item) for item in legend_items)

    # Находим ось X: первая снизу строка с символом '└' или '┴'
    axis_idx = len(lines) - 1
    for i in range(len(lines) - 1, -1, -1):
        if '└' in lines[i] or '┴' in lines[i]:
            axis_idx = i
            break

    margin = 2  # отступ от правой рамки графика
    start = width - legend_width - margin
    if start < 0:
        start = 0

    top = max(0, axis_idx - len(legend_items))
    for offset, item in enumerate(legend_items):
        row = top + offset
        if row >= axis_idx or row >= len(lines):
            break

        line = lines[row]
        if len(line) < width:
            line = line + " " * (width - len(line))

        max_len = max(0, width - start - margin)
        item = item[:max_len]

        line = line[:start] + item + line[start + len(item):]
        lines[row] = line

    return "\n".join(lines)


class RocScreen(Screen):
    """Экран ROC-анализа для исследовательских целей"""

    CSS = """
        #roc_list_panel {
            width: 70%;
            border: round $secondary;
            background: transparent;
            padding: 1 0 1 0;
            margin: 0;
            border-title-align: left;
            border-title-color: $secondary;
        }

        #roc_content_panel {
            width: 30%;
            border: round $secondary;
            background: transparent;
            padding: 0;
            margin: 0;
            border-title-align: left;
            border-title-color: $secondary;
        }

        #roc_list_panel:focus-within,
        #roc_content_panel:focus-within {
            border: round $accent;
            border-title-color: $accent;
            border-title-style: bold;
            background: transparent;
        }

        #roc_scroll {
            scrollbar-gutter: stable;
            scrollbar-size: 1 1;
            scrollbar-color: $accent 30%;
            scrollbar-color-hover: $accent 30%;
            scrollbar-background: transparent;
            background: transparent;
            padding: 0;
            margin: 0;
        }

        #roc_text {
            padding: 1 2 1 2;
            color: $text;
            text-wrap: wrap;
            width: 100%;
            background: transparent;
        }

        #roc_buttons_container {
            width: 100%;
            height: auto;
            align: left top;
            padding: 1 0 0 1;
        }

        #roc_buttons_container Button {
            width: 90%;
            height: 3;
            margin: 1 1;
            background: $background;
            color: $secondary;
            border: transparent;
            content-align: left middle;
            padding: 0 2;
        }

        #roc_buttons_container Button:focus {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
            content-align: left middle;
        }

        #roc_buttons_container Button:hover {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
        }

        #roc_buttons_container Button.variant-primary {
            border: round $accent;
            color: $accent;
        }

        #roc_buttons_container Button.variant-success {
            border: round green;
            color: green;
        }

        #roc_buttons_container Button.variant-default {
            border: round $secondary;
            color: $secondary;
        }

        #roc_buttons_container Button.spacer {
            height: 1;
            background: transparent;
            border: none;
            margin: 0;
            padding: 0;
            width: 100%;
        }

        #roc_buttons_container Button.spacer:focus {
            border: none;
            background: transparent;
        }

        #roc_buttons_container Button.spacer:hover {
            border: none;
            background: transparent;
        }

        #roc_buttons_container Button > Static {
            width: 100%;
            content-align: left middle;
        }
    """

    BINDINGS = [
        Binding("escape", "go_back", "Назад", show=True, priority=True),
        Binding("right", "focus_buttons", "К кнопкам", show=False),
        Binding("left", "focus_text", "К тексту", show=False),
        Binding("up", "move_up", "Вверх", show=False),
        Binding("down", "move_down", "Вниз", show=False),
        Binding("pageup", "scroll_view_up", "Вверх", show=True),
        Binding("pagedown", "scroll_view_down", "Вниз", show=True),
        Binding("enter", "activate_button", "Выбрать", show=False),
        Binding("1", "press_1", "Запустить", show=False),
        Binding("2", "press_2", "Открыть", show=False),
        Binding("3", "press_3", "В историю", show=False),
        Binding("4", "press_4", "История", show=False),
    ]

    def action_press_1(self):
        if self.show_history:
            self.action_open_editor()
        else:
            self.action_run_roc_analysis()

    def action_press_2(self):
        self.action_open_editor()

    def action_press_3(self):
        if not self.show_history:
            self.action_save_history()

    def action_press_4(self):
        if not self.show_history:
            self.action_open_history()


    class AnalysisDone(Message):
        """Сообщение о завершении анализа"""

        def __init__(self, report: str = "", error: str = None):
            super().__init__()
            self.report = report
            self.error = error

    def __init__(self, show_history: bool = False):
        super().__init__()
        self.show_history = show_history
        self.temp_file = None
        self.roc_content = ""
        self.view_buttons = []
        self.current_button_index = 0
        self._form_counter = 0
        self.history_path = Path.home() / ".arsenal_data" / "roc_history.txt"
        self._is_running = False
        self.roc_analysis = None
        self.report_history = []

    def _get_unique_id(self, base_id: str) -> str:
        self._form_counter += 1
        return f"{base_id}_{self._form_counter}"

    def compose(self) -> ComposeResult:
        with Horizontal():
            with Vertical(id="roc_list_panel", classes="panel-left") as list_panel:
                list_panel.border_title = "ROC-анализ"
                with VerticalScroll(id="roc_scroll"):
                    yield Static("", id="roc_text", markup=False)

            with Vertical(id="roc_content_panel", classes="panel-right") as content_panel:
                content_panel.border_title = "Действия"
                with VerticalScroll(id="roc_content_scroll"):
                    yield Vertical(id="roc_buttons_container")
        yield Footer(show_command_palette=False)

    def on_mount(self) -> None:
        if self.show_history:
            self.load_history()
            self.show_buttons()
        else:
            self.show_buttons()
            self.update_stats_info()

        try:
            roc_text = self.query_one("#roc_text", Static)
            roc_text.focus()
        except NoMatches:
            pass

    def update_stats_info(self):
        """Обновляет информацию о доступных данных для анализа."""
        data = self.app.results.data

        total_assessments = sum(len(p.get("assessments", [])) for p in data)
        total_outcomes = sum(len(p.get("outcomes", [])) for p in data)
        patients_with_outcomes = sum(1 for p in data if p.get("outcomes"))

        positive, negative = self._count_positive_negative()

        info = "Доступные данные для ROC-анализа:\n\n"
        info += f"  • Всего оценок: {total_assessments}\n"
        info += f"  • Оценок с положительным исходом: {positive}\n"
        info += f"  • Оценок без исхода (отрицательных): {negative}\n"
        info += f"  • Всего зарегистрированных исходов: {total_outcomes}\n"
        info += f"  • Пациентов с исходами: {patients_with_outcomes}\n\n"
        info += f"Период учета исходов: 6 месяцев после оценки.\n\n"

        if total_assessments < 30:
            info += "Для минимально надежного ROC-анализа рекомендуется обрабатывать данные не менее 30 оценок.\n"
        if total_outcomes == 0:
            info += "Нет данных об исходах. Добавьте их через раздел 'Опасные проявления'.\n"

        try:
            roc_text = self.query_one("#roc_text", Static)
            roc_text.update(info)
        except NoMatches:
            pass

    def _count_positive_negative(self):
        """Считает, сколько оценок дадут положительный/отрицательный исход при 6-месячном окне."""
        try:
            df = RocAnalysis(self.app.results).prepare_data(months=6)
            if len(df) == 0:
                return 0, 0
            positive = int(df['outcome'].sum())
            return positive, len(df) - positive
        except Exception:
            return 0, 0

    def show_buttons(self) -> None:
        """Показывает кнопки действий в правой панели."""
        try:
            container = self.query_one("#roc_buttons_container")

            for child in list(container.children):
                child.remove()

            self.view_buttons = []

            if self.show_history:
                # В режиме истории показываем кнопки для работы с файлом истории
                btn_open_id = self._get_unique_id("btn_open")
                btn_open = Button("1. Открыть в редакторе     ", variant="primary", id=btn_open_id)
                container.mount(btn_open)
                self.view_buttons.append(btn_open_id)

                btn_back_id = self._get_unique_id("btn_back")
                btn_back = Button("Esc. Назад                 ", variant="default", id=btn_back_id)
                container.mount(btn_back)
                self.view_buttons.append(btn_back_id)

                # Устанавливаем фокус на кнопку "Открыть в редакторе"
                self.current_button_index = 0
                self.call_after_refresh(lambda: btn_open.focus())
            else:
                btn_run_id = self._get_unique_id("btn_run_roc")
                btn_run = Button("1. Провести ROC-анализ   ", variant="primary", id=btn_run_id)
                container.mount(btn_run)
                self.view_buttons.append(btn_run_id)

                btn_open_id = self._get_unique_id("btn_open")
                btn_open = Button("2. Открыть в редакторе   ", variant="primary", id=btn_open_id)
                container.mount(btn_open)
                self.view_buttons.append(btn_open_id)

                btn_save_id = self._get_unique_id("btn_save_history")
                btn_save = Button("3. Записать в историю    ", variant="success", id=btn_save_id)
                container.mount(btn_save)
                self.view_buttons.append(btn_save_id)

                btn_history_id = self._get_unique_id("btn_open_history")
                btn_history = Button("4. Открыть историю       ", variant="primary", id=btn_history_id)
                container.mount(btn_history)
                self.view_buttons.append(btn_history_id)

                spacer_id = self._get_unique_id("spacer")
                container.mount(Button("", disabled=True, classes="spacer", id=spacer_id))

                btn_back_id = self._get_unique_id("btn_back")
                btn_back = Button("Esc. Назад               ", variant="default", id=btn_back_id)
                container.mount(btn_back)
                self.view_buttons.append(btn_back_id)

                # Устанавливаем фокус на кнопку "Провести ROC-анализ"
                self.current_button_index = 0
                self.call_after_refresh(lambda: btn_run.focus())

        except NoMatches:
            pass

    def action_run_roc_analysis(self):
        """Запускает ROC-анализ в фоне."""
        if self._is_running:
            self.app.custom_notify("Анализ уже выполняется...", severity="info")
            return

        def worker():
            try:
                self._is_running = True

                # Отправляем сообщение о начале анализа через call_from_thread
                self.app.call_from_thread(
                    self.app.custom_notify,
                    "Не закрывайте это окно. Данные по ROC-анализу загрузятся через минуту.",
                    "info"
                )

                self.roc_analysis = RocAnalysis(self.app.results)
                results = self.roc_analysis.run_analysis(months=6)

                if isinstance(results, dict) and 'error' in results:
                    self.post_message(self.AnalysisDone(error=results['error']))
                    return

                report = self.roc_analysis.generate_report(months=6)

                if not report:
                    report = "Отчет пуст. Проверьте данные для анализа."

                self.post_message(self.AnalysisDone(report=report))

            except Exception as e:
                import traceback
                error_msg = f"{str(e)}\n\n{traceback.format_exc()}"
                self.post_message(self.AnalysisDone(error=error_msg))

        threading.Thread(target=worker, daemon=True).start()

    @on(AnalysisDone)
    def handle_analysis_done(self, message: AnalysisDone):
        """Обработка завершения анализа в главном потоке."""
        if message.error:
            self.roc_content = f"Ошибка ROC-анализа:\n\n{message.error}"
            self.app.custom_notify("Ошибка ROC-анализа", severity="error")
        else:
            self.roc_content = message.report
            # Сохраняем во временный файл и в память, НО НЕ В ИСТОРИЮ
            self._save_report_to_history(message.report)
            self.app.custom_notify("ROC-анализ завершен",
                                   severity="info")

        self.update_roc_display()
        self._is_running = False

    def _save_report_to_history(self, report: str):
        """Сохраняет отчет в память и во временный файл (НЕ В ИСТОРИЮ)."""
        entry = {
            'datetime': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'period': "6 месяцев после оценки",
            'report': report,
        }
        self.report_history.append(entry)

        # Сохраняем во временный файл для просмотра
        try:
            temp_dir = Path(tempfile.gettempdir())
            self.temp_file = temp_dir / f"arsenal_roc_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            self.temp_file.write_text(report, encoding="utf-8")
        except Exception as e:
            print(f"Ошибка создания временного файла: {e}")

    def update_roc_display(self) -> None:
        """Обновляет отображение ROC-анализа в левой панели."""
        try:
            roc_text = self.query_one("#roc_text", Static)
            roc_text.update(self.roc_content)
        except NoMatches:
            pass

    def load_history(self) -> None:
        """Загружает историю ROC-анализа из файла."""
        try:
            if self.history_path.exists():
                self.roc_content = self.history_path.read_text(encoding="utf-8")
            else:
                self.roc_content = "Файл истории ROC-анализа не найден.\n\n"
                self.roc_content += "Сначала выполните ROC-анализ и запишите его в историю."
        except Exception as e:
            self.roc_content = f"Ошибка при чтении истории:\n\n{str(e)}"

        self.update_roc_display()

    # --- Действия кнопок ---

    def action_open_editor(self) -> None:
        """Открывает соответствующий файл во внешнем редакторе."""
        if self.show_history:
            # В режиме истории открываем файл истории
            if self.history_path.exists():
                open_file_externally(self.history_path)
                self.app.custom_notify("Файл истории открыт во внешнем редакторе", severity="info")
            else:
                self.app.custom_notify("Файл истории не найден", severity="warning")
        else:
            # В обычном режиме открываем временный файл с последним отчетом
            if self.temp_file and self.temp_file.exists():
                open_file_externally(self.temp_file)
                self.app.custom_notify("Файл ROC-анализа открыт во внешнем редакторе", severity="info")
            else:
                self.app.custom_notify("Сначала проведите ROC-анализ", severity="warning")

    def action_save_history(self) -> None:
        """Сохраняет текущий отчет в историю (по кнопке 3)"""
        if not self.roc_content or "Ошибка" in self.roc_content:
            self.app.custom_notify("Сначала проведите успешный ROC-анализ", severity="warning")
            return

        try:
            # Сохраняем в файл истории
            self.history_path.parent.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime('%Y.%m.%d %H:%M:%S')
            separator = "─" * 80

            entry = f"\n\n{separator}\n"
            entry += f"ЗАПИСЬ ROC-АНАЛИЗА ОТ {timestamp}\n"
            entry += f"{separator}\n\n"
            entry += self.roc_content

            if self.history_path.exists():
                with open(self.history_path, "a", encoding="utf-8") as f:
                    f.write(entry)
            else:
                with open(self.history_path, "w", encoding="utf-8") as f:
                    f.write(entry)

            self.app.custom_notify("ROC-анализ записан в историю", severity="success")

        except Exception as e:
            self.app.custom_notify(f"Ошибка сохранения истории: {e}", severity="error")

    def action_open_history(self) -> None:
        self.app.push_screen(RocScreen(show_history=True))

    def action_go_back(self) -> None:
        self.app.pop_screen()

    # --- Навигационные методы ---

    def action_focus_buttons(self) -> None:
        if self.view_buttons:
            self.current_button_index = 0
            try:
                self.query_one(f"#{self.view_buttons[0]}").focus()
            except NoMatches:
                pass

    def action_focus_text(self) -> None:
        try:
            roc_text = self.query_one("#roc_text", Static)
            roc_text.focus()
        except NoMatches:
            pass

    def action_move_up(self) -> None:
        focused = self.focused
        if isinstance(focused, Button) and self.view_buttons:
            current_index = -1
            for i, btn_id in enumerate(self.view_buttons):
                try:
                    if self.query_one(f"#{btn_id}").has_focus:
                        current_index = i
                        break
                except NoMatches:
                    pass

            if current_index >= 0:
                self.current_button_index = (current_index - 1) % len(self.view_buttons)
                try:
                    self.query_one(f"#{self.view_buttons[self.current_button_index]}").focus()
                except NoMatches:
                    pass

    def action_move_down(self) -> None:
        focused = self.focused
        if isinstance(focused, Button) and self.view_buttons:
            current_index = -1
            for i, btn_id in enumerate(self.view_buttons):
                try:
                    if self.query_one(f"#{btn_id}").has_focus:
                        current_index = i
                        break
                except NoMatches:
                    pass

            if current_index >= 0:
                self.current_button_index = (current_index + 1) % len(self.view_buttons)
                try:
                    self.query_one(f"#{self.view_buttons[self.current_button_index]}").focus()
                except NoMatches:
                    pass

    def action_scroll_view_up(self) -> None:
        try:
            scroll = self.query_one("#roc_scroll")
            scroll.scroll_up()
        except NoMatches:
            pass

    def action_scroll_view_down(self) -> None:
        try:
            scroll = self.query_one("#roc_scroll")
            scroll.scroll_down()
        except NoMatches:
            pass

    def action_activate_button(self) -> None:
        focused = self.focused
        if isinstance(focused, Button):
            event = Button.Pressed(focused)
            self.handle_button(event)

    # --- Обработчики событий ---

    @on(Button.Pressed)
    def handle_button(self, event: Button.Pressed) -> None:
        button_id = event.button.id

        if '_' in button_id:
            base_id = button_id.rsplit('_', 1)[0]
        else:
            base_id = button_id

        if base_id == "btn_run_roc":
            self.action_run_roc_analysis()
        elif base_id == "btn_open":
            self.action_open_editor()
        elif base_id == "btn_save_history":
            self.action_save_history()
        elif base_id == "btn_open_history":
            self.action_open_history()
        elif base_id == "btn_back":
            self.action_go_back()

    def on_key(self, event: events.Key) -> None:
        if event.key == "enter":
            focused = self.focused
            if isinstance(focused, Button):
                pass
            elif self.view_buttons:
                self.action_focus_buttons()
                event.stop()
                event.prevent_default()

        elif event.key == "up" or event.key == "down":
            focused = self.focused
            if isinstance(focused, Button) and self.view_buttons:
                if event.key == "up":
                    self.action_move_up()
                else:
                    self.action_move_down()
                event.stop()
                event.prevent_default()

        elif event.key == "right":
            self.action_focus_buttons()
            event.stop()
            event.prevent_default()

        elif event.key == "left":
            self.action_focus_text()
            event.stop()
            event.prevent_default()


class OutcomesMainScreen(Screen):
    """Главный экран управления исходами - с кнопками и пояснениями"""

    BINDINGS = [
        Binding("escape", "go_back_to_research", "Назад", show=True, priority=True),
        Binding("1", "press_view", "Просмотр", show=False),
        Binding("2", "press_maketab", "Шаблон", show=False),
        Binding("3", "press_import", "Импорт", show=False),
    ]

    def action_press_view(self):
        self.query_one("#btn_view").press()

    def action_press_maketab(self):
        self.query_one("#btn_maketab").press()

    def action_press_import(self):
        self.query_one("#btn_import").press()

    # Словарь описаний для кнопок
    OUTCOMES_DESCRIPTIONS = {
        "btn_import": "Для автоматической загрузки данных об опасных проявлениях пациентов (исходов) вам необходимо подготовить файлы в папке Документы.\n\nЕсли вы работаете в МИС ПБСТИН, для подготовки данных об ООД пациентов откройте раздел Отделение ➜ Регистрация Общественно опасных деяний, нажмите Еще ➜ Вывести список... ➜ Выводить в: Текстовый документ ➜ Ок ➜ Сохранить (Ctrl+S). Сохраните список в папку Документы и назовите файл [bold $accent]ООД.txt[/].\n\nДля подготовки данных о применении мер физического ограничения пациентов откройте раздел Отделение ➜ Физические стеснения пациента, нажмите Еще ➜ Вывести список... ➜ Выводить в: Текстовый документ ➜ Ок ➜ Сохранить (Ctrl+S). Сохраните список в папку Документы и назовите файл [bold $accent]вязки.txt[/].\n\nЕсли вы не работаете в МИС ПБСТИН, вы можете создать шаблон для загрузки данных, либо внести данные об исходах вручную для каждого пациента в меню [bold $accent]1. Просмотр исходов[/].",

        "btn_maketab": "Эта функция создает в папке Документы таблицу со списком пациентов, которым были проведены оценки, и полями для ввода данных об опасных проявлениях этих пациентов.\n\nНайдите в папке Документы файл [bold $accent]Арсенал - исходы.xlsx[/], внесите в него сведения об опасных проявлениях пациентов с датами. Если нужно внести несколько исходов для одного пациента, скопируйте его строку несколько раз. Затем произведите загрузку из файлов.\n\nДля пользователей MacOS:\nЕсли вы используете Numbers, после заполнения таблицы выберите Файл ➜ Экспорт в ➜ Excel…, сохраните как [bold $accent]Арсенал - исходы.xlsx[/], а затем произведите загрузку из файлов.",

        "btn_view": "Просмотр и управление данными об опасных проявлениях пациентов.",

        "btn_back": "Вернуться в раздел работы с данными."
    }

    def compose(self) -> ComposeResult:
        with Horizontal():
            # Левая панель - подсказки
            with Vertical(id="outcomes_main_hint_panel") as hint_col:
                hint_col.border_title = "Управление исходами"
                yield Static("", id="outcomes_main_hint_text")

            # Правая панель - кнопки
            with Vertical(id="outcomes_main_buttons_panel") as btn_col:
                btn_col.border_title = "Действия"
                with Vertical(id="outcomes_main_buttons_container"):
                    yield HoverButton("1. Просмотр исходов            ", id="btn_view")
                    yield HoverButton("2. Создать шаблон для данных   ", id="btn_maketab")
                    yield HoverButton("3. Загрузка из файлов          ", id="btn_import")
                    yield Button("", id="btn_spacer", disabled=True, classes="spacer")
                    yield HoverButton("Esc. Назад                     ", id="btn_back")
        yield Footer(show_command_palette=False)

    def on_descendant_focus(self, event: events.Focus) -> None:
        """Срабатывает при фокусе на любой кнопке"""
        target = getattr(event, "control", None) or getattr(event, "node", None)

        if isinstance(target, Button) and target.id:
            self._update_hint(target.id)

    def _update_hint(self, btn_id: str) -> None:
        """Обновление текста в левой панели"""
        description = self.OUTCOMES_DESCRIPTIONS.get(btn_id, "Выберите действие...")
        try:
            self.query_one("#outcomes_main_hint_text", Static).update(description)
        except NoMatches:
            pass

    def on_mount(self) -> None:
        self.query_one("#btn_view").focus()
        self._update_hint("btn_view")

    def on_mouse_move(self, event: events.MouseMove) -> None:
        result = self.get_widget_at(event.screen_x, event.screen_y)
        if result is None:
            return

        if isinstance(result, tuple):
            widget = result[0]
        else:
            widget = result

        current = widget
        while current:
            if isinstance(current, Button):
                if current.id and current.id.startswith("btn_") and self.focused != current:
                    current.focus()
                    self._update_hint(current.id)
                return
            current = current.parent

    def on_button_pressed(self, event: Button.Pressed) -> None:
        btn_id = event.button.id

        if btn_id == "btn_import":
            self._run_import_for_all()
        elif btn_id == "btn_maketab":
            self._create_outcomes_template()
        elif btn_id == "btn_view":
            self.app.push_screen(OutcomesScreen())
        elif btn_id == "btn_back":
            self.action_go_back_to_research()

    def _create_outcomes_template(self):
        """Создает Excel-шаблон для ввода исходов"""
        try:
            # Проверяем наличие openpyxl
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font, PatternFill
            except ImportError:
                self.app.custom_notify(
                    "Для создания таблицы необходимо установить библиотеку openpyxl.\n\n"
                    "Выполните в терминале команду:\n"
                    "pip install openpyxl",
                    severity="error"
                )
                return

            docs_dir = self.get_documents_dir()
            filepath = docs_dir / "Арсенал - исходы.xlsx"

            # Проверяем, существует ли уже файл
            if filepath.exists():
                # Показываем диалог подтверждения
                def handle_dialog_result(confirmed: bool):
                    if confirmed:
                        self._create_template_file(filepath)
                    else:
                        self.app.custom_notify("Создание шаблона отменено", severity="info")
                
                self.app.push_screen(ConfirmTemplateDialog(), handle_dialog_result)
            else:
                # Файла нет - создаем сразу
                self._create_template_file(filepath)

        except Exception as e:
            self.app.custom_notify(f"Ошибка создания шаблона: {e}", severity="error")
            import traceback
            traceback.print_exc()

    def _create_template_file(self, filepath: Path):
        """Создает файл шаблона"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            # Создаем новую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Исходы"

            # Заголовки
            headers = [
                "Фамилия",
                "Имя",
                "Отчество",
                "Год рождения",
                "UID пациента",
                "Дата исхода",
                "Тип исхода",
                "Комментарий"
            ]

            # Применяем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            # Получаем всех пациентов с оценками
            patients = self.app.results.data

            if not patients:
                self.app.custom_notify("Нет пациентов с оценками. Сначала проведите хотя бы одну оценку.", severity="warning")
                return

            # Сортируем пациентов по фамилии
            patients.sort(key=lambda p: p.get("last_name", "").lower())

            row_num = 2
            for patient in patients:
                last_name = patient.get("last_name", "")
                first_name = patient.get("first_name", "")
                patronymic = patient.get("patronymic", "")
                birth_year = patient.get("birth_year", "")
                uid = patient.get("uid", "")

                # Проверяем, есть ли уже исходы у пациента
                outcomes = patient.get("outcomes", [])
                
                if outcomes:
                    # Если есть исходы, для КАЖДОГО исхода создаем отдельную строку
                    # с полными данными пациента
                    for outcome in outcomes:
                        date = outcome.get("date", "")
                        outcome_type = outcome.get("type", "")
                        description = outcome.get("description", "")

                        # Заполняем данные пациента в каждой строке
                        ws.cell(row=row_num, column=1, value=last_name)
                        ws.cell(row=row_num, column=2, value=first_name)
                        ws.cell(row=row_num, column=3, value=patronymic)
                        ws.cell(row=row_num, column=4, value=birth_year)
                        ws.cell(row=row_num, column=5, value=uid)
                        
                        # Заполняем данные исхода
                        ws.cell(row=row_num, column=6, value=date)
                        ws.cell(row=row_num, column=7, value=outcome_type)
                        ws.cell(row=row_num, column=8, value=description)

                        row_num += 1
                else:
                    # Если исходов нет, создаем одну строку с данными пациента
                    # и оставляем пустые ячейки для заполнения
                    ws.cell(row=row_num, column=1, value=last_name)
                    ws.cell(row=row_num, column=2, value=first_name)
                    ws.cell(row=row_num, column=3, value=patronymic)
                    ws.cell(row=row_num, column=4, value=birth_year)
                    ws.cell(row=row_num, column=5, value=uid)
                    row_num += 1

            # Настраиваем ширину столбцов
            column_widths = {
                'A': 20,  # Фамилия
                'B': 15,  # Имя
                'C': 15,  # Отчество
                'D': 12,  # Год рождения
                'E': 12,  # UID
                'F': 15,  # Дата исхода
                'G': 15,  # Тип исхода
                'H': 40,  # Комментарий
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Устанавливаем фильтры
            ws.auto_filter.ref = ws.dimensions

            # Замораживаем первую строку
            ws.freeze_panes = "A2"

            # Сохраняем файл
            wb.save(filepath)

            # Открываем файл
            open_file_externally(filepath)

            self.app.custom_notify(f"Шаблон создан:\n{filepath}", severity="success")

        except Exception as e:
            self.app.custom_notify(f"Ошибка создания файла шаблона: {e}", severity="error")
            import traceback
            traceback.print_exc()

    def action_go_back_to_research(self) -> None:
        """Возврат в ResearchScreen"""
        self.app.pop_screen()

    def _run_import_for_all(self):
        """Импорт исходов для всех пациентов из файлов"""
        def confirm_import(do_import: bool):
            if do_import:
                self._do_import_all()

        self.app.push_screen(
            ConfirmDialog(
                "Загрузить данные об опасных проявлениях сейчас?\n\n"
                "Уже внесенные данные сохранятся и не будут продублированы."
            ),
            confirm_import
        )

    def _do_import_all(self):
        """Выполняет импорт исходов для всех пациентов"""
        try:
            data = self.app.results.data
            total_added = 0
            errors = []

            # Сначала импортируем из файлов ООД.txt и вязки.txt
            for patient in data:
                outcomes = patient.get("outcomes", [])
                existing = {(o.get("date"), o.get("type")) for o in outcomes}

                imported = self._import_from_files(patient)

                for outcome in imported:
                    key = (outcome["date"], outcome["type"])
                    if key not in existing:
                        try:
                            self.app.results.add_outcome(
                                patient_uid=patient.get("uid"),
                                outcome_type=outcome["type"],
                                date=outcome["date"],
                                description=outcome.get("description", ""),
                                source=outcome.get("source", "импорт из МИС")  # Используем source из outcome
                            )
                            total_added += 1
                            existing.add(key)
                        except Exception as e:
                            errors.append(f"{patient.get('last_name', '')}: {e}")

            # Затем импортируем из Excel-файла
            excel_imported = self._import_from_excel()
            for outcome in excel_imported:
                patient_uid = outcome.get("patient_uid")
                if not patient_uid:
                    continue

                patient = self.app.results.get_patient_by_uid(patient_uid)
                if not patient:
                    errors.append(f"Пациент с UID {patient_uid} не найден")
                    continue

                existing = {(o.get("date"), o.get("type")) for o in patient.get("outcomes", [])}
                key = (outcome["date"], outcome["type"])

                if key not in existing:
                    try:
                        self.app.results.add_outcome(
                            patient_uid=patient_uid,
                            outcome_type=outcome["type"],
                            date=outcome["date"],
                            description=outcome.get("description", ""),
                            source=outcome.get("source", "импорт из шаблона")  # Используем source из outcome
                        )
                        total_added += 1
                        existing.add(key)
                    except Exception as e:
                        errors.append(f"{patient.get('last_name', '')}: {e}")

            if total_added > 0:
                self.app.custom_notify(f"Добавлено {total_added} исходов", severity="success")
            else:
                if errors:
                    self.app.custom_notify(f"Ошибок: {len(errors)}, новых исходов не найдено", severity="info")
                else:
                    self.app.custom_notify("Новых исходов не найдено", severity="info")

        except Exception as e:
            self.app.custom_notify(f"Ошибка импорта: {e}", severity="error")
            import traceback
            traceback.print_exc()

    def _import_from_excel(self) -> list:
        """Импортирует исходы из Excel-файла Арсенал - исходы.xlsx
           Поддерживает .xlsx, .xls, и проверяет наличие .numbers
        """
        outcomes = []
        docs_dir = self.get_documents_dir()

        # Проверяем наличие файлов в порядке приоритета
        possible_files = [
            docs_dir / "Арсенал - исходы.xlsx",
            docs_dir / "Арсенал - исходы.xls",
            docs_dir / "Арсенал - исходы.xlsm",
        ]

        filepath = None
        for f in possible_files:
            if f.exists():
                filepath = f
                break

        if not filepath:
            # Проверяем, нет ли файла Numbers
            numbers_file = docs_dir / "Арсенал - исходы.numbers"
            if numbers_file.exists():
                self.app.custom_notify(
                    "Обнаружен файл в формате Numbers.\n\n"
                    "Для импорта данных:\n"
                    "1. Откройте файл в Numbers\n"
                    "2. Выберите Файл ➜ Экспорт в ➜ Excel…\n"
                    "3. Сохраните как 'Арсенал - исходы.xlsx'\n"
                    "4. Повторите импорт",
                    severity="warning"
                )
            else:
                self.app.custom_notify(
                    "Файл 'Арсенал - исходы.xlsx' не найден.\n\n"
                    "Сначала создайте шаблон (кнопка 2) и внесите данные.",
                    severity="warning"
                )
            return outcomes

        try:
            # Проверяем наличие openpyxl
            try:
                import openpyxl
            except ImportError:
                self.app.custom_notify(
                    "Для работы с Excel необходимо установить библиотеку openpyxl.\n\n"
                    "Выполните в терминале команду:\n"
                    "pip install openpyxl",
                    severity="error"
                )
                return outcomes

            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active

            # Определяем индексы колонок
            headers = {}
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers[cell_value] = col

            required_cols = ["UID пациента", "Дата исхода", "Тип исхода"]
            for col in required_cols:
                if col not in headers:
                    self.app.custom_notify(
                        f"В файле не найдена колонка '{col}'.\n\n"
                        "Проверьте, что вы используете правильный шаблон.",
                        severity="error"
                    )
                    return outcomes

            # Получаем текущую дату и время для источника
            from datetime import datetime
            now = datetime.now()
            source_with_date = f"импорт из шаблона {now.strftime('%Y.%m.%d %H:%M')}"

            # Читаем данные
            for row in range(2, ws.max_row + 1):
                uid = ws.cell(row=row, column=headers["UID пациента"]).value
                date = ws.cell(row=row, column=headers["Дата исхода"]).value
                outcome_type = ws.cell(row=row, column=headers["Тип исхода"]).value
                comment_col = headers.get("Комментарий")
                comment = ws.cell(row=row, column=comment_col).value if comment_col else ""

                # Пропускаем пустые строки
                if not uid or not date or not outcome_type:
                    continue

                # Преобразуем uid в строку
                uid = str(uid)

                # Преобразуем дату в строку
                if isinstance(date, datetime):
                    date_str = date.strftime('%Y.%m.%d')
                elif isinstance(date, str):
                    date_str = self._normalize_date(date)
                    if not date_str:
                        continue
                else:
                    continue

                # Добавляем проверку даты
                is_valid, error_msg = DataManager.validate_date(date_str)
                if not is_valid:
                    # Пропускаем строку с невалидной датой
                    continue

                # Приводим тип исхода к нужному формату
                outcome_type = str(outcome_type).strip()

                # Валидация и нормализация типа исхода
                outcome_type_lower = outcome_type.lower()
                if outcome_type_lower == "оод":
                    outcome_type_normalized = "ООД"
                elif outcome_type_lower == "вязки":
                    outcome_type_normalized = "вязки"
                else:
                    # Если тип не распознан, пропускаем строку
                    continue

                outcomes.append({
                    "patient_uid": uid,
                    "type": outcome_type_normalized,
                    "date": date_str,
                    "description": str(comment) if comment else "",
                    "source": source_with_date
                })

        except Exception as e:
            self.app.custom_notify(
                f"Ошибка чтения файла:\n{str(e)}\n\n"
                "Убедитесь, что файл не открыт в другой программе.",
                severity="error"
            )
            import traceback
            traceback.print_exc()

        return outcomes

    def _normalize_date(self, date_str: str) -> str:
        """Нормализует дату из разных форматов в YYYY.MM.DD и проверяет валидность"""
        date_str = date_str.strip()

        if re.match(r'^\d{2}\.\d{2}\.\d{4}$', date_str):
            parts = date_str.split('.')
            normalized = f"{parts[2]}.{parts[1]}.{parts[0]}"
        elif re.match(r'^\d{4}\.\d{2}\.\d{2}$', date_str):
            normalized = date_str
        elif re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
            parts = date_str.split('-')
            normalized = f"{parts[0]}.{parts[1]}.{parts[2]}"
        elif re.match(r'^\d{2}/\d{2}/\d{4}$', date_str):
            parts = date_str.split('/')
            normalized = f"{parts[2]}.{parts[1]}.{parts[0]}"
        else:
            try:
                from datetime import datetime
                for fmt in ['%d.%m.%Y', '%Y.%m.%d', '%Y-%m-%d', '%d/%m/%Y']:
                    try:
                        dt = datetime.strptime(date_str, fmt)
                        normalized = dt.strftime('%Y.%m.%d')
                        break
                    except ValueError:
                        continue
                else:
                    return None
            except Exception:
                return None

        # Проверяем валидность нормализованной даты
        is_valid, _ = DataManager.validate_date(normalized)
        return normalized if is_valid else None

    def _import_from_files(self, patient: dict) -> list:
        """Загружает исходы для пациента из файлов"""
        docs_dir = self.get_documents_dir()
        outcomes = []

        last_name = patient.get("last_name", "")
        first_name = patient.get("first_name", "")
        patronymic = patient.get("patronymic", "")
        birth_year = str(patient.get("birth_year", ""))

        # Получаем текущую дату и время для источника
        from datetime import datetime
        now = datetime.now()
        source_with_date = f"импорт из МИС {now.strftime('%Y.%m.%d %H:%M')}"

        for filename in ["ООД.txt", "вязки.txt"]:
            filepath = docs_dir / filename
            if not filepath.exists():
                continue

            outcome_type = "ООД" if filename == "ООД.txt" else "вязки"

            try:
                content = None
                for encoding in ['utf-8-sig', 'utf-8', 'cp1251']:
                    try:
                        with open(filepath, 'r', encoding=encoding) as f:
                            content = f.read()
                        break
                    except UnicodeDecodeError:
                        continue

                if content is None:
                    continue

                lines = content.splitlines()
                if not lines:
                    continue

                first_line = lines[0]
                if '\t' in first_line:
                    delimiter = '\t'
                else:
                    delimiter = ';'

                headers = first_line.split(delimiter)
                name_col_index = -1
                date_col_index = -1
                number_col_index = -1
                dept_col_index = -1
                restraint_col_index = -1

                for i, header in enumerate(headers):
                    header_lower = header.lower().strip()
                    if 'пациент' in header_lower or 'фио' in header_lower:
                        name_col_index = i
                    if 'дата' in header_lower:
                        date_col_index = i
                    if 'номер' in header_lower:
                        number_col_index = i
                    if 'отделение' in header_lower:
                        dept_col_index = i
                    if 'вид' in header_lower and ('физического' in header_lower or 'стеснения' in header_lower):
                        restraint_col_index = i

                if name_col_index == -1:
                    continue

                for line in lines[1:]:
                    if not line.strip():
                        continue

                    parts = line.split(delimiter)
                    if len(parts) <= name_col_index:
                        continue

                    name = parts[name_col_index].strip()
                    if not self._patient_matches_name(name, last_name, first_name, patronymic):
                        continue

                    date = self._extract_date_from_line(line, parts, date_col_index)
                    if not date:
                        continue

                    # Формируем описание
                    if filename == "вязки.txt":
                        # Для вязки.txt - только нужные поля
                        fields = []
                        
                        if date_col_index != -1 and len(parts) > date_col_index:
                            fields.append(parts[date_col_index].strip())
                        
                        if number_col_index != -1 and len(parts) > number_col_index:
                            fields.append(parts[number_col_index].strip())
                        
                        if name_col_index != -1 and len(parts) > name_col_index:
                            fields.append(parts[name_col_index].strip())
                        
                        if dept_col_index != -1 and len(parts) > dept_col_index:
                            fields.append(parts[dept_col_index].strip())
                        
                        if restraint_col_index != -1 and len(parts) > restraint_col_index:
                            restraint = parts[restraint_col_index].strip()
                            if restraint == "Мягкая вязка рук и ног в постели":
                                restraint = "Фиксация в постели"
                            fields.append(restraint)
                        
                        description = " ".join(fields)
                    else:
                        # Для ООД.txt - все данные, но с заменой табуляции на пробел
                        description = line.strip()
                        # Заменяем табуляции на пробелы
                        description = description.replace('\t', ' ')
                        # Сжимаем множественные пробелы в один
                        description = ' '.join(description.split())

                    outcomes.append({
                        "type": outcome_type,
                        "date": date,
                        "description": description,
                        "source": source_with_date  # Теперь здесь дата и время
                    })

            except Exception as e:
                print(f"Ошибка чтения {filename}: {e}")

        return outcomes

    def _patient_matches_name(self, name: str, last_name: str, first_name: str, patronymic: str) -> bool:
        """Проверяет, соответствует ли строка ФИО пациента"""
        if not name:
            return False

        name_lower = name.lower()

        # Проверяем все части ФИО
        parts = [last_name.lower(), first_name.lower()]
        if patronymic:
            parts.append(patronymic.lower())

        # Все части должны присутствовать в строке
        for part in parts:
            if part and part not in name_lower:
                return False

        return True

    def _extract_date_from_line(self, line: str, parts: list = None, date_col_index: int = -1) -> str:
        """Извлекает дату из строки и проверяет её валидность"""
        # Если есть колонка с датой
        if parts is not None and date_col_index >= 0 and len(parts) > date_col_index:
            date_str = parts[date_col_index].strip()
            if date_str:
                # Пробуем разные форматы
                # DD.MM.YYYY
                match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', date_str)
                if match:
                    normalized = f"{match.group(3)}.{match.group(2)}.{match.group(1)}"
                    is_valid, _ = DataManager.validate_date(normalized)
                    return normalized if is_valid else None

                # YYYY.MM.DD
                match = re.search(r'(\d{4})\.(\d{2})\.(\d{2})', date_str)
                if match:
                    normalized = f"{match.group(1)}.{match.group(2)}.{match.group(3)}"
                    is_valid, _ = DataManager.validate_date(normalized)
                    return normalized if is_valid else None

                # YYYY-MM-DD
                match = re.search(r'(\d{4})-(\d{2})-(\d{2})', date_str)
                if match:
                    normalized = f"{match.group(1)}.{match.group(2)}.{match.group(3)}"
                    is_valid, _ = DataManager.validate_date(normalized)
                    return normalized if is_valid else None

        # Ищем дату в строке
        patterns = [
            r'(\d{4})\.(\d{2})\.(\d{2})',  # 2025.06.15
            r'(\d{2})\.(\d{2})\.(\d{4})',  # 15.06.2025
            r'(\d{4})-(\d{2})-(\d{2})',  # 2025-06-15
        ]

        for pattern in patterns:
            match = re.search(pattern, line)
            if match:
                groups = match.groups()
                if len(groups) == 3:
                    if len(groups[0]) == 4:  # YYYY.MM.DD
                        normalized = f"{groups[0]}.{groups[1]}.{groups[2]}"
                    elif len(groups[2]) == 4:  # DD.MM.YYYY
                        normalized = f"{groups[2]}.{groups[1]}.{groups[0]}"
                    else:  # YYYY-MM-DD
                        normalized = f"{groups[0]}.{groups[1]}.{groups[2]}"

                    is_valid, _ = DataManager.validate_date(normalized)
                    return normalized if is_valid else None

        # Если дата не найдена или невалидна, используем текущую
        today = datetime.now().strftime('%Y.%m.%d')
        return today

    def get_documents_dir(self) -> Path:
        """Определяет путь к папке Документы"""
        import platform

        system = platform.system()
        home = Path.home()

        if system == "Windows":
            docs = home / "Documents"
            if docs.exists():
                return docs
            return home / "Мои документы"
        elif system == "Darwin":
            return home / "Documents"
        else:
            try:
                import subprocess
                result = subprocess.run(['xdg-user-dir', 'DOCUMENTS'],
                                        capture_output=True, text=True)
                if result.returncode == 0:
                    docs = Path(result.stdout.strip())
                    if docs.exists():
                        return docs
            except (FileNotFoundError, subprocess.SubprocessError):
                pass
            return home / "Документы"


class OutcomesScreen(Screen):
    """Экран управления исходами пациентов"""

    CSS = """
        #outcomes_list_panel {
            width: 50%;
            border: round $secondary;
            background: transparent;
            padding: 1 0 1 0;
            margin: 0;
            border-title-align: left;
            border-title-color: $secondary;
        }

        #outcomes_detail_panel {
            width: 50%;
            border: round $secondary;
            background: transparent;
            padding: 0;
            margin: 0;
            border-title-align: left;
            border-title-color: $secondary;
        }

        #outcomes_list_panel:focus-within,
        #outcomes_detail_panel:focus-within {
            border: round $accent;
            border-title-color: $accent;
            border-title-style: bold;
            background: transparent;
        }

        #outcomes_scroll {
            scrollbar-gutter: stable;
            scrollbar-size: 1 1;
            scrollbar-color: $accent 30%;
            scrollbar-color-hover: $accent 30%;
            scrollbar-background: transparent;
            background: transparent;
            padding: 0;
            margin: 0;
        }

        #outcomes_text {
            padding: 1 2 1 2;
            color: $text;
            text-wrap: wrap;
            width: 100%;
            background: transparent;
        }

        #outcomes_buttons_container {
            width: 100%;
            height: auto;
            align: left top;
            padding: 1 0 0 1;
        }

        #outcomes_buttons_container Button {
            width: 90%;
            height: 3;
            margin: 1 1;
            background: $background;
            color: $secondary;
            border: transparent;
            content-align: left middle;
            padding: 0 2;
        }

        #outcomes_buttons_container Button:focus {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
            content-align: left middle;
        }

        #outcomes_buttons_container Button:hover {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
        }

        #outcomes_buttons_container Button.variant-primary {
            border: round $accent;
            color: $accent;
        }

        #outcomes_buttons_container Button.variant-error {
            border: round red;
            color: red;
        }

        #outcomes_buttons_container Button.variant-success {
            border: round green;
            color: green;
        }

        #outcomes_buttons_container Button.spacer {
            height: 1;
            background: transparent;
            border: none;
            margin: 0;
            padding: 0;
            width: 100%;
        }

        #outcomes_buttons_container Button.spacer:focus {
            border: none;
            background: transparent;
        }

        #outcomes_buttons_container Button.spacer:hover {
            border: none;
            background: transparent;
        }

        #outcomes_buttons_container Button > Static {
            width: 100%;
            content-align: left middle;
        }
    """

    BINDINGS = [
        Binding("escape", "go_back", "Назад", show=True, priority=True),
        Binding("right", "focus_buttons", "К кнопкам", show=False),
        Binding("left", "focus_list", "К списку", show=False),
        Binding("up", "move_up", "Вверх", show=False),
        Binding("down", "move_down", "Вниз", show=False),
        Binding("pageup", "scroll_view_up", "Вверх", show=True),
        Binding("pagedown", "scroll_view_down", "Вниз", show=True),
        Binding("enter", "select_item", "Выбрать", show=True),
        Binding("1", "press_add", "Добавить", show=False),
        Binding("2", "press_delete", "Удалить", show=False),
        Binding("3", "press_write", "Записать", show=False),
    ]

    def action_press_add(self):
        if self.selected_patient:
            self.show_add_outcome_dialog()
        else:
            self.app.custom_notify("Сначала выберите пациента", severity="warning")

    def action_press_delete(self):
        if self.selected_patient:
            outcomes = self.selected_patient.get("outcomes", [])
            if outcomes:
                self.show_delete_outcome_dialog()
            else:
                self.app.custom_notify("Нет исходов для удаления", severity="info")
        else:
            self.app.custom_notify("Сначала выберите пациента", severity="warning")

    def action_press_write(self):
        if self.selected_patient:
            self.action_write_to_file()
        else:
            self.app.custom_notify("Сначала выберите пациента", severity="warning")

    def __init__(self):
        super().__init__()
        self.selected_patient = None
        self.selected_patient_uid = None  # Сохраняем UID выбранного пациента
        self.view_buttons = []
        self.current_button_index = 0
        self.mode = "list"
        self._is_on_list = True  # Флаг: фокус на списке или на кнопках

    def compose(self) -> ComposeResult:
        with Horizontal():
            # Левая панель - список пациентов
            with Vertical(id="outcomes_list_panel", classes="panel-left") as list_panel:
                list_panel.border_title = "Пациенты"
                with VerticalScroll(id="outcomes_scroll"):
                    yield ListView(id="patients_list")

            # Правая панель - детали или список исходов
            with Vertical(id="outcomes_detail_panel", classes="panel-right") as detail_panel:
                detail_panel.border_title = "Детали"
                with VerticalScroll(id="outcomes_scroll"):
                    yield Static("", id="outcomes_text", markup=False)
                    yield Vertical(id="outcomes_buttons_container")
        yield Footer(show_command_palette=False)

    def on_mount(self) -> None:
        self.refresh_patients_list()
        self.show_buttons()

        # Восстанавливаем выбранного пациента, если он был
        if self.selected_patient_uid:
            list_view = self.query_one("#patients_list", ListView)
            for i, item in enumerate(list_view.children):
                if hasattr(item, 'patient') and item.patient.get('uid') == self.selected_patient_uid:
                    list_view.index = i
                    self.selected_patient = item.patient
                    self.show_patient_detail(item.patient)
                    break

        try:
            list_view = self.query_one("#patients_list", ListView)
            list_view.focus()
            self._is_on_list = True
        except NoMatches:
            pass

    def refresh_patients_list(self) -> None:
        """Обновляет список пациентов"""
        list_view = self.query_one("#patients_list", ListView)

        # Запоминаем текущий индекс и UID выбранного пациента до обновления
        current_index = list_view.index if list_view.index is not None else 0
        current_patient_uid = None
        if self.selected_patient:
            current_patient_uid = self.selected_patient.get('uid')

        list_view.clear()

        patients = self.app.results.data
        if not patients:
            label = ListLabel("Нет пациентов в базе данных.")
            list_view.append(ListItem(label))
            return

        patients.sort(key=lambda p: p.get("last_name", "").lower())

        for patient in patients:
            last_name = patient.get("last_name", "")
            first_name = patient.get("first_name", "")
            patronymic = patient.get("patronymic", "")
            birth_year = patient.get("birth_year", "")
            outcomes_count = len(patient.get("outcomes", []))

            display = f"{last_name} {first_name} {patronymic} {birth_year} г.р."
            if outcomes_count > 0:
                display += f" [исходов: {outcomes_count}]"

            label = ListLabel(display)
            item = ListItem(label)
            item.patient = patient
            list_view.append(item)

        if list_view.children:
            # Восстанавливаем выбранного пациента по UID
            if current_patient_uid:
                for i, item in enumerate(list_view.children):
                    if hasattr(item, 'patient') and item.patient.get('uid') == current_patient_uid:
                        list_view.index = i
                        self.selected_patient = item.patient
                        return
            # Если не нашли или нет сохраненного UID, выбираем первого
            if list_view.index is None or list_view.index >= len(list_view.children):
                list_view.index = 0

    def show_patient_detail(self, patient: dict) -> None:
        """Показывает детали пациента и его исходы"""
        self.selected_patient = patient
        self.selected_patient_uid = patient.get('uid')  # Сохраняем UID

        last_name = patient.get("last_name", "")
        first_name = patient.get("first_name", "")
        patronymic = patient.get("patronymic", "")
        birth_year = patient.get("birth_year", "")

        outcomes = patient.get("outcomes", [])

        lines = []
        lines.append(f"Пациент: {last_name} {first_name} {patronymic} {birth_year} г.р.")
        lines.append(f"UID: {patient.get('uid', '')}")
        lines.append(f"Всего исходов: {len(outcomes)}")
        lines.append("")

        if outcomes:
            lines.append("Исходы:")
            lines.append("")
            sorted_outcomes = sorted(outcomes, key=lambda o: o.get("date", ""), reverse=False)
            for i, outcome in enumerate(sorted_outcomes, 1):
                outcome_type = outcome.get("type", "")
                date = outcome.get("date", "")
                description = outcome.get("description", "")
                source = outcome.get("source", "")

                lines.append(f"    {i}. {outcome_type} от {date}")
                if description:
                    lines.append(f"{description}")
                if source:
                    lines.append(f"Источник: {source}")
                lines.append("")
        else:
            lines.append("Нет зарегистрированных исходов.")

        # Обновляем текст в правой панели
        self.query_one("#outcomes_text", Static).update("\n".join(lines))

        self.show_buttons(patient)
        self.mode = "detail"

    def show_buttons(self, patient: dict = None) -> None:
        """Показывает кнопки действий"""
        try:
            container = self.query_one("#outcomes_buttons_container", Vertical)

            for child in list(container.children):
                child.remove()

            self.view_buttons = []

            if patient:
                btn_add_id = f"btn_add_outcome_{uuid.uuid4().hex[:6]}"
                btn_add = Button("1. Добавить исход                          ", variant="primary", id=btn_add_id)
                container.mount(btn_add)
                self.view_buttons.append(btn_add_id)

                if patient.get("outcomes"):
                    btn_delete_id = f"btn_delete_outcome_{uuid.uuid4().hex[:6]}"
                    btn_delete = Button("2. Удалить исход                           ", variant="error",
                                        id=btn_delete_id)
                    container.mount(btn_delete)
                    self.view_buttons.append(btn_delete_id)

                # НОВАЯ КНОПКА - Записать в файл пациента
                btn_write_id = f"btn_write_outcome_{uuid.uuid4().hex[:6]}"
                btn_write = Button("3. Записать в файл пациента                ", variant="success",
                                   id=btn_write_id)
                container.mount(btn_write)
                self.view_buttons.append(btn_write_id)
            else:
                btn_back_id = f"btn_back_{uuid.uuid4().hex[:6]}"
                btn_back = Button("Esc. Назад                ", variant="default", id=btn_back_id)
                container.mount(btn_back)
                self.view_buttons.append(btn_back_id)

            self.current_button_index = 0

        except NoMatches as e:
            print(f"show_buttons error: {e}")

    def action_focus_list(self) -> None:
        """Перемещает фокус на список пациентов"""
        try:
            list_view = self.query_one("#patients_list", ListView)
            list_view.focus()
            self._is_on_list = True
        except NoMatches:
            pass

    def action_focus_buttons(self) -> None:
        """Перемещает фокус на первую кнопку в правой панели"""
        if self.view_buttons:
            self.current_button_index = 0
            try:
                self.query_one(f"#{self.view_buttons[0]}").focus()
                self._is_on_list = False
            except NoMatches:
                pass

    def action_move_up(self) -> None:
        """Перемещение вверх по списку или кнопкам"""
        if self._is_on_list:
            # На списке - перемещение вверх по пунктам
            try:
                list_view = self.query_one("#patients_list", ListView)
                if list_view.index > 0:
                    list_view.index -= 1
            except NoMatches:
                pass
        else:
            # На кнопках - перемещение вверх между кнопками
            if not self.view_buttons:
                return

            focused = self.focused
            if not isinstance(focused, Button) or focused.id not in self.view_buttons:
                # Если фокус не на кнопке, переключаем на последнюю
                try:
                    self.query_one(f"#{self.view_buttons[-1]}").focus()
                except NoMatches:
                    pass
                return

            # Находим индекс текущей кнопки
            current_index = -1
            for i, btn_id in enumerate(self.view_buttons):
                try:
                    if self.query_one(f"#{btn_id}").has_focus:
                        current_index = i
                        break
                except NoMatches:
                    pass

            if current_index > 0:
                self.current_button_index = current_index - 1
                try:
                    self.query_one(f"#{self.view_buttons[self.current_button_index]}").focus()
                except NoMatches:
                    pass

    def action_move_down(self) -> None:
        """Перемещение вниз по списку или кнопкам"""
        if self._is_on_list:
            # На списке - перемещение вниз по пунктам
            try:
                list_view = self.query_one("#patients_list", ListView)
                if list_view.index < len(list_view.children) - 1:
                    list_view.index += 1
            except NoMatches:
                pass
        else:
            # На кнопках - перемещение вниз между кнопками
            if not self.view_buttons:
                return

            focused = self.focused
            if not isinstance(focused, Button) or focused.id not in self.view_buttons:
                # Если фокус не на кнопке, переключаем на первую
                try:
                    self.query_one(f"#{self.view_buttons[0]}").focus()
                except NoMatches:
                    pass
                return

            # Находим индекс текущей кнопки
            current_index = -1
            for i, btn_id in enumerate(self.view_buttons):
                try:
                    if self.query_one(f"#{btn_id}").has_focus:
                        current_index = i
                        break
                except NoMatches:
                    pass

            if current_index < len(self.view_buttons) - 1:
                self.current_button_index = current_index + 1
                try:
                    self.query_one(f"#{self.view_buttons[self.current_button_index]}").focus()
                except NoMatches:
                    pass

    def action_scroll_view_up(self) -> None:
        try:
            scroll = self.query_one("#outcomes_scroll")
            scroll.scroll_up()
        except NoMatches:
            pass

    def action_scroll_view_down(self) -> None:
        try:
            scroll = self.query_one("#outcomes_scroll")
            scroll.scroll_down()
        except NoMatches:
            pass

    def action_select_item(self) -> None:
        """Выбор пациента из списка (Enter)"""
        if not self._is_on_list:
            return

        try:
            list_view = self.query_one("#patients_list", ListView)
            if list_view.index is not None and list_view.index < len(list_view.children):
                item = list_view.children[list_view.index]
                if hasattr(item, 'patient'):
                    self.show_patient_detail(item.patient)
        except NoMatches:
            pass

    def action_go_back(self) -> None:
        """Возврат в OutcomesMainScreen"""
        self.app.pop_screen()

    # --- Обработчики событий ---

    def on_key(self, event: events.Key) -> None:
        """Обработка клавиш для навигации"""
        if event.key == "up":
            self.action_move_up()
            event.stop()
            event.prevent_default()
            return
        elif event.key == "down":
            self.action_move_down()
            event.stop()
            event.prevent_default()
            return
        elif event.key == "right":
            # Стрелка вправо - переключение на правую панель
            if self._is_on_list:
                self.action_focus_buttons()
                event.stop()
                event.prevent_default()
                return
        elif event.key == "left":
            # Стрелка влево - переключение на левую панель
            if not self._is_on_list:
                self.action_focus_list()
                event.stop()
                event.prevent_default()
                return
        elif event.key == "enter":
            # Enter - выбор элемента в списке
            if self._is_on_list:
                self.action_select_item()
                event.stop()
                event.prevent_default()
                return

    @on(ListView.Highlighted, "#patients_list")
    def handle_list_highlighted(self, event: ListView.Highlighted) -> None:
        """Обновляет маркеры и детали при навигации по списку"""
        if event.item and hasattr(event.item, 'patient'):
            # Обновляем маркеры
            try:
                list_view = self.query_one("#patients_list", ListView)
                for i, child in enumerate(list_view.children):
                    if hasattr(child, 'children') and child.children:
                        label = child.children[0]
                        if isinstance(label, ListLabel):
                            label.is_highlighted = (i == list_view.index)
            except NoMatches:
                pass

            # Обновляем детали в правой панели
            self.show_patient_detail(event.item.patient)

    @on(ListView.Selected, "#patients_list")
    def handle_list_selected(self, event: ListView.Selected) -> None:
        """Обработка Enter или двойного клика по элементу списка"""
        if event.item and hasattr(event.item, 'patient'):
            self.show_patient_detail(event.item.patient)

    @on(Button.Pressed)
    def handle_button(self, event: Button.Pressed) -> None:
        """Обработчик нажатия кнопок"""
        button_id = event.button.id

        if button_id.startswith("btn_add_outcome"):
            self.show_add_outcome_dialog()
        elif button_id.startswith("btn_delete_outcome"):
            self.show_delete_outcome_dialog()
        elif button_id.startswith("btn_write_outcome"):
            self.action_write_to_file()
        elif button_id.startswith("btn_back"):
            self.action_go_back()

    def show_add_outcome_dialog(self):
        if not self.selected_patient:
            self.app.custom_notify("Сначала выберите пациента", severity="warning")
            return

        def after_dialog(result: dict):
            if result:
                try:
                    patient_uid = self.selected_patient.get("uid")
                    outcome = self.app.results.add_outcome(
                        patient_uid=patient_uid,
                        outcome_type=result["type"],
                        date=result["date"],
                        description=result["description"],
                        source="ручной ввод"
                    )
                    self.app.custom_notify(f"Исход добавлен: {result['type']} от {result['date']}", severity="success")

                    # Обновляем список и показываем детали выбранного пациента
                    self.refresh_patients_list()
                    self.show_patient_detail(self.selected_patient)

                    # Устанавливаем фокус на кнопку "Добавить исход"
                    self.action_focus_buttons()

                except ValueError as e:
                    self.app.custom_notify(str(e), severity="warning")
                    self.show_buttons(self.selected_patient)
                    # Восстанавливаем фокус на кнопку
                    self.action_focus_buttons()
                except Exception as e:
                    self.app.custom_notify(f"Ошибка: {e}", severity="error")
                    self.show_buttons(self.selected_patient)
                    self.action_focus_buttons()
            else:
                # Если диалог отменен, возвращаем фокус на кнопку
                self.show_buttons(self.selected_patient)
                self.action_focus_buttons()

        self.app.push_screen(AddOutcomeDialog(self.selected_patient), after_dialog)

    def show_delete_outcome_dialog(self):
        if not self.selected_patient:
            return

        outcomes = self.selected_patient.get("outcomes", [])
        if not outcomes:
            self.app.custom_notify("Нет исходов для удаления", severity="info")
            return

        def after_dialog(result: dict):
            if result:
                try:
                    patient_uid = self.selected_patient.get("uid")
                    outcome_id = result.get("id")
                    deleted = self.app.results.delete_outcome(patient_uid, outcome_id)
                    if deleted:
                        self.app.custom_notify("Исход удален", severity="success")
                        self.refresh_patients_list()
                        self.show_patient_detail(self.selected_patient)
                        # Восстанавливаем фокус на кнопку "Добавить исход"
                        self.action_focus_buttons()
                    else:
                        self.app.custom_notify("Не удалось удалить исход", severity="error")
                        self.show_buttons(self.selected_patient)
                        self.action_focus_buttons()
                except Exception as e:
                    self.app.custom_notify(f"Ошибка: {e}", severity="error")
                    self.show_buttons(self.selected_patient)
                    self.action_focus_buttons()
            else:
                self.show_buttons(self.selected_patient)
                self.action_focus_buttons()

        self.app.push_screen(DeleteOutcomeDialog(self.selected_patient), after_dialog)

    def action_write_to_file(self) -> None:
        """Записывает информацию об исходах в файл пациента"""
        if not self.selected_patient:
            self.app.custom_notify("Сначала выберите пациента", severity="warning")
            return

        try:
            patient = self.selected_patient
            last_name = patient.get("last_name", "")
            first_name = patient.get("first_name", "")
            patronymic = patient.get("patronymic", "")
            birth_year = str(patient.get("birth_year", ""))

            # Формируем имя файла
            filename_parts = [last_name, first_name, patronymic, birth_year]
            filename_base = " ".join(filter(None, filename_parts))
            filename_safe = filename_base.replace(" ", "_")
            filename = f"{filename_safe}.txt"
            filepath = self.app.results.reports_dir / filename

            # Генерируем содержимое для записи
            content = self.generate_outcomes_content(patient)

            # Записываем в файл
            if filepath.exists():
                # Если файл существует, добавляем разделитель и новое содержание
                with open(filepath, "a", encoding="utf-8") as f:
                    f.write("\n\n" + content)
            else:
                # Создаем новый файл
                filepath.write_text(content, encoding="utf-8")

            self.app.custom_notify(f"Данные об исходах записаны в файл:\n{filename}", severity="success")

        except Exception as e:
            self.app.custom_notify(f"Ошибка записи в файл: {e}", severity="error")
            import traceback
            traceback.print_exc()

    def generate_outcomes_content(self, patient: dict) -> str:
        """Генерирует содержимое для записи об исходах в файл пациента"""
        last_name = patient.get("last_name", "")
        first_name = patient.get("first_name", "")
        patronymic = patient.get("patronymic", "")
        birth_year = patient.get("birth_year", "")

        outcomes = patient.get("outcomes", [])
        current_date = datetime.now().strftime('%Y.%m.%d')

        lines = []
        lines.append(f"Пациент {last_name} {first_name} {patronymic} {birth_year} г.р.")
        lines.append("╭" + "─" * 78 + "╮")
        lines.append(f"Данные об опасных проявлениях на {current_date}")
        lines.append("")

        if outcomes:
            sorted_outcomes = sorted(outcomes, key=lambda o: o.get("date", ""), reverse=False)
            for i, outcome in enumerate(sorted_outcomes, 1):
                outcome_type = outcome.get("type", "")
                date = outcome.get("date", "")
                description = outcome.get("description", "")
                source = outcome.get("source", "")

                lines.append(f"  {i}. {outcome_type} от {date}")
                if description:
                    wrapped = textwrap.fill(description, width=74)
                    for line in wrapped.split("\n"):
                        lines.append(f"     {line}")
                if source:
                    lines.append(f"     Источник: {source}")
                lines.append("")
        else:
            lines.append("Нет зарегистрированных исходов.")

        lines.append("")
        lines.append("╰" + "─" * 78 + "╯")

        return "\n".join(lines)

class AddOutcomeDialog(ModalScreen[dict]):
    """Диалог добавления исхода"""

    # Отключаем встроенную прокрутку ModalScreen
    SCROLL_VIEW = False

    CSS = """
    AddOutcomeDialog {
        align: center middle;
        background: transparent;
        overflow-y: auto;
    }

    AddOutcomeDialog VerticalScroll {
        scrollbar-gutter: stable;
        scrollbar-size: 1 1;
        scrollbar-color: $accent 30%;
        scrollbar-color-hover: $accent 30%;
        scrollbar-background: transparent;
    }

    #outcome_add_dialog {
        width: 55;
        height: auto;
        background: $background;
        border: round $accent;
        padding: 2 3;
    }

    #outcome_add_dialog .title-label {
        width: 100%;
        height: 2;
        content-align: center middle;
        color: $accent;
        text-style: bold;
        margin-bottom: 1;
    }

    #outcome_add_dialog .field-label {
        margin-top: 1;
        margin-bottom: 0;
        text-style: bold;
        color: $text;
        height: 1;
    }

    #outcome_add_dialog Input {
        margin-top: 0;
        margin-bottom: 0;
        width: 100%;
        background: $background;
        border: transparent;
        color: $text;
        height: 3;
        padding: 0 1;
    }

    #outcome_add_dialog Input:focus {
        border: round $accent;
        background: transparent;
    }

    #outcome_add_dialog TextArea {
        margin-top: 0;
        margin-bottom: 0;
        width: 100%;
        background: $background;
        border: transparent;
        color: $text;
        height: 5;
        min-height: 3;
        max-height: 8;
        padding: 0 1;
    }

    #outcome_add_dialog TextArea:focus {
        border: round $accent;
        background: transparent;
    }

    #outcome_type_container {
        width: 100%;
        height: 4;
        align: center middle;
        margin-top: 0;
        margin-bottom: 0;
    }

    #outcome_type_container Button {
        width: 20;
        height: 3;
        margin: 0 2;
        background: $background;
        color: $secondary;
        border: round $secondary;
    }

    #outcome_type_container Button:focus {
        border: round $accent;
        color: $accent;
        background: transparent;
        text-style: bold;
    }

    #outcome_type_container Button.selected {
        border: round $accent;
        color: $accent;
        background: transparent;
        text-style: bold;
    }

    #outcome_add_dialog .dialog_buttons {
        margin-top: 2;
        align: center middle;
        height: 3;
    }

    #outcome_add_dialog .dialog_buttons Button {
        width: 16;
        height: 3;
        margin: 0 1;
        background: $background;
        color: $secondary;
        border: round $secondary;
    }

    #outcome_add_dialog .dialog_buttons Button:focus {
        border: round $accent;
        color: $accent;
        background: transparent;
        text-style: bold;
    }

    #outcome_add_dialog .dialog_buttons Button.variant-primary {
        border: round $accent;
        color: $accent;
    }

    #outcome_add_dialog .dialog_buttons Button.variant-error {
        border: round red;
        color: red;
    }
    """

    BINDINGS = [
        Binding("escape", "cancel", "Отмена", show=False, priority=True),
    ]

    def __init__(self, patient: dict):
        super().__init__()
        self.patient = patient
        self.selected_type = "ООД"
        # Список полей ввода для навигации
        self.input_fields = []
        self.current_field_index = 0

    def compose(self) -> ComposeResult:
        with Vertical(id="outcome_add_dialog"):
            yield Label("Добавление исхода", classes="title-label")
            yield Label(f"{self.patient.get('last_name', '')} {self.patient.get('first_name', '')} {self.patient.get('patronymic', '')}", classes="field-label")

            yield Label("Тип исхода:", classes="field-label")
            with Horizontal(id="outcome_type_container"):
                btn_ood = Button("ООД", variant="primary", id="btn_type_ood")
                btn_vyazki = Button("вязки", id="btn_type_vyazki")
                yield btn_ood
                yield btn_vyazki

            yield Label("Дата (ГГГГ.ММ.ДД):", classes="field-label")
            date_input = Input(placeholder=datetime.now().strftime('%Y.%m.%d'), id="date_input")
            yield date_input
            self.input_fields.append("date_input")

            yield Label("Описание (необязательно):", classes="field-label")
            desc_input = TextArea("", id="description_input")
            yield desc_input
            self.input_fields.append("description_input")

            with Horizontal(classes="dialog_buttons"):
                btn_add = Button("Добавить", variant="primary", id="btn_add")
                btn_cancel = Button("Отмена", variant="error", id="btn_cancel")
                yield btn_add
                yield btn_cancel
            self.input_fields.append("btn_add")
            self.input_fields.append("btn_cancel")

    def on_mount(self) -> None:
        self.query_one("#date_input").focus()
        self.query_one("#btn_type_ood").add_class("selected")

    def on_key(self, event: events.Key) -> None:
        """Обработка клавиш для навигации между полями"""
        if event.key == "escape":
            self.action_cancel()
            event.prevent_default()
            event.stop()
            return

        focused = self.focused

        if event.key == "down":
            if isinstance(focused, (Input, TextArea, Button)):
                self._focus_next_field(focused)
                event.prevent_default()
                event.stop()
                return
        elif event.key == "up":
            if isinstance(focused, (Input, TextArea, Button)):
                self._focus_prev_field(focused)
                event.prevent_default()
                event.stop()
                return

    def _focus_next_field(self, current_widget):
        """Переход к следующему полю"""
        current_id = getattr(current_widget, 'id', None)
        if current_id is None:
            return

        try:
            current_index = self.input_fields.index(current_id)
            next_index = (current_index + 1) % len(self.input_fields)
            next_id = self.input_fields[next_index]
            next_widget = self.query_one(f"#{next_id}")
            if next_widget:
                next_widget.focus()
        except (ValueError, NoMatches):
            pass

    def _focus_prev_field(self, current_widget):
        """Переход к предыдущему полю"""
        current_id = getattr(current_widget, 'id', None)
        if current_id is None:
            return

        try:
            current_index = self.input_fields.index(current_id)
            prev_index = (current_index - 1) % len(self.input_fields)
            prev_id = self.input_fields[prev_index]
            prev_widget = self.query_one(f"#{prev_id}")
            if prev_widget:
                prev_widget.focus()
        except (ValueError, NoMatches):
            pass

    @on(Button.Pressed, "#btn_type_ood")
    def handle_type_ood(self, event: Button.Pressed) -> None:
        self.selected_type = "ООД"
        self.query_one("#btn_type_ood").add_class("selected")
        self.query_one("#btn_type_vyazki").remove_class("selected")

    @on(Button.Pressed, "#btn_type_vyazki")
    def handle_type_vyazki(self, event: Button.Pressed) -> None:
        self.selected_type = "вязки"
        self.query_one("#btn_type_vyazki").add_class("selected")
        self.query_one("#btn_type_ood").remove_class("selected")

    @on(Button.Pressed, "#btn_add")
    def handle_add(self, event: Button.Pressed) -> None:
        try:
            date_input = self.query_one("#date_input", Input)
            date = date_input.value.strip()

            if not date:
                date = datetime.now().strftime('%Y.%m.%d')

            # Проверяем дату
            is_valid, error_msg = DataManager.validate_date(date)
            if not is_valid:
                self.app.custom_notify(error_msg, severity="error")
                return

            desc = self.query_one("#description_input", TextArea)
            description = desc.text.strip()

            self.dismiss({
                "type": self.selected_type,
                "date": date,
                "description": description
            })

        except Exception as e:
            self.app.custom_notify(f"Ошибка: {e}", severity="error")

    @on(Button.Pressed, "#btn_cancel")
    def handle_cancel(self, event: Button.Pressed) -> None:
        self.dismiss(None)

    def action_cancel(self) -> None:
        """Закрыть диалог без сохранения"""
        self.dismiss(None)


class DeleteOutcomeDialog(ModalScreen[dict]):
    """Диалог выбора и подтверждения удаления исхода"""

    # Отключаем встроенную прокрутку ModalScreen
    SCROLL_VIEW = False

    CSS = """
    DeleteOutcomeDialog {
        align: center middle;
        background: transparent;
    }

    #delete_dialog {
        width: 62;
        height: auto;
        max-height: 85%;
        background: $background;
        border: round $accent;
        padding: 2 3;
    }

    #delete_dialog .title-label {
        width: 100%;
        height: 2;
        content-align: center middle;
        color: $accent;
        text-style: bold;
        margin-bottom: 1;
    }

    #delete_dialog .subtitle-label {
        width: 100%;
        height: 1;
        content-align: center middle;
        color: $text;
        margin-bottom: 1;
    }

    #delete_outcome_list {
        height: auto;
        max-height: 16;
        border: round $secondary;
        background: transparent;
    }

    #delete_outcome_list:focus {
        border: round $accent;
        background: transparent;
    }

    #delete_outcome_list ListItem {
        padding-left: 1;
        background: $background;
        color: $text;
    }

    #delete_outcome_list > ListItem.--highlight {
        background: $background;
    }

    #delete_outcome_list > ListItem.--highlight > Static {
        color: $accent;
        text-style: bold;
    }

    #delete_dialog .dialog_buttons {
        margin-top: 1;
        align: center middle;
        height: 3;
    }

    #delete_dialog .dialog_buttons Button {
        width: 20;
        height: 3;
        margin: 0 1;
        background: $background;
        color: $secondary;
        border: round $secondary;
    }

    #delete_dialog .dialog_buttons Button:focus {
        border: round $accent;
        color: $accent;
        background: transparent;
        text-style: bold;
    }

    #delete_dialog .dialog_buttons Button.variant-error {
        border: round red;
        color: red;
    }

    #delete_dialog .dialog_buttons Button.variant-primary {
        border: round $accent;
        color: $accent;
    }
    """

    BINDINGS = [
        Binding("escape", "cancel", "Отмена", show=False, priority=True),
        Binding("enter", "confirm_delete", "Удалить", show=False),
    ]

    def __init__(self, patient: dict):
        super().__init__()
        self.patient = patient
        self._outcomes = []

    def compose(self) -> ComposeResult:
        self._outcomes = self.patient.get("outcomes", [])

        with Vertical(id="delete_dialog"):
            yield Label("Удаление исхода", classes="title-label")
            fio = f"{self.patient.get('last_name', '')} {self.patient.get('first_name', '')} {self.patient.get('patronymic', '')}"
            yield Label(fio, classes="subtitle-label")

            yield ListView(id="delete_outcome_list")

            with Horizontal(classes="dialog_buttons"):
                yield Button("Удалить (Enter)", variant="error", id="btn_delete")
                yield Button("Отмена (Esc)", variant="primary", id="btn_cancel")

    def on_mount(self) -> None:
        """Заполняем список исходов после примонтирования ListView в DOM"""
        try:
            list_view = self.query_one("#delete_outcome_list", ListView)

            if not self._outcomes:
                label = ListLabel("Нет исходов для удаления")
                list_view.append(ListItem(label))
            else:
                for outcome in self._outcomes:
                    outcome_type = outcome.get("type", "")
                    date = outcome.get("date", "")
                    text = f"{outcome_type} от {date}"

                    label = ListLabel(text)
                    item = ListItem(label)
                    item.outcome = outcome
                    list_view.append(item)

            if list_view.children:
                list_view.index = 0
                self._update_highlight(list_view)
            list_view.focus()
        except NoMatches:
            pass

    @on(ListView.Highlighted, "#delete_outcome_list")
    def handle_list_highlighted(self, event: ListView.Highlighted) -> None:
        """Обновляет маркер выделения при навигации"""
        try:
            list_view = self.query_one("#delete_outcome_list", ListView)
            self._update_highlight(list_view)
        except NoMatches:
            pass

    def _update_highlight(self, list_view: ListView) -> None:
        """Обновляет маркеры █ у всех элементов списка"""
        for i, child in enumerate(list_view.children):
            if hasattr(child, "children") and child.children:
                label = child.children[0]
                if isinstance(label, ListLabel):
                    label.is_highlighted = (i == list_view.index)

    @on(ListView.Selected, "#delete_outcome_list")
    def handle_list_selected(self, event: ListView.Selected) -> None:
        """Выбор элемента через Enter — подтверждаем удаление"""
        self.action_confirm_delete()

    def action_confirm_delete(self) -> None:
        """Удалить выбранный исход"""
        try:
            list_view = self.query_one("#delete_outcome_list", ListView)
            if list_view.index is not None and list_view.index < len(list_view.children):
                item = list_view.children[list_view.index]
                if hasattr(item, "outcome"):
                    self.dismiss({"id": item.outcome.get("id")})
                    return
        except NoMatches:
            pass
        self.app.custom_notify("Сначала выберите исход из списка", severity="warning")

    @on(Button.Pressed, "#btn_delete")
    def handle_delete(self, event: Button.Pressed) -> None:
        """Кнопка «Удалить»"""
        self.action_confirm_delete()

    @on(Button.Pressed, "#btn_cancel")
    def handle_cancel(self, event: Button.Pressed) -> None:
        """Кнопка «Отмена»"""
        self.dismiss(None)

    def action_cancel(self) -> None:
        """Закрыть диалог без удаления"""
        self.dismiss(None)

    def on_key(self, event: events.Key) -> None:
        """Обработка клавиш — только escape для закрытия"""
        if event.key == "escape":
            self.action_cancel()
            event.prevent_default()
            event.stop()



class OpinionsMainScreen(Screen):
    """Главный экран управления прогнозами специалистов"""

    BINDINGS = [
        Binding("escape", "go_back_to_data", "Назад", show=True, priority=True),
        Binding("1", "press_view", "Просмотр", show=False),
        Binding("2", "press_maketab", "Шаблон", show=False),
        Binding("3", "press_import", "Импорт", show=False),
    ]

    def action_press_view(self):
        self.query_one("#btn_view").press()

    def action_press_maketab(self):
        self.query_one("#btn_maketab").press()

    def action_press_import(self):
        self.query_one("#btn_import").press()

    OPINIONS_DESCRIPTIONS = {
        "btn_import": "Для автоматической загрузки прогнозов специалистов вам необходимо подготовить файл с данными.\n\nСоздайте шаблон для данных с помощью кнопки 2, внесите в него имеющиеся прогнозы, затем выполните загрузку кнопкой 3.\n\nФайл должен называться [bold $accent]Арсенал - прогнозы.xlsx[/] и находиться в папке Документы.\n\nДля пользователей MacOS:\nЕсли вы используете Numbers, после заполнения таблицы выберите Файл ➜ Экспорт в ➜ Excel…, сохраните как [bold $accent]Арсенал - прогнозы.xlsx[/], а затем произведите загрузку из шаблона.",

        "btn_maketab": "Эта функция создает в папке Документы таблицу со списком пациентов, которым были проведены оценки, и полями для ввода прогнозов специалистов.\n\nНайдите в папке Документы файл [bold $accent]Арсенал - прогнозы.xlsx[/], внесите в него сведения о прогнозах специалистов с датами. Если нужно внести несколько прогнозов для одного пациента, скопируйте его строку несколько раз. Затем произведите загрузку из шаблона.",

        "btn_view": "Просмотр и управление прогнозами специалистов по каждому пациенту.",

        "btn_back": "Вернуться в раздел работы с данными."
    }

    def compose(self) -> ComposeResult:
        with Horizontal():
            with Vertical(id="outcomes_main_hint_panel") as hint_col:
                hint_col.border_title = "Мнения специалистов"
                yield Static("", id="outcomes_main_hint_text")

            with Vertical(id="outcomes_main_buttons_panel") as btn_col:
                btn_col.border_title = "Действия"
                with Vertical(id="outcomes_main_buttons_container"):
                    yield HoverButton("1. Просмотр прогнозов            ", id="btn_view")
                    yield HoverButton("2. Создать шаблон для данных     ", id="btn_maketab")
                    yield HoverButton("3. Загрузка из шаблона           ", id="btn_import")
                    yield Button("", id="btn_spacer", disabled=True, classes="spacer")
                    yield HoverButton("Esc. Назад                       ", id="btn_back")
        yield Footer(show_command_palette=False)

    def on_descendant_focus(self, event: events.Focus) -> None:
        target = getattr(event, "control", None) or getattr(event, "node", None)
        if isinstance(target, Button) and target.id:
            self._update_hint(target.id)

    def _update_hint(self, btn_id: str) -> None:
        description = self.OPINIONS_DESCRIPTIONS.get(btn_id, "Выберите действие...")
        try:
            self.query_one("#outcomes_main_hint_text", Static).update(description)
        except NoMatches:
            pass

    def on_mount(self) -> None:
        self.query_one("#btn_view").focus()
        self._update_hint("btn_view")

    def on_mouse_move(self, event: events.MouseMove) -> None:
        result = self.get_widget_at(event.screen_x, event.screen_y)
        if result is None:
            return
        if isinstance(result, tuple):
            widget = result[0]
        else:
            widget = result
        current = widget
        while current:
            if isinstance(current, Button):
                if current.id and current.id.startswith("btn_") and self.focused != current:
                    current.focus()
                    self._update_hint(current.id)
                return
            current = current.parent

    def on_button_pressed(self, event: Button.Pressed) -> None:
        btn_id = event.button.id
        if btn_id == "btn_import":
            self._run_import_opinions()
        elif btn_id == "btn_maketab":
            self._create_opinions_template()
        elif btn_id == "btn_view":
            self.app.push_screen(OpinionsScreen())
        elif btn_id == "btn_back":
            self.action_go_back_to_data()

    def _create_opinions_template(self):
        """Создает Excel-шаблон для ввода прогнозов специалистов"""
        try:
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font, PatternFill
            except ImportError:
                self.app.custom_notify(
                    "Для создания таблицы необходимо установить библиотеку openpyxl.\n\n"
                    "Выполните в терминале команду:\n"
                    "pip install openpyxl",
                    severity="error"
                )
                return

            docs_dir = self.get_documents_dir()
            filepath = docs_dir / "Арсенал - прогнозы.xlsx"

            if filepath.exists():
                def handle_dialog_result(confirmed: bool):
                    if confirmed:
                        self._create_opinions_template_file(filepath)
                    else:
                        self.app.custom_notify("Создание шаблона отменено", severity="info")
                self.app.push_screen(ConfirmTemplateDialog(), handle_dialog_result)
            else:
                self._create_opinions_template_file(filepath)

        except Exception as e:
            self.app.custom_notify(f"Ошибка создания шаблона: {e}", severity="error")
            import traceback
            traceback.print_exc()

    def _create_opinions_template_file(self, filepath: Path):
        """Создает файл шаблона для прогнозов"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Прогнозы"

            headers = [
                "Фамилия",
                "Имя",
                "Отчество",
                "Год рождения",
                "UID пациента",
                "Дата прогноза",
                "Оценка риска (1-10)",
                "Специалист",
                "Комментарий"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            patients = self.app.results.data
            if not patients:
                self.app.custom_notify("Нет пациентов с оценками. Сначала проведите хотя бы одну оценку.", severity="warning")
                return

            patients.sort(key=lambda p: p.get("last_name", "").lower())

            row_num = 2
            for patient in patients:
                last_name = patient.get("last_name", "")
                first_name = patient.get("first_name", "")
                patronymic = patient.get("patronymic", "")
                birth_year = patient.get("birth_year", "")
                uid = patient.get("uid", "")

                opinions = patient.get("opinions", [])

                if opinions:
                    for opinion in opinions:
                        date = opinion.get("date", "")
                        risk_score = opinion.get("risk_score", "")
                        expert = opinion.get("expert", "")
                        comment = opinion.get("comment", "")

                        ws.cell(row=row_num, column=1, value=last_name)
                        ws.cell(row=row_num, column=2, value=first_name)
                        ws.cell(row=row_num, column=3, value=patronymic)
                        ws.cell(row=row_num, column=4, value=birth_year)
                        ws.cell(row=row_num, column=5, value=uid)
                        ws.cell(row=row_num, column=6, value=date)
                        ws.cell(row=row_num, column=7, value=risk_score)
                        ws.cell(row=row_num, column=8, value=expert)
                        ws.cell(row=row_num, column=9, value=comment)
                        row_num += 1
                else:
                    ws.cell(row=row_num, column=1, value=last_name)
                    ws.cell(row=row_num, column=2, value=first_name)
                    ws.cell(row=row_num, column=3, value=patronymic)
                    ws.cell(row=row_num, column=4, value=birth_year)
                    ws.cell(row=row_num, column=5, value=uid)
                    row_num += 1

            column_widths = {
                'A': 20, 'B': 15, 'C': 15, 'D': 12,
                'E': 12, 'F': 15, 'G': 18, 'H': 25, 'I': 40
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

            wb.save(filepath)
            open_file_externally(filepath)
            self.app.custom_notify(f"Шаблон для прогнозов создан:\n{filepath}", severity="success")

        except Exception as e:
            self.app.custom_notify(f"Ошибка создания файла шаблона: {e}", severity="error")
            import traceback
            traceback.print_exc()

    def _run_import_opinions(self):
        """Импорт прогнозов специалистов из файлов"""
        def confirm_import(do_import: bool):
            if do_import:
                self._do_import_opinions()

        self.app.push_screen(
            ConfirmDialog(
                "Загрузить прогнозы специалистов сейчас?\n\n"
                "Уже внесенные прогнозы сохранятся и не будут продублированы."
            ),
            confirm_import
        )

    def _do_import_opinions(self):
        """Выполняет импорт прогнозов специалистов"""
        try:
            total_added = 0
            errors = []

            excel_imported = self._import_opinions_from_excel()

            for opinion in excel_imported:
                patient_uid = opinion.get("patient_uid")
                if not patient_uid:
                    continue

                patient = self.app.results.get_patient_by_uid(patient_uid)
                if not patient:
                    errors.append(f"Пациент с UID {patient_uid} не найден")
                    continue

                existing = {(o.get("date"), o.get("expert")) for o in patient.get("opinions", [])}
                key = (opinion["date"], opinion["expert"])

                if key not in existing:
                    try:
                        self.app.results.add_opinion(
                            patient_uid=patient_uid,
                            expert=opinion["expert"],
                            risk_score=opinion["risk_score"],
                            date=opinion["date"],
                            comment=opinion.get("comment", "")
                        )
                        total_added += 1
                        existing.add(key)
                    except Exception as e:
                        errors.append(f"{patient.get('last_name', '')}: {e}")

            if total_added > 0:
                self.app.custom_notify(f"Добавлено {total_added} прогнозов", severity="success")
            else:
                if errors:
                    self.app.custom_notify(f"Ошибок: {len(errors)}, новых прогнозов не найдено", severity="info")
                else:
                    self.app.custom_notify("Новых прогнозов не найдено", severity="info")

        except Exception as e:
            self.app.custom_notify(f"Ошибка импорта: {e}", severity="error")
            import traceback
            traceback.print_exc()

    def _import_opinions_from_excel(self) -> list:
        """Импортирует прогнозы из Excel-файла"""
        opinions = []
        docs_dir = self.get_documents_dir()

        possible_files = [
            docs_dir / "Арсенал - прогнозы.xlsx",
            docs_dir / "Арсенал - прогнозы.xls",
            docs_dir / "Арсенал - прогнозы.xlsm",
        ]

        filepath = None
        for f in possible_files:
            if f.exists():
                filepath = f
                break

        if not filepath:
            numbers_file = docs_dir / "Арсенал - прогнозы.numbers"
            if numbers_file.exists():
                self.app.custom_notify(
                    "Обнаружен файл в формате Numbers.\n\n"
                    "Для импорта данных:\n"
                    "1. Откройте файл в Numbers\n"
                    "2. Выберите Файл ➜ Экспорт в ➜ Excel…\n"
                    "3. Сохраните как 'Арсенал - прогнозы.xlsx'\n"
                    "4. Повторите импорт",
                    severity="warning"
                )
            else:
                self.app.custom_notify(
                    "Файл 'Арсенал - прогнозы.xlsx' не найден.\n\n"
                    "Сначала создайте шаблон (кнопка 2) и внесите данные.",
                    severity="warning"
                )
            return opinions

        try:
            try:
                import openpyxl
            except ImportError:
                self.app.custom_notify(
                    "Для работы с Excel необходимо установить библиотеку openpyxl.\n\n"
                    "Выполните в терминале команду:\n"
                    "pip install openpyxl",
                    severity="error"
                )
                return opinions

            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active

            headers = {}
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers[cell_value] = col

            required_cols = ["UID пациента", "Дата прогноза", "Оценка риска (1-10)", "Специалист"]
            for col in required_cols:
                if col not in headers:
                    self.app.custom_notify(
                        f"В файле не найдена колонка '{col}'.\n\n"
                        "Проверьте, что вы используете правильный шаблон.",
                        severity="error"
                    )
                    return opinions

            for row in range(2, ws.max_row + 1):
                uid = ws.cell(row=row, column=headers["UID пациента"]).value
                date = ws.cell(row=row, column=headers["Дата прогноза"]).value
                risk_score = ws.cell(row=row, column=headers["Оценка риска (1-10)"]).value
                expert = ws.cell(row=row, column=headers["Специалист"]).value
                comment_col = headers.get("Комментарий")
                comment = ws.cell(row=row, column=comment_col).value if comment_col else ""

                if not uid or not date or risk_score is None or not expert:
                    continue

                uid = str(uid)

                if isinstance(date, datetime):
                    date_str = date.strftime('%Y.%m.%d')
                elif isinstance(date, str):
                    date_str = self._normalize_date(date)
                    if not date_str:
                        continue
                else:
                    continue

                # Добавляем проверку даты
                is_valid, error_msg = DataManager.validate_date(date_str)
                if not is_valid:
                    # Пропускаем строку с невалидной датой
                    continue

                try:
                    risk_score = int(risk_score)
                    if risk_score < 1 or risk_score > 10:
                        continue
                except (ValueError, TypeError):
                    continue

                expert = str(expert).strip()
                if not expert:
                    continue

                opinions.append({
                    "patient_uid": uid,
                    "expert": expert,
                    "risk_score": risk_score,
                    "date": date_str,
                    "comment": str(comment) if comment else ""
                })

        except Exception as e:
            self.app.custom_notify(
                f"Ошибка чтения файла:\n{str(e)}\n\n"
                "Убедитесь, что файл не открыт в другой программе.",
                severity="error"
            )
            import traceback
            traceback.print_exc()

        return opinions

    def _normalize_date(self, date_str: str) -> str:
        """Нормализует дату из разных форматов в YYYY.MM.DD и проверяет валидность"""
        date_str = date_str.strip()

        if re.match(r'^\d{2}\.\d{2}\.\d{4}$', date_str):
            parts = date_str.split('.')
            normalized = f"{parts[2]}.{parts[1]}.{parts[0]}"
        elif re.match(r'^\d{4}\.\d{2}\.\d{2}$', date_str):
            normalized = date_str
        elif re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
            parts = date_str.split('-')
            normalized = f"{parts[0]}.{parts[1]}.{parts[2]}"
        elif re.match(r'^\d{2}/\d{2}/\d{4}$', date_str):
            parts = date_str.split('/')
            normalized = f"{parts[2]}.{parts[1]}.{parts[0]}"
        else:
            try:
                from datetime import datetime
                for fmt in ['%d.%m.%Y', '%Y.%m.%d', '%Y-%m-%d', '%d/%m/%Y']:
                    try:
                        dt = datetime.strptime(date_str, fmt)
                        normalized = dt.strftime('%Y.%m.%d')
                        break
                    except ValueError:
                        continue
                else:
                    return None
            except Exception:
                return None

        # Проверяем валидность нормализованной даты
        is_valid, _ = DataManager.validate_date(normalized)
        return normalized if is_valid else None

    def get_documents_dir(self) -> Path:
        """Определяет путь к папке Документы"""
        import platform
        system = platform.system()
        home = Path.home()

        if system == "Windows":
            docs = home / "Documents"
            if docs.exists():
                return docs
            return home / "Мои документы"
        elif system == "Darwin":
            return home / "Documents"
        else:
            try:
                import subprocess
                result = subprocess.run(['xdg-user-dir', 'DOCUMENTS'],
                                        capture_output=True, text=True)
                if result.returncode == 0:
                    docs = Path(result.stdout.strip())
                    if docs.exists():
                        return docs
            except (FileNotFoundError, subprocess.SubprocessError):
                pass
            return home / "Документы"

    def action_go_back_to_data(self) -> None:
        self.app.pop_screen()


class OpinionsScreen(Screen):
    """Экран просмотра прогнозов специалистов"""

    CSS = """
        #opinions_list_panel {
            width: 50%;
            border: round $secondary;
            background: transparent;
            padding: 1 0 1 0;
            margin: 0;
            border-title-align: left;
            border-title-color: $secondary;
        }

        #opinions_detail_panel {
            width: 50%;
            border: round $secondary;
            background: transparent;
            padding: 0;
            margin: 0;
            border-title-align: left;
            border-title-color: $secondary;
        }

        #opinions_list_panel:focus-within,
        #opinions_detail_panel:focus-within {
            border: round $accent;
            border-title-color: $accent;
            border-title-style: bold;
            background: transparent;
        }

        #opinions_scroll {
            scrollbar-gutter: stable;
            scrollbar-size: 1 1;
            scrollbar-color: $accent 30%;
            scrollbar-color-hover: $accent 30%;
            scrollbar-background: transparent;
            background: transparent;
            padding: 0;
            margin: 0;
        }

        #opinions_text {
            padding: 1 2 1 2;
            color: $text;
            text-wrap: wrap;
            width: 100%;
            background: transparent;
        }

        #opinions_buttons_container {
            width: 100%;
            height: auto;
            align: left top;
            padding: 1 0 0 1;
        }

        #opinions_buttons_container Button {
            width: 90%;
            height: 3;
            margin: 1 1;
            background: $background;
            color: $secondary;
            border: transparent;
            content-align: left middle;
            padding: 0 2;
        }

        #opinions_buttons_container Button:focus {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
            content-align: left middle;
        }

        #opinions_buttons_container Button:hover {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
        }

        #opinions_buttons_container Button.variant-primary {
            border: round $accent;
            color: $accent;
        }

        #opinions_buttons_container Button.variant-error {
            border: round red;
            color: red;
        }

        #opinions_buttons_container Button.spacer {
            height: 1;
            background: transparent;
            border: none;
            margin: 0;
            padding: 0;
            width: 100%;
        }

        #opinions_buttons_container Button.spacer:focus {
            border: none;
            background: transparent;
        }

        #opinions_buttons_container Button.spacer:hover {
            border: none;
            background: transparent;
        }

        #opinions_buttons_container Button > Static {
            width: 100%;
            content-align: left middle;
        }
    """

    BINDINGS = [
        Binding("escape", "go_back", "Назад", show=True, priority=True),
        Binding("right", "focus_buttons", "К кнопкам", show=False),
        Binding("left", "focus_list", "К списку", show=False),
        Binding("up", "move_up", "Вверх", show=False),
        Binding("down", "move_down", "Вниз", show=False),
        Binding("pageup", "scroll_view_up", "Вверх", show=True),
        Binding("pagedown", "scroll_view_down", "Вниз", show=True),
        Binding("enter", "select_item", "Выбрать", show=True),
        Binding("1", "press_add", "Добавить", show=False),
        Binding("2", "press_delete", "Удалить", show=False),
        Binding("3", "press_write", "Записать", show=False),
    ]

    def action_press_add(self):
        if self.selected_patient:
            self.show_add_opinion_dialog()
        else:
            self.app.custom_notify("Сначала выберите пациента", severity="warning")

    def action_press_delete(self):
        if self.selected_patient:
            opinions = self.selected_patient.get("opinions", [])
            if opinions:
                self.show_delete_opinion_dialog()
            else:
                self.app.custom_notify("Нет прогнозов для удаления", severity="info")
        else:
            self.app.custom_notify("Сначала выберите пациента", severity="warning")

    def action_press_write(self):
        if self.selected_patient:
            self.action_write_to_file()
        else:
            self.app.custom_notify("Сначала выберите пациента", severity="warning")

    def __init__(self):
        super().__init__()
        self.selected_patient = None
        self.selected_patient_uid = None  # Сохраняем UID выбранного пациента
        self.view_buttons = []
        self.current_button_index = 0
        self._is_on_list = True

    def compose(self) -> ComposeResult:
        with Horizontal():
            with Vertical(id="opinions_list_panel", classes="panel-left") as list_panel:
                list_panel.border_title = "Пациенты"
                with VerticalScroll(id="opinions_scroll"):
                    yield ListView(id="patients_list")

            with Vertical(id="opinions_detail_panel", classes="panel-right") as detail_panel:
                detail_panel.border_title = "Прогнозы"
                with VerticalScroll(id="opinions_scroll"):
                    yield Static("", id="opinions_text", markup=False)
                    yield Vertical(id="opinions_buttons_container")
        yield Footer(show_command_palette=False)

    def on_mount(self) -> None:
        self.refresh_patients_list()
        self.show_buttons()

        # Восстанавливаем выбранного пациента, если он был
        if self.selected_patient_uid:
            list_view = self.query_one("#patients_list", ListView)
            for i, item in enumerate(list_view.children):
                if hasattr(item, 'patient') and item.patient.get('uid') == self.selected_patient_uid:
                    list_view.index = i
                    self.selected_patient = item.patient
                    self.show_patient_detail(item.patient)
                    break

        try:
            list_view = self.query_one("#patients_list", ListView)
            list_view.focus()
            self._is_on_list = True
        except NoMatches:
            pass

    def refresh_patients_list(self) -> None:
        list_view = self.query_one("#patients_list", ListView)

        # Запоминаем текущий индекс и UID выбранного пациента до обновления
        current_index = list_view.index if list_view.index is not None else 0
        current_patient_uid = None
        if self.selected_patient:
            current_patient_uid = self.selected_patient.get('uid')

        list_view.clear()

        patients = self.app.results.data
        if not patients:
            label = ListLabel("Нет пациентов в базе данных.")
            list_view.append(ListItem(label))
            return

        patients.sort(key=lambda p: p.get("last_name", "").lower())

        for patient in patients:
            last_name = patient.get("last_name", "")
            first_name = patient.get("first_name", "")
            patronymic = patient.get("patronymic", "")
            birth_year = patient.get("birth_year", "")
            opinions_count = len(patient.get("opinions", []))

            display = f"{last_name} {first_name} {patronymic} {birth_year} г.р."
            if opinions_count > 0:
                display += f" [прогнозов: {opinions_count}]"

            label = ListLabel(display)
            item = ListItem(label)
            item.patient = patient
            list_view.append(item)

        if list_view.children:
            # Восстанавливаем выбранного пациента по UID
            if current_patient_uid:
                for i, item in enumerate(list_view.children):
                    if hasattr(item, 'patient') and item.patient.get('uid') == current_patient_uid:
                        list_view.index = i
                        self.selected_patient = item.patient
                        return
            # Если не нашли или нет сохраненного UID, выбираем первого
            if list_view.index is None or list_view.index >= len(list_view.children):
                list_view.index = 0

    def show_patient_detail(self, patient: dict) -> None:
        self.selected_patient = patient
        self.selected_patient_uid = patient.get('uid')  # Сохраняем UID

        last_name = patient.get("last_name", "")
        first_name = patient.get("first_name", "")
        patronymic = patient.get("patronymic", "")
        birth_year = patient.get("birth_year", "")

        opinions = patient.get("opinions", [])

        lines = []
        lines.append(f"Пациент: {last_name} {first_name} {patronymic} {birth_year} г.р.")
        lines.append(f"UID: {patient.get('uid', '')}")
        lines.append(f"Всего прогнозов: {len(opinions)}")
        lines.append("")

        if opinions:
            lines.append("Прогнозы специалистов:")
            lines.append("")
            sorted_opinions = sorted(opinions, key=lambda o: o.get("date", ""), reverse=False)
            for i, opinion in enumerate(sorted_opinions, 1):
                risk = opinion.get("risk_score", 0)
                expert = opinion.get("expert", "")
                date = opinion.get("date", "")
                comment = opinion.get("comment", "")

                # Отображаем риски от 1 до 10
                stars = "█" * risk + "░" * (10 - risk)
                lines.append(f"  {i}. {risk}/10 {stars} - {expert} от {date}")
                if comment:
                    lines.append(f"     {comment}")
                lines.append("")
        else:
            lines.append("Нет зарегистрированных прогнозов.")

        self.query_one("#opinions_text", Static).update("\n".join(lines))
        self.show_buttons(patient)

    def show_buttons(self, patient: dict = None) -> None:
        """Показывает кнопки действий"""
        try:
            container = self.query_one("#opinions_buttons_container", Vertical)

            for child in list(container.children):
                child.remove()

            self.view_buttons = []

            if patient:
                btn_add_id = f"btn_add_opinion_{uuid.uuid4().hex[:6]}"
                btn_add = Button("1. Добавить прогноз                        ", variant="primary", id=btn_add_id)
                container.mount(btn_add)
                self.view_buttons.append(btn_add_id)

                if patient.get("opinions"):
                    btn_delete_id = f"btn_delete_opinion_{uuid.uuid4().hex[:6]}"
                    btn_delete = Button("2. Удалить прогноз                         ", variant="error",
                                        id=btn_delete_id)
                    container.mount(btn_delete)
                    self.view_buttons.append(btn_delete_id)

                btn_write_id = f"btn_write_opinion_{uuid.uuid4().hex[:6]}"
                btn_write = Button("3. Записать в файл пациента                ", variant="success",
                                   id=btn_write_id)
                container.mount(btn_write)
                self.view_buttons.append(btn_write_id)
            else:
                btn_back_id = f"btn_back_{uuid.uuid4().hex[:6]}"
                btn_back = Button("Esc. Назад                ", variant="default", id=btn_back_id)
                container.mount(btn_back)
                self.view_buttons.append(btn_back_id)

            self.current_button_index = 0

        except NoMatches as e:
            print(f"show_buttons error: {e}")

    def action_focus_list(self) -> None:
        try:
            list_view = self.query_one("#patients_list", ListView)
            list_view.focus()
            self._is_on_list = True
        except NoMatches:
            pass

    def action_focus_buttons(self) -> None:
        if self.view_buttons:
            self.current_button_index = 0
            try:
                self.query_one(f"#{self.view_buttons[0]}").focus()
                self._is_on_list = False
            except NoMatches:
                pass

    def action_move_up(self) -> None:
        if self._is_on_list:
            try:
                list_view = self.query_one("#patients_list", ListView)
                if list_view.index > 0:
                    list_view.index -= 1
            except NoMatches:
                pass
        else:
            if not self.view_buttons:
                return
            focused = self.focused
            if not isinstance(focused, Button) or focused.id not in self.view_buttons:
                try:
                    self.query_one(f"#{self.view_buttons[-1]}").focus()
                except NoMatches:
                    pass
                return
            current_index = -1
            for i, btn_id in enumerate(self.view_buttons):
                try:
                    if self.query_one(f"#{btn_id}").has_focus:
                        current_index = i
                        break
                except NoMatches:
                    pass
            if current_index > 0:
                self.current_button_index = current_index - 1
                try:
                    self.query_one(f"#{self.view_buttons[self.current_button_index]}").focus()
                except NoMatches:
                    pass

    def action_move_down(self) -> None:
        if self._is_on_list:
            try:
                list_view = self.query_one("#patients_list", ListView)
                if list_view.index < len(list_view.children) - 1:
                    list_view.index += 1
            except NoMatches:
                pass
        else:
            if not self.view_buttons:
                return
            focused = self.focused
            if not isinstance(focused, Button) or focused.id not in self.view_buttons:
                try:
                    self.query_one(f"#{self.view_buttons[0]}").focus()
                except NoMatches:
                    pass
                return
            current_index = -1
            for i, btn_id in enumerate(self.view_buttons):
                try:
                    if self.query_one(f"#{btn_id}").has_focus:
                        current_index = i
                        break
                except NoMatches:
                    pass
            if current_index < len(self.view_buttons) - 1:
                self.current_button_index = current_index + 1
                try:
                    self.query_one(f"#{self.view_buttons[self.current_button_index]}").focus()
                except NoMatches:
                    pass

    def action_scroll_view_up(self) -> None:
        try:
            scroll = self.query_one("#opinions_scroll")
            scroll.scroll_up()
        except NoMatches:
            pass

    def action_scroll_view_down(self) -> None:
        try:
            scroll = self.query_one("#opinions_scroll")
            scroll.scroll_down()
        except NoMatches:
            pass

    def action_select_item(self) -> None:
        if not self._is_on_list:
            return
        try:
            list_view = self.query_one("#patients_list", ListView)
            if list_view.index is not None and list_view.index < len(list_view.children):
                item = list_view.children[list_view.index]
                if hasattr(item, 'patient'):
                    self.show_patient_detail(item.patient)
        except NoMatches:
            pass

    def action_go_back(self) -> None:
        self.app.pop_screen()

    def on_key(self, event: events.Key) -> None:
        if event.key == "up":
            self.action_move_up()
            event.stop()
            event.prevent_default()
            return
        elif event.key == "down":
            self.action_move_down()
            event.stop()
            event.prevent_default()
            return
        elif event.key == "right":
            if self._is_on_list:
                self.action_focus_buttons()
                event.stop()
                event.prevent_default()
                return
        elif event.key == "left":
            if not self._is_on_list:
                self.action_focus_list()
                event.stop()
                event.prevent_default()
                return
        elif event.key == "enter":
            if self._is_on_list:
                self.action_select_item()
                event.stop()
                event.prevent_default()
                return

    @on(ListView.Highlighted, "#patients_list")
    def handle_list_highlighted(self, event: ListView.Highlighted) -> None:
        if event.item and hasattr(event.item, 'patient'):
            try:
                list_view = self.query_one("#patients_list", ListView)
                for i, child in enumerate(list_view.children):
                    if hasattr(child, 'children') and child.children:
                        label = child.children[0]
                        if isinstance(label, ListLabel):
                            label.is_highlighted = (i == list_view.index)
            except NoMatches:
                pass
            self.show_patient_detail(event.item.patient)

    @on(ListView.Selected, "#patients_list")
    def handle_list_selected(self, event: ListView.Selected) -> None:
        if event.item and hasattr(event.item, 'patient'):
            self.show_patient_detail(event.item.patient)

    @on(Button.Pressed)
    def handle_button(self, event: Button.Pressed) -> None:
        """Обработчик нажатия кнопок"""
        button_id = event.button.id

        if button_id.startswith("btn_add_opinion"):
            self.show_add_opinion_dialog()
        elif button_id.startswith("btn_delete_opinion"):
            self.show_delete_opinion_dialog()
        elif button_id.startswith("btn_write_opinion"):
            self.action_write_to_file()
        elif button_id.startswith("btn_back"):
            self.action_go_back()

    def show_add_opinion_dialog(self):
        if not self.selected_patient:
            self.app.custom_notify("Сначала выберите пациента", severity="warning")
            return

        def after_dialog(result: dict):
            if result:
                try:
                    patient_uid = self.selected_patient.get("uid")
                    patient = self.app.results.get_patient_by_uid(patient_uid)
                    if not patient:
                        self.app.custom_notify("Пациент не найден", severity="error")
                        # Восстанавливаем отображение кнопок
                        self.show_buttons(self.selected_patient)
                        return

                    if "opinions" not in patient:
                        patient["opinions"] = []

                    opinion = {
                        "id": f"op_{uuid.uuid4().hex[:6]}",
                        "expert": result["expert"],
                        "risk_score": result["risk_score"],
                        "date": result["date"],
                        "comment": result["comment"],
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    patient["opinions"].append(opinion)
                    self.app.results._save_data()

                    self.app.custom_notify(f"Прогноз добавлен: {result['expert']} - {result['risk_score']}/10",
                                           severity="success")

                    # Обновляем список и показываем детали выбранного пациента
                    self.refresh_patients_list()
                    self.show_patient_detail(self.selected_patient)

                    # Устанавливаем фокус на кнопку "Добавить прогноз"
                    self.action_focus_buttons()

                except Exception as e:
                    self.app.custom_notify(f"Ошибка: {e}", severity="error")
                    self.show_buttons(self.selected_patient)
            else:
                # Если диалог отменен, возвращаем фокус на кнопку
                self.show_buttons(self.selected_patient)
                self.action_focus_buttons()

        self.app.push_screen(AddOpinionDialog(self.selected_patient), after_dialog)

    def show_delete_opinion_dialog(self):
        if not self.selected_patient:
            return

        opinions = self.selected_patient.get("opinions", [])
        if not opinions:
            self.app.custom_notify("Нет прогнозов для удаления", severity="info")
            return

        def after_dialog(result: dict):
            if result:
                try:
                    patient_uid = self.selected_patient.get("uid")
                    patient = self.app.results.get_patient_by_uid(patient_uid)
                    if not patient:
                        return

                    opinion_id = result.get("id")
                    opinions = patient.get("opinions", [])
                    initial_length = len(opinions)
                    patient["opinions"] = [o for o in opinions if o.get("id") != opinion_id]

                    if len(patient["opinions"]) < initial_length:
                        self.app.results._save_data()
                        self.app.custom_notify("Прогноз удален", severity="success")
                        self.refresh_patients_list()
                        self.show_patient_detail(self.selected_patient)
                        # Восстанавливаем фокус на кнопку "Добавить прогноз"
                        self.action_focus_buttons()
                    else:
                        self.app.custom_notify("Не удалось удалить прогноз", severity="error")
                        self.show_buttons(self.selected_patient)
                except Exception as e:
                    self.app.custom_notify(f"Ошибка: {e}", severity="error")
                    self.show_buttons(self.selected_patient)
            else:
                self.show_buttons(self.selected_patient)
                self.action_focus_buttons()

        self.app.push_screen(DeleteOpinionDialog(self.selected_patient), after_dialog)

    def action_write_to_file(self) -> None:
        """Записывает информацию о прогнозах в файл пациента"""
        if not self.selected_patient:
            self.app.custom_notify("Сначала выберите пациента", severity="warning")
            return

        try:
            patient = self.selected_patient
            last_name = patient.get("last_name", "")
            first_name = patient.get("first_name", "")
            patronymic = patient.get("patronymic", "")
            birth_year = str(patient.get("birth_year", ""))

            # Формируем имя файла
            filename_parts = [last_name, first_name, patronymic, birth_year]
            filename_base = " ".join(filter(None, filename_parts))
            filename_safe = filename_base.replace(" ", "_")
            filename = f"{filename_safe}.txt"
            filepath = self.app.results.reports_dir / filename

            # Генерируем содержимое для записи
            content = self.generate_opinions_content(patient)

            # Записываем в файл
            if filepath.exists():
                with open(filepath, "a", encoding="utf-8") as f:
                    f.write("\n\n" + content)
            else:
                filepath.write_text(content, encoding="utf-8")

            self.app.custom_notify(f"Данные о прогнозах записаны в файл:\n{filename}", severity="success")

        except Exception as e:
            self.app.custom_notify(f"Ошибка записи в файл: {e}", severity="error")
            import traceback
            traceback.print_exc()

    def generate_opinions_content(self, patient: dict) -> str:
        """Генерирует содержимое для записи о прогнозах в файл пациента"""
        last_name = patient.get("last_name", "")
        first_name = patient.get("first_name", "")
        patronymic = patient.get("patronymic", "")
        birth_year = patient.get("birth_year", "")

        opinions = patient.get("opinions", [])
        current_date = datetime.now().strftime('%Y.%m.%d')

        lines = []
        lines.append(f"Пациент {last_name} {first_name} {patronymic} {birth_year} г.р.")
        lines.append("╭" + "─" * 78 + "╮")
        lines.append(f"Данные о прогнозах специалистов на {current_date}")
        lines.append("")

        if opinions:
            sorted_opinions = sorted(opinions, key=lambda o: o.get("date", ""), reverse=False)
            for i, opinion in enumerate(sorted_opinions, 1):
                risk = opinion.get("risk_score", 0)
                expert = opinion.get("expert", "")
                date = opinion.get("date", "")
                comment = opinion.get("comment", "")

                stars = "█" * risk + "░" * (10 - risk)
                lines.append(f"  {i}. {risk}/10 {stars} - {expert} от {date}")
                if comment:
                    wrapped = textwrap.fill(comment, width=74)
                    for line in wrapped.split("\n"):
                        lines.append(f"     {line}")
                lines.append("")
        else:
            lines.append("Нет зарегистрированных прогнозов.")

        lines.append("")
        lines.append("╰" + "─" * 78 + "╯")

        return "\n".join(lines)


class AddOpinionDialog(ModalScreen[dict]):
    """Диалог добавления прогноза специалиста"""

    SCROLL_VIEW = False

    CSS = """
    AddOpinionDialog {
        align: center middle;
        background: transparent;
        overflow-y: auto;
    }

    #opinion_add_dialog {
        width: 55;
        height: auto;
        background: $background;
        border: round $accent;
        padding: 2 3;
    }

    #opinion_add_dialog .title-label {
        width: 100%;
        height: 2;
        content-align: center middle;
        color: $accent;
        text-style: bold;
        margin-bottom: 1;
    }

    #opinion_add_dialog .field-label {
        margin-top: 1;
        margin-bottom: 0;
        text-style: bold;
        color: $text;
        height: 1;
    }

    #opinion_add_dialog Input {
        margin-top: 0;
        margin-bottom: 0;
        width: 100%;
        background: $background;
        border: transparent;
        color: $text;
        height: 3;
        padding: 0 1;
    }

    #opinion_add_dialog Input:focus {
        border: round $accent;
        background: transparent;
    }

    #opinion_add_dialog .dialog_buttons {
        margin-top: 2;
        align: center middle;
        height: 3;
    }

    #opinion_add_dialog .dialog_buttons Button {
        width: 16;
        height: 3;
        margin: 0 1;
        background: $background;
        color: $secondary;
        border: round $secondary;
    }

    #opinion_add_dialog .dialog_buttons Button:focus {
        border: round $accent;
        color: $accent;
        background: transparent;
        text-style: bold;
    }

    #opinion_add_dialog .dialog_buttons Button.variant-primary {
        border: round $accent;
        color: $accent;
    }

    #opinion_add_dialog .dialog_buttons Button.variant-error {
        border: round red;
        color: red;
    }
    """

    BINDINGS = [
        Binding("escape", "cancel", "Отмена", show=False, priority=True),
    ]

    def __init__(self, patient: dict):
        super().__init__()
        self.patient = patient
        self.input_fields = []
        self.current_field_index = 0

    def compose(self) -> ComposeResult:
        with Vertical(id="opinion_add_dialog"):
            yield Label("Добавление прогноза специалиста", classes="title-label")
            yield Label(
                f"{self.patient.get('last_name', '')} {self.patient.get('first_name', '')} {self.patient.get('patronymic', '')}",
                classes="field-label")

            yield Label("Оценка риска (1-10):", classes="field-label")
            risk_input = Input(placeholder="5", id="risk_input", max_length=2)
            yield risk_input
            self.input_fields.append("risk_input")

            yield Label("Дата (ГГГГ.ММ.ДД):", classes="field-label")
            date_input = Input(placeholder=datetime.now().strftime('%Y.%m.%d'), id="date_input")
            yield date_input
            self.input_fields.append("date_input")

            yield Label("Специалист:", classes="field-label")
            expert_input = Input(placeholder="Фамилия И. О.", id="expert_input")
            yield expert_input
            self.input_fields.append("expert_input")

            yield Label("Комментарий (необязательно):", classes="field-label")
            comment_input = Input(placeholder="Оценка методом \"Пронзающий взгляд\"", id="comment_input")
            yield comment_input
            self.input_fields.append("comment_input")

            with Horizontal(classes="dialog_buttons"):
                btn_add = Button("Добавить", variant="primary", id="btn_add")
                btn_cancel = Button("Отмена", variant="error", id="btn_cancel")
                yield btn_add
                yield btn_cancel
            self.input_fields.append("btn_add")
            self.input_fields.append("btn_cancel")

    def on_mount(self) -> None:
        self.query_one("#risk_input").focus()

    def on_key(self, event: events.Key) -> None:
        """Обработка клавиш для навигации между полями"""
        if event.key == "escape":
            self.action_cancel()
            event.prevent_default()
            event.stop()
            return

        focused = self.focused

        if event.key == "down":
            if isinstance(focused, (Input, Button)):
                self._focus_next_field(focused)
                event.prevent_default()
                event.stop()
                return
        elif event.key == "up":
            if isinstance(focused, (Input, Button)):
                self._focus_prev_field(focused)
                event.prevent_default()
                event.stop()
                return

    def _focus_next_field(self, current_widget):
        """Переход к следующему полю"""
        current_id = getattr(current_widget, 'id', None)
        if current_id is None:
            return

        try:
            current_index = self.input_fields.index(current_id)
            next_index = (current_index + 1) % len(self.input_fields)
            next_id = self.input_fields[next_index]
            next_widget = self.query_one(f"#{next_id}")
            if next_widget:
                next_widget.focus()
        except (ValueError, NoMatches):
            pass

    def _focus_prev_field(self, current_widget):
        """Переход к предыдущему полю"""
        current_id = getattr(current_widget, 'id', None)
        if current_id is None:
            return

        try:
            current_index = self.input_fields.index(current_id)
            prev_index = (current_index - 1) % len(self.input_fields)
            prev_id = self.input_fields[prev_index]
            prev_widget = self.query_one(f"#{prev_id}")
            if prev_widget:
                prev_widget.focus()
        except (ValueError, NoMatches):
            pass

    @on(Button.Pressed, "#btn_add")
    def handle_add(self, event: Button.Pressed) -> None:
        try:
            risk_input = self.query_one("#risk_input", Input)
            risk_value = risk_input.value.strip()

            if not risk_value:
                self.app.custom_notify("Введите оценку риска (1-10)", severity="error")
                return

            try:
                risk_score = int(risk_value)
                if risk_score < 1 or risk_score > 10:
                    self.app.custom_notify("Оценка риска должна быть от 1 до 10", severity="error")
                    return
            except ValueError:
                self.app.custom_notify("Введите целое число от 1 до 10", severity="error")
                return

            date_input = self.query_one("#date_input", Input)
            date = date_input.value.strip()
            if not date:
                date = datetime.now().strftime('%Y.%m.%d')

            is_valid, error_msg = DataManager.validate_date(date)
            if not is_valid:
                self.app.custom_notify(error_msg, severity="error")
                return

            expert_input = self.query_one("#expert_input", Input)
            expert = expert_input.value.strip()
            if not expert:
                self.app.custom_notify("Введите фамилию специалиста", severity="error")
                return

            comment_input = self.query_one("#comment_input", Input)
            comment = comment_input.value.strip()

            # Если комментарий пустой, не сохраняем его
            if not comment:
                comment = ""

            self.dismiss({
                "expert": expert,
                "risk_score": risk_score,
                "date": date,
                "comment": comment
            })

        except Exception as e:
            self.app.custom_notify(f"Ошибка: {e}", severity="error")

    @on(Button.Pressed, "#btn_cancel")
    def handle_cancel(self, event: Button.Pressed) -> None:
        self.dismiss(None)

    def action_cancel(self) -> None:
        self.dismiss(None)

class DeleteOpinionDialog(ModalScreen[dict]):
    """Диалог выбора и подтверждения удаления прогноза"""

    SCROLL_VIEW = False

    CSS = """
    DeleteOpinionDialog {
        align: center middle;
        background: transparent;
    }

    #delete_opinion_dialog {
        width: 62;
        height: auto;
        max-height: 85%;
        background: $background;
        border: round $accent;
        padding: 2 3;
    }

    #delete_opinion_dialog .title-label {
        width: 100%;
        height: 2;
        content-align: center middle;
        color: $accent;
        text-style: bold;
        margin-bottom: 1;
    }

    #delete_opinion_dialog .subtitle-label {
        width: 100%;
        height: 1;
        content-align: center middle;
        color: $text;
        margin-bottom: 1;
    }

    #delete_opinion_list {
        height: auto;
        max-height: 16;
        border: round $secondary;
        background: transparent;
    }

    #delete_opinion_list:focus {
        border: round $accent;
        background: transparent;
    }

    #delete_opinion_list ListItem {
        padding-left: 1;
        background: $background;
        color: $text;
    }

    #delete_opinion_list > ListItem.--highlight {
        background: $background;
    }

    #delete_opinion_list > ListItem.--highlight > Static {
        color: $accent;
        text-style: bold;
    }

    #delete_opinion_dialog .dialog_buttons {
        margin-top: 1;
        align: center middle;
        height: 3;
    }

    #delete_opinion_dialog .dialog_buttons Button {
        width: 20;
        height: 3;
        margin: 0 1;
        background: $background;
        color: $secondary;
        border: round $secondary;
    }

    #delete_opinion_dialog .dialog_buttons Button:focus {
        border: round $accent;
        color: $accent;
        background: transparent;
        text-style: bold;
    }

    #delete_opinion_dialog .dialog_buttons Button.variant-error {
        border: round red;
        color: red;
    }

    #delete_opinion_dialog .dialog_buttons Button.variant-primary {
        border: round $accent;
        color: $accent;
    }
    """

    BINDINGS = [
        Binding("escape", "cancel", "Отмена", show=False, priority=True),
        Binding("enter", "confirm_delete", "Удалить", show=False),
    ]

    def __init__(self, patient: dict):
        super().__init__()
        self.patient = patient
        self._opinions = []

    def compose(self) -> ComposeResult:
        self._opinions = self.patient.get("opinions", [])

        with Vertical(id="delete_opinion_dialog"):
            yield Label("Удаление прогноза специалиста", classes="title-label")
            fio = f"{self.patient.get('last_name', '')} {self.patient.get('first_name', '')} {self.patient.get('patronymic', '')}"
            yield Label(fio, classes="subtitle-label")

            yield ListView(id="delete_opinion_list")

            with Horizontal(classes="dialog_buttons"):
                yield Button("Удалить (Enter)", variant="error", id="btn_delete")
                yield Button("Отмена (Esc)", variant="primary", id="btn_cancel")

    def on_mount(self) -> None:
        try:
            list_view = self.query_one("#delete_opinion_list", ListView)

            if not self._opinions:
                label = ListLabel("Нет прогнозов для удаления")
                list_view.append(ListItem(label))
            else:
                for opinion in self._opinions:
                    risk = opinion.get("risk_score", 0)
                    stars = "█" * risk + "░" * (10 - risk)
                    expert = opinion.get("expert", "")
                    date = opinion.get("date", "")
                    text = f"{risk}/10 {stars} - {expert} от {date}"

                    label = ListLabel(text)
                    item = ListItem(label)
                    item.opinion = opinion
                    list_view.append(item)

            if list_view.children:
                list_view.index = 0
                self._update_highlight(list_view)
            list_view.focus()
        except NoMatches:
            pass

    @on(ListView.Highlighted, "#delete_opinion_list")
    def handle_list_highlighted(self, event: ListView.Highlighted) -> None:
        try:
            list_view = self.query_one("#delete_opinion_list", ListView)
            self._update_highlight(list_view)
        except NoMatches:
            pass

    def _update_highlight(self, list_view: ListView) -> None:
        for i, child in enumerate(list_view.children):
            if hasattr(child, "children") and child.children:
                label = child.children[0]
                if isinstance(label, ListLabel):
                    label.is_highlighted = (i == list_view.index)

    @on(ListView.Selected, "#delete_opinion_list")
    def handle_list_selected(self, event: ListView.Selected) -> None:
        self.action_confirm_delete()

    def action_confirm_delete(self) -> None:
        try:
            list_view = self.query_one("#delete_opinion_list", ListView)
            if list_view.index is not None and list_view.index < len(list_view.children):
                item = list_view.children[list_view.index]
                if hasattr(item, "opinion"):
                    self.dismiss({"id": item.opinion.get("id")})
                    return
        except NoMatches:
            pass
        self.app.custom_notify("Сначала выберите прогноз из списка", severity="warning")

    @on(Button.Pressed, "#btn_delete")
    def handle_delete(self, event: Button.Pressed) -> None:
        self.action_confirm_delete()

    @on(Button.Pressed, "#btn_cancel")
    def handle_cancel(self, event: Button.Pressed) -> None:
        self.dismiss(None)

    def action_cancel(self) -> None:
        self.dismiss(None)

    def on_key(self, event: events.Key) -> None:
        if event.key == "escape":
            self.action_cancel()
            event.prevent_default()
            event.stop()




class DataScreen(Screen):
    """Экран управления данными (импорт/экспорт)"""

    BINDINGS = [
        Binding("escape", "app.pop_screen", "Назад"),
        Binding("1", "press_1", "Черновик", show=False),
        Binding("2", "press_2", "Экспорт", show=False),
        Binding("3", "press_3", "Импорт", show=False),
        Binding("4", "press_4", "Старая версия", show=False),
        Binding("5", "press_5", "Таблица", show=False),
        Binding("6", "press_6", "Статистика", show=False),
        Binding("7", "press_7", "Исходы", show=False),
        Binding("8", "press_8", "Прогнозы", show=False),
        Binding("9", "press_9", "ROC", show=False),
    ]

    def action_press_1(self):
        self.query_one("#btn_open_draft").press()

    def action_press_2(self):
        self.query_one("#btn_export").press()

    def action_press_3(self):
        self.query_one("#btn_import").press()

    def action_press_4(self):
        self.query_one("#btn_import_old").press()

    def action_press_5(self):
        self.query_one("#btn_export_table").press()

    def action_press_6(self):
        self.query_one("#btn_stat").press()

    def action_press_7(self):
        self.query_one("#btn_outcomes").press()

    def action_press_8(self):
        self.query_one("#btn_opinion").press()

    def action_press_9(self):
        self.query_one("#btn_roc").press()

    def compose(self) -> ComposeResult:
        with Horizontal():
            with Vertical(id="data_hint_panel") as hint_col:
                hint_col.border_title = "Работа с данными"
                yield Static("", id="data_hint_text")

            with Vertical(id="data_buttons_panel") as btn_col:
                btn_col.border_title = "Действия"
                with Vertical(id="data_buttons_container"):
                    # Используем HoverButton
                    yield HoverButton("1. Продолжить оценку из черновика      ", id="btn_open_draft")
                    yield HoverButton("2. Экспорт данных с этого компьютера   ", id="btn_export")
                    yield HoverButton("3. Импорт данных на этот компьютер     ", id="btn_import")
                    yield HoverButton("4. Импорт данных из старой версии      ", id="btn_import_old")
                    yield HoverButton("5. Вывод таблицы оценок                ", id="btn_export_table")
                    yield HoverButton("6. Статистика оценок                   ", id="btn_stat")
                    yield HoverButton("7. Опасные проявления (исходы)         ", id="btn_outcomes")
                    yield HoverButton("8. Обработка прогнозов специалистов    ", id="btn_opinion")
                    yield HoverButton("9. ROC-анализ                          ", id="btn_roc")
                    yield Button("", id="btn_spacer2", disabled=True, classes="spacer")
                    yield HoverButton("Esc. Назад                             ", id="btn_back")
        yield Footer(show_command_palette=False)

    # Словарь описаний для кнопок
    DATA_DESCRIPTIONS = {
        "btn_open_draft": "Эта функция позволяет открыть любой сохраненный черновик "
                          "(первичной или повторной оценки) и продолжить работу с того места, "
                          "где вы остановились.\n\n"
                          "После успешного завершения оценки черновик будет автоматически удален.",

        "btn_export": "Эта функция создает папку \"Арсенал - данные\" в папке Документы и копирует туда все ваши данные для переноса на другой компьютер.\n\n"
                      "В папку копируются:\n"
                      "1. Все файлы пациентов из папки arsenal_reports,\n"
                      "2. Файл базы данных arsenal_database.json.\n\n"
                      "После завершения экспорта вы можете скопировать папку \"Арсенал - данные\" на другой компьютер и выполнить там импорт данных.",

        "btn_import": "Эта функция объединяет данные из папки \"Арсенал - данные\", "
                      "которая должна находиться в папке Документы, с вашими текущими данными.\n\n"
                      "Для импорта нужно:\n"
                      "1. Провести экспорт данных на другом компьютере и перенести папку \"Арсенал - данные\" в папку Документы на этом компьютере;\n"
                      "2. Запустить эту функцию, она объединяет файлы пациентов и базу данных с теми, что уже есть на этом компьютере.\n\n"
                      "После импорта все данные будут доступны в программе.",

        "btn_import_old": "Эта функция предназначена для переноса данных из программы на Bash "
                          "(версия v1.0) в новый формат, используемый данной программой.\n\n"
                          "Для импорта нужно:\n"
                          "1. Скопировать папку с данными, экспортированными из старой программы, в папку Документы;\n"
                          "2. Запустить эту функцию.\n\n"
                          "Программа прочитает файл 'Журнал.txt' из старой версии и преобразует его "
                          "в формат новой базы данных, а также скопирует файлы пациентов.\n\n"
                          "Если вы использовали версию программы v0.х, не все данные могут быть перенесены корректно.",

        "btn_export_table": "Эта функция создает в папке Документы Excel-файл arsenal_data.xlsx со всеми данными о проведенных оценках.\n\n"
                            "Для повторных оценок указывается в том числе дата первичной оценки, на основе которой она была проведена, вместо оценок по каждому фактору приводятся пересчитанные баллы, учитывающие вновь установленную стадию изменения.\n\n"
                            "В таблице каждая строка соответствует одной оценке. Используйте сортировку и фильтры для анализа данных.",

        "btn_stat": "Вывод статистики по всем оценкам, которые сохранены на этом компьютере, с возможностью сохранить данные в файл и просмотреть ранее сохраненные данные.",

        "btn_outcomes": "Функции для управления данными об опасных проявлениях пациентов, позволяющие загрузить эти данные из файлов, либо внести их вручную. Эти данные в том числе необходимы для проведения ROC-анализа.",

        "btn_roc": "ROC-анализ (Receiver Operating Characteristic) позволяет оценить прогностическую способность методики Арсенал на вашей выборке пациентов. Этот анализ проводится путем сопоставления результатов проведенных оценок с тем, отмечалось ли у пациента в последующие 6 месяцев опасное поведение.\n\nДля проведения ROC-анализа необходимы значительное количество оценок (не менее 30) и данные об опасных проявлениях пациентов (исходах).",

        "btn_opinion": "Функции для учета и обработки мнений специалистов (прогнозов и оценок риска) для сопоставления с оценками по методике Арсенал и данными об опасных проявлениях пациентов (исходами).",

        "btn_back": "Вернуться в главное меню."
    }

    def on_descendant_focus(self, event: events.Focus) -> None:
        """Срабатывает при фокусе на любой кнопке"""
        target = getattr(event, "control", None) or getattr(event, "node", None)

        if isinstance(target, Button) and target.id:
            self._update_hint(target.id)

    def _update_hint(self, btn_id: str) -> None:
        """Обновление текста в левой панели"""
        description = self.DATA_DESCRIPTIONS.get(btn_id, "Выберите действие...")
        try:
            self.query_one("#data_hint_text", Static).update(description)
        except NoMatches:
            pass

    def on_mount(self) -> None:
        """При монтировании фокусируем первую кнопку"""
        self.query_one("#btn_open_draft").focus()
        self._update_hint("btn_open_draft")

    def on_mouse_move(self, event: events.MouseMove) -> None:
        """При движении мыши проверяем, не наведена ли она на кнопку"""
        result = self.get_widget_at(event.screen_x, event.screen_y)

        if result is None:
            return

        widget, region = result

        button = self._find_button_parent(widget)

        if button and button.id:
            if button.id.startswith("btn_") and self.focused != button:
                button.focus()

    def _find_button_parent(self, widget):
        """Рекурсивно ищем родителя-кнопку"""
        current = widget
        while current:
            if isinstance(current, Button):
                return current
            current = current.parent
        return None

    def on_button_pressed(self, event: Button.Pressed) -> None:
        """Обработка нажатия кнопок"""
        btn_id = event.button.id

        if btn_id == "btn_open_draft":
            self.open_draft()
        elif btn_id == "btn_export":
            self.export_data()
        elif btn_id == "btn_import":
            self.import_data()
        elif btn_id == "btn_import_old":
            self.import_old_data()
        elif btn_id == "btn_export_table":
            self.export_table(anonymous=False)
        elif btn_id == "btn_stat":
            self.app.push_screen(StatsScreen(show_history=False))
        elif btn_id == "btn_outcomes":
            self.app.push_screen(OutcomesMainScreen())
        elif btn_id == "btn_roc":
            self.app.push_screen(RocScreen(show_history=False))
        elif btn_id == "btn_opinion":
            self.opinion()
        elif btn_id == "btn_back":
            self.app.pop_screen()

    def open_draft(self) -> None:
        """Открывает диалог выбора любого черновика"""
        drafts = self.app.drafts.get_all_drafts()

        if not drafts:
            self.app.custom_notify("Нет сохраненных черновиков", severity="info")
            return

        def handle_draft(draft_data):
            if draft_data:
                try:
                    print(f"DEBUG: Выбран черновик с ID: {draft_data.get('draft_id')}")
                    assessment_type = draft_data.get("assessment_type")

                    if assessment_type == "первичная":
                        self.app.push_screen(Rate1Screen(draft_data=draft_data))
                    elif assessment_type == "повторная":
                        self.app.push_screen(Rate2Screen(draft_data=draft_data))
                    else:
                        self.app.custom_notify(f"Неизвестный тип черновика: {assessment_type}", severity="error")
                except Exception as e:
                    self.app.custom_notify(f"Ошибка при открытии черновика: {e}", severity="error")
                    import traceback
                    traceback.print_exc()

        try:
            self.app.push_screen(OpenDraftDialog(drafts), handle_draft)
        except Exception as e:
            self.app.custom_notify(f"Ошибка открытия диалога черновиков: {e}", severity="error")
            import traceback
            traceback.print_exc()

    # --- Вспомогательные функции ---

    def find_existing_file(self, target_name: str) -> Path:
        """
        Ищет существующий файл в reports_dir, нормализуя имена для сравнения.
        """
        reports_dir = self.app.results.reports_dir
        target_normalized = self.app.results.normalize_filename(target_name)

        for existing_file in reports_dir.glob("*.txt"):
            existing_normalized = self.app.results.normalize_filename(existing_file.name)
            if existing_normalized == target_normalized:
                return existing_file

        return None

    def get_documents_dir(self) -> Path:
        """Определяет путь к папке Документы в зависимости от ОС"""
        import platform

        system = platform.system()
        home = Path.home()

        if system == "Windows":
            # В Windows Documents обычно находится в профиле пользователя
            docs = home / "Documents"
            if docs.exists():
                return docs
            # Альтернативный вариант
            return home / "Мои документы"
        elif system == "Darwin":  # macOS
            return home / "Documents"
        else:  # Linux
            # Пробуем xdg-user-dir
            try:
                import subprocess
                result = subprocess.run(['xdg-user-dir', 'DOCUMENTS'],
                                        capture_output=True, text=True)
                if result.returncode == 0:
                    docs = Path(result.stdout.strip())
                    if docs.exists():
                        return docs
            except (FileNotFoundError, subprocess.SubprocessError):
                pass
            # По умолчанию
            return home / "Документы"

    def notify_success(self, message: str):
        """Показать уведомление об успехе"""
        self.app.custom_notify(message, severity="success")

    def notify_error(self, message: str):
        """Показать уведомление об ошибке"""
        self.app.custom_notify(message, severity="error")

    def notify_info(self, message: str):
        """Показать информационное уведомление"""
        self.app.custom_notify(message, severity="info")

    # --- Функция экспорта ---

    def export_data(self) -> None:
        """Экспорт данных с нормализацией имен файлов"""
        try:
            docs_dir = self.get_documents_dir()
            export_dir = docs_dir / "Арсенал - данные"

            # Создаем папку для экспорта
            export_dir.mkdir(exist_ok=True)

            # 1. Копируем файлы заключений с нормализацией имен
            reports_dir = self.app.results.reports_dir
            if reports_dir.exists():
                for txt_file in reports_dir.glob("*.txt"):
                    # Нормализуем имя для экспорта
                    normalized_name = self.app.results.normalize_filename(txt_file.name)
                    dest_file = export_dir / normalized_name
                    shutil.copy2(txt_file, dest_file)

            # 2. Копируем базу данных
            db_path = self.app.results.db_path
            if db_path.exists():
                dest_db = export_dir / "arsenal_database.json"
                shutil.copy2(db_path, dest_db)

            self.notify_success(f"Данные экспортированы в:\n{export_dir}")

        except Exception as e:
            self.notify_error(f"Ошибка экспорта: {e}")

    # --- Функция импорта ---

    def import_data(self) -> None:
        """Импорт данных из папки Документы/Арсенал - данные"""
        try:
            docs_dir = self.get_documents_dir()
            import_dir = docs_dir / "Арсенал - данные"

            if not import_dir.exists():
                self.notify_error(f"Папка не найдена:\n{import_dir}\n\n"
                                 "Сначала скопируйте папку 'Арсенал - данные' в папку Документы.")
                return

            # 1. Импортируем файлы заключений
            self.import_report_files(import_dir)

            # 2. Импортируем и объединяем базу данных
            self.merge_database(import_dir)

            self.notify_success("Импорт завершен успешно!\n"
                               "Данные объединены с вашей базой.")

        except Exception as e:
            self.notify_error(f"Ошибка импорта: {e}")
            import traceback
            traceback.print_exc()

    def import_report_files(self, import_dir: Path) -> None:
        """Импортирует текстовые файлы с нормализацией имен"""
        reports_dir = self.app.results.reports_dir
        reports_dir.mkdir(exist_ok=True)

        imported_files = list(import_dir.glob("*.txt"))

        for src_file in imported_files:
            if src_file.name == "arsenal_database.json":
                continue

            # Нормализуем имя исходного файла
            normalized_name = self.app.results.normalize_filename(src_file.name)

            # Ищем существующий файл по нормализованному имени
            existing_file = self.find_existing_file(normalized_name)

            if existing_file is None:
                # Копируем с нормализованным именем
                dst_file = reports_dir / normalized_name
                shutil.copy2(src_file, dst_file)
                self.notify_info(f"Скопирован новый файл: {normalized_name}")
            else:
                # Объединяем содержимое
                self.merge_text_files(src_file, existing_file)

    def merge_text_files(self, src_file: Path, dst_file: Path) -> None:
        """Объединяет два текстовых файла, избегая дублирования"""
        try:
            # Читаем оба файла
            src_content = src_file.read_text(encoding="utf-8").strip()
            dst_content = dst_file.read_text(encoding="utf-8").strip()

            # Если исходный файл пуст - ничего не делаем
            if not src_content:
                return

            # Проверяем, не содержится ли уже исходный контент в целевом
            if src_content in dst_content:
                return

            # Добавляем разделитель и исходный контент
            with open(dst_file, "a", encoding="utf-8") as f:
                f.write("\n\n" + src_content)

        except Exception as e:
            print(f"Ошибка при слиянии файлов {src_file.name}: {e}")

    def merge_database(self, import_dir: Path) -> None:
        """Объединяет две базы данных JSON"""
        src_db = import_dir / "arsenal_database.json"

        if not src_db.exists():
            self.notify_info("Файл базы данных не найден в импортируемой папке.\n"
                             "Будут импортированы только текстовые файлы.")
            return

        # Загружаем импортируемую базу
        with open(src_db, "r", encoding="utf-8") as f:
            import_data = json.load(f)

        # Загружаем текущую базу (или создаем новую)
        current_data = self.app.results.data

        # Объединяем
        merged_data = self.merge_json_databases(current_data, import_data)

        # Сохраняем
        with open(self.app.results.db_path, "w", encoding="utf-8") as f:
            json.dump(merged_data, f, ensure_ascii=False, indent=4)

        # Обновляем данные в памяти
        self.app.results.data = merged_data

        # Подсчет статистики
        outcomes_count = sum(len(p.get("outcomes", [])) for p in merged_data)
        opinions_count = sum(len(p.get("opinions", [])) for p in merged_data)

        self.notify_success(
            f"Импорт базы данных завершен!\n"
            f"Всего пациентов: {len(merged_data)}\n"
            f"Всего исходов: {outcomes_count}\n"
            f"Всего прогнозов: {opinions_count}"
        )

    def merge_json_databases(self, current: list, imported: list) -> list:
        """
        Объединяет две базы данных JSON.
        Ключ уникальности: комбинация last_name, first_name, patronymic, birth_year
        """
        patients_dict = {}
        for patient in current:
            key = (
                patient.get("last_name", ""),
                patient.get("first_name", ""),
                patient.get("patronymic", ""),
                patient.get("birth_year", "")
            )
            patients_dict[key] = patient

        for imported_patient in imported:
            key = (
                imported_patient.get("last_name", ""),
                imported_patient.get("first_name", ""),
                imported_patient.get("patronymic", ""),
                imported_patient.get("birth_year", "")
            )

            if key in patients_dict:
                existing = patients_dict[key]
                self.merge_assessments(existing, imported_patient)
            else:
                current.append(imported_patient)

        return current

    def merge_assessments(self, existing: dict, imported: dict) -> None:
        """Объединяет списки оценок, исходов и прогнозов для одного пациента"""

        # 1. Оценки (по assessment_id)
        existing_ids = {a.get("assessment_id", "") for a in existing.get("assessments", [])}
        for assessment in imported.get("assessments", []):
            if assessment.get("assessment_id", "") not in existing_ids:
                existing.setdefault("assessments", []).append(assessment)

        # 2. Исходы (по дате + типу)
        existing_outcome_keys = set()
        for o in existing.get("outcomes", []):
            key = (o.get("date", ""), o.get("type", ""))
            existing_outcome_keys.add(key)

        for outcome in imported.get("outcomes", []):
            key = (outcome.get("date", ""), outcome.get("type", ""))
            if key not in existing_outcome_keys:
                existing.setdefault("outcomes", []).append(outcome)
                existing_outcome_keys.add(key)

        # 3. Прогнозы (по дате + специалисту)
        existing_opinion_keys = set()
        for o in existing.get("opinions", []):
            key = (o.get("date", ""), o.get("expert", ""))
            existing_opinion_keys.add(key)

        for opinion in imported.get("opinions", []):
            key = (opinion.get("date", ""), opinion.get("expert", ""))
            if key not in existing_opinion_keys:
                existing.setdefault("opinions", []).append(opinion)
                existing_opinion_keys.add(key)

    def cleanup_duplicate_files(self):
        """Очищает дубликаты файлов, возникшие из-за проблем с нормализацией"""
        reports_dir = self.app.results.reports_dir

        # Группируем файлы по нормализованному имени
        files_by_name = {}
        for filepath in reports_dir.glob("*.txt"):
            normalized = self.app.results.normalize_filename(filepath.name)
            if normalized not in files_by_name:
                files_by_name[normalized] = []
            files_by_name[normalized].append(filepath)

        # Обрабатываем дубликаты
        for norm_name, file_list in files_by_name.items():
            if len(file_list) > 1:
                # Выбираем основной файл (тот, у которого имя в NFC форме)
                primary = None
                for f in file_list:
                    if f.name == norm_name:
                        primary = f
                        break
                if not primary:
                    primary = file_list[0]
                    # Переименовываем в нормализованное имя
                    new_path = reports_dir / norm_name
                    primary.rename(new_path)
                    primary = new_path

                # Объединяем остальные
                for dup in file_list:
                    if dup != primary:
                        self.merge_text_files(dup, primary)
                        dup.unlink()

                self.notify_info(f"Объединены дубликаты для {norm_name}")

    # --- Функция импорта из старой версии ---

    def import_old_data(self) -> None:
        """Импорт данных из старой версии программы (Bash)"""
        try:
            docs_dir = self.get_documents_dir()

            # Ищем папку со старыми данными
            # В старой версии данные могли быть в разных местах
            possible_paths = [
                docs_dir / "Арсенал-данные",  # Из data.sh
                docs_dir / "Арсенал данные",
                docs_dir / "arsenal-data",
                Path("/usr/local/bin/arsenal/Заключения")  # Стандартный путь старой версии
            ]

            old_data_dir = None
            for path in possible_paths:
                if path.exists() and (path / "Журнал.txt").exists():
                    old_data_dir = path
                    break

            if not old_data_dir:
                self.notify_error("Папка с данными старой версии не найдена.\n"
                                 "Скопируйте папку с данными (например, 'Арсенал-данные') "
                                 "в папку Документы.")
                return

            # 1. Импортируем текстовые файлы
            self.import_old_report_files(old_data_dir)

            # 2. Импортируем и конвертируем журнал
            self.convert_and_import_journal(old_data_dir)

            self.notify_success("Импорт из старой версии завершен!")

        except Exception as e:
            self.notify_error(f"Ошибка импорта из старой версии: {e}")
            import traceback
            traceback.print_exc()

    def import_old_report_files(self, old_dir: Path) -> None:
        """Импортирует файлы заключений из старой версии, конвертируя имена"""
        reports_dir = self.app.results.reports_dir
        reports_dir.mkdir(exist_ok=True)

        # В старой версии файлы назывались по шаблону "Фамилия_Имя_Отчество_1999.txt"
        # Сейчас формат тот же, так что просто копируем, если файла нет
        for txt_file in old_dir.glob("*.txt"):
            if txt_file.name == "Журнал.txt":
                continue

            dst_file = reports_dir / txt_file.name

            if not dst_file.exists():
                shutil.copy2(txt_file, dst_file)
            else:
                # Объединяем содержимое
                self.merge_text_files(txt_file, dst_file)

    def convert_and_import_journal(self, old_dir: Path) -> None:
        """Конвертирует Журнал.txt в формат новой базы данных и добавляет"""
        journal_path = old_dir / "Журнал.txt"

        if not journal_path.exists():
            return

        # Читаем журнал
        content = journal_path.read_text(encoding="utf-8")
        lines = content.splitlines()

        # Парсим журнал и создаем структуру данных
        converted_data = self.parse_old_journal(lines)

        if not converted_data:
            self.notify_info("Не удалось извлечь данные из журнала.")
            return

        # Объединяем с текущей базой
        current_data = self.app.results.data
        merged_data = self.merge_json_databases(current_data, converted_data)

        # Сохраняем
        with open(self.app.results.db_path, "w", encoding="utf-8") as f:
            json.dump(merged_data, f, ensure_ascii=False, indent=4)

        self.app.results.data = merged_data

    def parse_old_journal(self, lines: list) -> list:
        """
        Парсит старый Журнал.txt и возвращает список пациентов в новом формате.

        Формат строки журнала (таблица с разделителями | ):
        Фамилия | Имя | Отчество | г.р. | Оценка (тип) | Специалист | Дата оценки |
        оцен1 | стад1 | оцен2 | стад2 | оцен3 | стад3 | оцен4 | стад4 |
        оцен5 | стад5 | оцен6 | стад6 | оцен7 | стад7 | оцен8 | стад8 |
        ком1 | ком2 | ком3 | ком4 | ком5 | ком6 | ком7 | ком8 | Заключение
        """
        patients_dict = {}  # Ключ: (фамилия, имя, год)

        # Пропускаем первую строку с заголовками
        data_lines = [line for line in lines if line.strip() and not line.startswith("Фамилия")]

        for line in data_lines:
            line = line.strip()
            if not line:
                continue

            # Разделяем по |, но учитываем, что поля могут содержать пробелы
            parts = [part.strip() for part in line.split("|")]

            # Минимальное количество полей: 8 факторов * 2 (оценка+стадия) + основные поля
            if len(parts) < 7 + 16:  # 7 основных + 16 полей оценок/стадий
                continue

            # --- Основные данные ---
            last_name = parts[0].strip()
            first_name = parts[1].strip()
            patronymic = parts[2].strip()
            birth_year = parts[3].strip()
            assessment_type = parts[4].strip().lower()  # "первичная" или "повторная"
            rater = parts[5].strip()
            date_str = parts[6].strip()

            # Пропускаем, если нет обязательных полей
            if not last_name or not first_name or not birth_year:
                continue

            # --- Оценки и стадии (8 факторов) ---
            scores = []
            stages = []

            # Индексы: оцен1 (7), стад1 (8), оцен2 (9), стад2 (10), ... оцен8 (21), стад8 (22)
            for i in range(8):
                score_idx = 7 + i * 2
                stage_idx = 8 + i * 2

                # Оценка
                try:
                    score = int(parts[score_idx]) if parts[score_idx] else 0
                except (ValueError, IndexError):
                    score = 0
                scores.append(score)

                # Стадия (может быть пустой или содержать число)
                stage = None
                if score > 0 and len(parts) > stage_idx:
                    stage_str = parts[stage_idx].strip()
                    if stage_str and stage_str.isdigit():
                        stage_val = int(stage_str)
                        if 0 <= stage_val <= 4:
                            stage = stage_val

                # Для повторных оценок стадия всегда есть (иначе это ошибка)
                if assessment_type == "повторная" and score > 0 and stage is None:
                    # Если стадия не указана, ставим по умолчанию 0
                    stage = 0

                stages.append(stage)

            # --- Комментарии к факторам (8 полей) ---
            comments = []
            for i in range(8):
                comment_idx = 7 + 16 + i  # 7 основных + 16 оценок/стадий
                if len(parts) > comment_idx:
                    comment = parts[comment_idx].strip()
                    comments.append(comment)
                else:
                    comments.append("")

            # --- Заключение (последнее поле) ---
            conclusion_idx = 7 + 16 + 8  # 7 основных + 16 оценок/стадий + 8 комментариев
            conclusion = ""
            if len(parts) > conclusion_idx:
                # Всё, что осталось после заключительного разделителя - это заключение
                # (может содержать символы | внутри себя, поэтому объединяем остаток)
                remaining = parts[conclusion_idx:]
                if remaining:
                    conclusion = "|".join(remaining).strip()

            # Ключ пациента
            patient_key = (last_name, first_name, birth_year)

            # Создаем или получаем пациента
            if patient_key not in patients_dict:
                patient = {
                    "uid": str(uuid.uuid4())[:8],
                    "last_name": last_name,
                    "first_name": first_name,
                    "patronymic": patronymic,
                    "birth_year": int(birth_year) if birth_year.isdigit() else 0,
                    "fiogr": f"{last_name} {first_name} {patronymic} {birth_year}".strip(),
                    "assessments": []
                }
                patients_dict[patient_key] = patient
            else:
                patient = patients_dict[patient_key]

            # Создаем оценку
            assessment = {
                "assessment_id": self.generate_assessment_id(date_str, assessment_type, patient["assessments"]),
                "type": assessment_type,
                "date": date_str,
                "rater": rater,
                "total_score": sum(scores),
                "factors": {},
                "conclusion": conclusion
            }

            # Заполняем факторы
            for i, score in enumerate(scores, 1):
                stage_val = stages[i-1]

                # Для первичной оценки: если оценка 0, стадия не определена (5)
                # Для повторной оценки: стадия всегда должна быть, если оценка > 0
                if assessment_type == "первичная":
                    if score == 0:
                        stage_for_db = 5  # Нет стадии
                    else:
                        stage_for_db = stage_val if stage_val is not None else 0
                else:  # повторная
                    if score == 0:
                        stage_for_db = 5
                    else:
                        # Для повторной оценки стадия должна быть, если её нет - ставим 0
                        stage_for_db = stage_val if stage_val is not None else 0

                assessment["factors"][f"f{i}"] = {
                    "score": score,
                    "stage": stage_for_db,
                    "comment": comments[i-1] if i-1 < len(comments) else ""
                }

            # Добавляем оценку, если её ещё нет
            existing_ids = {a.get("assessment_id", "") for a in patient["assessments"]}
            if assessment["assessment_id"] not in existing_ids:
                patient["assessments"].append(assessment)

        return list(patients_dict.values())

    def generate_assessment_id(self, date_str: str, assessment_type: str, existing_assessments: list) -> str:
        """Генерирует уникальный ID оценки на основе даты и существующих ID"""
        if date_str:
            # Преобразуем дату из формата YYYY.MM.DD в YYYYMMDD_HHMM
            try:
                # Удаляем точки и разбиваем
                date_parts = date_str.replace('.', '').strip()
                if len(date_parts) == 8:  # YYYYMMDD
                    base_id = f"{date_parts}_0000"
                else:
                    # Если формат другой, просто используем дату как есть
                    base_id = date_str.replace('.', '')
            except (AttributeError, ValueError, TypeError):
                # Если возникла ошибка при обработке даты, используем текущее время
                base_id = datetime.now().strftime("%Y%m%d_%H%M")
        else:
            base_id = datetime.now().strftime("%Y%m%d_%H%M")

        # Проверяем уникальность
        counter = 0
        candidate = base_id
        while any(a.get("assessment_id") == candidate for a in existing_assessments):
            counter += 1
            if len(base_id) > 8:
                candidate = f"{base_id[:8]}_{counter:04d}"
            else:
                candidate = f"{base_id}_{counter:04d}"

        return candidate

    def export_table(self, anonymous: bool = False) -> None:
        """Создает Excel-таблицу с данными оценок"""
        try:
            # Загружаем базу данных
            db_path = self.app.results.db_path
            if not db_path.exists():
                self.notify_error("База данных не найдена. Сначала проведите хотя бы одну оценку.")
                return

            with open(db_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            if not data:
                self.notify_error("База данных пуста. Сначала проведите хотя бы одну оценку.")
                return

            # Создаем Excel-файл
            wb = Workbook()
            ws = wb.active
            ws.title = "Оценки Арсенал"

            # Определяем заголовки столбцов
            if anonymous:
                headers = [
                    "uid пациента",
                    "ID оценки",
                    "Тип оценки",
                    "Дата оценки",
                    "Специалист",
                    "Дата первичной оценки",
                    "Сумма баллов"
                ]
            else:
                headers = [
                    "Фамилия",
                    "Имя",
                    "Отчество",
                    "Год рождения",
                    "ФИО и г.р.",
                    "uid пациента",
                    "ID оценки",
                    "Тип оценки",
                    "Дата оценки",
                    "Специалист",
                    "Дата первичной оценки",
                    "Сумма баллов"
                ]

            # Добавляем заголовки для 8 факторов
            for i in range(1, 9):
                factor_name = FACTOR_NAMES.get(i, f"Фактор {i}")
                headers.append(f"{i}. {factor_name} - оценка")
                headers.append(f"{i}. {factor_name} - стадия")
                headers.append(f"{i}. {factor_name} - комментарий")

            headers.append("Заключение")

            # Применяем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            # Собираем все оценки
            row_num = 2
            stage_names = ["предобдумывание", "обдумывание", "подготовка", "действие", "удержание"]

            for patient in data:
                patient_uid = patient.get("uid", "")
                last_name = patient.get("last_name", "")
                first_name = patient.get("first_name", "")
                patronymic = patient.get("patronymic", "")
                birth_year = patient.get("birth_year", "")
                fiogr = patient.get("fiogr", f"{last_name} {first_name} {patronymic} {birth_year}".strip())

                # Сортируем оценки по дате
                assessments = sorted(
                    patient.get("assessments", []),
                    key=lambda a: a.get("date", "")
                )

                # Выводим каждую оценку в отдельную строку
                for assessment in assessments:
                    assessment_type = assessment.get("type", "")
                    assessment_id = assessment.get("assessment_id", "")
                    date = assessment.get("date", "")
                    rater = assessment.get("rater", "")
                    total_score = assessment.get("total_score", 0)
                    conclusion = assessment.get("conclusion", "")

                    # Для повторной оценки ищем дату первичной оценки
                    primary_date = ""
                    if assessment_type == "повторная":
                        # Ищем первичную оценку, которая была до этой повторной
                        for a in assessments:
                            if a.get("type") == "первичная" and a.get("date", "") <= date:
                                primary_date = a.get("date", "")
                                break

                    # Заполняем основные поля
                    col = 1
                    if not anonymous:
                        ws.cell(row=row_num, column=col, value=last_name)
                        col += 1
                        ws.cell(row=row_num, column=col, value=first_name)
                        col += 1
                        ws.cell(row=row_num, column=col, value=patronymic)
                        col += 1
                        ws.cell(row=row_num, column=col, value=birth_year)
                        col += 1
                        ws.cell(row=row_num, column=col, value=fiogr)
                        col += 1

                    ws.cell(row=row_num, column=col, value=patient_uid)
                    col += 1
                    ws.cell(row=row_num, column=col, value=assessment_id)
                    col += 1
                    ws.cell(row=row_num, column=col, value=assessment_type)
                    col += 1
                    ws.cell(row=row_num, column=col, value=date)
                    col += 1
                    ws.cell(row=row_num, column=col, value=rater)
                    col += 1
                    ws.cell(row=row_num, column=col, value=primary_date)
                    col += 1
                    ws.cell(row=row_num, column=col, value=total_score)
                    col += 1

                    # Факторы
                    factors = assessment.get("factors", {})
                    for i in range(1, 9):
                        f_key = f"f{i}"
                        factor = factors.get(f_key, {})
                        score = factor.get("score", 0)
                        stage = factor.get("stage")
                        comment = factor.get("comment", "")

                        # Преобразуем стадию в текст
                        stage_text = ""
                        if stage is not None:
                            if stage == 5:
                                stage_text = "нет"
                            elif 0 <= stage <= 4:
                                stage_text = stage_names[stage]
                            else:
                                stage_text = str(stage)

                        ws.cell(row=row_num, column=col, value=score)
                        col += 1
                        ws.cell(row=row_num, column=col, value=stage_text)
                        col += 1
                        ws.cell(row=row_num, column=col, value=comment)
                        col += 1

                    ws.cell(row=row_num, column=col, value=conclusion)

                    row_num += 1

            # Настраиваем ширину столбцов
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20

            # Устанавливаем фильтры
            ws.auto_filter.ref = ws.dimensions

            # Замораживаем первую строку
            ws.freeze_panes = "A2"

            # Сохраняем файл
            docs_dir = self.get_documents_dir()
            if anonymous:
                filename = "arsenal_data_anonymous.xlsx"
            else:
                filename = "arsenal_data.xlsx"

            filepath = docs_dir / filename

            # Если файл существует, добавляем дату к имени
            if filepath.exists():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                name_parts = filename.rsplit('.', 1)
                filepath = docs_dir / f"{name_parts[0]}_{timestamp}.{name_parts[1]}"

            wb.save(filepath)

            # Открываем файл
            open_file_externally(filepath)

            if anonymous:
                self.notify_success(f"Таблица с обезличенными данными создана:\n{filepath}")
            else:
                self.notify_success(f"Таблица с данными создана:\n{filepath}")

        except ImportError:
            self.notify_error(
                "Для работы с Excel необходимо установить библиотеку openpyxl.\n\n"
                "Выполните в терминале команду:\n"
                "pip install openpyxl"
            )
        except Exception as e:
            self.notify_error(f"Ошибка создания таблицы: {e}")
            import traceback
            traceback.print_exc()

    def opinion(self):
        """Открывает экран управления прогнозами специалистов"""
        self.app.push_screen(OpinionsMainScreen())


# Экран настроек

# Словарь описаний для левой панели
SETBUTTON_DESCRIPTIONS = {
    "btn_info": "[bold $accent]Версия программы v2.5.0[/]\n\nПрограмма методики оценки динамики риска опасного поведения «МОДРОП Арсенал» разработана для компьютеров с операционными системами Linux, Windows и MacOS, распространяется под лицензией GNU General Public License v3.0, либо любой более поздней версии, то есть может свободно копироваться, использоваться и изменяться со ссылкой на автора и сохранением открытой лицензии. Полный текст лицензии размещен на сайте https://www.gnu.org/licenses/\n\nДанные о государственной регистрации программы доступны на сайте Федерального института промышленной собственности https://fips.ru, свидетельство No 2026619934.\nПрямая ссылка (клик при нажатом Ctrl) https://fips.ru/EGD/bcbdf1ab-a333-4286-a6ac-2d5a54a750fb\n\nПо всем вопросам, касающимся установки, настройки и работы с программой, приветствуется обращение к автору по адресам электронной почты shadrov@pbstin.ru, shadrovv@gmail.com.\n\n(c) 2024-2026 Шадров Василий Валерьевич\n\n[bold $accent]Автор выражает глубочайшую благодарность всем специалистам, оказавшим содействие в создании и развитии методики Арсенал и программы для нее и не может не перечислить следующих людей с особой признательностью.[/]\n\n[bold $accent]Александр Николаевич Колесник [/](Санкт-Петербургская ПБСТИН)\n[bold $accent]Анастасия Александровна Ульянич [/](Санкт-Петербургский государственный университет)\n[bold $accent]Анна Сергеевна Шадрова  [/](Санкт-Петербургская ПБСТИН)\n[bold $accent]Гаянэ Аршалуисовна Вартанян [/](Санкт-Петербургский государственный университет)\n[bold $accent]Иван Станиславович Григорьев [/](ПБ Святого Николая Чудотворца)\n[bold $accent]Игорь Иванович Чижиков [/](Санкт-Петербургская ПБСТИН)\n[bold $accent]Лидия Николаевна Казакова [/](Санкт-Петербургский государственный университет)\n[bold $accent]Талия Станиславовна Богомолова [/](Санкт-Петербургская ПБСТИН)\n[bold $accent]Audrey Gordon [/](University of Saskatchewan)\n[bold $accent]Carlo C. DiClemente [/](University of Rhode Island, UMBC)\n[bold $accent]James O. Prochaska [/](University of Rhode Island)\n[bold $accent]Liang Wenfeng [/](DeepSeek AI)\n[bold $accent]Stephen Wong [/](University of Nottingham, University of Saskatchewan)\n[bold $accent]Will McGugan [/](Textualize)",
    "btn_themeLT": "Отображение цветов зависит от вашего терминала. Если они отображаются некорректно, попробуйте изменить настройку цветовых схем или отображения жирного шрифта ярким цветом в параметрах терминала.\n\n[teal on #ffffff]╭─ [/][bold teal on #ffffff]Просмотр темы[/][teal on #ffffff] ──────────────────────────────────────────────────────────────────╮\n│                                                                                  │\n│      [/][black on #ffffff]М О Д Р О П   А р с е н а л    ╭────2025.05.25───╮ [/][teal on #ffffff]╭──────────────────────╮ │\n│ [/][black on #ffffff]Факторы                             │ оцен  стад      │ [/][teal on #ffffff]│ [/][bold teal on #ffffff]1. Кнопка в фокусе[/][teal on #ffffff]   │ │\n│ [/][black on #ffffff]\\[1] Агрессия                        │ 0 ▏             │ [/][teal on #ffffff]╰──────────────────────╯ │\n│ [/][black on #ffffff]\\[2] Когнитивные и другие симптомы   │ 1 ▒   ▁     пре │[/][teal on #ffffff]                          │\n│ [/][black on #ffffff]\\[3] Контроль над эмоциями           │ 2 ▓▓  ▁▂    обд │   [/][bold #666666 on #ffffff]2. Кнопка вне фокуса[/][teal on #ffffff]   │\n│ [/][black on #ffffff]\\[4] Контроль над поведением         │ 3 ███ ▁▂▄   под │[/][teal on #ffffff]                          │\n│ [/][black on #ffffff]\\[5] Злоупотребление веществами      │ 3 ███ ▁▂▄▆  дей │ [/][teal on #ffffff]╭──────────────────────╮ │\n│ [/][black on #ffffff]\\[6] Приверженность режиму и лечению │ 2 ▓▓  ▁▂▄▆█ уде │[/][teal on #ffffff] │ [/][black on #ffffff]Поле ввода текста[/][teal on #ffffff]    │ │\n│ [/][black on #ffffff]\\[7] Личностные установки            │ 1 ▒   ▁     пре │ [/][teal on #ffffff]╰──────────────────────╯ │\n│ [/][black on #ffffff]\\[8] Окружение, быт и планы          │ 0 ▏             │[/][teal on #ffffff]                          │\n│ [/][black on #ffffff]                                    ╰─────────────────╯                          [/][teal on #ffffff]│\n│   [/][bold black on #ffffff] █ Выделенный пункт списка[/][black on #ffffff]        ╭────┬────────────────────────╮[/][teal on #ffffff]              │\n│    [/][black on #ffffff]Не выделенный пункт списка       │ 12 │████████████            │ [/][teal on #ffffff]             │\n│                                     [/][black on #ffffff]╰────┴────────────────────────╯[/][teal on #ffffff]              │\n╰──────────────────────────────────────────────────────────────────────────────────╯\n[/][bold #008080 on #DCDCDC] esc [/][bold black on #DCDCDC]Выход  [/][bold #008080 on #DCDCDC]f2 [/][bold black on #DCDCDC]Назад  [/][bold #008080 on #DCDCDC]f3 [/][bold black on #DCDCDC]Вперед                                                     [/]",
    "btn_themeLB": "Отображение цветов зависит от вашего терминала. Если они отображаются некорректно, попробуйте изменить настройку цветовых схем или отображения жирного шрифта ярким цветом в параметрах терминала.\n\n[blue on #ffffff]╭─ [/][bold blue on #ffffff]Просмотр темы[/][blue on #ffffff] ──────────────────────────────────────────────────────────────────╮\n│                                                                                  │\n│      [/][black on #ffffff]М О Д Р О П   А р с е н а л    ╭────2025.05.25───╮ [/][blue on #ffffff]╭──────────────────────╮ │\n│ [/][black on #ffffff]Факторы                             │ оцен  стад      │ [/][blue on #ffffff]│ [/][bold blue on #ffffff]1. Кнопка в фокусе[/][blue on #ffffff]   │ │\n│ [/][black on #ffffff]\\[1] Агрессия                        │ 0 ▏             │ [/][blue on #ffffff]╰──────────────────────╯ │\n│ [/][black on #ffffff]\\[2] Когнитивные и другие симптомы   │ 1 ▒   ▁     пре │[/][blue on #ffffff]                          │\n│ [/][black on #ffffff]\\[3] Контроль над эмоциями           │ 2 ▓▓  ▁▂    обд │   [/][bold #666666 on #ffffff]2. Кнопка вне фокуса[/][blue on #ffffff]   │\n│ [/][black on #ffffff]\\[4] Контроль над поведением         │ 3 ███ ▁▂▄   под │[/][blue on #ffffff]                          │\n│ [/][black on #ffffff]\\[5] Злоупотребление веществами      │ 3 ███ ▁▂▄▆  дей │ [/][blue on #ffffff]╭──────────────────────╮ │\n│ [/][black on #ffffff]\\[6] Приверженность режиму и лечению │ 2 ▓▓  ▁▂▄▆█ уде │[/][blue on #ffffff] │ [/][black on #ffffff]Поле ввода текста[/][blue on #ffffff]    │ │\n│ [/][black on #ffffff]\\[7] Личностные установки            │ 1 ▒   ▁     пре │ [/][blue on #ffffff]╰──────────────────────╯ │\n│ [/][black on #ffffff]\\[8] Окружение, быт и планы          │ 0 ▏             │[/][blue on #ffffff]                          │\n│ [/][black on #ffffff]                                    ╰─────────────────╯                          [/][blue on #ffffff]│\n│   [/][bold black on #ffffff] █ Выделенный пункт списка[/][black on #ffffff]        ╭────┬────────────────────────╮[/][blue on #ffffff]              │\n│    [/][black on #ffffff]Не выделенный пункт списка       │ 12 │████████████            │ [/][blue on #ffffff]             │\n│                                     [/][black on #ffffff]╰────┴────────────────────────╯[/][blue on #ffffff]              │\n╰──────────────────────────────────────────────────────────────────────────────────╯\n[/][bold #0000FF on #b5b5b5] esc [/][bold black on #b5b5b5]Выход  [/][bold #0000FF on #b5b5b5]f2 [/][bold black on #b5b5b5]Назад  [/][bold #0000FF on #b5b5b5]f3 [/][bold black on #b5b5b5]Вперед                                                     [/]",
    "btn_themeLS": "Отображение цветов зависит от вашего терминала. Если они отображаются некорректно, попробуйте изменить настройку цветовых схем или отображения жирного шрифта ярким цветом в параметрах терминала.\n\n[steelblue on #ffffff]╭─ [/][bold steelblue on #ffffff]Просмотр темы[/][steelblue on #ffffff] ──────────────────────────────────────────────────────────────────╮\n│                                                                                  │\n│      [/][black on #ffffff]М О Д Р О П   А р с е н а л    ╭────2025.05.25───╮ [/][steelblue on #ffffff]╭──────────────────────╮ │\n│ [/][black on #ffffff]Факторы                             │ оцен  стад      │ [/][steelblue on #ffffff]│ [/][bold steelblue on #ffffff]1. Кнопка в фокусе[/][steelblue on #ffffff]   │ │\n│ [/][black on #ffffff]\\[1] Агрессия                        │ 0 ▏             │ [/][steelblue on #ffffff]╰──────────────────────╯ │\n│ [/][black on #ffffff]\\[2] Когнитивные и другие симптомы   │ 1 ▒   ▁     пре │[/][steelblue on #ffffff]                          │\n│ [/][black on #ffffff]\\[3] Контроль над эмоциями           │ 2 ▓▓  ▁▂    обд │   [/][bold #666666 on #ffffff]2. Кнопка вне фокуса[/][steelblue on #ffffff]   │\n│ [/][black on #ffffff]\\[4] Контроль над поведением         │ 3 ███ ▁▂▄   под │[/][steelblue on #ffffff]                          │\n│ [/][black on #ffffff]\\[5] Злоупотребление веществами      │ 3 ███ ▁▂▄▆  дей │ [/][steelblue on #ffffff]╭──────────────────────╮ │\n│ [/][black on #ffffff]\\[6] Приверженность режиму и лечению │ 2 ▓▓  ▁▂▄▆█ уде │[/][steelblue on #ffffff] │ [/][black on #ffffff]Поле ввода текста[/][steelblue on #ffffff]    │ │\n│ [/][black on #ffffff]\\[7] Личностные установки            │ 1 ▒   ▁     пре │ [/][steelblue on #ffffff]╰──────────────────────╯ │\n│ [/][black on #ffffff]\\[8] Окружение, быт и планы          │ 0 ▏             │[/][steelblue on #ffffff]                          │\n│ [/][black on #ffffff]                                    ╰─────────────────╯                          [/][steelblue on #ffffff]│\n│   [/][bold black on #ffffff] █ Выделенный пункт списка[/][black on #ffffff]        ╭────┬────────────────────────╮[/][steelblue on #ffffff]              │\n│    [/][black on #ffffff]Не выделенный пункт списка       │ 12 │████████████            │ [/][steelblue on #ffffff]             │\n│                                     [/][black on #ffffff]╰────┴────────────────────────╯[/][steelblue on #ffffff]              │\n╰──────────────────────────────────────────────────────────────────────────────────╯\n[/][bold #4682B4 on #DCDCDC] esc [/][bold black on #DCDCDC]Выход  [/][bold #4682B4 on #DCDCDC]f2 [/][bold black on #DCDCDC]Назад  [/][bold #4682B4 on #DCDCDC]f3 [/][bold black on #DCDCDC]Вперед                                                     [/]",
    "btn_themeLM": "Отображение цветов зависит от вашего терминала. Если они отображаются некорректно, попробуйте изменить настройку цветовых схем или отображения жирного шрифта ярким цветом в параметрах терминала.\n\n[black on #DCDCDC]╭─ [/][bold black on #DCDCDC]Просмотр темы[/][black on #DCDCDC] ──────────────────────────────────────────────────────────────────╮\n│                                                                                  │\n│      [/][black on #DCDCDC]М О Д Р О П   А р с е н а л    ╭────2025.05.25───╮ [/][black on #DCDCDC]╭──────────────────────╮ │\n│ [/][black on #DCDCDC]Факторы                             │ оцен  стад      │ [/][black on #DCDCDC]│ [/][bold black on #DCDCDC]1. Кнопка в фокусе[/][black on #DCDCDC]   │ │\n│ [/][black on #DCDCDC]\\[1] Агрессия                        │ 0 ▏             │ [/][black on #DCDCDC]╰──────────────────────╯ │\n│ [/][black on #DCDCDC]\\[2] Когнитивные и другие симптомы   │ 1 ▒   ▁     пре │[/][black on #DCDCDC]                          │\n│ [/][black on #DCDCDC]\\[3] Контроль над эмоциями           │ 2 ▓▓  ▁▂    обд │   [/][bold #404040 on #DCDCDC]2. Кнопка вне фокуса[/][black on #DCDCDC]   │\n│ [/][black on #DCDCDC]\\[4] Контроль над поведением         │ 3 ███ ▁▂▄   под │[/][black on #DCDCDC]                          │\n│ [/][black on #DCDCDC]\\[5] Злоупотребление веществами      │ 3 ███ ▁▂▄▆  дей │ [/][black on #DCDCDC]╭──────────────────────╮ │\n│ [/][black on #DCDCDC]\\[6] Приверженность режиму и лечению │ 2 ▓▓  ▁▂▄▆█ уде │[/][black on #DCDCDC] │ [/][black on #DCDCDC]Поле ввода текста[/][black on #DCDCDC]    │ │\n│ [/][black on #DCDCDC]\\[7] Личностные установки            │ 1 ▒   ▁     пре │ [/][black on #DCDCDC]╰──────────────────────╯ │\n│ [/][black on #DCDCDC]\\[8] Окружение, быт и планы          │ 0 ▏             │[/][black on #DCDCDC]                          │\n│ [/][black on #DCDCDC]                                    ╰─────────────────╯                          [/][black on #DCDCDC]│\n│   [/][bold black on #DCDCDC] █ Выделенный пункт списка[/][black on #DCDCDC]        ╭────┬────────────────────────╮[/][black on #DCDCDC]              │\n│    [/][black on #DCDCDC]Не выделенный пункт списка       │ 12 │████████████            │ [/][black on #DCDCDC]             │\n│                                     [/][black on #DCDCDC]╰────┴────────────────────────╯[/][black on #DCDCDC]              │\n╰──────────────────────────────────────────────────────────────────────────────────╯\n[/][bold black on #b5b5b5] esc [/][bold black on #b5b5b5]Выход  [/][bold black on #b5b5b5]f2 [/][bold black on #b5b5b5]Назад  [/][bold black on #b5b5b5]f3 [/][bold black on #b5b5b5]Вперед                                                     [/]",
    "btn_themeDQ": "Отображение цветов зависит от вашего терминала. Если они отображаются некорректно, попробуйте изменить настройку цветовых схем или отображения жирного шрифта ярким цветом в параметрах терминала.\n\n[tan on #2F4F4F]╭─ [/][bold tan on #2F4F4F]Просмотр темы[/][tan on #2F4F4F] ──────────────────────────────────────────────────────────────────╮\n│                                                                                  │\n│      [/][#DCDCDC on #2F4F4F]М О Д Р О П   А р с е н а л    ╭────2025.05.25───╮ [/][tan on #2F4F4F]╭──────────────────────╮ │\n│ [/][#DCDCDC on #2F4F4F]Факторы                             │ оцен  стад      │ [/][tan on #2F4F4F]│ [/][bold tan on #2F4F4F]1. Кнопка в фокусе[/][tan on #2F4F4F]   │ │\n│ [/][#DCDCDC on #2F4F4F]\\[1] Агрессия                        │ 0 ▏             │ [/][tan on #2F4F4F]╰──────────────────────╯ │\n│ [/][#DCDCDC on #2F4F4F]\\[2] Когнитивные и другие симптомы   │ 1 ▒   ▁     пре │[/][tan on #2F4F4F]                          │\n│ [/][#DCDCDC on #2F4F4F]\\[3] Контроль над эмоциями           │ 2 ▓▓  ▁▂    обд │   [/][bold #DCDCDC on #2F4F4F]2. Кнопка вне фокуса[/][tan on #2F4F4F]   │\n│ [/][#DCDCDC on #2F4F4F]\\[4] Контроль над поведением         │ 3 ███ ▁▂▄   под │[/][tan on #2F4F4F]                          │\n│ [/][#DCDCDC on #2F4F4F]\\[5] Злоупотребление веществами      │ 3 ███ ▁▂▄▆  дей │ [/][tan on #2F4F4F]╭──────────────────────╮ │\n│ [/][#DCDCDC on #2F4F4F]\\[6] Приверженность режиму и лечению │ 2 ▓▓  ▁▂▄▆█ уде │[/][tan on #2F4F4F] │ [/][#DCDCDC on #2F4F4F]Поле ввода текста[/][tan on #2F4F4F]    │ │\n│ [/][#DCDCDC on #2F4F4F]\\[7] Личностные установки            │ 1 ▒   ▁     пре │ [/][tan on #2F4F4F]╰──────────────────────╯ │\n│ [/][#DCDCDC on #2F4F4F]\\[8] Окружение, быт и планы          │ 0 ▏             │[/][tan on #2F4F4F]                          │\n│ [/][#DCDCDC on #2F4F4F]                                    ╰─────────────────╯                          [/][tan on #2F4F4F]│\n│   [/][bold #DCDCDC on #2F4F4F] █ Выделенный пункт списка[/][#DCDCDC on #2F4F4F]        ╭────┬────────────────────────╮[/][tan on #2F4F4F]              │\n│    [/][#DCDCDC on #2F4F4F]Не выделенный пункт списка       │ 12 │████████████            │ [/][tan on #2F4F4F]             │\n│                                     [/][#DCDCDC on #2F4F4F]╰────┴────────────────────────╯[/][tan on #2F4F4F]              │\n╰──────────────────────────────────────────────────────────────────────────────────╯\n[/][bold tan on #2c2c2c] esc [/][bold white on #2c2c2c]Выход  [/][bold tan on #2c2c2c]f2 [/][bold white on #2c2c2c]Назад  [/][bold tan on #2c2c2c]f3 [/][bold white on #2c2c2c]Вперед                                                     [/]",
    "btn_themeDM": "Отображение цветов зависит от вашего терминала. Если они отображаются некорректно, попробуйте изменить настройку цветовых схем или отображения жирного шрифта ярким цветом в параметрах терминала.\n\n[goldenrod on #3b3b3b]╭─ [/][bold goldenrod on #3b3b3b]Просмотр темы[/][goldenrod on #3b3b3b] ──────────────────────────────────────────────────────────────────╮\n│                                                                                  │\n│      [/][white on #3b3b3b]М О Д Р О П   А р с е н а л    ╭────2025.05.25───╮ [/][goldenrod on #3b3b3b]╭──────────────────────╮ │\n│ [/][white on #3b3b3b]Факторы                             │ оцен  стад      │ [/][goldenrod on #3b3b3b]│ [/][bold goldenrod on #3b3b3b]1. Кнопка в фокусе[/][goldenrod on #3b3b3b]   │ │\n│ [/][white on #3b3b3b]\\[1] Агрессия                        │ 0 ▏             │ [/][goldenrod on #3b3b3b]╰──────────────────────╯ │\n│ [/][white on #3b3b3b]\\[2] Когнитивные и другие симптомы   │ 1 ▒   ▁     пре │[/][goldenrod on #3b3b3b]                          │\n│ [/][white on #3b3b3b]\\[3] Контроль над эмоциями           │ 2 ▓▓  ▁▂    обд │   [/][bold #DCDCDC on #3b3b3b]2. Кнопка вне фокуса[/][goldenrod on #3b3b3b]   │\n│ [/][white on #3b3b3b]\\[4] Контроль над поведением         │ 3 ███ ▁▂▄   под │[/][goldenrod on #3b3b3b]                          │\n│ [/][white on #3b3b3b]\\[5] Злоупотребление веществами      │ 3 ███ ▁▂▄▆  дей │ [/][goldenrod on #3b3b3b]╭──────────────────────╮ │\n│ [/][white on #3b3b3b]\\[6] Приверженность режиму и лечению │ 2 ▓▓  ▁▂▄▆█ уде │[/][goldenrod on #3b3b3b] │ [/][white on #3b3b3b]Поле ввода текста[/][goldenrod on #3b3b3b]    │ │\n│ [/][white on #3b3b3b]\\[7] Личностные установки            │ 1 ▒   ▁     пре │ [/][goldenrod on #3b3b3b]╰──────────────────────╯ │\n│ [/][white on #3b3b3b]\\[8] Окружение, быт и планы          │ 0 ▏             │[/][goldenrod on #3b3b3b]                          │\n│ [/][white on #3b3b3b]                                    ╰─────────────────╯                          [/][goldenrod on #3b3b3b]│\n│   [/][bold white on #3b3b3b] █ Выделенный пункт списка[/][white on #3b3b3b]        ╭────┬────────────────────────╮[/][goldenrod on #3b3b3b]              │\n│    [/][white on #3b3b3b]Не выделенный пункт списка       │ 12 │████████████            │ [/][goldenrod on #3b3b3b]             │\n│                                     [/][white on #3b3b3b]╰────┴────────────────────────╯[/][goldenrod on #3b3b3b]              │\n╰──────────────────────────────────────────────────────────────────────────────────╯\n[/][bold #DAA520 on #2c2c2c] esc [/][bold white on #2c2c2c]Выход  [/][bold #DAA520 on #2c2c2c]f2 [/][bold white on #2c2c2c]Назад  [/][bold #DAA520 on #2c2c2c]f3 [/][bold white on #2c2c2c]Вперед                                                     [/]",
    "btn_themeDC": "Отображение цветов зависит от вашего терминала. Если они отображаются некорректно, попробуйте изменить настройку цветовых схем или отображения жирного шрифта ярким цветом в параметрах терминала.\n\n[white on #111111]╭─ [/][bold white on #111111]Просмотр темы[/][white on #111111] ──────────────────────────────────────────────────────────────────╮\n│                                                                                  │\n│      [/][white on #111111]М О Д Р О П   А р с е н а л    ╭────2025.05.25───╮ [/][white on #111111]╭──────────────────────╮ │\n│ [/][white on #111111]Факторы                             │ оцен  стад      │ [/][white on #111111]│ [/][bold white on #111111]1. Кнопка в фокусе[/][white on #111111]   │ │\n│ [/][white on #111111]\\[1] Агрессия                        │ 0 ▏             │ [/][white on #111111]╰──────────────────────╯ │\n│ [/][white on #111111]\\[2] Когнитивные и другие симптомы   │ 1 ▒   ▁     пре │[/][white on #111111]                          │\n│ [/][white on #111111]\\[3] Контроль над эмоциями           │ 2 ▓▓  ▁▂    обд │   [/][bold #DCDCDC on #111111]2. Кнопка вне фокуса[/][white on #111111]   │\n│ [/][white on #111111]\\[4] Контроль над поведением         │ 3 ███ ▁▂▄   под │[/][white on #111111]                          │\n│ [/][white on #111111]\\[5] Злоупотребление веществами      │ 3 ███ ▁▂▄▆  дей │ [/][white on #111111]╭──────────────────────╮ │\n│ [/][white on #111111]\\[6] Приверженность режиму и лечению │ 2 ▓▓  ▁▂▄▆█ уде │[/][white on #111111] │ [/][white on #111111]Поле ввода текста[/][white on #111111]    │ │\n│ [/][white on #111111]\\[7] Личностные установки            │ 1 ▒   ▁     пре │ [/][white on #111111]╰──────────────────────╯ │\n│ [/][white on #111111]\\[8] Окружение, быт и планы          │ 0 ▏             │[/][white on #111111]                          │\n│ [/][white on #111111]                                    ╰─────────────────╯                          [/][white on #111111]│\n│   [/][bold white on #111111] █ Выделенный пункт списка[/][white on #111111]        ╭────┬────────────────────────╮[/][white on #111111]              │\n│    [/][white on #111111]Не выделенный пункт списка       │ 12 │████████████            │ [/][white on #111111]             │\n│                                     [/][white on #111111]╰────┴────────────────────────╯[/][white on #111111]              │\n╰──────────────────────────────────────────────────────────────────────────────────╯\n[/][bold white on #111111] esc [/][bold #DCDCDC on #111111]Выход  [/][bold white on #111111]f2 [/][bold #DCDCDC on #111111]Назад  [/][bold white on #111111]f3 [/][bold #DCDCDC on #111111]Вперед                                                     [/]",
    "btn_themeKL": "Отображение цветов зависит от вашего терминала. Если они отображаются некорректно, попробуйте изменить настройку цветовых схем или отображения жирного шрифта ярким цветом в параметрах терминала.\n\n[#0363a2 on #DCDCDC]╭─ [/][bold #0363a2 on #DCDCDC]Просмотр темы[/][#0363a2 on #DCDCDC] ──────────────────────────────────────────────────────────────────╮\n│                                                                                  │\n│      [/][black on #DCDCDC]М О Д Р О П   А р с е н а л    ╭────2025.05.25───╮ [/][#0363a2 on #DCDCDC]╭──────────────────────╮ │\n│ [/][black on #DCDCDC]Факторы                             │ оцен  стад      │ [/][#0363a2 on #DCDCDC]│ [/][bold #0363a2 on #DCDCDC]1. Кнопка в фокусе[/][#0363a2 on #DCDCDC]   │ │\n│ [/][black on #DCDCDC]\\[1] Агрессия                        │ 0 ▏             │ [/][#0363a2 on #DCDCDC]╰──────────────────────╯ │\n│ [/][black on #DCDCDC]\\[2] Когнитивные и другие симптомы   │ 1 ▒   ▁     пре │[/][#0363a2 on #DCDCDC]                          │\n│ [/][black on #DCDCDC]\\[3] Контроль над эмоциями           │ 2 ▓▓  ▁▂    обд │   [/][bold #404040 on #DCDCDC]2. Кнопка вне фокуса[/][#0363a2 on #DCDCDC]   │\n│ [/][black on #DCDCDC]\\[4] Контроль над поведением         │ 3 ███ ▁▂▄   под │[/][#0363a2 on #DCDCDC]                          │\n│ [/][black on #DCDCDC]\\[5] Злоупотребление веществами      │ 3 ███ ▁▂▄▆  дей │ [/][#0363a2 on #DCDCDC]╭──────────────────────╮ │\n│ [/][black on #DCDCDC]\\[6] Приверженность режиму и лечению │ 2 ▓▓  ▁▂▄▆█ уде │[/][#0363a2 on #DCDCDC] │ [/][black on #DCDCDC]Поле ввода текста[/][#0363a2 on #DCDCDC]    │ │\n│ [/][black on #DCDCDC]\\[7] Личностные установки            │ 1 ▒   ▁     пре │ [/][#0363a2 on #DCDCDC]╰──────────────────────╯ │\n│ [/][black on #DCDCDC]\\[8] Окружение, быт и планы          │ 0 ▏             │[/][#0363a2 on #DCDCDC]                          │\n│ [/][black on #DCDCDC]                                    ╰─────────────────╯                          [/][#0363a2 on #DCDCDC]│\n│   [/][bold black on #DCDCDC] █ Выделенный пункт списка[/][black on #DCDCDC]        ╭────┬────────────────────────╮[/][#0363a2 on #DCDCDC]              │\n│    [/][black on #DCDCDC]Не выделенный пункт списка       │ 12 │████████████            │ [/][#0363a2 on #DCDCDC]             │\n│                                     [/][black on #DCDCDC]╰────┴────────────────────────╯[/][#0363a2 on #DCDCDC]              │\n╰──────────────────────────────────────────────────────────────────────────────────╯\n[/][bold #0363a2 on #b5b5b5] esc [/][bold black on #b5b5b5]Выход  [/][bold #0363a2 on #b5b5b5]f2 [/][bold black on #b5b5b5]Назад  [/][bold #0363a2 on #b5b5b5]f3 [/][bold black on #b5b5b5]Вперед                                                     [/]",
    "btn_exit": "Возврат в основное меню.",
}

class SetScreen(Screen):
    BINDINGS = [
        Binding("escape", "app.pop_screen", "Назад"),
        Binding("1", "press_1", "О программе", show=False),
        Binding("2", "press_2", "Стяжкин", show=False),
        Binding("3", "press_3", "Голубев", show=False),
        Binding("4", "press_4", "Тарасевич", show=False),
        Binding("5", "press_5", "Колесник", show=False),
        Binding("6", "press_6", "Чекунова", show=False),
        Binding("7", "press_7", "Макушев", show=False),
        Binding("8", "press_8", "Шалек", show=False),
        Binding("9", "press_9", "Львов", show=False),
    ]

    def action_press_1(self):
        self.query_one("#btn_info").press()

    def action_press_2(self):
        self.query_one("#btn_themeLT").press()

    def action_press_3(self):
        self.query_one("#btn_themeLB").press()

    def action_press_4(self):
        self.query_one("#btn_themeLS").press()

    def action_press_5(self):
        self.query_one("#btn_themeKL").press()

    def action_press_6(self):
        self.query_one("#btn_themeLM").press()

    def action_press_7(self):
        self.query_one("#btn_themeDQ").press()

    def action_press_8(self):
        self.query_one("#btn_themeDM").press()

    def action_press_9(self):
        self.query_one("#btn_themeDC").press()

    def compose(self) -> ComposeResult:
        with Horizontal():
            # Левая панель для подсказок
            with Vertical(id="set_hint_panel") as hint_col:
                hint_col.border_title = "МОДРОП Арсенал"
                yield Static("", id="set_hint_text")

            # Правая панель для кнопок
            with Vertical(id="set_buttons_panel") as btn_col:
                btn_col.border_title = "Настройки"
                with Vertical(id="set_buttons_container"):
                    # Используем HoverButton
                    yield HoverButton("1. О программе      ", id="btn_info")
                    yield HoverButton("2. Тема Стяжкин     ", id="btn_themeLT")
                    yield HoverButton("3. Тема Голубев     ", id="btn_themeLB")
                    yield HoverButton("4. Тема Тарасевич   ", id="btn_themeLS")
                    yield HoverButton("5. Тема Колесник    ", id="btn_themeKL")
                    yield HoverButton("6. Тема Чекунова    ", id="btn_themeLM")
                    yield HoverButton("7. Тема Макушев     ", id="btn_themeDQ")
                    yield HoverButton("8. Тема Шалек       ", id="btn_themeDM")
                    yield HoverButton("9. Тема Львов       ", id="btn_themeDC")
                    yield HoverButton("Esc. Назад          ", id="btn_exit")
        yield Footer(show_command_palette=False)

    def on_mouse_move(self, event: events.MouseMove) -> None:
        """При движении мыши обновляем подсказку"""
        try:
            # Получаем виджет под курсором
            result = self.get_widget_at(event.screen_x, event.screen_y)
            print(f"Result: {result}")  # Отладка

            if result is None:
                return

            # result может быть кортежем (widget, region) или просто виджетом
            if isinstance(result, tuple):
                widget = result[0]
            else:
                widget = result

            print(f"Widget: {widget}")  # Отладка

            # Ищем кнопку
            current = widget
            while current:
                print(f"Current: {current}")  # Отладка
                if isinstance(current, Button):
                    print(f"Found button: {current.id}")  # Отладка
                    if current.id and current.id.startswith("btn_"):
                        self._update_hint(current.id)
                        # Перемещаем фокус на кнопку
                        if self.focused != current:
                            current.focus()
                    return
                current = current.parent
        except Exception as e:
            print(f"Error: {e}")
            import traceback
            traceback.print_exc()

    @on(events.MouseMove)
    def handle_mouse_move(self, event: events.MouseMove) -> None:
        """Обработчик движения мыши"""
        print(f"Mouse move at: {event.x}, {event.y}")
        result = self.get_widget_at(event.screen_x, event.screen_y)
        print(f"Result: {result}")

    def on_descendant_focus(self, event: events.Focus) -> None:
        """Срабатывает, когда любая кнопка внутри экрана получает фокус"""
        target = getattr(event, "control", None) or getattr(event, "node", None)

        if isinstance(target, Button) and target.id:
            self._update_hint(target.id)

    def _update_hint(self, btn_id: str) -> None:
        """Обновление текста в левой панели"""
        description = SETBUTTON_DESCRIPTIONS.get(btn_id, "Выберите настройку")
        # Убеждаемся, что виджет существует, прежде чем обновлять
        try:
            self.query_one("#set_hint_text", Static).update(description)
        except NoMatches:
            pass

    def on_mount(self) -> None:
        # Фокусируем первую кнопку при старте
        self.query_one("#btn_info").focus()
        # И сразу принудительно ставим для неё текст
        self._update_hint("btn_info")

    # --- ЛОГИКА НАЖАТИЯ КНОПОК ---
    def on_button_pressed(self, event: Button.Pressed) -> None:
        # ID кнопки: Имя темы в словаре THEMES
        theme_mapping = {
            "btn_themeLT": "themeLT",
            "btn_themeLB": "themeLB",
            "btn_themeLS": "themeLS",
            "btn_themeLM": "themeLM",
            "btn_themeDQ": "themeDQ",
            "btn_themeDM": "themeDM",
            "btn_themeDC": "themeDC",
            "btn_themeKL": "themeKL",
        }

        btn_id = event.button.id

        if btn_id in theme_mapping:
            target_theme = theme_mapping[btn_id]

            if target_theme in THEMES:
                self.app.theme = target_theme

                # Сохраняем тему через централизованный менеджер данных приложения
                self.app.data_manager.save_theme(target_theme)

        elif btn_id == "btn_exit":
            self.app.pop_screen()
        elif btn_id == "btn_info":
            pass

# Экран меню

# Словарь описаний для левой панели
BUTTON_DESCRIPTIONS = {
    "btn_manual": "Методика Арсенал разработана для оценки динамики риска опасного поведения у пациентов с психическими расстройствами.\n\nВ меню справа собраны основные разделы программы для работы с методикой.\n\nПри работе с программой используйте сочетания клавиш:\n[bold $accent]Ctrl -[/] и [bold $accent]Ctrl +[/] (на MacOS [bold $accent]Cmd -[/] и [bold $accent]Cmd +[/]) для изменения масштаба отображения,\n[bold $accent]стрелки[/] для навигации и пролистывания,\n[bold $accent]Tab[/] и [bold $accent]Shift Tab[/] для перемещения фокуса вперед и назад,\n[bold $accent]PgUp[/] и [bold $accent]PgDn[/] для быстрого пролистывания,\n[bold $accent]Enter[/] для выбора, открытия,\n[bold $accent]цифры 1 - 9[/] для быстрого нажатия на соответствующие пронумерованные кнопки без перемещения фокуса,\n[bold $accent]Esc[/] для выхода.\n\nПрименимые на конкретном экране клавиши и их сочетания указаны на нижней панели, они также срабатывают по нажатию кнопкой мыши.",
    "btn_rate1": "Раздел для проведения первичной оценки. Она включает в себя оценку представленности у пациента восьми факторов риска, стадии изменения по каждому из них, составление комментариев и заключения.\n\nЧтобы провести оценку вам нужно собрать информацию о пациенте и провести его расспрос. Для удобства можете сформировать и распечатать форму для заметок.\n\nПри проведении оценки вам будут даны инструкции к каждому шагу, но, возможно, сначала стоит ознакомиться с полным руководством к методике.",
    "btn_rate2": "Раздел для проведения повторной оценки.\n\nПри каждой повторной оценке специалисту необходимо сначала ознакомиться с заключением по первичной оценке. В том случае, если у пациента произошло значимое ухудшение по какому-либо из факторов, например совершено новое ООД или обнаружились новые проблемы по факторам, которые первично оценены как отсутствующие, то стоит еще раз провести первичную оценку и заново определить представленность всех факторов и стадий изменения по ним. Если такого значимого ухудшения нет, достаточно провести повторную оценку только стадий изменения по факторам, первично определенным как присутствующие, на основании данных о текущем состоянии и поведении пациента и его прицельного расспроса. Каждая такая оценка сопровождается комментарием, в котором следует указать, на основании каких наблюдений она установлена.\n\nЧтобы провести повторную оценку вам потребуются сохраненные данные о первичной оценке и расспрос пациента. Для удобства можете сформировать и распечатать форму для заметок.\n\nПри проведении оценки вам будут даны инструкции к каждому шагу, но, возможно, сначала стоит ознакомиться с полным руководством к методике.",
    "btn_list": "Раздел для просмотра сохраненных оценок по каждому пациенту.\n\nДанная функция позволяет просматривать именно сохраненные файлы пациентов с внесенными в него оценками. Есть возможность открыть файл во внешнем редакторе для печати или удалить. При редактировании заключений имейте в виду, что вы вносите изменение только в файл пациента, а исходные данные оценки, остаются неизменными.\n\nНе удаляйте заключения, созданные в самых ранних выпусках программы версии v0.x, т. к. не все данные по этим оценкам могут быть сохранены в журнале.\n\nПодробнее о способах обработки данных смотрите в разделе руководства \"Программа - работа с данными\".",
    "btn_look": "Функция просмотра и поиска всех проведенных оценок.\n\nВ отличие от функции просмотра файлов пациентов эта функция отображает все записи об оценках, которые сохранены в журнале (базе данных). Эти данные доступны для просмотра и открытия для печати во внешнем редакторе.\n\nЕсли выбранная оценка не внесена в файл пациента, ее можно сохранить в него.\n\nПодробнее о способах обработки данных смотрите в разделе руководства \"Программа - работа с данными\".",
    "btn_form": "Функция создания формы для ведения заметок во время сбора информации о пациенте и его расспроса. Можно создать форму на одну или на две страницы. Форма открывается во внешнем редакторе, там ее можно изменить и распечатать.",
    "btn_data": "Раздел с функциями обработки данных.\n\nДоступны продолжение проведения оценок, сохраненных как черновики, перенос данных между компьютерами, вывод таблицы с данными о всех оценках, а также исследовательские функции статистики и анализа данных.",
    "btn_set": "Раздел с информацией и настройками внешнего вида программы.",
    "btn_exit": "Завершение работы с программой.",
}

class MenuScreen(Screen):

    BINDINGS = [
        Binding("escape", "request_quit", "Выход", show=True, priority=True),
        Binding("1", "press_1", "Руководство", show=False),
        Binding("2", "press_2", "Первичная оценка", show=False),
        Binding("3", "press_3", "Повторная оценка", show=False),
        Binding("4", "press_4", "Файлы пациентов", show=False),
        Binding("5", "press_5", "Все оценки", show=False),
        Binding("6", "press_6", "Форма", show=False),
        Binding("7", "press_7", "Данные", show=False),
        Binding("8", "press_8", "Настройки", show=False),
    ]

    def action_request_quit(self):
        self.app.action_request_quit()

    def action_press_1(self):
        self.query_one("#btn_manual").press()

    def action_press_2(self):
        self.query_one("#btn_rate1").press()

    def action_press_3(self):
        self.query_one("#btn_rate2").press()

    def action_press_4(self):
        self.query_one("#btn_list").press()

    def action_press_5(self):
        self.query_one("#btn_look").press()

    def action_press_6(self):
        self.query_one("#btn_form").press()

    def action_press_7(self):
        self.query_one("#btn_data").press()

    def action_press_8(self):
        self.query_one("#btn_set").press()

    def compose(self) -> ComposeResult:
        with Horizontal():
            # Левая панель для подсказок
            with Vertical(id="menu_hint_panel") as hint_col:
                hint_col.border_title = "МОДРОП Арсенал"
                yield Static("", id="menu_hint_text")

            # Правая панель для кнопок
            with Vertical(id="menu_buttons_panel") as btn_col:
                btn_col.border_title = "Действия"
                with Vertical(id="menu_buttons_container"):
                    # Используем HoverButton
                    yield HoverButton("1. Читать руководство           ", id="btn_manual")
                    yield HoverButton("2. Первичная оценка             ", id="btn_rate1")
                    yield HoverButton("3. Повторная оценка             ", id="btn_rate2")
                    yield HoverButton("4. Файлы пациентов              ", id="btn_list")
                    yield HoverButton("5. Все оценки                   ", id="btn_look")
                    yield HoverButton("6. Форма для заметок            ", id="btn_form")
                    yield HoverButton("7. Работа с данными             ", id="btn_data")
                    yield HoverButton("8. Настройки                    ", id="btn_set")
                    yield HoverButton("Esc. Выход                      ", id="btn_exit")
        yield Footer(show_command_palette=False)

    def on_mouse_move(self, event: events.MouseMove) -> None:
        """При движении мыши обновляем подсказку"""
        try:
            # Получаем виджет под курсором
            result = self.get_widget_at(event.screen_x, event.screen_y)
            print(f"Result: {result}")  # Отладка

            if result is None:
                return

            # result может быть кортежем (widget, region) или просто виджетом
            if isinstance(result, tuple):
                widget = result[0]
            else:
                widget = result

            print(f"Widget: {widget}")  # Отладка

            # Ищем кнопку
            current = widget
            while current:
                print(f"Current: {current}")  # Отладка
                if isinstance(current, Button):
                    print(f"Found button: {current.id}")  # Отладка
                    if current.id and current.id.startswith("btn_"):
                        self._update_hint(current.id)
                        # Перемещаем фокус на кнопку
                        if self.focused != current:
                            current.focus()
                    return
                current = current.parent
        except Exception as e:
            print(f"Error: {e}")
            import traceback
            traceback.print_exc()

    @on(events.MouseMove)
    def handle_mouse_move(self, event: events.MouseMove) -> None:
        """Обработчик движения мыши"""
        print(f"Mouse move at: {event.x}, {event.y}")
        result = self.get_widget_at(event.screen_x, event.screen_y)
        print(f"Result: {result}")

    def on_descendant_focus(self, event: events.Focus) -> None:
        """Срабатывает, когда любая кнопка внутри экрана получает фокус"""
        target = getattr(event, "control", None) or getattr(event, "node", None)

        if isinstance(target, Button) and target.id:
            self._update_hint(target.id)

    def _update_hint(self, btn_id: str) -> None:
        """Обновление текста в левой панели"""
        description = BUTTON_DESCRIPTIONS.get(btn_id, "Выберите раздел...")
        # Убеждаемся, что виджет существует, прежде чем обновлять
        try:
            self.query_one("#menu_hint_text", Static).update(description)
        except NoMatches:
            pass

    def on_mount(self) -> None:
        # Фокусируем первую кнопку при старте
        self.query_one("#btn_manual").focus()
        # И сразу принудительно ставим для неё текст
        self._update_hint("btn_manual")

    # --- ЛОГИКА НАЖАТИЯ КНОПОК ---
    def on_button_pressed(self, event: Button.Pressed) -> None:
        screens = {
            "btn_manual": ManualScreen,
            "btn_rate1": Rate1Screen,
            "btn_rate2": Rate2Screen,
            "btn_list": ListScreen,
            "btn_look": LookScreen,
            "btn_data": DataScreen,
            "btn_set": SetScreen,
        }

        btn_id = event.button.id
        if btn_id in screens:
            self.app.push_screen(screens[btn_id]())
        elif btn_id == "btn_exit":
            self.app.action_request_quit()
        elif btn_id == "btn_form":
            self.show_form_dialog()

    def show_form_dialog(self):
        """Показывает диалог создания формы для заметок"""

        class FormDialog(ModalScreen[str]):
            """Внутренний класс диалога создания формы"""

            CSS = """
            FormDialog {
                align: center middle;
                background: transparent;
            }

            FormDialog > Vertical {
                width: auto;
                height: auto;
                background: transparent;
            }

            #dialog {
                width: 76;
                height: 11;
                background: $background;
                border: round $accent;
                padding: 1 2;
            }

            #question {
                width: 100%;
                height: 3;
                content-align: center middle;
                color: $text;
                text-style: bold;
            }

            #buttons {
                width: 100%;
                height: 4;
                align: center middle;
            }

            #buttons Button {
                width: 22;
                height: 3;
                margin: 0 1;
                background: $background;
                color: $secondary;
                border: none;
                content-align: center middle;
            }

            #buttons Button:focus {
                border: round $accent;
                color: $accent;
                background: transparent;
                text-style: bold;
            }

            #buttons Button.variant-primary {
                border: round $accent;
                color: $accent;
            }

            #buttons Button.variant-error {
                border: round red;
                color: red;
            }
            """

            BINDINGS = [
                Binding("escape", "cancel", "Отмена", show=False),
            ]

            def compose(self) -> ComposeResult:
                with Vertical():
                    with Vertical(id="dialog"):
                        yield Label("Создать форму для заметок?", id="question")
                        with Horizontal(id="buttons"):
                            yield Button("На одну страницу", variant="primary", id="one")
                            yield Button("На две страницы", variant="primary", id="two")
                            yield Button("Отмена", variant="error", id="cancel")

            def on_mount(self) -> None:
                self.query_one("#one").focus()

            def on_key(self, event: events.Key) -> None:
                if event.key == "left":
                    self.focus_previous()
                    event.stop()
                elif event.key == "right":
                    self.focus_next()
                    event.stop()

            def on_button_pressed(self, event: Button.Pressed) -> None:
                if event.button.id == "one":
                    self.dismiss("one")
                elif event.button.id == "two":
                    self.dismiss("two")
                elif event.button.id == "cancel":
                    self.dismiss(None)

            def action_cancel(self) -> None:
                self.dismiss(None)

        def handle_dialog_result(result: str):
            """Обработка результата диалога"""
            if result in ["one", "two"]:
                self.create_and_open_form(result == "one")

        # Показываем диалог
        self.app.push_screen(FormDialog(), handle_dialog_result)

    def create_and_open_form(self, one_page: bool) -> None:
        """Создает пустую форму и открывает во внешнем редакторе"""
        if one_page:
            content = self.generate_one_page_form()
        else:
            content = self.generate_two_page_form()

        # Создаем временный файл
        temp_dir = Path(tempfile.gettempdir())
        form_path = temp_dir / "Форма-Арсенал.txt"

        # Записываем содержимое
        form_path.write_text(content, encoding="utf-8")

        # Открываем во внешнем редакторе
        open_file_externally(form_path)

        # Показываем уведомление
        self.app.custom_notify("Форма создана и открыта в редакторе", severity="info")

    def generate_one_page_form(self) -> str:
        """Генерирует пустую форму на одной странице"""
        return """Пациент (ф.)_____________ (и.)_____________ (о.)__________________ (г.р.)______

Оценка:   первичная / повторная

╭[1] Агрессия ─────────────────────────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[2] Когнитивные и другие симптомы ────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[3] Контроль над эмоциями ────────────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[4] Контроль над поведением ──────────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[5] Злоупотребление веществами ───────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[6] Приверженность режиму и лечению ──────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[7] Личностные установки ─────────────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[8] Окружение, быт и планы ───────────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[9] Заключение по оценке ─────────────────────────────────────────────────────╮
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯"""

    def generate_two_page_form(self) -> str:
        """Генерирует пустую форму на двух страницах"""
        return """Пациент (ф.)_____________ (и.)_____________ (о.)__________________ (г.р.)______

Оценка:   первичная / повторная

╭[1] Агрессия ─────────────────────────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[2] Когнитивные и другие симптомы ────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[3] Контроль над эмоциями ────────────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[4] Контроль над поведением ──────────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[5] Злоупотребление веществами ───────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯




╭[6] Приверженность режиму и лечению ──────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[7] Личностные установки ─────────────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[8] Окружение, быт и планы ───────────────────────────────────────────────────╮
│ оценка: 0 1 2 3        стадия: пред. обд. подг. дейст. удер.                 │
│ Комментарий:                                                                 │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯
╭[9] Заключение по оценке ─────────────────────────────────────────────────────╮
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
│                                                                              │
╰──────────────────────────────────────────────────────────────────────────────╯"""

# --- Приложение ---
class arsenal(App):

    CSS = """
        Screen {
            background: $background;
            color: $text;
        }

        ConfirmDeleteDialog {
            align: center middle;
            background: transparent;
        }

        #dialog_confirm {
            width: 50;
            height: auto;
            background: $background;
            border: round $accent;
            padding: 1 2;
        }

        #confirm_msg {
            text-align: center;
            width: 100%;
            height: 4;
            color: $text;
            text-style: bold;
        }

        #confirm_buttons {
            width: 100%;
            align: center middle;
            height: 3;
        }

        #confirm_buttons Button {
            width: 14;
            height: 3;
            margin: 0 2;
            background: $background;
            color: $secondary;
            border: none;
        }

        #confirm_buttons Button:focus {
            border: round $accent;
            color: $accent;
            background: transparent;
            text-style: bold;
        }

        #set_hint_panel {
            width: 70%;
            border: round $accent;
            padding: 2 2 1 4;
            background: $background;
            border-title-style: bold;
        }

        #set_buttons_panel {
            width: 30%;
            border: round $secondary;
            background: $background;
        }

        #set_hint_text {
            height: auto;
            width: 100%;
            color: $text;
        }

        /* Подсветка панелей при фокусе внутри них */
        #set_hint_panel:focus-within, #set_buttons_panel:focus-within {
            border: round $accent;
            border-title-color: $accent;
            border-title-style: bold;
            background: transparent;
        }

        /* Контейнер для кнопок внутри правой панели */
        #set_buttons_container {
            width: 100%;
            height: auto;
            align: left top;
            padding: 1 0 0 1;
        }

        #set_buttons_container Button {
            width: 35;
            height: 3;
            background: $background;
            color: $secondary;
            border: none;
            content-align: left middle;
            padding: 0 2;
            margin: 0;
        }

        #set_buttons_container Button > Static {
            width: 100%;
            content-align: left middle;
        }

        #set_buttons_container Button:focus {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
            content-align: left middle;
        }

        #set_buttons_container Button:hover {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
        }

        #menu_hint_panel {
            width: 60%;
            border: round $accent;
            padding: 1 2;
            background: $background;
            border-title-style: bold;
        }

        #menu_buttons_panel {
            width: 40%;
            border: round $secondary;
            background: $background;
        }

        #menu_hint_text {
            height: auto;
            width: 100%;
            color: $text;
            padding: 1 2;
        }

        /* Подсветка панелей при фокусе внутри них */
        #menu_hint_panel:focus-within, #menu_buttons_panel:focus-within {
            border: round $accent;
            border-title-color: $accent;
            border-title-style: bold;
            background: transparent;
        }

        /* Контейнер для кнопок внутри правой панели */
        #menu_buttons_container {
            width: 100%;
            height: auto;
            align: left top;
            padding: 1 0 0 1;
        }

        #menu_buttons_container Button {
            width: 47;
            height: 3;
            background: $background;
            color: $secondary;
            border: none;
            content-align: left middle;
            padding: 0 2;
            margin: 0;
        }

        #menu_buttons_container Button > Static {
            width: 100%;
            content-align: left middle;
        }

        #menu_buttons_container Button:focus {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
            content-align: left middle;
        }

        #menu_buttons_container Button:hover {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
        }

        VerticalScroll, ListView {
            scrollbar-gutter: stable;
            scrollbar-size: 1 1;
            scrollbar-color: $accent 30%;
            scrollbar-color-hover: $accent 30%;
            scrollbar-background: transparent;
            background: $background;
            padding: 0;
            margin: 0;
        }

        /* Хорошие панели                               - как образец */
        #list_panel {
            width: 36%;
            border: round $secondary;
            padding: 0;
            margin: 0;
            border-title-align: left;
            border-title-color: $secondary;
            background: transparent;
        }

        #detail_panel {
            width: 64%;
            border: round $secondary;
            padding: 0;
            border-title-align: left;
            border-title-color: $secondary;
        }

        /* Фокусы панелей */
        #list_panel:focus-within, #detail_panel:focus-within {
            border: round $accent;
            border-title-color: $accent;
            border-title-style: bold;
            background: transparent;
        }

        /* Настройка контента */
        #details {
            padding: 1 2;
            width: 100%;
            color: $text;
        }

        ListView {
            scrollbar-gutter: stable;
            background: transparent;
            color: $text;
            padding: 0;
            margin: 0;
            border: none;
        }

        /* Убираем padding у самого скроллбара */
        .scrollbar--vertical {
            margin: 0;
            padding: 0;
        }

        ListItem {
            padding-left: 1;
            padding-right: 1;
            background: transparent;
            color: $text;
        }

        ListItem > Static {
            color: $text;
            text-style: none;
            background: transparent;
        }

        ListItem.spacer {
            height: 1;
            background: transparent;
            border: none;
        }

        ListItem.spacer:hover {
            background: transparent;
        }

        /* Отменяем стандартный сине-серый фокус Textual для списка */
        ListView:focus > ListItem.--highlight {
            background: transparent;
        }

        ListView > ListItem.--highlight,
        ListView:focus > ListItem.--highlight {
            background: transparent;
        }
        ListView > ListItem.--highlight ListLabel {
            text-style: bold;
        }

        QuitScreen {
            align: center middle;
            background: transparent;
        }

        #dialog {
            width: 50;
            height: 10;
            background: $background;
            border: round $accent;
            padding: 1 2;
        }

        #question {
            width: 100%;
            height: 3;
            content-align: center middle;
            color: $text;
            text-style: bold;
        }

        #buttons {
            width: 100%;
            height: 3;
            align: center middle;
        }

        QuitScreen Button {
            width: 14;
            height: 3;
            margin: 0 2;
            background: $background;
            color: $secondary;
            border: none;
        }

        QuitScreen Button:focus {
            border: round $accent;
            color: $accent;
            background: transparent;
            text-style: bold;
        }

        #report_actions {
            height: 5;
            align: center middle;
            border-top: round $secondary;
            dock: bottom;
            padding: 1;
        }

        #btn_open_office {
            width: 40;
            background: $background;
            color: $secondary;
            border: round $secondary;
        }

        #btn_open_office:focus {
            border: round $accent;
            color: $accent;
            background: transparent;
        }

        #details {
            padding: 1 2;
            color: $text;
            text-wrap: wrap;
        }

        /* Центрируем диалог */
        ReportDialog {
            align: center middle;
        }

        #dialog_box {
            width: 50;
            height: auto;
            background: $background;
            border: round $accent;
            align: center middle;
            padding: 1 2;
        }

        #dialog_title {
            text-align: center;
            width: 100%;
            margin-bottom: 1;
            text-style: bold;
            color: $text;
        }

        #dialog_box Button {
            width: 100%;
            margin: 0 0 1 0;
        }

        /* Панели для первичной оценки */
        #guide_panel {
            width: 60%;
            border: round $secondary;
            background: $background;
            padding: 1 0 0 1;
            border-title-align: left;
            border-title-color: $text;
            border-title-style: bold;
        }

        #input_panel {
            width: 40%;
            border: round $secondary;
            background: $background;
            padding: 1 0 0 1;
            border-title-align: left;
        }

        /* Фокусы панелей */
        #guide_panel:focus-within, #input_panel:focus-within {
            border: round $accent;
            border-title-color: $accent;
            border-title-style: bold;
            background: transparent;
        }

        #guide_title, #input_title {
            padding: 0 1;
            background: $background;
            color: $text;
            height: 1;
        }

        #guide_scroll {
            scrollbar-gutter: stable;
            scrollbar-size: 1 1;
            scrollbar-color: $accent 30%;
            scrollbar-color-hover: $accent 30%;
            scrollbar-background: transparent;
            background: transparent;
        }

        #guide_text {
            color: $text;
            text-wrap: wrap;
            padding-right: 1;  /* отступ от полосы прокрутки */
        }

        #input_scroll {
            scrollbar-gutter: stable;
            scrollbar-size: 1 1;
            scrollbar-color: $accent;
            scrollbar-color-hover: $accent;
            scrollbar-background: transparent;
            background: transparent;
            margin: 0;
        }

        #input_container {
            width: 100%;
            padding: 0;
            margin: 0;
        }

        #input_container Label {
            margin-top: 0;
            margin-bottom: 0;  /* убираем отступ снизу */
            text-style: bold;
            color: $text;
            padding: 0;
        }

        /* Стили для полей ввода */
        #input_container Input {
            margin: 0 0 0 0;
            width: 100%;
            background: $background;
            border: transparent;
            color: $text;
        }

        #input_container Input:focus {
            border: round $accent;
            background: transparent;
        }

        #input_container Input.placeholder {
            color: $text-muted;  /* бледно-серый для placeholder */
        }

        #input_container RadioSet {
            margin: 0;
            width: 100%;
            border: round $secondary;
            padding: 1 1 0 3;
            background: $background;
            height: 20;
            min-height: 20;
        }

        #input_container RadioSet:focus {
            border: round $accent;
            background: transparent;
        }

        MyRadioButton {
            width: 100%;
            height: 3;
            padding: 0;
            margin: 0;
            background: $background;
            border: none;
        }

        /* Внутренний контейнер с заранее зарезервированной прозрачной рамкой */
        MyRadioButton > .radio-button--container {
            padding: 0 2;
            width: 100%;
            height: 100%;
            border: transparent;
        }

        /* Фокус: меняем на рамку $accent */
        MyRadioButton:focus > .radio-button--container {
            background: $background;
            border: round $accent;
            color: $accent;
        }

        MyRadioButton:focus-within .toggle--button {
            color: $accent;
            background: transparent;
        }

        MyRadioButton:hover > .radio-button--container {
            background: transparent;
        }

        MyRadioButton.--on > .radio-button--container {
            text-style: bold;
            color: $accent;
        }

        /* Стили для TextArea (комментарий) */
        #input_container TextArea {
            margin: 0 0 0 0;
            width: 100%;
            border: round $secondary;
            background: $background;
            color: $text;
            height: 20;
            min-height: 20;
            scrollbar-gutter: stable;
            scrollbar-size: 1 1;
            scrollbar-color: $accent 30%;
            scrollbar-color-hover: $accent;
            scrollbar-background: transparent;
            padding: 0;
        }

        #input_container TextArea:focus {
            border: round $accent;
            background: transparent;
        }

        /* Стили для кнопок */
        #input_container Button {
            width: 100%;
            margin: 1 0;
            background: $background;
            color: $secondary;
            border: round $secondary;
        }

        #input_container Button:focus {
            border: round $accent;
            color: $accent;
            background: transparent;
            text-style: bold;
        }

        #input_container Button.variant-primary {
            border: round $accent;
            color: $accent;
        }

        #input_container Button.variant-success {
            border: round green;
            color: green;
        }

        .comment-area, .conclusion-area {
            height: 20;
        }

        /* Стили для диалога подтверждения выхода из мастера */
        ConfirmDialog {
            align: center middle;
            background: transparent;
        }

        ConfirmDialog #dialog {
            width: 50;
            height: 10;
            background: $background;
            border: round $accent;
            padding: 1 2;
        }

        ConfirmDialog #question {
            width: 100%;
            height: 3;
            content-align: center middle;
            color: $text;
            text-style: bold;
        }

        ConfirmDialog #buttons {
            width: 100%;
            height: 3;
            align: center middle;
        }

        ConfirmDialog Button {
            width: 14;
            height: 3;
            margin: 0 2;
            background: $background;
            color: $secondary;
            border: none;
        }

        ConfirmDialog Button:focus {
            border: round $accent;
            color: $accent;
            background: transparent;
            text-style: bold;
        }

        /* Диалог подтверждения выхода */
        #dialog {
            width: 50;
            height: 10;
            background: $background;
            align: center middle;
            border: round $accent;
            padding: 1 2;
        }

        #question {
            width: 100%;
            height: 3;
            content-align: center middle;
            color: $text;
            text-style: bold;
        }

        #buttons {
            width: 100%;
            height: 3;
            align: center middle;
        }

        #dialog Button {
            width: 14;
            height: 3;
            margin: 0 2;
            background: $background;
            color: $secondary;
            border: none;
        }

        #dialog Button:focus {
            border: round $accent;
            color: $accent;
            background: transparent;
            text-style: bold;
        }

        /* Диалог подтверждения сохранения и завершения оценки */
        ConfirmSaveDialog {
            align: center middle;
            background: transparent;
        }

        ConfirmSaveDialog #dialog {
            width: 50;
            height: 10;
            background: $background;
            border: round $accent;
            padding: 1 2;
        }

        ConfirmSaveDialog #question {
            width: 100%;
            height: 3;
            content-align: center middle;
            color: $text;
            text-style: bold;
        }

        ConfirmSaveDialog #buttons {
            width: 100%;
            height: 3;
            align: center middle;
        }

        ConfirmSaveDialog Button {
            width: 14;
            height: 3;
            margin: 0 2;
            background: $background;
            color: $secondary;
            border: none;
        }

        ConfirmSaveDialog Button:focus {
            border: round $accent;
            color: $accent;
            background: transparent;
            text-style: bold;
        }

        /* Слой для уведомлений */
        NotificationLayer {
            layer: notifications;
            align: center top;
            width: auto;
            height: auto;
            background: transparent;  /* Прозрачный фон */
        }

        .notification {
            background: $background;
            color: $text;
            border: round $accent;
            padding: 1 2;
            margin: 1;
            width: auto;  /* Автоширина по содержимому */
            max-width: 80%;  /* Максимальная ширина */
            height: auto;
            text-align: center;
            layer: notifications;
            box-sizing: border-box;
        }

        .notification.error {
            border: round red;
        }

        .notification.warning {
            border: round red;
        }

        .notification.success {
            border: round green;
        }

        .notification.info {
            border: round $accent;
        }

        #data_hint_panel {
            width: 60%;
            border: round $accent;
            padding: 1 2;
            background: $background;
            border-title-style: bold;
        }

        #data_buttons_panel {
            width: 40%;
            border: round $secondary;
            background: $background;
        }

        #data_hint_text {
            height: auto;
            width: 100%;
            color: $text;
            padding: 1 2;
        }

        /* Подсветка панелей при фокусе внутри них */
        #data_hint_panel:focus-within, #data_buttons_panel:focus-within {
            border: round $accent;
            border-title-color: $accent;
            border-title-style: bold;
            background: transparent;
        }

        /* Контейнер для кнопок внутри правой панели */
        #data_buttons_container {
            width: 100%;
            height: auto;
            align: left top;
            padding: 1 0 0 1;
        }

        #data_buttons_container Button {
            width: 47;
            height: 3;
            background: $background;
            color: $secondary;
            border: transparent;  /* Прозрачная рамка вместо none */
            border: round transparent;  /* Прозрачная круглая рамка */
            content-align: left middle;
            padding: 0 2;
            margin: 0;
        }

        #data_buttons_container Button:focus {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;  /* При фокусе рамка становится видимой */
            content-align: left middle;
        }

        #data_buttons_container Button:hover {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
        }

        #data_buttons_container Button.spacer {
            height: 1;
            background: transparent;
            border: none;
        }

        #data_buttons_container Button > Static {
            width: 100%;
            content-align: left middle;
        }

        #data_buttons_container Button:focus {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
            content-align: left middle;
        }

        /* Стили для LookScreen */
        #look_list_panel {
            width: 70%;
            border: round $secondary;
            background: transparent;
            padding: 1 0 1 0;
            margin: 0;
            border-title-align: left;
        }

        #look_content_panel {
            width: 30%;
            border: round $secondary;
            background: transparent;
            padding: 0;
            margin: 0;
            border-title-align: left;
        }

        #look_list_panel:focus-within, #look_content_panel:focus-within {
            border: round $accent;
            border-title-color: $accent;
            border-title-style: bold;
            background: transparent;
        }

        #list_scroll {
            scrollbar-gutter: stable;
            scrollbar-size: 1 1;
            background: transparent;
            padding: 0;
            margin: 0;
        }

        #assessments_list {
            scrollbar-gutter: stable;
            background: transparent;
            color: $text;
            padding: 0;
            margin: 0;
        }

        #assessments_list ListItem {
            padding: 0 1 0 1;
            background: $background;
            color: $text;
        }

        #assessments_list ListItem > Static {
            color: $text;
            text-style: none;
            padding: 0 1 0 1;
        }

        #assessments_list > ListItem.--highlight > Static {
            color: $accent;
            text-style: bold;
        }

        /* Стили для полей ввода в LookScreen - такие же, как в Rate1Screen */
        #content_container {
            padding: 0 1;
        }

        #content_container Input {
            margin: 0 0 0 0;
            width: 100%;
            background: $background;
            border: transparent;
            color: $text;
        }

        #content_container Input:focus {
            border: round $accent;
            background: transparent;
        }

        #content_container Input.placeholder {
            color: $text-muted;
        }

        #content_container Label {
            margin-top: 0;
            margin-bottom: 0;
            text-style: bold;
            color: $secondary;
            padding: 0;
        }

        #content_container Button {
            width: 100%;
            margin: 1 0;
            background: $background;
            color: $secondary;
            border: round $secondary;
        }

        #content_container Button:focus {
            border: round $accent;
            color: $accent;
            background: transparent;
            text-style: bold;
        }

        #content_container Button.variant-primary {
            border: round $accent;
            color: $accent;
        }

        #content_container Button.variant-success {
            border: round green;
            color: green;
        }

        #content_container Button.variant-default {
            border: round $secondary;
            color: $secondary;
        }

        /* Стили для контейнера кнопок в режиме просмотра */
        #buttons_container {
            width: 100%;
            padding: 0 0 0 1;  /* Отступы справа и слева 1 знак */
        }

        #buttons_container Button {
            width: 100%;
            margin: 1 0;
            background: $background;
            color: $secondary;
            border: transparent;
        }

        #buttons_container Button:focus {
            border: round $accent;
            color: $accent;
            background: transparent;
            text-style: bold;
        }

        #buttons_container Button:hover {
            border: round $accent;
            color: $accent;
            background: transparent;
            text-style: bold;
        }

        #view_buttons {
            width: 100%;
            height: auto;
            margin-top: 1;
            align: center middle;
        }

        #view_buttons Button {
            width: 100%;
            margin: 1 0;
        }

        #view_text {
            padding: 1 2;
            color: $text;
            text-wrap: wrap;
            width: 100%;
        }

        #buttons_container Button.variant-primary {
            border: round $accent;
            color: $accent;
        }

        .menu-btn.hover {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
        }

        /* Класс для placeholder */
        .placeholder-text {
            color: $text-muted;
        }


        /* --- OutcomesMainScreen --- */
        #outcomes_main_hint_panel {
            width: 60%;
            border: round $accent;
            padding: 1 2;
            background: $background;
            border-title-style: bold;
        }

        #outcomes_main_buttons_panel {
            width: 40%;
            border: round $secondary;
            background: $background;
        }

        #outcomes_main_hint_text {
            height: auto;
            width: 100%;
            color: $text;
            padding: 1 2;
        }

        #outcomes_main_hint_panel:focus-within,
        #outcomes_main_buttons_panel:focus-within {
            border: round $accent;
            border-title-color: $accent;
            border-title-style: bold;
            background: transparent;
        }

        #outcomes_main_buttons_container {
            width: 100%;
            height: auto;
            align: left top;
            padding: 1 0 0 1;
        }

        #outcomes_main_buttons_container Button {
            width: 47;
            height: 3;
            background: $background;
            color: $secondary;
            border: round transparent;
            content-align: left middle;
            padding: 0 2;
            margin: 0;
        }

        #outcomes_main_buttons_container Button:focus {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
            content-align: left middle;
        }

        #outcomes_main_buttons_container Button:hover {
            background: transparent;
            color: $accent;
            text-style: bold;
            border: round $accent;
        }

        #outcomes_main_buttons_container Button.spacer {
            height: 1;
            background: transparent;
            border: none;
        }

        #outcomes_main_buttons_container Button > Static {
            width: 100%;
            content-align: left middle;
        }

        PeriodDialog {
            align: center middle;
        }

        #period_dialog_container {
            width: 66;
            height: auto;
            padding: 1 2;
            background: $surface;
            border: round $primary;
        }

        #period_dialog_title {
            text-align: center;
            text-style: bold;
            margin-bottom: 1;
        }

        #period_radios {
            margin: 1 0;
        }

        #period_radios RadioButton {
            margin-bottom: 1;
        }

        PeriodDialog Horizontal {
            height: auto;
            align-horizontal: center;
            margin-top: 1;
        }

        #btn_period_apply {
            margin-right: 1;
        }

        Footer {
            background: $panel;
            text-style: bold;
        }

    """

    BINDINGS = [
        Binding("escape", "request_quit", "Выход"),
        Binding("up", "focus_previous", show=False),
        Binding("down", "focus_next", show=False),
    ]

    def __init__(self):
        super().__init__()
        self.results = DataManager()
        self.drafts = DraftManager(self.results)
        self.notification_layer = None

    def compose(self):
        yield NotificationLayer()
        yield MenuScreen()

    def on_mount(self) -> None:
        # 1. Инициализируем менеджер данных (он сам создаст папки, если их нет)
        self.data_manager = DataManager()

        # 2. Регистрируем ваши кастомные темы
        for theme_obj in THEMES.values():
            self.register_theme(theme_obj)

        # 3. Загружаем тему, используя путь из data_manager
        # Передаем дефолтную тему и словарь THEMES для проверки валидности
        self.theme = self.data_manager.load_saved_theme(
            default_theme="themeLT",
            available_themes=THEMES
        )

        # Отображаем главный экран
        self.push_screen(MenuScreen())

    def custom_notify(self, message: str, severity: str = "info"):
        """Показать кастомное уведомление"""
        notification = CustomNotification(message, severity)

        # Добавляем уведомление непосредственно в главный экран
        screen = self.screen
        if screen:
            screen.mount(notification)

        # Автоматическое удаление через 5 секунд
        def remove_notification():
            try:
                notification.remove()
            except (NoMatches, Exception):
                pass

        self.set_timer(5, remove_notification)

    def action_request_quit(self) -> None:
        def check_quit(quit_app: bool) -> None:
            if quit_app:
                self.exit()
        self.push_screen(QuitScreen(), check_quit)

def main():
    # Проверка: запущен ли скрипт в интерактивном терминале
    # Если нет (например, клик по AppImage), пробуем открыть терминал
    if not sys.stdin.isatty():
        try:
            # Для Linux: пытаемся запустить через x-terminal-emulator
            # sys.argv[0] — это путь к самому AppImage
            subprocess.run(['x-terminal-emulator', '-e', sys.argv[0]])
            sys.exit(0)
        except FileNotFoundError:
            # Если x-terminal-emulator нет, можно попробовать xterm или gnome-terminal
            pass

    app = arsenal()
    app.run()

if __name__ == "__main__":
    main()
